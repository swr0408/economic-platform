"""
NBS プレスリリース Excel ダウンロード ユーティリティ

NBS data.stats.gov.cn API が WAF でブロックされているため、
www.stats.gov.cn/sj/zxfb/ のプレスリリースページに添付された
Excel ファイルからデータを取得する。

各プレスリリースの Excel 構造:
  - 工業生産: 1シート「表」、R4=鉱工業生産YoY、R61=集積回路
  - CPI: 1シート「CPI」、R4=CPI総合 MoM/YoY
  - PPI: 1シート「Sheet1」、R4=PPI出厂 MoM/YoY
  - 固定資産投資: 1シート「表」、R4=固定資産投資 YoY(累計)
  - GDP: 3シート、表1=GDP YoY、表2=GDP QoQ 時系列
  - 小売売上: 1シート「表」、R5=社会消費品小売総額 MoM/YoY
  - PMI: 2シート「制造業」「非制造业」、時系列
  - 不動産: 3シート、表1=不動産開発・販売
  - 貿易: NBS プレスリリースには含まれない（税関データ）
"""
import io
import logging
import re
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

BASE_URL = "https://www.stats.gov.cn/sj/zxfb/"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
}

# プレスリリース タイトル → カテゴリ キーワード
RELEASE_KEYWORDS = {
    "industrial_production": "规模以上工业增加值",
    "cpi": "居民消费价格",
    "ppi": "工业生产者出厂价格",
    "fixed_asset": "固定资产投资",
    "gdp": "国内生产总值",
    "retail": "社会消费品零售",
    "pmi": "采购经理指数",
    "real_estate": "房地产市场基本情况",
    "unemployment": "国民经济",  # 総合記事に失業率が含まれる
}

# Excel タイトルから対象月を抽出するパターン
# "2026年3月份规模以上工业..." → (2026, 3)
# "2026年1-3月份全国固定资产..." → (2026, 3)
# "2026年一季度国内生产总值..." → (2026, Q1)
RE_MONTH = re.compile(r"(\d{4})年(\d{1,2})月")
RE_RANGE = re.compile(r"(\d{4})年\d{1,2}[—-](\d{1,2})月")
RE_QUARTER = re.compile(r"(\d{4})年(一|二|三|四)季度")

QUARTER_MAP = {"一": 1, "二": 2, "三": 3, "四": 4}
QUARTER_END_MONTH = {1: 3, 2: 6, 3: 9, 4: 12}


RE_QUARTER_NOYEAR = re.compile(r"(一|二|三|四)季度")
# 年なし月次タイトル（「国民经济」総合レポート等）。
# "5月份国民经济..." → 5 / "1—4月份国民经济..." → 末尾の "4月份" にマッチ=4。
# 年付きパターンが全滅した場合のみ使う（年は発表日基準で推定）。
RE_MONTH_NOYEAR = re.compile(r"(\d{1,2})月份")


def _parse_release_period(title: str) -> Optional[Tuple[int, int]]:
    """プレスリリースタイトルから (year, month) を抽出"""
    # "2026年1-3月份..." → (2026, 3)
    m = RE_RANGE.search(title)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    # "2026年一季度..." → (2026, 3)
    m = RE_QUARTER.search(title)
    if m:
        year = int(m.group(1))
        q = QUARTER_MAP.get(m.group(2))
        if q:
            return (year, QUARTER_END_MONTH[q])
    # "2026年3月份..." → (2026, 3)
    m = RE_MONTH.search(title)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    # "一季度国民经济..." → (current_year, 3) - 年なし四半期タイトル
    m = RE_QUARTER_NOYEAR.search(title)
    if m:
        q = QUARTER_MAP.get(m.group(1))
        if q:
            return (datetime.now(JST).year, QUARTER_END_MONTH[q])
    # "5月份国民经济..." / "1—4月份国民经济..." - 年なし月次タイトル
    # 累計("1—4月份")は末月をデータ月とする。年は発表月基準で推定
    # （対象月>発表月なら前年データ＝12月分が翌年1月発表など）。
    m = RE_MONTH_NOYEAR.search(title)
    if m:
        month = int(m.group(1))
        if 1 <= month <= 12:
            now = datetime.now(JST)
            year = now.year
            if month > now.month:
                year -= 1
            return (year, month)
    return None


def find_recent_releases(
    category: str,
    max_pages: int = 2,
    limit: int = 6,
) -> List[Dict[str, Any]]:
    """プレスリリース一覧から指定カテゴリの直近 N 件を新→旧順で返す

    Args:
        category: RELEASE_KEYWORDS のキー (例: "cpi", "ppi")
        max_pages: 検索するページ数
        limit: 取得する記事数の上限

    Returns:
        [{"title": str, "url": str, "period": (year, month)}, ...] 新しい順
    """
    keyword = RELEASE_KEYWORDS.get(category)
    if not keyword:
        logger.warning(f"[NBS-PR] Unknown category: {category}")
        return []

    seen_urls = set()
    results: List[Dict[str, Any]] = []

    for page in range(max_pages):
        if len(results) >= limit:
            break
        try:
            if page == 0:
                page_url = BASE_URL
            else:
                page_url = f"{BASE_URL}index_{page}.html"

            resp = requests.get(page_url, headers=HEADERS, timeout=15)
            resp.encoding = "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")

            for a in soup.find_all("a", href=True):
                title = a.get_text(strip=True)
                if keyword not in title:
                    continue

                href = a["href"]
                # 相対URL→絶対URL
                if href.startswith("./"):
                    full_url = BASE_URL + href[2:]
                elif href.startswith("http"):
                    full_url = href
                else:
                    full_url = BASE_URL + href

                if full_url in seen_urls:
                    continue
                seen_urls.add(full_url)

                period = _parse_release_period(title)
                results.append({
                    "title": title,
                    "url": full_url,
                    "period": period,
                })
                if len(results) >= limit:
                    break
        except Exception as e:
            logger.warning(f"[NBS-PR] Failed to fetch release list page {page}: {e}")

    return results


def find_latest_release(category: str, max_pages: int = 2) -> Optional[Dict[str, Any]]:
    """直近1件を取得（後方互換）"""
    results = find_recent_releases(category, max_pages=max_pages, limit=1)
    return results[0] if results else None


def download_excel_from_release(release_url: str) -> Optional[bytes]:
    """プレスリリースページから Excel ファイルをダウンロード

    Returns:
        Excel ファイルのバイトデータ、または None
    """
    try:
        resp = requests.get(release_url, headers=HEADERS, timeout=15)
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")

        # Excel リンクを検索
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if ".xlsx" in href or ".xls" in href:
                # 相対URL→絶対URL
                if href.startswith("./"):
                    # release_url のディレクトリ部分を取得
                    base_dir = release_url.rsplit("/", 1)[0] + "/"
                    excel_url = base_dir + href[2:]
                elif href.startswith("http"):
                    excel_url = href
                else:
                    base_dir = release_url.rsplit("/", 1)[0] + "/"
                    excel_url = base_dir + href

                time.sleep(0.5)
                excel_resp = requests.get(excel_url, headers=HEADERS, timeout=30)
                if excel_resp.status_code == 200:
                    logger.info(f"[NBS-PR] Downloaded Excel: {excel_url} ({len(excel_resp.content)} bytes)")
                    return excel_resp.content
                else:
                    logger.warning(f"[NBS-PR] Excel download failed: {excel_url} -> {excel_resp.status_code}")

        logger.warning(f"[NBS-PR] No Excel link found in: {release_url}")
        return None

    except Exception as e:
        logger.error(f"[NBS-PR] Failed to download Excel from {release_url}: {e}")
        return None


def parse_xlsx(data: bytes, sheet_name: Optional[str] = None) -> List[List[Any]]:
    """xlsx バイトデータをパースし、2Dリストで返す"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def parse_xls(data: bytes, sheet_name: Optional[str] = None, sheet_index: int = 0) -> List[List[Any]]:
    """xls バイトデータをパースし、2Dリストで返す"""
    import xlrd
    wb = xlrd.open_workbook(file_contents=data)
    if sheet_name and sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(sheet_index)
    rows = []
    for r in range(ws.nrows):
        rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    return rows


def parse_excel(data: bytes, sheet_name: Optional[str] = None, sheet_index: int = 0) -> List[List[Any]]:
    """xlsx/xls を自動判定してパース"""
    # xlsx magic bytes: PK (50 4B)
    if data[:2] == b"PK":
        return parse_xlsx(data, sheet_name=sheet_name)
    else:
        return parse_xls(data, sheet_name=sheet_name, sheet_index=sheet_index)


def _safe_float(val: Any) -> Optional[float]:
    """値を float に変換（None、空文字、"…" は None）"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ("…", "—", "-", ""):
        return None
    # "0.7(百分点)" のようなケース
    m = re.match(r"([+-]?\d+\.?\d*)", s)
    if m:
        return float(m.group(1))
    return None


def _data_exists_in_db(indicator: str, date_str: str) -> bool:
    """nbs_monthly_data に (indicator, date) の行があるか確認"""
    try:
        from core.database import SessionLocal
        from sqlalchemy import text

        with SessionLocal() as session:
            row = session.execute(text("""
                SELECT 1 FROM nbs_monthly_data
                WHERE indicator = :ind AND date = :dt
                LIMIT 1
            """), {"ind": indicator, "dt": date_str}).fetchone()
        return row is not None
    except Exception as e:
        logger.debug(f"[NBS-PR] DB existence check failed for {indicator} {date_str}: {e}")
        return False


def fetch_and_upsert_from_press_release(
    category: str,
    extractor_fn,
    max_pages: int = 2,
    lookback_count: int = 6,
    primary_indicator: Optional[str] = None,
) -> Dict[str, int]:
    """プレスリリースからデータを取得し、DB に UPSERT

    取りこぼし耐性のため、直近 `lookback_count` 件の発表を新→旧順で巡回し、
    `primary_indicator` 指定時はDB に既に当該月のデータがあれば Excel DL をスキップする。
    `primary_indicator` 未指定時は従来通り最新1件のみ処理（後方互換）。

    Args:
        category: RELEASE_KEYWORDS のキー
        extractor_fn: Excel データ (bytes) と period (year, month) を受け取り、
                      {db_indicator: {date_str: value}} を返す関数
        max_pages: 検索ページ数
        lookback_count: primary_indicator 指定時に遡る発表数（既定: 6 = 半年分）
        primary_indicator: DB存在チェック用の主指標ID。未指定なら最新1件のみ取得

    Returns:
        {db_indicator: upsert_count, ...} 全リリースのUPSERT合算
    """
    from services.china.nbs_db_utils import upsert_nbs_data

    if primary_indicator is None:
        releases = find_recent_releases(category, max_pages=max_pages, limit=1)
    else:
        releases = find_recent_releases(category, max_pages=max_pages, limit=lookback_count)

    if not releases:
        logger.warning(f"[NBS-PR] No release found for category: {category}")
        return {}

    aggregated: Dict[str, int] = {}
    processed = 0

    for release in releases:
        title = release["title"]
        period = release["period"]

        # DB存在チェックで早期スキップ
        if primary_indicator and period:
            target_date = f"{period[0]:04d}-{period[1]:02d}-01"
            if _data_exists_in_db(primary_indicator, target_date):
                logger.debug(f"[NBS-PR] Skip (already in DB): {title} ({primary_indicator} @ {target_date})")
                continue

        logger.info(f"[NBS-PR] Fetching release: {title} (period={period})")

        excel_data = download_excel_from_release(release["url"])
        if not excel_data:
            continue

        try:
            extracted = extractor_fn(excel_data, period)
        except Exception as e:
            logger.error(f"[NBS-PR] Extraction failed for {category} ({title}): {e}")
            import traceback
            traceback.print_exc()
            continue

        for db_indicator, data in extracted.items():
            if not data:
                continue
            count = upsert_nbs_data(db_indicator, data, source="api")
            aggregated[db_indicator] = aggregated.get(db_indicator, 0) + count
            logger.info(f"[NBS-PR] {db_indicator}: upserted {count} records from {title}")

        processed += 1

    if processed > 1:
        logger.info(f"[NBS-PR] {category}: backfilled {processed} releases, {aggregated}")

    return aggregated


def scrape_unemployment_from_html(release_url: str) -> Dict[str, float]:
    """「国民経済」プレスリリースの HTML テキストから失業率を抽出

    失業率は Excel に含まれないため、HTML テキストからスクレイピング。

    Returns:
        {"total": float, "youth": float} (見つからない場合は空dict)
    """
    try:
        resp = requests.get(release_url, headers=HEADERS, timeout=15)
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")
        content = soup.find("div", class_="TRS_Editor") or soup
        text = content.get_text("", strip=False)

        result = {}

        # 全国城镇调查失业率: "3月份，全国城镇调查失业率为5.4%"
        m = re.search(r"全国城镇调查失业率为(\d+\.?\d*)%", text)
        if m:
            result["total"] = float(m.group(1))

        # 若年失業率: パターンが変わりやすいためスキップ可
        # "16-24岁劳动力调查失业率为XX.X%"
        m = re.search(r"(?:16[—-]24岁|青年).*?失业率为(\d+\.?\d*)%", text)
        if m:
            result["youth"] = float(m.group(1))

        return result

    except Exception as e:
        logger.error(f"[NBS-PR] Unemployment scraping failed: {e}")
        return {}

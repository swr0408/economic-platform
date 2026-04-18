"""
中国 外貨準備（Foreign Exchange Reserves）サービス

データ方式:
- 国家外汇管理局（SAFE）のExcel公表ファイルから「外汇储备 Foreign currency reserves」行を取得
- ハードコードURLで2015〜最新年をカバー + SAFEページから新年を自動検出
- 単位: 亿美元（100 million USD）
- 前年比（YoY%）は12ヶ月前の値から算出

Excel構造:
  XLSX (2016+): Row 4 = 月ヘッダー（YYYY.MM）、Row 8 = 外汇储备、偶数列=USD値
  XLS (pre-2016): Row 2 = 月ヘッダー、Row 4 = 外汇储备、1列ずつ

データソース: https://www.safe.gov.cn/en/ForexReserves/index.html

Redisキャッシュ: china:cn_forex_reserves:data (TTL: 24時間)
ファイルキャッシュ: backend/data/cache/china/policy/cn_forex_reserves_cache.json
"""
import io
import json
import os
import re
import logging
import calendar
from datetime import date, datetime
from typing import Optional, Dict, Any, List, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# -----------------------------------------------------------------
# キャッシュ設定
# -----------------------------------------------------------------
REDIS_KEY = "china:cn_forex_reserves:data"
REDIS_TTL = 24 * 3600  # 24 hours

_FILE_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "cache", "china", "policy"
)
FILE_CACHE_PATH = os.path.join(_FILE_CACHE_DIR, "cn_forex_reserves_cache.json")

# -----------------------------------------------------------------
# サイト定数
# -----------------------------------------------------------------
SAFE_INDEX_URL = "https://www.safe.gov.cn/en/ForexReserves/index.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.safe.gov.cn/",
}

# 外汇储备キーワード
FOREX_RESERVES_KEYWORDS = ["外汇储备", "Foreign currency reserves"]

# ハードコードされた過去年のExcelファイルURL
# 注: 各年の「Official Reserve Assets」Excelを1つずつ登録
# 年が変わったらSAFEページからの自動検出で補完される
HARDCODED_URLS = [
    # 2015.06-2015.12
    "https://www.safe.gov.cn/en/file/file/20180517/f6b4e78a7ce4478094aa1500049f60b7.xls?n=Official%20reserve%20assets%EF%BC%882015.06-2015.12%EF%BC%89",
    # 2016
    "https://www.safe.gov.cn/en/file/file/20180517/54673fdd29784f43add070617ebe17ec.xlsx?n=Official%20reserve%20assets%EF%BC%882016%EF%BC%89",
    # 2017
    "https://www.safe.gov.cn/en/file/file/20180517/fb97784244b64558beb1b21b7a4def4b.xlsx?n=Official%20reserve%20assets%EF%BC%882017%EF%BC%89",
    # 2018
    "https://www.safe.gov.cn/en/file/file/20190123/e2b1acc52c734c0da3d26c693d74c885.xls?n=xlsx",
    # 2019
    "https://www.safe.gov.cn/en/file/file/20210207/ee2bb9252b934a7bad758e79c52b5b5f.xls?n=xlsx",
    # 2020 (listed on https://www.safe.gov.cn/en/2018/0517/1433.html)
    # Note: 2019 and 2020 may share the same file or separate file.
    # The SAFE page groups multiple years per file; later files override.
    # 2021
    "https://www.safe.gov.cn/en/file/file/20220128/b62ccc0d28dc401ba16712f592ecd2a0.xlsx?n=xlsx",
    # 2022
    "https://www.safe.gov.cn/en/file/file/20230131/99ae51feb9114aad94ef12715ab95cda.xlsx?n=xlsx",
    # 2023
    "https://www.safe.gov.cn/en/file/file/20240124/612f20d0024b41ec9d8ece9250b922cf.xlsx?n=xlsx",
    # 2024
    "https://www.safe.gov.cn/en/file/file/20250107/cccae56ba9204e43821b62c8441bbea8.xlsx?n=xlsx",
    # 2025 (full year published Feb 2026)
    "https://www.safe.gov.cn/en/file/file/20260207/532af9a5edf040fb95117264b66a748c.xlsx?n=xlsx",
    # 2026 (interim, published Feb 2026 - Jan data)
    "https://www.safe.gov.cn/en/file/file/20260226/c737c4ea7b0b4519abc8beefe3a2894b.xlsx?n=xlsx",
    # 2026 (updated Apr 2026 - Jan-Mar data)
    "https://www.safe.gov.cn/en/file/file/20260407/ad48d41728cc40e9aa3f9080dee8113c.xlsx",
]


# =============================================================================
# ユーティリティ
# =============================================================================

def _parse_date_str(date_str: Any) -> Optional[date]:
    """
    "2026.01" / "2026.1" / 2026.01 (float) → date(2026, 1, 31)
    月末日を使用。
    """
    if date_str is None:
        return None
    s = str(date_str).strip()
    m = re.match(r'^(\d{4})\.(\d{1,2})$', s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            _, last_day = calendar.monthrange(year, month)
            return date(year, month, last_day)
    return None


def _clean_value(val: Any) -> Optional[float]:
    """
    Excelセル値を float に変換。
    文字列の \xa0（ノーブレークスペース）や先頭のアポストロフィを除去。
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('\xa0', '').replace('\u3000', '').lstrip("'")
    # 数字・小数点・マイナスのみ残す
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# =============================================================================
# XLSX パーサー (2016+): 偶数列がUSD値、奇数列がSDR値
# =============================================================================

def _parse_xlsx(content: bytes) -> List[Tuple[date, float]]:
    """XLSXファイルから (date, value_usd) リストを返す"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # 日付ヘッダー行と外汇储备行を動的検索
        header_row = None
        forex_row = None

        for row_idx in range(1, min(ws.max_row + 1, 15)):
            # 日付ヘッダー行: いずれかの列に "YYYY.MM" パターン
            if header_row is None:
                for col_idx in range(2, min(ws.max_column + 1, 30)):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break

            # 外汇储备行
            if forex_row is None:
                cell_a = ws.cell(row=row_idx, column=1).value
                if cell_a:
                    cell_a_str = str(cell_a).strip()
                    for kw in FOREX_RESERVES_KEYWORDS:
                        if kw in cell_a_str:
                            forex_row = row_idx
                            break

        if header_row is None or forex_row is None:
            logger.warning(f"[FOREX] XLSX: header_row={header_row}, forex_row={forex_row} - parse failed")
            return []

        results = []
        for col_idx in range(2, ws.max_column + 1):
            date_val = ws.cell(row=header_row, column=col_idx).value
            d = _parse_date_str(date_val)
            if not d:
                continue

            forex_val = ws.cell(row=forex_row, column=col_idx).value
            val = _clean_value(forex_val)
            if val is not None:
                results.append((d, val))

        return results

    except Exception as e:
        logger.warning(f"[FOREX] XLSX parse error: {e}")
        return []


# =============================================================================
# XLS パーサー (pre-2016): 1列ずつ
# =============================================================================

def _parse_xls(content: bytes) -> List[Tuple[date, float]]:
    """XLSファイルから (date, value_usd) リストを返す"""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)

        header_row = None
        forex_row = None

        for row_idx in range(0, min(ws.nrows, 15)):
            # 日付ヘッダー行
            if header_row is None:
                for col_idx in range(1, min(ws.ncols, 20)):
                    v = ws.cell_value(row_idx, col_idx)
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break

            # 外汇储备行
            if forex_row is None:
                cell_a = ws.cell_value(row_idx, 0)
                if cell_a:
                    cell_a_str = str(cell_a).strip()
                    for kw in FOREX_RESERVES_KEYWORDS:
                        if kw in cell_a_str:
                            forex_row = row_idx
                            break

        if header_row is None or forex_row is None:
            logger.warning(f"[FOREX] XLS: header_row={header_row}, forex_row={forex_row} - parse failed")
            return []

        results = []
        for col_idx in range(1, ws.ncols):
            date_val = ws.cell_value(header_row, col_idx)
            d = _parse_date_str(date_val)
            if not d:
                continue

            forex_val = ws.cell_value(forex_row, col_idx)
            val = _clean_value(forex_val)
            if val is not None:
                results.append((d, val))

        return results

    except Exception as e:
        logger.warning(f"[FOREX] XLS parse error: {e}")
        return []


def _download_and_parse(url: str) -> List[Tuple[date, float]]:
    """URLをダウンロードしてXLS/XLSXをパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        content = resp.content
        # URL末尾の拡張子で判定（クエリパラメータ除去）
        clean_url = url.split('?')[0].lower()
        if clean_url.endswith('.xlsx'):
            return _parse_xlsx(content)
        else:
            return _parse_xls(content)
    except Exception as e:
        logger.warning(f"[FOREX] Download failed {url}: {e}")
        return []


# =============================================================================
# ページスクレイプ: SAFE ForexReserves ページから新年のXLSXを検出
# =============================================================================

def _discover_new_urls() -> List[str]:
    """
    SAFEの ForexReserves/index.html をスクレイプし、
    .xls/.xlsx リンクを取得する。
    ハードコード済みURLと重複しない新しいURLのみ返す。

    SAFEサイト構造:
      インデックスページ → "Official Reserve Assets (20XX)" リンク → 詳細ページ → Excelダウンロードリンク
    インデックスページにはExcelリンクがないため、詳細ページまで辿る必要がある。
    """
    try:
        resp = requests.get(SAFE_INDEX_URL, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            logger.warning(f"[FOREX] SAFE page returned {resp.status_code}")
            return []
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")

        # ハードコードURLからファイルパス部分を抽出（比較用）
        hardcoded_paths = set()
        for url in HARDCODED_URLS:
            path = url.split('?')[0]
            m = re.search(r'/file/file/(.+)', path)
            if m:
                hardcoded_paths.add(m.group(1))

        new_urls = []

        # Step 1: インデックスページの直接Excelリンクを検索
        for a in soup.find_all("a", href=True):
            href = a["href"]
            clean = href.split('?')[0].lower()
            if clean.endswith(('.xls', '.xlsx')):
                full_url = href if href.startswith("http") else f"https://www.safe.gov.cn{href}"
                path = full_url.split('?')[0]
                m = re.search(r'/file/file/(.+)', path)
                if m and m.group(1) not in hardcoded_paths:
                    if full_url not in new_urls:
                        new_urls.append(full_url)

        # Step 2: "Official Reserve Assets" 詳細ページリンクを辿り、Excelを探す
        detail_links = []
        for a in soup.find_all("a", href=True):
            text = a.get_text(strip=True).lower()
            if "official reserve" in text or "reserve assets" in text:
                href = a["href"]
                if href.startswith("/en/") and href.endswith(".html"):
                    full_url = f"https://www.safe.gov.cn{href}"
                    if full_url not in detail_links:
                        detail_links.append(full_url)

        for detail_url in detail_links:
            try:
                detail_resp = requests.get(detail_url, headers=HEADERS, timeout=15)
                if detail_resp.status_code != 200:
                    continue
                detail_resp.encoding = "utf-8"
                detail_soup = BeautifulSoup(detail_resp.text, "html.parser")

                for a in detail_soup.find_all("a", href=True):
                    href = a["href"]
                    clean = href.split('?')[0].lower()
                    if clean.endswith(('.xls', '.xlsx')):
                        full_url = href if href.startswith("http") else f"https://www.safe.gov.cn{href}"
                        path = full_url.split('?')[0]
                        m = re.search(r'/file/file/(.+)', path)
                        if m and m.group(1) not in hardcoded_paths:
                            if full_url not in new_urls:
                                new_urls.append(full_url)
            except Exception as e:
                logger.warning(f"[FOREX] Failed to scrape detail page {detail_url}: {e}")

        if new_urls:
            logger.info(f"[FOREX] Discovered {len(new_urls)} new URLs from SAFE page")

        return new_urls

    except Exception as e:
        logger.warning(f"[FOREX] SAFE page scrape failed: {e}")
        return []


# =============================================================================
# 全データ取得・マージ
# =============================================================================

def fetch_all_data() -> Dict[str, float]:
    """
    ハードコードURL + 動的検出URLからデータ取得
    Returns: {date_str: value_usd} (亿美元)
    """
    all_data: Dict[str, float] = {}

    # Step 1: ハードコードURLから取得
    for url in HARDCODED_URLS:
        records = _download_and_parse(url)
        for d, val in records:
            date_str = d.strftime("%Y-%m-%d")
            all_data[date_str] = val
        if records:
            logger.info(f"[FOREX] {url.split('/')[-1][:30]}: {len(records)} records")

    # Step 2: SAFEページから新しいURLを検出して取得
    new_urls = _discover_new_urls()
    for url in new_urls:
        records = _download_and_parse(url)
        for d, val in records:
            date_str = d.strftime("%Y-%m-%d")
            all_data[date_str] = val
        if records:
            logger.info(f"[FOREX] New: {url.split('/')[-1][:30]}: {len(records)} records")

    logger.info(f"[FOREX] Total: {len(all_data)} months")
    return all_data


def _compute_yoy(data_map: Dict[str, float]) -> Dict[str, Optional[float]]:
    """
    12ヶ月前の値から前年比（YoY%）を算出
    Returns: {date_str: yoy_percent}
    """
    yoy_map: Dict[str, Optional[float]] = {}
    for date_str, current_val in data_map.items():
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        # 12ヶ月前の日付を生成
        prev_year = dt.year - 1
        prev_month = dt.month
        _, prev_last_day = calendar.monthrange(prev_year, prev_month)
        prev_date_str = f"{prev_year}-{prev_month:02d}-{prev_last_day:02d}"

        if prev_date_str in data_map:
            prev_val = data_map[prev_date_str]
            if prev_val and prev_val > 0:
                yoy = ((current_val - prev_val) / prev_val) * 100
                yoy_map[date_str] = round(yoy, 2)
            else:
                yoy_map[date_str] = None
        else:
            yoy_map[date_str] = None

    return yoy_map


# =============================================================================
# サービスクラス
# =============================================================================

class CnForeignExchangeReservesService:
    """中国 外貨準備（Foreign Exchange Reserves）サービス"""

    def __init__(self):
        self._redis = None

    def _get_redis(self):
        if self._redis is None:
            try:
                import redis
                self._redis = redis.Redis(host="localhost", port=6379, db=0, socket_timeout=2)
                self._redis.ping()
            except Exception:
                self._redis = None
        return self._redis

    def _from_redis(self) -> Optional[Dict[str, Any]]:
        r = self._get_redis()
        if not r:
            return None
        try:
            raw = r.get(REDIS_KEY)
            if raw:
                return json.loads(raw)
        except Exception:
            pass
        return None

    def _to_redis(self, payload: Dict[str, Any]) -> None:
        r = self._get_redis()
        if not r:
            return
        try:
            r.setex(REDIS_KEY, REDIS_TTL, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _from_file(self) -> Optional[Dict[str, Any]]:
        if not os.path.exists(FILE_CACHE_PATH):
            return None
        try:
            with open(FILE_CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _to_file(self, payload: Dict[str, Any]) -> None:
        os.makedirs(_FILE_CACHE_DIR, exist_ok=True)
        try:
            with open(FILE_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"[FOREX] File cache write error: {e}")

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """FMP（CN）から次回発表日を取得"""
        try:
            from services.usa.fmp_next_release_utils import get_next_release_from_fmp
            result = get_next_release_from_fmp("cn_foreign_exchange_reserves", country="CN")
            if result:
                return result
        except Exception as e:
            logger.debug(f"[FOREX] FMP next_release failed: {e}")
        return None

    def _build_payload(self, data_map: Dict[str, float]) -> Dict[str, Any]:
        yoy_map = _compute_yoy(data_map)

        records = []
        for date_str in sorted(data_map.keys()):
            records.append({
                "date": date_str,
                "value": data_map[date_str],
                "yoy": yoy_map.get(date_str),
            })

        latest = records[-1] if records else None
        next_release = self._get_next_release()

        return {
            "data": records,
            "latest": latest,
            "metadata": {
                "unit": "亿美元",
                "indicator": "Foreign Exchange Reserves",
                "source": "SAFE (State Administration of Foreign Exchange)",
                "total_records": len(records),
                "last_fetched": datetime.now(JST).isoformat(),
            },
            "next_release": next_release,
        }

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """外貨準備データを取得（キャッシュ優先）"""
        if not force_refresh:
            cached = self._from_redis()
            if cached:
                return cached
            cached = self._from_file()
            if cached:
                self._to_redis(cached)
                return cached

        data_map = fetch_all_data()
        if not data_map:
            cached = self._from_file()
            if cached:
                return cached
            return {"data": [], "latest": None, "metadata": {}, "next_release": None}

        payload = self._build_payload(data_map)
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def invalidate_cache(self) -> Dict[str, Any]:
        r = self._get_redis()
        if r:
            try:
                r.delete(REDIS_KEY)
            except Exception:
                pass
        if os.path.exists(FILE_CACHE_PATH):
            os.remove(FILE_CACHE_PATH)
        return {"success": True, "message": "Foreign Exchange Reserves cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        r = self._get_redis()
        redis_exists = False
        redis_ttl = None
        if r:
            try:
                redis_exists = bool(r.exists(REDIS_KEY))
                redis_ttl = r.ttl(REDIS_KEY)
            except Exception:
                pass
        file_exists = os.path.exists(FILE_CACHE_PATH)
        file_mtime = None
        if file_exists:
            file_mtime = datetime.fromtimestamp(
                os.path.getmtime(FILE_CACHE_PATH), tz=JST
            ).isoformat()
        return {
            "redis": {"exists": redis_exists, "ttl_seconds": redis_ttl},
            "file": {"exists": file_exists, "last_modified": file_mtime},
        }


# シングルトン
cn_foreign_exchange_reserves_service = CnForeignExchangeReservesService()

"""
中国人民銀行（PBOC）社会融資総量（Aggregate Financing to the Real Economy）サービス

データ方式:
- 親ページ（統計数据）から「社会融資規模」XLS/XLSXリンクを動的に発見
- Flow（月次フロー）: 100億元単位、月次棒グラフ用
- Stock（残高）: 兆元単位、YoY(%) 付き、折れ線グラフ用
- ExcelのないFMP欠損月はFMP APIの Total Social Financing actualで補完

Flow XLS/XLSX構造:
  Row 1: "Aggregate Financing to the Real Economy (Flow)"
  Row 3: "Unit: 100 million Yuan"
  Row 9: 列ヘッダー（Month, AFRE(flow), ...）
  Row 11+: YYYY.MM, 数値, ...

Stock XLS/XLSX構造（横形式）:
  Row 1: "Aggregate Financing to the Real Economy (Stock)"
  Row 3: "Unit: Trillion Yuan"
  Row 4: 月ヘッダー（2021.1, 2021.2, ...） ← 偶数列がStock, 奇数列がGrowth Rate
  Row 5-6: サブヘッダー ("Stock", "Growth Rate(%)")
  Row 8+: 各行に指標（AFRE stock = Row 8）

データソース: https://camlmac.pbc.gov.cn/diaochatongjisi/116219/116319/index.html

Redisキャッシュ: china:cn_aggregate_financing:data (TTL: 24時間)
ファイルキャッシュ: backend/data/cache/china/policy/cn_aggregate_financing_cache.json
"""
import io
import json
import os
import re
import logging
import calendar
from datetime import date, datetime
from typing import Optional, Dict, Any, List, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# -----------------------------------------------------------------
# キャッシュ設定
# -----------------------------------------------------------------
REDIS_KEY = "china:cn_aggregate_financing:data"
REDIS_TTL = 24 * 3600  # 24 hours

_FILE_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "cache", "china", "policy"
)
FILE_CACHE_PATH = os.path.join(_FILE_CACHE_DIR, "cn_aggregate_financing_cache.json")

# -----------------------------------------------------------------
# サイト定数
# -----------------------------------------------------------------
PBOC_HOST = "https://camlmac.pbc.gov.cn"
BASE_URL = f"{PBOC_HOST}/diaochatongjisi"
PARENT_INDEX_URL = f"{BASE_URL}/116219/116319/index.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://camlmac.pbc.gov.cn/",
}

# Flow/Stock 識別キーワード
FLOW_KEYWORDS = ["社会融资规模增量", "Aggregate Financing to the Real Economy (Flow)", "AFRE(flow)"]
STOCK_KEYWORDS = ["社会融资规模存量", "Aggregate Financing to the Real Economy (Stock)", "AFRE(stock)"]

MIN_YEAR = 2013

# ハードコードされた過去年のExcelファイルURL
# PBOCの過去データは動的発見できないため固定URLを使用
HARDCODED_FLOW_URLS = [
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2018/04/2018041810462448165.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2018/04/2018041810585985661.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2018/04/2018041811054763138.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2019/08/2019081915081468438.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2020/03/2020030610323558317.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2021/05/2021051317183722377.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2022/03/2022031709275051766.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2024/01/2024011510325120042.xlsx",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2024/02/2024021917193120091.xlsx",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2025/02/2025021419540681589.xlsx",
    "https://www.pbc.gov.cn/diaochatongjisi/attachDir/2025/11/2025111913535789644.xlsx",
]

HARDCODED_STOCK_URLS = [
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2018/04/2018041810585987735.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2018/04/2018041811054734868.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2019/05/2019050709591895742.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2019/05/2019053116170320732.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2020/03/2020030610323497759.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2021/05/2021051317183748976.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2022/03/2022031709274991948.xls",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2024/01/2024011510325140978.xlsx",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2024/02/2024021917193153911.xlsx",
    "https://www.pbc.gov.cn/diaochatongjisi/fileDir/resource/cms/2025/02/2025021419540645809.xlsx",
    # 2025-11 interim file (Jan-Oct 2025 only; Nov/Dec cells were empty at publication time)
    "https://www.pbc.gov.cn/diaochatongjisi/attachDir/2025/11/2025111913535786100.xlsx",
    # 2025 full-year file published Feb 2026 (all 12 months including Nov/Dec 2025)
    # NOTE: 毎年2月頃にPBOCが前年度まとめXLSXを公開する。翌年にはこのURLに相当するファイルを追加すること
    "https://camlmac.pbc.gov.cn/diaochatongjisi/attachDir/2026/02/2026021616264071771.xlsx",
]


# =============================================================================
# ユーティリティ
# =============================================================================

def _abs_url(href: str) -> str:
    if href.startswith("http"):
        return href
    elif href.startswith("/"):
        return f"{PBOC_HOST}{href}"
    else:
        return f"{BASE_URL}/{href}"


def _parse_date_str(date_str: Any) -> Optional[date]:
    """
    "2025.01" / "2025.1" / 2025.01 (float) → date(2025, 1, 31)
    月末日を使用。
    """
    if date_str is None:
        return None
    s = str(date_str).strip()
    m = re.match(r'^(\d{4})\.(\d{1,2})$', s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        try:
            _, last_day = calendar.monthrange(year, month)
            return date(year, month, last_day)
        except ValueError:
            return None
    return None


def _fetch_soup(url: str, timeout: int = 15) -> Optional[BeautifulSoup]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code != 200:
            return None
        resp.encoding = "utf-8"
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        logger.debug(f"[AFRE] Fetch failed {url}: {e}")
        return None


# =============================================================================
# スクレイピング
# =============================================================================

def _discover_year_urls(min_year: int = MIN_YEAR) -> Dict[int, str]:
    """
    親ページから各年の「社融規模（shrzgm）」ページURLを発見。
    社会融資総量は货币统计概览（hbtjgl）ではなく shrzgm サブページにある。

    URL パターン（2種類）:
      新形式: /diaochatongjisi/116219/116319/{YYYY}ntjsj/shrzgm/index.html  (2026〜)
      旧形式: /diaochatongjisi/116219/116319/{numeric_id}/{sub_numeric_id}/index.html (〜2025)

    旧形式の場合: 年ページのリンクから shrzgm サブページを動的に探索する。
    """
    soup = _fetch_soup(PARENT_INDEX_URL)
    if soup is None:
        logger.warning("[AFRE] Failed to fetch parent index page")
        return {}

    year_map: Dict[int, str] = {}

    # --- パターン1: {YYYY}ntjsj/shrzgm/ (新形式、直接リンク) ---
    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = re.search(r'/(\d{4})ntjsj/shrzgm/', href)
        if m:
            year = int(m.group(1))
            if year >= min_year and year not in year_map:
                year_map[year] = _abs_url(href)

    # --- パターン2: 年ページ ({YYYY}ntjsj/ または 数値ID) を辿って shrzgm を探索 ---
    # 年ページのリンクを収集（新形式・旧形式の両方）
    year_index_urls: Dict[int, str] = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        # 新形式: /116319/2025ntjsj/index.html
        m_new = re.search(r'/(\d{4})ntjsj/index\.html', href)
        if m_new:
            year = int(m_new.group(1))
            if year >= min_year and year not in year_index_urls:
                year_index_urls[year] = _abs_url(href)
            continue
        # 旧形式: /116319/{numeric_id}/index.html (テキストから年を推定)
        m_old = re.search(r'/116319/(\d{5,7})/index\.html', href)
        if m_old:
            text = a.get_text(strip=True)
            # リンクテキストから年を抽出 ("2025年统计数据" → 2025)
            y_match = re.search(r'(\d{4})', text)
            if y_match:
                year = int(y_match.group(1))
                if year >= min_year and year not in year_index_urls:
                    year_index_urls[year] = _abs_url(href)

    for year, year_index_url in sorted(year_index_urls.items()):
        if year in year_map:
            continue  # すでに直接発見済み
        # 年ページを取得してサブページリンクを探す
        year_soup = _fetch_soup(year_index_url)
        if year_soup is None:
            continue
        # shrzgm サブページを探す（"shrzgm" またはテキストに "社会融资" を含む）
        for a in year_soup.find_all("a", href=True):
            sub_href = a["href"]
            sub_text = a.get_text(strip=True)
            is_shrzgm = (
                "shrzgm" in sub_href.lower()
                or "社会融资" in sub_text
                or "社会融資" in sub_text
                or "Aggregate Financing" in sub_text
            )
            if is_shrzgm and (sub_href.endswith("index.html") or sub_href.endswith("/")):
                shrzgm_url = _abs_url(sub_href)
                test_soup = _fetch_soup(shrzgm_url)
                if test_soup and test_soup.find("div", class_="titp20"):
                    year_map[year] = shrzgm_url
                    logger.info(f"[AFRE] Found shrzgm for {year} via year index: {shrzgm_url}")
                    break

    logger.info(f"[AFRE] Discovered {len(year_map)} years with shrzgm: {sorted(year_map.keys())}")
    return year_map


def _find_afre_links(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
    """
    shrzgm ページから「Aggregate Financing (Flow)」と「(Stock)」の
    XLS/XLSXリンクを取得する。
    (flow_url, stock_url) のタプルを返す。
    """
    flow_url = None
    stock_url = None

    for div in soup.find_all("div", class_="titp20"):
        text = div.get_text()
        # "Province" を含む場合はスキップ（省別データ）
        if "Province" in text or "province" in text:
            continue
        is_flow = "Flow" in text or "flow" in text.lower()
        is_stock = "Stock" in text or "stock" in text.lower()

        if not is_flow and not is_stock:
            continue

        tr = div.find_parent("tr")
        if tr is None:
            continue

        for a_tag in tr.find_all("a", href=True):
            href = a_tag["href"]
            if href.lower().endswith(".xlsx") or href.lower().endswith(".xls"):
                url = _abs_url(href)
                if is_flow and flow_url is None:
                    flow_url = url
                elif is_stock and stock_url is None:
                    stock_url = url
                break

    return flow_url, stock_url


# =============================================================================
# Flow パーサー（月次フロー、100億元単位）
# =============================================================================

def _parse_flow_xlsx(content: bytes) -> Dict[str, float]:
    """XLSX Flow ファイルをパース → {date_str: afre_flow} (100億元単位)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # ヘッダー行（"Month" が col0 にある行）と AFRE(flow) 列を動的検索
        header_row = None
        afre_col = None

        for row_idx in range(1, min(ws.max_row + 1, 15)):
            cell0 = ws.cell(row=row_idx, column=1).value
            if cell0 and "Month" in str(cell0):
                header_row = row_idx
                # "AFRE(flow)" 列を探す（col 2 = AFRE(flow) のはず）
                for col_idx in range(1, min(ws.max_column + 1, 15)):
                    cv = ws.cell(row=row_idx, column=col_idx).value
                    if cv and ("AFRE" in str(cv) or "flow" in str(cv).lower()):
                        afre_col = col_idx
                        break
                if afre_col is None:
                    afre_col = 2  # デフォルト: col B
                break

        if header_row is None:
            logger.warning("[AFRE] XLSX Flow: header row not found")
            return {}

        result = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            afre_val = ws.cell(row=row_idx, column=afre_col).value
            d = _parse_date_str(date_val)
            if d and afre_val is not None:
                try:
                    result[d.strftime("%Y-%m-%d")] = float(afre_val)
                except (ValueError, TypeError):
                    pass

        return result

    except Exception as e:
        logger.warning(f"[AFRE] XLSX Flow parse error: {e}")
        return {}


def _clean_numeric(val: Any) -> Optional[float]:
    """
    XLS セルの数値を float に変換。
    文字列に含まれる数字以外の文字（単位など）を除去して変換する。
    例: '30793亿' → 30793.0, '100.0' → 100.0
    """
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    # 文字列の場合: 数字・小数点・マイナス以外を除去
    s = re.sub(r"[^\d.\-]", "", str(val).strip())
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_flow_xls(content: bytes) -> Dict[str, float]:
    """
    XLS Flow ファイルをパース → {date_str: afre_flow}

    複数セクションを持つXLSに対応:
    - 最初のデータセクション（YYYY.MM形式の行が連続する最初のブロック）のみ読む
    - 旧XLSでは複数テーブルが縦に並んでいるが、最初のセクションが最新調整済みデータ

    対応構造:
    A. col 0 = "Month" ヘッダー, col 1 = AFRE(flow) → 通常形式
    B. col 0 = "YYYY.MM" データ行, col 1 = AFRE(flow) 値（旧XLS複合ヘッダー形式）
    """
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)

        # AFRE(flow) 列を探す（最初の15行のいずれかに "AFRE" + "flow" が含まれる）
        afre_col = None
        for row_idx in range(min(ws.nrows, 15)):
            for col_idx in range(min(ws.ncols, 10)):
                cv = ws.cell_value(row_idx, col_idx)
                if cv and "AFRE" in str(cv) and "flow" in str(cv).lower():
                    # 説明文行（長い文章）は除外
                    if len(str(cv)) < 50:
                        afre_col = col_idx
                        break
            if afre_col is not None:
                break

        if afre_col is None:
            # デフォルト: col 1
            afre_col = 1

        # データ行を読む（YYYY.MM形式の最初の連続ブロックのみ）
        result = {}
        in_data_block = False
        consecutive_empty = 0

        for row_idx in range(0, ws.nrows):
            date_val = ws.cell_value(row_idx, 0)
            d = _parse_date_str(date_val)
            if d:
                in_data_block = True
                consecutive_empty = 0
                afre_val = ws.cell_value(row_idx, afre_col)
                num = _clean_numeric(afre_val)
                # 妥当な範囲: 1,000〜200,000百億元（大きすぎる/小さすぎる値は除外）
                if num is not None and 100 < num < 200000:
                    date_str = d.strftime("%Y-%m-%d")
                    if date_str not in result:  # 最初に見つかった値を優先
                        result[date_str] = num
            elif in_data_block:
                consecutive_empty += 1
                # データブロック終了後に空行が3行以上続いたら、次のブロックへ
                # ただし最初のブロックのみ読む（2ブロック目以降は skip）
                if consecutive_empty >= 3 and result:
                    # 最初のブロックを読み終えた → 終了
                    break

        return result

    except Exception as e:
        logger.warning(f"[AFRE] XLS Flow parse error: {e}")
        return {}


# =============================================================================
# Stock パーサー（月末残高、兆元単位 + YoY%）
# =============================================================================

def _parse_stock_xlsx(content: bytes) -> Dict[str, Dict[str, Optional[float]]]:
    """
    XLSX Stock ファイルをパース（横形式）
    → {date_str: {"stock": float, "yoy": float}} (stock単位: 兆元)
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # 月ヘッダー行を探す（"2025.1" のような浮動小数点または文字列が並ぶ行）
        date_row = None
        afre_row = None  # AFRE(stock) データ行

        for row_idx in range(1, min(ws.max_row + 1, 15)):
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    s = str(v).strip()
                    if re.match(r'^\d{4}\.\d{1,2}$', s):
                        date_row = row_idx
                        break
            if date_row is not None:
                break

        if date_row is None:
            logger.warning("[AFRE] XLSX Stock: date row not found")
            return {}

        # AFRE(stock) 行を探す（date_row 以降で最初の大きな数値がある行）
        for row_idx in range(date_row + 1, min(ws.max_row + 1, 15)):
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(v, (int, float)) and float(v) > 50:  # 兆元なので50以上
                    afre_row = row_idx
                    break
            if afre_row is not None:
                break

        if afre_row is None:
            logger.warning("[AFRE] XLSX Stock: afre data row not found")
            return {}

        # 月ヘッダーから列インデックスをマッピング
        # 構造: col1=月, col2=Growth Rate, col3=月, col4=Growth Rate, ...
        result: Dict[str, Dict[str, Optional[float]]] = {}

        for col_idx in range(1, ws.max_column + 1):
            date_val = ws.cell(row=date_row, column=col_idx).value
            d = _parse_date_str(date_val)
            if not d:
                continue

            stock_val = ws.cell(row=afre_row, column=col_idx).value
            # Growth Rate は同じ月の次の列
            yoy_val = ws.cell(row=afre_row, column=col_idx + 1).value

            date_str = d.strftime("%Y-%m-%d")
            try:
                stock = float(stock_val) if stock_val is not None else None
            except (ValueError, TypeError):
                stock = None
            try:
                yoy = float(yoy_val) if yoy_val is not None else None
            except (ValueError, TypeError):
                yoy = None

            if stock is not None:
                result[date_str] = {"stock": stock, "yoy": yoy}

        return result

    except Exception as e:
        logger.warning(f"[AFRE] XLSX Stock parse error: {e}")
        return {}


def _parse_stock_xls(content: bytes) -> Dict[str, Dict[str, Optional[float]]]:
    """XLS Stock ファイルをパース（横形式）"""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)

        date_row = None
        afre_row = None

        for row_idx in range(min(ws.nrows, 12)):
            for col_idx in range(min(ws.ncols, 30)):
                v = ws.cell_value(row_idx, col_idx)
                if v:
                    s = str(v).strip()
                    if re.match(r'^\d{4}\.\d{1,2}$', s):
                        date_row = row_idx
                        break
            if date_row is not None:
                break

        if date_row is None:
            logger.warning("[AFRE] XLS Stock: date row not found")
            return {}

        # AFRE(stock) 行を探す（float型または文字列形式の数値に対応）
        for row_idx in range(date_row + 1, min(ws.nrows, 15)):
            for col_idx in range(min(ws.ncols, 30)):
                v = ws.cell_value(row_idx, col_idx)
                try:
                    num = float(str(v).strip().replace('\xa0', '').replace(',', ''))
                    if num > 50:
                        afre_row = row_idx
                        break
                except (ValueError, TypeError):
                    pass
            if afre_row is not None:
                break

        if afre_row is None:
            logger.warning("[AFRE] XLS Stock: afre data row not found")
            return {}

        result: Dict[str, Dict[str, Optional[float]]] = {}

        for col_idx in range(ws.ncols):
            date_val = ws.cell_value(date_row, col_idx)
            d = _parse_date_str(date_val)
            if not d:
                continue

            stock_val = ws.cell_value(afre_row, col_idx)
            yoy_val = ws.cell_value(afre_row, col_idx + 1) if col_idx + 1 < ws.ncols else None

            date_str = d.strftime("%Y-%m-%d")
            try:
                stock = float(str(stock_val).strip().replace('\xa0', '').replace(',', '')) if stock_val else None
                if stock is not None and stock <= 0:
                    stock = None
            except (ValueError, TypeError):
                stock = None
            try:
                yoy = float(str(yoy_val).strip().replace('\xa0', '').replace(',', '')) if yoy_val else None
            except (ValueError, TypeError):
                yoy = None

            if stock is not None:
                result[date_str] = {"stock": stock, "yoy": yoy}

        return result

    except Exception as e:
        logger.warning(f"[AFRE] XLS Stock parse error: {e}")
        return {}


def _download_and_parse_flow(url: str) -> Dict[str, float]:
    """Flow URLをダウンロードしてパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        if url.lower().endswith(".xlsx"):
            return _parse_flow_xlsx(resp.content)
        else:
            return _parse_flow_xls(resp.content)
    except Exception as e:
        logger.warning(f"[AFRE] Flow download failed {url}: {e}")
        return {}


def _download_and_parse_stock(url: str) -> Dict[str, Dict[str, Optional[float]]]:
    """Stock URLをダウンロードしてパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        if url.lower().endswith(".xlsx"):
            return _parse_stock_xlsx(resp.content)
        else:
            return _parse_stock_xls(resp.content)
    except Exception as e:
        logger.warning(f"[AFRE] Stock download failed {url}: {e}")
        return {}


# =============================================================================
# FMP補完: Excelに未掲載の最新月を FMP から取得
# =============================================================================

def _fetch_fmp_flow_actuals() -> Dict[str, float]:
    """
    FMP から Total Social Financing の actual 値を取得
    → {date_str: flow_value} (単位: 百億元 = FMP値 × 10)
    FMP値は「十億元」単位なので × 10 で「100億元」単位に換算

    注: 402エラー回避のため過去6ヶ月に限定
    """
    try:
        from datetime import timedelta
        from services.calendar.fmp_service import fmp_service

        today = date.today()
        # 過去6ヶ月 + 今後2ヶ月（長期間は402エラーになる）
        events = fmp_service.fetch_calendar(
            today - timedelta(days=180),
            today + timedelta(days=60),
            country="CN",
        )

        result: Dict[str, float] = {}
        for event in events:
            if "Total Social Financing" not in event.get("event", ""):
                continue
            actual = event.get("actual")
            if actual is None:
                continue

            # イベント名から参照月を抽出: "Total Social Financing (Jan)" → 1月
            event_name = event.get("event", "")
            month_match = re.search(r'\((\w+)\)', event_name)
            if not month_match:
                continue

            # 発表日から年を推定
            event_date_str = event.get("date", "")[:10]
            try:
                event_dt = datetime.strptime(event_date_str, "%Y-%m-%d")
            except ValueError:
                continue

            month_abbr = month_match.group(1)
            try:
                month_num = datetime.strptime(month_abbr, "%b").month
            except ValueError:
                continue

            # 発表月の前月が参照月 (1月データは2月に発表)
            # ただし年末処理: 12月データは翌年1月発表
            ref_year = event_dt.year
            if month_num > event_dt.month:
                ref_year -= 1

            _, last_day = calendar.monthrange(ref_year, month_num)
            ref_date = date(ref_year, month_num, last_day)
            date_str = ref_date.strftime("%Y-%m-%d")

            # FMP単位(10億元) × 10 = 100億元単位
            result[date_str] = float(actual) * 10

        return result

    except Exception as e:
        logger.warning(f"[AFRE] FMP flow actuals failed: {e}")
        return {}


# =============================================================================
# 全データ取得・マージ
# =============================================================================

def fetch_all_data() -> Dict[str, Dict]:
    """
    ハードコードされた過去URL + 動的発見（最新年）でデータ取得・マージ
    Returns: {date_str: {"date", "flow", "stock", "yoy"}}

    データ取得順:
    1. HARDCODED_FLOW_URLS / HARDCODED_STOCK_URLS から過去データを取得
    2. 動的に発見した shrzgm ページから最新年データを取得（あれば上書き）
    3. FMP から未収録の最新月を補完
    """
    flow_all: Dict[str, float] = {}
    stock_all: Dict[str, Dict[str, Optional[float]]] = {}

    # --- Step 1: ハードコードURLから過去データ ---
    logger.info(f"[AFRE] Parsing {len(HARDCODED_FLOW_URLS)} hardcoded flow URLs")
    for url in HARDCODED_FLOW_URLS:
        flow_data = _download_and_parse_flow(url)
        logger.info(f"[AFRE] Flow {url.split('/')[-1]}: {len(flow_data)} records")
        flow_all.update(flow_data)

    logger.info(f"[AFRE] Parsing {len(HARDCODED_STOCK_URLS)} hardcoded stock URLs")
    for url in HARDCODED_STOCK_URLS:
        stock_data = _download_and_parse_stock(url)
        logger.info(f"[AFRE] Stock {url.split('/')[-1]}: {len(stock_data)} records")
        stock_all.update(stock_data)

    logger.info(f"[AFRE] After hardcoded: flow={len(flow_all)} months, stock={len(stock_all)} months")

    # --- Step 2: 動的発見（全年の shrzgm ページ、HARDCODEDの補完・上書き）---
    # MIN_YEAR以降を全探索することで、HARDCODEDに未登録の年のXLSXも自動取得できる
    year_map = _discover_year_urls(min_year=MIN_YEAR)
    for year in sorted(year_map.keys()):
        soup = _fetch_soup(year_map[year])
        if soup is None:
            logger.warning(f"[AFRE] Could not fetch shrzgm for year {year}")
            continue

        flow_url, stock_url = _find_afre_links(soup)

        if flow_url:
            logger.info(f"[AFRE] Dynamic flow year {year}: {flow_url}")
            flow_data = _download_and_parse_flow(flow_url)
            flow_all.update(flow_data)
        else:
            logger.warning(f"[AFRE] No dynamic Flow file for year {year}")

        if stock_url:
            logger.info(f"[AFRE] Dynamic stock year {year}: {stock_url}")
            stock_data = _download_and_parse_stock(stock_url)
            stock_all.update(stock_data)
        else:
            logger.warning(f"[AFRE] No dynamic Stock file for year {year}")

    # --- Step 3: FMP で最新月を補完 ---
    fmp_flows = _fetch_fmp_flow_actuals()
    logger.info(f"[AFRE] FMP補完候補: {len(fmp_flows)} months")
    for date_str, fmp_val in fmp_flows.items():
        if date_str not in flow_all:
            flow_all[date_str] = fmp_val
            logger.info(f"[AFRE] FMP補完 {date_str}: {fmp_val:.0f} 百億元")

    # --- マージ: すべての日付を統合 ---
    all_dates = set(flow_all.keys()) | set(stock_all.keys())
    result: Dict[str, Dict] = {}

    for date_str in sorted(all_dates):
        flow_val = flow_all.get(date_str)
        stock_rec = stock_all.get(date_str, {})
        result[date_str] = {
            "date": date_str,
            "flow": flow_val,
            "stock": stock_rec.get("stock"),
            "yoy": stock_rec.get("yoy"),
        }

    logger.info(f"[AFRE] Total merged: {len(result)} months")
    return result


# =============================================================================
# サービスクラス
# =============================================================================

class CnAggregateFinancingService:
    """中国 社会融資総量（AFRE）サービス"""

    def __init__(self):
        self._redis = None

    def _get_redis(self):
        if self._redis is None:
            try:
                import redis
                self._redis = redis.from_url(os.getenv("REDIS_URL", "redis://localhost:6379/0"), socket_timeout=2, socket_connect_timeout=2)
                self._redis.ping()
            except Exception:
                self._redis = None
        return self._redis

    def _from_redis(self) -> Optional[Dict[str, Any]]:
        r = self._get_redis()
        if not r:
            return None
        try:
            raw = r.get(REDIS_KEY)
            if raw:
                return json.loads(raw)
        except Exception:
            pass
        return None

    def _to_redis(self, payload: Dict[str, Any]) -> None:
        r = self._get_redis()
        if not r:
            return
        try:
            r.setex(REDIS_KEY, REDIS_TTL, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _from_file(self) -> Optional[Dict[str, Any]]:
        if not os.path.exists(FILE_CACHE_PATH):
            return None
        try:
            with open(FILE_CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _to_file(self, payload: Dict[str, Any]) -> None:
        os.makedirs(_FILE_CACHE_DIR, exist_ok=True)
        try:
            with open(FILE_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"[AFRE] File cache write error: {e}")

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """FMP（CN）から次回発表日を取得"""
        try:
            from services.usa.fmp_next_release_utils import get_next_release_from_fmp
            result = get_next_release_from_fmp("cn_aggregate_financing_to_the_real_economy", country="CN")
            if result:
                return result
        except Exception as e:
            logger.debug(f"[AFRE] FMP next_release failed: {e}")
        return None

    def _build_payload(self, data_map: Dict[str, Dict]) -> Dict[str, Any]:
        records = sorted(data_map.values(), key=lambda x: x["date"])
        latest = records[-1] if records else None
        next_release = self._get_next_release()

        return {
            "data": records,
            "latest": latest,
            "metadata": {
                "unit_flow": "100億元",
                "unit_stock": "兆元",
                "unit_yoy": "%",
                "indicator": "Aggregate Financing to the Real Economy",
                "source": "PBOC",
                "total_records": len(records),
                "last_fetched": datetime.now(JST).isoformat(),
            },
            "next_release": next_release,
        }

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """社会融資総量データを取得（キャッシュ優先）"""
        if not force_refresh:
            cached = self._from_redis()
            if cached:
                return cached
            cached = self._from_file()
            if cached:
                self._to_redis(cached)
                return cached

        data_map = fetch_all_data()
        if not data_map:
            cached = self._from_file()
            if cached:
                return cached
            return {"data": [], "latest": None, "metadata": {}, "next_release": None}

        payload = self._build_payload(data_map)
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def update_latest(self) -> Dict[str, Any]:
        """
        当月データを取得して既存キャッシュとマージ。

        欠損検出ロジック:
        - stockデータの最後の月と今月を比較
        - stock欠損が3ヶ月以上ある場合 → fetch_all_data() で全再構築
          （新しいXLSXが公開されていれば自動取得できる）
        - 3ヶ月未満の欠損 → shrzgm動的発見 + FMP補完
        """
        existing_payload = self._from_redis() or self._from_file()
        existing_map: Dict[str, Dict] = {}
        if existing_payload:
            for rec in existing_payload.get("data", []):
                existing_map[rec["date"]] = rec

        today = datetime.now(JST).date()

        # --- stockの欠損月数を確認 ---
        stock_months_missing = 0
        if existing_map:
            records_with_stock = [
                r for r in existing_map.values()
                if r.get("stock") is not None
            ]
            if records_with_stock:
                last_stock_date = max(r["date"] for r in records_with_stock)
                last_stock_dt = datetime.strptime(last_stock_date, "%Y-%m-%d")
                # 今月との差分（月数）
                stock_months_missing = (
                    (today.year - last_stock_dt.year) * 12
                    + (today.month - last_stock_dt.month)
                    - 1  # 今月分はまだ発表されていない可能性
                )
                logger.info(
                    f"[AFRE] Last stock month: {last_stock_date}, "
                    f"missing approx {stock_months_missing} months"
                )

        # --- stock欠損が3ヶ月以上 → 全再構築（新XLSXが公開されていれば自動取得） ---
        if stock_months_missing >= 3:
            logger.info("[AFRE] Stock gap >= 3 months, running full fetch_all_data()")
            data_map = fetch_all_data()
            if data_map:
                payload = self._build_payload(data_map)
                self._to_redis(payload)
                self._to_file(payload)
                return payload

        # --- 通常更新: shrzgm動的発見 + 最新ハードコードURL + FMP補完 ---
        target_min = today.year - 1 if today.month <= 2 else today.year

        def _update_map_flow(flow_data: Dict[str, float]) -> None:
            for date_str, val in flow_data.items():
                if date_str not in existing_map:
                    existing_map[date_str] = {"date": date_str, "flow": None, "stock": None, "yoy": None}
                existing_map[date_str]["flow"] = val

        def _update_map_stock(stock_data: Dict[str, Dict]) -> None:
            for date_str, rec in stock_data.items():
                if date_str not in existing_map:
                    existing_map[date_str] = {"date": date_str, "flow": None, "stock": None, "yoy": None}
                if rec.get("stock") is not None:
                    existing_map[date_str]["stock"] = rec["stock"]
                if rec.get("yoy") is not None:
                    existing_map[date_str]["yoy"] = rec["yoy"]

        # 最新ハードコードURLから補完（最後の2件、重複カバー用）
        for url in HARDCODED_FLOW_URLS[-2:]:
            _update_map_flow(_download_and_parse_flow(url))
        for url in HARDCODED_STOCK_URLS[-2:]:
            _update_map_stock(_download_and_parse_stock(url))

        # 動的発見（最新年の shrzgm ページ）
        year_map = _discover_year_urls(min_year=target_min)
        for year in sorted(year_map.keys()):
            if year < target_min:
                continue
            soup = _fetch_soup(year_map[year])
            if soup is None:
                continue
            flow_url, stock_url = _find_afre_links(soup)
            if flow_url:
                _update_map_flow(_download_and_parse_flow(flow_url))
            if stock_url:
                _update_map_stock(_download_and_parse_stock(stock_url))

        # FMP補完（flowのみ）
        fmp_flows = _fetch_fmp_flow_actuals()
        for date_str, fmp_val in fmp_flows.items():
            if date_str not in existing_map or existing_map[date_str].get("flow") is None:
                if date_str not in existing_map:
                    existing_map[date_str] = {"date": date_str, "flow": None, "stock": None, "yoy": None}
                existing_map[date_str]["flow"] = fmp_val

        payload = self._build_payload(existing_map)
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def invalidate_cache(self) -> Dict[str, Any]:
        r = self._get_redis()
        if r:
            try:
                r.delete(REDIS_KEY)
            except Exception:
                pass
        if os.path.exists(FILE_CACHE_PATH):
            os.remove(FILE_CACHE_PATH)
        return {"success": True, "message": "AFRE cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        r = self._get_redis()
        redis_exists = False
        redis_ttl = None
        if r:
            try:
                redis_exists = bool(r.exists(REDIS_KEY))
                redis_ttl = r.ttl(REDIS_KEY)
            except Exception:
                pass
        file_exists = os.path.exists(FILE_CACHE_PATH)
        file_mtime = None
        if file_exists:
            file_mtime = datetime.fromtimestamp(
                os.path.getmtime(FILE_CACHE_PATH), tz=JST
            ).isoformat()
        return {
            "redis": {"exists": redis_exists, "ttl_seconds": redis_ttl},
            "file": {"exists": file_exists, "last_modified": file_mtime},
        }


# シングルトン
cn_aggregate_financing_service = CnAggregateFinancingService()

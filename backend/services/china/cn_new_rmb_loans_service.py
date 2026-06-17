"""
新規人民元貸出（New RMB Loans）サービス

指標:
- cn_new_rmb_loans: 新規人民元貸出

データソース:
1. PBOC Excel「金融機構人民幣信貸收支表 Sources & Uses of Credit Funds of Financial Institutions (RMB)」
   - 「各项贷款 Total Loans」行の月末残高（亿元）
   - 残高差分から月次フロー（新増）を計算、前年同月差分からYoYを計算
2. FMP: "New Yuan Loans" / "New Loans" イベント → 月次フロー補完
   - FMP値は十億元単位。亿元に変換: × 10

Excel構造:
  Row 5: ヘッダー（Col0="项目 Item", Col1..12="2025.01"..."2025.12"）
  Row 26: "各项贷款 Total Loans" = 全項貸款残高（亿元）
  各ファイルは1年分の月次データ

発表スケジュール: 毎月9～17日 15:00 CST

キャッシュ方式: Redis(24h) + ファイルキャッシュ
"""
import io
import json
import os
import re
import calendar
import logging
from datetime import date, datetime
from typing import Optional, Dict, Any, List, Tuple
from zoneinfo import ZoneInfo

import requests

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
REDIS_KEY = "china:cn_new_rmb_loans:data"
REDIS_TTL = 24 * 3600

_FILE_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "cache", "china", "policy"
)
FILE_CACHE_PATH = os.path.join(_FILE_CACHE_DIR, "cn_new_rmb_loans_cache.json")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.pbc.gov.cn/",
}

# ハードコードURL（Sources & Uses of Credit Funds RMB）
# 古い順（2013年～）
HARDCODED_URLS = [
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112717062845500.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112717012995549.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716571458044.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716532399602.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716493731571.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716461352389.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716413424004.xlsx",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716370636596.xls",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716335546885.xlsx",
    "https://www.pbc.gov.cn/en/attachDir/2025/11/2025112716295144724.xlsx",
    # 2026年版
    "https://www.pbc.gov.cn/diaochatongjisi/attachDir/2026/01/2026011517314370061.xlsx",
]

# FMP イベントパターン
FMP_EVENT_PATTERNS = ["New Yuan Loans", "New Loans"]
FMP_COUNTRY = "CN"

# Total Loans キーワード
TOTAL_LOANS_KEYWORDS = ["各项贷款", "Total Loans"]


# =============================================================================
# Excel パーサー
# =============================================================================

def _parse_date_str(date_str: Any) -> Optional[date]:
    """'2025.01' → date(2025, 1, 31) 月末日"""
    if date_str is None:
        return None
    s = str(date_str).strip()
    m = re.match(r'^(\d{4})\.(\d{1,2})$', s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            _, last_day = calendar.monthrange(year, month)
            return date(year, month, last_day)
    return None


def _resolve_header_dates(raw_dates: List[Any]) -> List[Optional[date]]:
    """
    ヘッダー行の日付リストを解析し、曖昧な値（例: '2024.1' → 1月 or 10月）を
    列順序の連続性から解決する。

    PBOCのExcelでは日付ヘッダーは左から1月, 2月, ...12月の順。
    '2024.1' は列位置的に10月（col 11）のはずが、単純パースだと1月になるバグを修正。

    解決方法:
    - まず全日付を単純パースし、月の重複を検出
    - 重複がある場合、列位置ベースで月を再割当（連続月を前提）
    """
    # Step 1: 単純パース
    parsed = [_parse_date_str(d) for d in raw_dates]

    # Step 2: 重複月の検出
    seen_months: Dict[str, int] = {}  # "YYYY-MM" -> count
    has_duplicate = False
    for d in parsed:
        if d is None:
            continue
        key = f"{d.year}-{d.month:02d}"
        seen_months[key] = seen_months.get(key, 0) + 1
        if seen_months[key] > 1:
            has_duplicate = True

    if not has_duplicate:
        return parsed

    # Step 3: 重複あり → 列位置ベースで月を再割当
    # 最初の有効な日付から年を取得し、連続月を仮定
    first_date = next((d for d in parsed if d is not None), None)
    if first_date is None:
        return parsed

    # 最初の列位置を特定
    first_idx = next(i for i, d in enumerate(parsed) if d is not None)

    # 最初の日付の月を基準に連続月を割当
    base_year = first_date.year
    base_month = first_date.month

    resolved = []
    for i, raw in enumerate(raw_dates):
        if raw is None or _parse_date_str(raw) is None:
            resolved.append(None)
            continue
        # 列位置からの相対月
        offset = i - first_idx
        month = base_month + offset
        year = base_year
        while month > 12:
            month -= 12
            year += 1
        while month < 1:
            month += 12
            year -= 1
        _, last_day = calendar.monthrange(year, month)
        resolved.append(date(year, month, last_day))

    logger.info(f"[NewRMBLoans] Resolved ambiguous dates: {[str(d) for d in resolved if d]}")
    return resolved


def _parse_xlsx(content: bytes) -> List[Tuple[date, float]]:
    """XLSXから (date, total_loans_stock) を返す"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        header_row = None
        loans_row = None

        for row_idx in range(1, min(ws.max_row + 1, 40)):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a is None:
                continue
            cell_a_str = str(cell_a).strip()

            # ヘッダー行: "20XX.XX" パターン
            if header_row is None:
                for col_idx in range(2, min(ws.max_column + 1, 20)):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break

            # 各項貸款行（「一、各项贷款」ではなく先頭の「各项贷款」= Total Loans）
            if loans_row is None:
                for kw in TOTAL_LOANS_KEYWORDS:
                    if kw in cell_a_str:
                        # 「一、各项贷款」のようにサブヘッダーではなくトップレベルの行を取得
                        # Row 26 付近の "各项贷款 Total Loans"
                        loans_row = row_idx
                        break

        if header_row is None or loans_row is None:
            logger.warning(f"[NewRMBLoans] XLSX: header={header_row}, loans={loans_row}")
            return []

        # ヘッダー行の全日付を一括読み取り → 曖昧日付を列順で解決
        raw_dates = [
            ws.cell(row=header_row, column=col_idx).value
            for col_idx in range(2, ws.max_column + 1)
        ]
        resolved_dates = _resolve_header_dates(raw_dates)

        results = []
        for i, d in enumerate(resolved_dates):
            col_idx = i + 2  # raw_datesのインデックス0 → col 2
            loan_val = ws.cell(row=loans_row, column=col_idx).value
            if d and loan_val is not None:
                try:
                    results.append((d, float(loan_val)))
                except (ValueError, TypeError):
                    pass

        return results
    except Exception as e:
        logger.warning(f"[NewRMBLoans] XLSX parse error: {e}")
        return []


def _parse_xls(content: bytes) -> List[Tuple[date, float]]:
    """XLSから (date, total_loans_stock) を返す"""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)

        header_row = None
        loans_row = None

        for row_idx in range(0, min(ws.nrows, 40)):
            cell_a = ws.cell_value(row_idx, 0)
            cell_a_str = str(cell_a).strip() if cell_a else ""

            if header_row is None:
                for col_idx in range(1, min(ws.ncols, 20)):
                    v = ws.cell_value(row_idx, col_idx)
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break

            if loans_row is None:
                for kw in TOTAL_LOANS_KEYWORDS:
                    if kw in cell_a_str:
                        loans_row = row_idx
                        break

        if header_row is None or loans_row is None:
            logger.warning(f"[NewRMBLoans] XLS: header={header_row}, loans={loans_row}")
            return []

        # ヘッダー行の全日付を一括読み取り → 曖昧日付を列順で解決
        raw_dates = [
            ws.cell_value(header_row, col_idx)
            for col_idx in range(1, ws.ncols)
        ]
        resolved_dates = _resolve_header_dates(raw_dates)

        results = []
        for i, d in enumerate(resolved_dates):
            col_idx = i + 1  # raw_datesのインデックス0 → col 1
            loan_val = ws.cell_value(loans_row, col_idx)
            if d and loan_val:
                try:
                    results.append((d, float(loan_val)))
                except (ValueError, TypeError):
                    pass

        return results
    except Exception as e:
        logger.warning(f"[NewRMBLoans] XLS parse error: {e}")
        return []


def _download_and_parse(url: str) -> List[Tuple[date, float]]:
    """URLをダウンロードしてパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        if url.lower().endswith('.xlsx'):
            return _parse_xlsx(resp.content)
        else:
            return _parse_xls(resp.content)
    except Exception as e:
        logger.warning(f"[NewRMBLoans] Download failed {url}: {e}")
        return []


# =============================================================================
# FMP 補完
# =============================================================================

_MONTH_ABBR_TO_NUM = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_data_month(event_name: str, announcement_dt: datetime) -> Optional[date]:
    """
    FMPイベント名からデータ対象月を解析する。
    例: "New Yuan Loans (Mar)" announced 2026-04-13 → date(2026, 3, 31)

    イベント名に月略称がない場合は発表月の前月を使用。
    """
    m = re.search(r'\((\w{3})\)', event_name)
    if m:
        month_str = m.group(1).lower()
        month_num = _MONTH_ABBR_TO_NUM.get(month_str)
        if month_num:
            # 年は発表日から推定（12月データは翌年1月に発表されるケースを考慮）
            year = announcement_dt.year
            if month_num > announcement_dt.month:
                year -= 1
            _, last_day = calendar.monthrange(year, month_num)
            return date(year, month_num, last_day)

    # フォールバック: 発表月の前月
    prev_month = announcement_dt.month - 1
    prev_year = announcement_dt.year
    if prev_month < 1:
        prev_month = 12
        prev_year -= 1
    _, last_day = calendar.monthrange(prev_year, prev_month)
    return date(prev_year, prev_month, last_day)


def _fetch_fmp_flow_actuals() -> Dict[str, float]:
    """
    FMPから新增人民元貸出のフローデータを取得
    まずDBから検索し、結果が少なければFMP APIから直接取得してフォールバック。
    Returns: {date_str: flow_value_in_亿元}
    """
    result = {}

    # Step 1: DBから取得
    try:
        from core.database import SessionLocal
        from sqlalchemy import text

        with SessionLocal() as session:
            # SQLインジェクション対策: 値は直埋めせずバインドパラメータで渡す
            conditions = []
            pattern_params = {}
            for i, pattern in enumerate(FMP_EVENT_PATTERNS):
                conditions.append(f"event ILIKE :pat{i}")
                pattern_params[f"pat{i}"] = f"%{pattern}%"

            query = text(f"""
                SELECT datetime_utc, event, actual
                FROM economic_calendar_events
                WHERE country = 'CN'
                  AND ({' OR '.join(conditions)})
                  AND actual IS NOT NULL
                  AND datetime_utc >= NOW() - INTERVAL '24 months'
                ORDER BY datetime_utc DESC
            """)
            rows = session.execute(query, pattern_params).fetchall()

            for row in rows:
                dt_utc, event_name, actual = row
                if dt_utc and actual is not None:
                    data_date = _parse_data_month(event_name, dt_utc)
                    if data_date:
                        date_str = data_date.strftime("%Y-%m-%d")
                        if date_str not in result:
                            result[date_str] = float(actual) * 10

    except Exception as e:
        logger.warning(f"[NewRMBLoans] FMP DB fetch error: {e}")

    # Step 2: DBの結果が少ない場合、FMP APIから直接取得
    if len(result) < 3:
        try:
            from services.calendar.fmp_service import fmp_service
            from datetime import timedelta

            today = date.today()
            events = fmp_service.fetch_calendar(
                today - timedelta(days=365),
                today + timedelta(days=30),
                country=FMP_COUNTRY,
            )

            for event in events:
                if event.get("country") != FMP_COUNTRY:
                    continue
                if event.get("actual") is None:
                    continue
                event_name = event.get("event", "")
                # "New Yuan Loans" のみ使用（"New Loans" は重複）
                if "new yuan loans" not in event_name.lower():
                    continue

                dt_utc, _ = fmp_service.parse_datetime(event.get("date", ""))
                if dt_utc:
                    data_date = _parse_data_month(event_name, dt_utc)
                    if data_date:
                        date_str = data_date.strftime("%Y-%m-%d")
                        if date_str not in result:
                            # FMP値は十億元 → 亿元に変換（× 10）
                            result[date_str] = float(event["actual"]) * 10

            logger.info(f"[NewRMBLoans] FMP API fallback: {len(result)} months total")
        except Exception as e:
            logger.warning(f"[NewRMBLoans] FMP API fallback error: {e}")

    logger.info(f"[NewRMBLoans] FMP flow data: {len(result)} months")
    return result


def _get_fmp_next_release() -> Optional[Dict[str, Any]]:
    """FMPから次回発表日を取得"""
    try:
        from services.china.fmp_next_release_utils import get_next_release_by_pattern
        return get_next_release_by_pattern("New Yuan Loans", country=FMP_COUNTRY)
    except Exception as e:
        logger.warning(f"[NewRMBLoans] Next release error: {e}")
        return None


# =============================================================================
# データ構築
# =============================================================================

def _build_data() -> List[Dict[str, Any]]:
    """
    Excel残高データ + FMPフローデータ を結合して時系列データを構築

    Returns: [{date, stock, flow, stock_yoy}, ...]
    """
    # 1. 全Excelファイルから残高データを取得
    stock_map: Dict[str, float] = {}  # date_str → 残高（亿元）
    for url in HARDCODED_URLS:
        records = _download_and_parse(url)
        for d, val in records:
            date_str = d.strftime("%Y-%m-%d")
            stock_map[date_str] = val

    logger.info(f"[NewRMBLoans] Excel stock data: {len(stock_map)} months")

    # 2. FMPからフローデータを取得
    fmp_flow_map = _fetch_fmp_flow_actuals()

    # 3. 残高差分からフローを計算
    sorted_dates = sorted(stock_map.keys())

    result = []
    for i, date_str in enumerate(sorted_dates):
        stock = stock_map[date_str]

        # フロー = 当月残高 - 前月残高（前月がある場合）
        flow = None
        if i > 0:
            prev_date = sorted_dates[i - 1]
            prev_stock = stock_map[prev_date]
            flow = stock - prev_stock

        # FMPフローで上書き（最新月はExcelが遅れるためFMPを優先）
        if date_str in fmp_flow_map:
            flow = fmp_flow_map[date_str]

        # 残高前年比
        stock_yoy = None
        # 12ヶ月前を探す
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        prev_year_month = dt.month
        prev_year = dt.year - 1
        _, prev_last_day = calendar.monthrange(prev_year, prev_year_month)
        prev_year_date = f"{prev_year}-{prev_year_month:02d}-{prev_last_day:02d}"
        if prev_year_date in stock_map:
            prev_stock_yoy = stock_map[prev_year_date]
            if prev_stock_yoy > 0:
                stock_yoy = round(((stock - prev_stock_yoy) / prev_stock_yoy) * 100, 2)

        result.append({
            "date": date_str,
            "stock": round(stock, 2),
            "flow": round(flow, 2) if flow is not None else None,
            "stock_yoy": stock_yoy,
        })

    # FMPにあってExcelにないフローのみの月を追加
    existing_dates = {r["date"] for r in result}
    for date_str, flow_val in fmp_flow_map.items():
        if date_str not in existing_dates:
            result.append({
                "date": date_str,
                "stock": None,
                "flow": round(flow_val, 2),
                "stock_yoy": None,
            })

    result.sort(key=lambda x: x["date"])
    return result


# =============================================================================
# サービスクラス
# =============================================================================

class CnNewRmbLoansService:
    """新規人民元貸出サービス"""

    def __init__(self):
        self._redis = None

    def _get_redis(self):
        if self._redis is None:
            try:
                import redis
                self._redis = redis.from_url(os.getenv("REDIS_URL", "redis://localhost:6379/0"), socket_timeout=2, socket_connect_timeout=2)
                self._redis.ping()
            except Exception:
                self._redis = None
        return self._redis

    def _from_redis(self) -> Optional[Dict[str, Any]]:
        r = self._get_redis()
        if not r:
            return None
        try:
            raw = r.get(REDIS_KEY)
            if raw:
                return json.loads(raw)
        except Exception:
            pass
        return None

    def _to_redis(self, payload: Dict[str, Any]) -> None:
        r = self._get_redis()
        if not r:
            return
        try:
            r.setex(REDIS_KEY, REDIS_TTL, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _from_file(self) -> Optional[Dict[str, Any]]:
        if not os.path.exists(FILE_CACHE_PATH):
            return None
        try:
            with open(FILE_CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _to_file(self, payload: Dict[str, Any]) -> None:
        os.makedirs(_FILE_CACHE_DIR, exist_ok=True)
        try:
            with open(FILE_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"[NewRMBLoans] File cache write error: {e}")

    def _build_payload(self) -> Dict[str, Any]:
        data = _build_data()
        # FMP補完で stock=None の月が末尾に追加されることがある。
        # latest が stock=None だとダッシュボードの null 検出が
        # 毎リクエスト再読込をトリガーするため、stock を持つ直近月を採用。
        latest = next(
            (r for r in reversed(data) if r.get("stock") is not None),
            data[-1] if data else None,
        )

        next_release = _get_fmp_next_release()

        return {
            "data": data,
            "latest": latest,
            "metadata": {
                "unit_stock": "亿元",
                "unit_flow": "亿元",
                "indicator": "New RMB Loans / Total Loans",
                "source": "PBOC Excel + FMP",
                "total_records": len(data),
                "last_fetched": datetime.now(JST).isoformat(),
            },
            "next_release": next_release,
        }

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """データ取得（キャッシュ優先）"""
        if not force_refresh:
            cached = self._from_redis()
            if cached:
                return cached
            cached = self._from_file()
            if cached:
                self._to_redis(cached)
                return cached

        payload = self._build_payload()
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def invalidate_cache(self) -> Dict[str, Any]:
        r = self._get_redis()
        if r:
            try:
                r.delete(REDIS_KEY)
            except Exception:
                pass
        if os.path.exists(FILE_CACHE_PATH):
            os.remove(FILE_CACHE_PATH)
        return {"success": True, "message": "New RMB Loans cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        r = self._get_redis()
        redis_exists = False
        redis_ttl = None
        if r:
            try:
                redis_exists = bool(r.exists(REDIS_KEY))
                redis_ttl = r.ttl(REDIS_KEY)
            except Exception:
                pass
        file_exists = os.path.exists(FILE_CACHE_PATH)
        file_mtime = None
        if file_exists:
            file_mtime = datetime.fromtimestamp(
                os.path.getmtime(FILE_CACHE_PATH), tz=JST
            ).isoformat()
        return {
            "redis": {"exists": redis_exists, "ttl_seconds": redis_ttl},
            "file": {"exists": file_exists, "last_modified": file_mtime},
        }


# シングルトン
cn_new_rmb_loans_service = CnNewRmbLoansService()

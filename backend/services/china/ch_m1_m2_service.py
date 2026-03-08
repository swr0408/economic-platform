"""
中国人民銀行（PBOC）M1/M2（貨幣供応量）サービス

データ方式:
- 親ページ（统计数据）から年別「货币统计概览」URLを動的に発見
- 各年の货币统計概覧ページから「货币供应量」XLS/XLSXリンクを取得

ファイル構造（XLSX: 2021年以降）:
  Sheet1:
    Row 6:  ヘッダー（2025.01, 2025.02, ...）
    Row 8:  M2水準値（億元）
    Row 10: M1水準値（億元）
    Row 17: 前年ヘッダー（2024.01, ...）
    Row 18: M1前年実績（億元）
    Row 21: M1前年比（小数: 0.033 = 3.3%）
  Sheet2: 当月月次追加分（同構造）

ファイル構造（XLS: 〜2020年頃）:
  Sheet1:
    Row 5:  ヘッダー
    Row 7:  M2水準値
    Row 9:  M1水準値
    Row 18: 前年ヘッダー（M2用）
    Row 19: M2前年比（小数）

M2 YoY: 水準値から計算（前年同月との差分/前年同月）
M1 YoY: ファイル内のGrowth Rateを使用（XLSXのRow 21、XLSのRow 19）

データソース: https://camlmac.pbc.gov.cn/diaochatongjisi/116219/116319/index.html

Redisキャッシュ: china:ch_m1_m2:data (TTL: 24時間)
ファイルキャッシュ: backend/data/cache/china/policy/ch_m1_m2_cache.json
"""
import io
import json
import os
import re
import logging
import calendar
from datetime import date, datetime
from typing import Optional, Dict, Any, List, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# -----------------------------------------------------------------
# キャッシュ設定
# -----------------------------------------------------------------
REDIS_KEY = "china:ch_m1_m2:data"
REDIS_TTL = 24 * 3600  # 24 hours

_FILE_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "cache", "china", "policy"
)
FILE_CACHE_PATH = os.path.join(_FILE_CACHE_DIR, "ch_m1_m2_cache.json")

# -----------------------------------------------------------------
# サイト定数
# -----------------------------------------------------------------
PBOC_HOST = "https://camlmac.pbc.gov.cn"
BASE_URL = f"{PBOC_HOST}/diaochatongjisi"
PARENT_INDEX_URL = f"{BASE_URL}/116219/116319/index.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://camlmac.pbc.gov.cn/",
}

M1M2_KEYWORDS = ["货币供应量", "Money Supply"]
ARC_KEYWORDS = ["公布日程预告", "Advance Release Calendar", "ARC"]
MIN_YEAR = 2010


# =============================================================================
# ユーティリティ
# =============================================================================

def _abs_url(href: str) -> str:
    if href.startswith("http"):
        return href
    elif href.startswith("/"):
        return f"{PBOC_HOST}{href}"
    else:
        return f"{BASE_URL}/{href}"


def _parse_date_str(date_str: Any, prev_month: Optional[int] = None) -> Optional[date]:
    """
    "2025.01" / "2025.1" / 2025.01 (float) → date(2025, 1, 31)
    月末日を使用。

    Excelでは "2024.10" が float 2024.1 として保存されるため、
    前月のコンテキスト (prev_month) を使って月を推定する。
    例: prev_month=9 の場合、"2024.1" → 月10 (10月) として解釈
    """
    if date_str is None:
        return None
    s = str(date_str).strip()
    m = re.match(r'^(\d{4})\.(\d{1,2})$', s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        # "YYYY.1" が float由来の場合、October (10) の可能性を検討
        # prev_month が 9 またはそれ以降であれば month=10 と解釈
        if month == 1 and prev_month is not None and prev_month >= 9:
            month = 10
        try:
            _, last_day = calendar.monthrange(year, month)
            return date(year, month, last_day)
        except ValueError:
            return None
    return None


def _fetch_soup(url: str, timeout: int = 15) -> Optional[BeautifulSoup]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code != 200:
            return None
        resp.encoding = "utf-8"
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        logger.debug(f"[M1M2] Fetch failed {url}: {e}")
        return None


# =============================================================================
# スクレイピング
# =============================================================================

def _discover_year_urls(min_year: int = MIN_YEAR) -> Dict[int, str]:
    """親ページから各年の「货币统计概览」URLを発見"""
    soup = _fetch_soup(PARENT_INDEX_URL)
    if soup is None:
        logger.warning("[M1M2] Failed to fetch parent index page")
        return {}

    year_map: Dict[int, str] = {}
    current_year: Optional[int] = None

    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True)
        href = a["href"]
        m = re.match(r'^(\d{4})年统计数据$', text)
        if m:
            current_year = int(m.group(1))
            continue
        if text == "货币统计概览" and current_year is not None:
            if current_year >= min_year and current_year not in year_map:
                year_map[current_year] = _abs_url(href)

    logger.info(f"[M1M2] Discovered {len(year_map)} years: {sorted(year_map.keys())}")
    return year_map


def _find_m1m2_link(soup: BeautifulSoup) -> Optional[str]:
    """货币统计概览ページから「货币供应量」XLS/XLSXリンクを取得"""
    for div in soup.find_all("div", class_="titp20"):
        if not any(kw in div.get_text() for kw in M1M2_KEYWORDS):
            continue
        tr = div.find_parent("tr")
        if tr is None:
            continue
        for a_tag in tr.find_all("a", href=True):
            href = a_tag["href"]
            if href.lower().endswith(".xlsx") or href.lower().endswith(".xls"):
                return _abs_url(href)
    return None


def _get_arc_url(soup: BeautifulSoup) -> Optional[str]:
    """货币统計概覧ページからARC XLSリンクを取得"""
    for div in soup.find_all("div", class_="titp20"):
        if not any(kw in div.get_text() for kw in ARC_KEYWORDS):
            continue
        tr = div.find_parent("tr")
        if tr is None:
            continue
        for a_tag in tr.find_all("a", href=True):
            href = a_tag["href"]
            if href.lower().endswith(".xlsx") or href.lower().endswith(".xls"):
                return _abs_url(href)
    return None


# =============================================================================
# XLS/XLSX パーサー
#
# 戻り値: Dict[date_str, {"m2": float|None, "m1": float|None}]
# =============================================================================

def _parse_xlsx_sheet(ws) -> Dict[str, Dict[str, Optional[float]]]:
    """
    XLSXの1シートをパース。

    2つのヘッダーグループがある:
    - 主ヘッダー (Row 6): 当年データ → M2水準値 / M1水準値
    - 副ヘッダー (Row 17): 前年再計算データ（M1定義変更対応）
        - Row 18: 再計算済み前年M1残高（新定義ベース）→ m1_restated
        - Row 21: 再計算済み前年M1前年比 → m1_yoy (旧 vs 新定義の調整済みYoY)

    2025年以降、PBOCはM1定義を変更。
    - 副ヘッダーの m1_restated は当年のM1 YoY計算の基底値として使用する
    - 副ヘッダーの m1_yoy は前年（再計算後）の公式YoY値として使用する

    戻り値: {date_str: {"m2": val, "m1": val, "m1_yoy": val, "m1_restated": val}}
    m1_restated: 前年の再計算済みM1残高（当年YoY計算に使用）
    """
    result: Dict[str, Dict[str, Optional[float]]] = {}

    # すべてのヘッダー行（日付パターン YYYY.MM を含む行）を収集
    header_rows: list = []
    m2_row = None
    m1_row = None
    m1_yoy_row = None      # 小数（0.033）形式のYoY行
    m1_restated_row = None # 副ヘッダー内のM1残高行（余额・Balance）

    for row_idx in range(1, min(ws.max_row + 1, 30)):
        has_date_col = False
        for col_idx in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                if not has_date_col:
                    header_rows.append(row_idx)
                    has_date_col = True

        # 各行のA/B/C列でキーワードを確認（ヘッダー後）
        if header_rows:
            for ci in range(1, 4):
                cell_v = ws.cell(row=row_idx, column=ci).value
                if not cell_v:
                    continue
                cell_str = str(cell_v)
                # M2水準値行（主ヘッダーエリア）
                if m2_row is None and ("M2" in cell_str or "货币和准货币" in cell_str):
                    m2_row = row_idx
                # M1水準値行（M0・流通中を除外）
                if m1_row is None and ("M1" in cell_str or "货币（M1）" in cell_str or "货币(" in cell_str):
                    if "M0" not in cell_str and "流通" not in cell_str and "Circulation" not in cell_str:
                        m1_row = row_idx
                # M1 Growth Rate / YOY 行（小数値を持つ行を優先）
                if m1_yoy_row is None and (
                    "Growth Rate" in cell_str or "YOY" in cell_str.upper()
                    or "增长率" in cell_str or "同比" in cell_str
                ):
                    has_decimal = any(
                        ws.cell(row=row_idx, column=cc).value is not None
                        and isinstance(ws.cell(row=row_idx, column=cc).value, (int, float))
                        and abs(float(ws.cell(row=row_idx, column=cc).value)) < 2.0
                        for cc in range(1, min(ws.max_column + 1, 25))
                    )
                    if has_decimal:
                        m1_yoy_row = row_idx
                # 副ヘッダーエリアのM1残高行（余额/Balance）
                # 副ヘッダーが検出された後（len(header_rows)>=2）に探す
                if m1_restated_row is None and len(header_rows) >= 2:
                    if "余额" in cell_str or ("Balance" in cell_str and "100 Million" not in cell_str):
                        # この行に大きな数値（M1水準値レベル）があるか確認
                        has_large_val = any(
                            ws.cell(row=row_idx, column=cc).value is not None
                            and isinstance(ws.cell(row=row_idx, column=cc).value, (int, float))
                            and float(ws.cell(row=row_idx, column=cc).value) > 100000
                            for cc in range(1, min(ws.max_column + 1, 25))
                        )
                        if has_large_val:
                            m1_restated_row = row_idx

    if not header_rows or m2_row is None:
        logger.debug(f"[M1M2] XLSX sheet parse failed: headers={header_rows}, m2={m2_row}, m1={m1_row}")
        return result

    # --- 主ヘッダー（最初のヘッダー行）: M2/M1水準値を取得 ---
    primary_header_row = header_rows[0]
    prev_month: Optional[int] = None
    for col_idx in range(1, ws.max_column + 1):
        date_val = ws.cell(row=primary_header_row, column=col_idx).value
        d = _parse_date_str(date_val, prev_month)
        if not d:
            continue
        prev_month = d.month
        date_str = d.strftime("%Y-%m-%d")

        m2_val = ws.cell(row=m2_row, column=col_idx).value if m2_row else None
        m1_val = ws.cell(row=m1_row, column=col_idx).value if m1_row else None

        try:
            m2 = float(m2_val) if m2_val is not None else None
        except (ValueError, TypeError):
            m2 = None
        try:
            m1 = float(m1_val) if m1_val is not None else None
        except (ValueError, TypeError):
            m1 = None

        if m2 is not None or m1 is not None:
            if date_str not in result:
                result[date_str] = {"m2": None, "m1": None, "m1_yoy": None, "m1_restated": None}
            if m2 is not None:
                result[date_str]["m2"] = m2
            if m1 is not None:
                result[date_str]["m1"] = m1

    # --- 副ヘッダー（2番目以降のヘッダー行）---
    # 前年再計算M1残高 (m1_restated) と 前年公式M1 YoY を取得
    if len(header_rows) >= 2:
        secondary_header_row = header_rows[1]
        prev_month = None
        for col_idx in range(1, ws.max_column + 1):
            date_val = ws.cell(row=secondary_header_row, column=col_idx).value
            d = _parse_date_str(date_val, prev_month)
            if not d:
                continue
            prev_month = d.month
            date_str = d.strftime("%Y-%m-%d")

            # M1残高（再計算済み）
            m1_restated = None
            if m1_restated_row is not None:
                rv = ws.cell(row=m1_restated_row, column=col_idx).value
                try:
                    rv_f = float(rv) if rv is not None else None
                    if rv_f is not None and rv_f > 100000:
                        m1_restated = rv_f
                except (ValueError, TypeError):
                    pass

            # M1 YoY（官方値）
            m1_yoy = None
            if m1_yoy_row is not None:
                yv = ws.cell(row=m1_yoy_row, column=col_idx).value
                try:
                    m1_yoy = float(yv) if yv is not None else None
                    if m1_yoy is not None and abs(m1_yoy) < 2.0:
                        m1_yoy = round(m1_yoy * 100, 2)
                except (ValueError, TypeError):
                    m1_yoy = None

            if m1_restated is not None or m1_yoy is not None:
                if date_str not in result:
                    result[date_str] = {"m2": None, "m1": None, "m1_yoy": None, "m1_restated": None}
                if m1_restated is not None:
                    result[date_str]["m1_restated"] = m1_restated
                if m1_yoy is not None:
                    result[date_str]["m1_yoy"] = m1_yoy

    return result


def _parse_xlsx(content: bytes) -> Dict[str, Dict[str, Optional[float]]]:
    """XLSXファイルをパース（全シートを処理してマージ）"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        merged: Dict[str, Dict[str, Optional[float]]] = {}
        for ws in wb.worksheets:
            sheet_result = _parse_xlsx_sheet(ws)
            for date_str, rec in sheet_result.items():
                if date_str not in merged:
                    merged[date_str] = {"m2": None, "m1": None, "m1_yoy": None, "m1_restated": None}
                # 新しい値で上書き（Noneは上書きしない）
                for k in ["m2", "m1", "m1_yoy", "m1_restated"]:
                    if rec.get(k) is not None:
                        merged[date_str][k] = rec[k]
        return merged
    except Exception as e:
        logger.warning(f"[M1M2] XLSX parse error: {e}")
        return {}


def _parse_xls_sheet(ws) -> Dict[str, Dict[str, Optional[float]]]:
    """XLSの1シートをパース"""
    result: Dict[str, Dict[str, Optional[float]]] = {}
    header_row = None
    m2_row = None
    m1_row = None
    m2_yoy_row = None   # XLS形式: Row 19 が "M2" + 小数YoY

    for row_idx in range(min(ws.nrows, 30)):
        # ヘッダー行
        if header_row is None:
            for col_idx in range(min(ws.ncols, 25)):
                v = ws.cell_value(row_idx, col_idx)
                if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                    header_row = row_idx
                    break
        if header_row is None:
            continue

        for ci in range(3):
            cell_v = ws.cell_value(row_idx, ci)
            if not cell_v:
                continue
            cell_str = str(cell_v).strip()
            if m2_row is None and ("M2" in cell_str or "货币和准货币" in cell_str):
                m2_row = row_idx
            if m1_row is None and "M1" in cell_str and "M0" not in cell_str and "流通" not in cell_str and "Circulation" not in cell_str:
                m1_row = row_idx
            # XLS: "M2" 単独セルが YoY 行の識別子
            if m2_yoy_row is None and cell_str == "M2":
                # 数値が入っていて小数（YoY）らしい列があるか確認
                for cc in range(3, min(ws.ncols, 25)):
                    vv = ws.cell_value(row_idx, cc)
                    if isinstance(vv, float) and abs(vv) < 2.0:
                        m2_yoy_row = row_idx
                        break

    if header_row is None or m2_row is None:
        logger.debug(f"[M1M2] XLS sheet parse failed: header={header_row}, m2={m2_row}")
        return result

    prev_month: Optional[int] = None
    for col_idx in range(ws.ncols):
        date_val = ws.cell_value(header_row, col_idx)
        d = _parse_date_str(date_val, prev_month)
        if not d:
            continue
        prev_month = d.month
        date_str = d.strftime("%Y-%m-%d")

        m2_val = ws.cell_value(m2_row, col_idx) if m2_row is not None else None
        m1_val = ws.cell_value(m1_row, col_idx) if m1_row is not None else None
        m2_yoy_val = ws.cell_value(m2_yoy_row, col_idx) if m2_yoy_row is not None else None

        try:
            m2 = float(m2_val) if m2_val else None
        except (ValueError, TypeError):
            m2 = None
        try:
            m1 = float(m1_val) if m1_val else None
        except (ValueError, TypeError):
            m1 = None
        try:
            m2_yoy = float(m2_yoy_val) if m2_yoy_val else None
            if m2_yoy is not None and abs(m2_yoy) < 2.0:
                m2_yoy = round(m2_yoy * 100, 2)
        except (ValueError, TypeError):
            m2_yoy = None

        if m2 is not None or m1 is not None:
            if date_str not in result:
                result[date_str] = {"m2": None, "m1": None, "m2_yoy": m2_yoy}
            else:
                if m2_yoy is not None:
                    result[date_str]["m2_yoy"] = m2_yoy
            if m2 is not None:
                result[date_str]["m2"] = m2
            if m1 is not None:
                result[date_str]["m1"] = m1

    return result


def _parse_xls(content: bytes) -> Dict[str, Dict[str, Optional[float]]]:
    """XLSファイルをパース（全シートをマージ）"""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        merged: Dict[str, Dict[str, Optional[float]]] = {}
        for si in range(wb.nsheets):
            ws = wb.sheet_by_index(si)
            sheet_result = _parse_xls_sheet(ws)
            for date_str, rec in sheet_result.items():
                if date_str not in merged:
                    merged[date_str] = {"m2": None, "m1": None, "m2_yoy": None}
                for k in list(rec.keys()):
                    if rec.get(k) is not None:
                        merged[date_str][k] = rec[k]
        return merged
    except Exception as e:
        logger.warning(f"[M1M2] XLS parse error: {e}")
        return {}


def _download_and_parse(url: str) -> Dict[str, Dict[str, Optional[float]]]:
    """URLをダウンロードしてXLS/XLSXをパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        content = resp.content
        if url.lower().endswith('.xlsx'):
            return _parse_xlsx(content)
        else:
            return _parse_xls(content)
    except Exception as e:
        logger.warning(f"[M1M2] Download failed {url}: {e}")
        return {}


# =============================================================================
# 前年比計算（水準値から）
# =============================================================================

def _compute_yoy(all_data: Dict[str, Dict]) -> Dict[str, Dict]:
    """
    水準値から M2/M1 の前年比を計算して補完する。
    - m2_yoy: M2水準値から前年同月との差分/前年同月 × 100
    - m1_yoy: M1水準値から前年同月との差分/前年同月 × 100
    ファイルにYoYがある場合はそちらを優先。

    M1 YoY計算の基底値（前年M1）:
    - m1_restated（PBOCが新定義で再計算した前年M1残高）が利用可能な場合はそちらを使用
    - ない場合は従来の前年m1値を使用
    2025年1月以降のM1はPBOCが定義を変更しているため、
    m1_restatedを使うことで定義変更の影響を排除した正確なYoYを算出できる。
    """
    for date_str, rec in all_data.items():
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            _, last_day_prev = calendar.monthrange(dt.year - 1, dt.month)
            prev_year_str = f"{dt.year - 1:04d}-{dt.month:02d}-{last_day_prev:02d}"
        except ValueError:
            continue

        if prev_year_str not in all_data:
            continue

        prev = all_data[prev_year_str]

        # M2 YoY（ファイルに値がない場合は計算）
        if rec.get("m2_yoy") is None and rec.get("m2") and prev.get("m2"):
            try:
                rec["m2_yoy"] = round((rec["m2"] - prev["m2"]) / prev["m2"] * 100, 2)
            except (ZeroDivisionError, TypeError):
                pass

        # M1 YoY（ファイルに値がない場合は計算）
        # 前年M1の基底値: m1_restated（再計算値）> m1（旧定義値）の優先順位
        if rec.get("m1_yoy") is None and rec.get("m1"):
            prev_m1 = prev.get("m1_restated") or prev.get("m1")
            if prev_m1:
                try:
                    rec["m1_yoy"] = round((rec["m1"] - prev_m1) / prev_m1 * 100, 2)
                except (ZeroDivisionError, TypeError):
                    pass

    return all_data


# =============================================================================
# ARC 次回発表日
# =============================================================================

def _parse_arc_next_release(arc_url: str) -> Optional[date]:
    """ARCファイルから「Money supply」の次回発表日を返す"""
    try:
        resp = requests.get(arc_url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        content = resp.content
        today = datetime.now(JST).date()

        if arc_url.lower().endswith(".xlsx"):
            import openpyxl
            wb_xl = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws_xl = wb_xl.active
            rows = [list(r) for r in ws_xl.iter_rows(values_only=True)]
        else:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=content)
            ws_xls = wb_xls.sheet_by_index(0)
            rows = []
            for r in range(ws_xls.nrows):
                row = []
                for c in range(ws_xls.ncols):
                    val = ws_xls.cell_value(r, c)
                    if ws_xls.cell_type(r, c) == 3 and val > 40000:
                        val = xlrd.xldate_as_datetime(val, wb_xls.datemode).date()
                    row.append(val)
                rows.append(row)

        # "Depository corporations survey" / "Money supply" / "货币供应量" 行を探す
        # ARCファイルでは "Depository corporations survey:" という表記を使用
        ms_row_idx = None
        for i, row in enumerate(rows):
            first = str(row[0]).strip().strip("'") if row[0] else ""
            if (
                "Depository corporations survey" in first
                or "Money supply" in first
                or "货币供应量" in first
            ):
                ms_row_idx = i
                break

        if ms_row_idx is None:
            logger.debug(f"[M1M2] ARC: indicator row not found in {len(rows)} rows")
            return None

        # "Release Date" 行を探す（指標行の後5行以内）
        for i in range(ms_row_idx + 1, min(ms_row_idx + 6, len(rows))):
            first = str(rows[i][0]).strip().strip("'") if rows[i][0] else ""
            if "Release Date" in first:
                release_dates = []
                for cell in rows[i][1:]:
                    if isinstance(cell, date):
                        release_dates.append(cell)
                future = [d for d in release_dates if d >= today]
                if future:
                    return min(future)
                break

    except Exception as e:
        logger.warning(f"[M1M2] ARC parse error: {e}")

    return None


def get_next_release() -> Optional[date]:
    """次回発表日を取得"""
    year_map = _discover_year_urls(min_year=datetime.now().year)
    current_year = datetime.now().year
    overview_url = year_map.get(current_year) or year_map.get(current_year - 1)
    if overview_url is None:
        return None

    soup = _fetch_soup(overview_url)
    if soup is None:
        return None

    arc_url = _get_arc_url(soup)
    if arc_url is None:
        return None

    return _parse_arc_next_release(arc_url)


# =============================================================================
# 全データ取得・マージ
# =============================================================================

def _normalize_record(rec: Dict) -> Dict:
    """データレコードを正規化（m1_yoyキーを統一）"""
    return {
        "m2": rec.get("m2"),
        "m1": rec.get("m1"),
        "m2_yoy": rec.get("m2_yoy"),
        "m1_yoy": rec.get("m1_yoy"),
    }


def fetch_all_data() -> Dict[str, Dict]:
    """
    親ページから全年データを取得してマージ
    Returns: {date_str: {"date", "m2", "m1", "m2_yoy", "m1_yoy"}, ...}
    """
    year_map = _discover_year_urls(min_year=MIN_YEAR)
    if not year_map:
        logger.error("[M1M2] fetch_all_data: no years discovered")
        return {}

    all_data: Dict[str, Dict] = {}

    for year in sorted(year_map.keys()):
        overview_url = year_map[year]
        soup = _fetch_soup(overview_url)
        if soup is None:
            logger.warning(f"[M1M2] Could not fetch overview for year {year}")
            continue

        file_url = _find_m1m2_link(soup)
        if not file_url:
            logger.warning(f"[M1M2] No M1/M2 file for year {year}")
            continue

        logger.info(f"[M1M2] Fetching year {year}: {file_url}")
        parsed = _download_and_parse(file_url)

        for date_str, rec in parsed.items():
            if date_str not in all_data:
                all_data[date_str] = {"date": date_str, "m2": None, "m1": None, "m2_yoy": None, "m1_yoy": None, "m1_restated": None}
            # 新しい値で上書き（Noneは上書きしない）
            for k in ["m2", "m1", "m2_yoy", "m1_yoy", "m1_restated"]:
                if rec.get(k) is not None:
                    all_data[date_str][k] = rec[k]

        logger.info(f"[M1M2] Year {year}: {len(parsed)} records parsed")

    # 水準値から前年比を補完計算
    all_data = _compute_yoy(all_data)

    return all_data


def fetch_latest_data() -> Dict[str, Dict]:
    """当年（+ 年越し直後は前年も）のデータのみ取得"""
    current_year = datetime.now().year
    today = datetime.now(JST).date()
    target_years = {current_year - 1, current_year} if today.month <= 2 else {current_year}
    min_year = min(target_years)

    year_map = _discover_year_urls(min_year=min_year)
    if not year_map:
        return {}

    result: Dict[str, Dict] = {}

    for year in sorted(year_map.keys()):
        if year not in target_years:
            continue
        overview_url = year_map[year]
        soup = _fetch_soup(overview_url)
        if soup is None:
            continue

        file_url = _find_m1m2_link(soup)
        if not file_url:
            continue

        logger.info(f"[M1M2] Fetching latest year {year}: {file_url}")
        parsed = _download_and_parse(file_url)

        for date_str, rec in parsed.items():
            if date_str not in result:
                result[date_str] = {"date": date_str, "m2": None, "m1": None, "m2_yoy": None, "m1_yoy": None, "m1_restated": None}
            for k in ["m2", "m1", "m2_yoy", "m1_yoy", "m1_restated"]:
                if rec.get(k) is not None:
                    result[date_str][k] = rec[k]

    return result


# =============================================================================
# サービスクラス
# =============================================================================

class ChM1M2Service:
    """中国 M1/M2 貨幣供応量サービス"""

    def __init__(self):
        self._redis = None

    def _get_redis(self):
        if self._redis is None:
            try:
                import redis
                self._redis = redis.Redis(host="localhost", port=6379, db=0, socket_timeout=2)
                self._redis.ping()
            except Exception:
                self._redis = None
        return self._redis

    def _from_redis(self) -> Optional[Dict[str, Any]]:
        r = self._get_redis()
        if not r:
            return None
        try:
            raw = r.get(REDIS_KEY)
            if raw:
                return json.loads(raw)
        except Exception:
            pass
        return None

    def _to_redis(self, payload: Dict[str, Any]) -> None:
        r = self._get_redis()
        if not r:
            return
        try:
            r.setex(REDIS_KEY, REDIS_TTL, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _from_file(self) -> Optional[Dict[str, Any]]:
        if not os.path.exists(FILE_CACHE_PATH):
            return None
        try:
            with open(FILE_CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _to_file(self, payload: Dict[str, Any]) -> None:
        os.makedirs(_FILE_CACHE_DIR, exist_ok=True)
        try:
            with open(FILE_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"[M1M2] File cache write error: {e}")

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """
        次回発表日を取得。FMP（CN）を優先し、失敗時はPBOC ARCにフォールバック。
        FMPレスポンス例:
          {'date': '2026-03-14', 'datetime_jst': '...', 'time_jst': '18:00', 'label': 'M2 Money Supply YoY (Feb)'}
        """
        # 1. FMPから取得（CN対応）
        try:
            from services.usa.fmp_next_release_utils import get_next_release_from_fmp
            result = get_next_release_from_fmp("cn_m1_m2", country="CN")
            if result:
                return result
        except Exception as e:
            logger.debug(f"[M1M2] FMP next_release failed: {e}")

        # 2. PBOC ARC にフォールバック
        try:
            next_release_date = get_next_release()
            if next_release_date:
                return {
                    "date": next_release_date.isoformat(),
                    "label": f"M1/M2 Money Supply ({next_release_date.strftime('%b')})",
                }
        except Exception as e:
            logger.debug(f"[M1M2] ARC next_release failed: {e}")

        return None

    def _build_payload(self, data_map: Dict[str, Dict]) -> Dict[str, Any]:
        # m1_restated はフロントエンドに不要な内部計算用フィールドのため除外
        clean_records = []
        for rec in sorted(data_map.values(), key=lambda x: x["date"]):
            clean_records.append({
                "date": rec["date"],
                "m2": rec.get("m2"),
                "m1": rec.get("m1"),
                "m2_yoy": rec.get("m2_yoy"),
                "m1_yoy": rec.get("m1_yoy"),
            })
        records = clean_records
        latest = records[-1] if records else None

        # FMPから次回発表日を取得（CN対応）
        next_release = self._get_next_release()

        return {
            "data": records,
            "latest": latest,
            "metadata": {
                "unit": "億元",
                "unit_yoy": "%",
                "indicator": "M1/M2 Money Supply",
                "source": "PBOC",
                "total_records": len(records),
                "last_fetched": datetime.now(JST).isoformat(),
            },
            "next_release": next_release,
        }

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """M1/M2 データを取得（キャッシュ優先）"""
        if not force_refresh:
            cached = self._from_redis()
            if cached:
                return cached
            cached = self._from_file()
            if cached:
                self._to_redis(cached)
                return cached

        data_map = fetch_all_data()
        if not data_map:
            cached = self._from_file()
            if cached:
                return cached
            return {"data": [], "latest": None, "metadata": {}, "next_release": None}

        payload = self._build_payload(data_map)
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def update_latest(self) -> Dict[str, Any]:
        """当年データを取得して既存キャッシュとマージ → キャッシュ再構築"""
        existing_payload = self._from_redis() or self._from_file()
        existing_map: Dict[str, Dict] = {}
        if existing_payload:
            for rec in existing_payload.get("data", []):
                existing_map[rec["date"]] = rec

        latest_map = fetch_latest_data()
        for date_str, rec in latest_map.items():
            if date_str not in existing_map:
                existing_map[date_str] = {"date": date_str, "m2": None, "m1": None, "m2_yoy": None, "m1_yoy": None, "m1_restated": None}
            for k in ["m2", "m1", "m2_yoy", "m1_yoy", "m1_restated"]:
                if rec.get(k) is not None:
                    existing_map[date_str][k] = rec[k]

        existing_map = _compute_yoy(existing_map)

        payload = self._build_payload(existing_map)
        self._to_redis(payload)
        self._to_file(payload)

        latest = payload.get("latest") or {}
        m2_val = latest.get("m2")
        m1_val = latest.get("m1")
        logger.info(
            f"[M1M2] Updated: total={len(existing_map)}, "
            f"latest={latest.get('date')}, "
            f"M2={m2_val:.0f}" if m2_val else "[M1M2] Updated"
        )
        return payload

    def invalidate_cache(self) -> Dict[str, Any]:
        r = self._get_redis()
        if r:
            try:
                r.delete(REDIS_KEY)
            except Exception:
                pass
        if os.path.exists(FILE_CACHE_PATH):
            os.remove(FILE_CACHE_PATH)
        return {"success": True, "message": "M1M2 cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        r = self._get_redis()
        redis_exists = False
        redis_ttl = None
        if r:
            try:
                redis_exists = bool(r.exists(REDIS_KEY))
                redis_ttl = r.ttl(REDIS_KEY)
            except Exception:
                pass
        file_exists = os.path.exists(FILE_CACHE_PATH)
        file_mtime = None
        if file_exists:
            file_mtime = datetime.fromtimestamp(
                os.path.getmtime(FILE_CACHE_PATH), tz=JST
            ).isoformat()
        return {
            "redis": {"exists": redis_exists, "ttl_seconds": redis_ttl},
            "file": {"exists": file_exists, "last_modified": file_mtime},
        }


# シングルトン
ch_m1_m2_service = ChM1M2Service()

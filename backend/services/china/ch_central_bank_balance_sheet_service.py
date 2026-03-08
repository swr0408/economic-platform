"""
中国人民銀行（PBOC）中央銀行バランスシート（中銀総資産）サービス

データ方式:
- 親ページ（统计数据）から年別「货币统计概览」URLを動的に発見
- 各年の货币统計概覧ページから「货币当局资产负债表」XLS/XLSXリンクを取得
- 「总资产 Total Assets」行の月次値を抽出
- 単位: 亿元人民币（1億元）

スクレイピング方針:
- URLはハードコードせず、すべて親ページから動的に辿る
- 年が変わっても追従可能

対象: 总资产（Total Assets of Monetary Authority）

データソース: https://camlmac.pbc.gov.cn/diaochatongjisi/116219/116319/index.html

Redisキャッシュ: china:ch_cbs:data (TTL: 24時間)
ファイルキャッシュ: backend/data/cache/china/policy/ch_cbs_cache.json
"""
import io
import json
import os
import re
import logging
import calendar
from datetime import date, datetime
from typing import Optional, Dict, Any, List, Tuple
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
from sqlalchemy import text

try:
    from core.database import SessionLocal
except ImportError:
    from backend.core.database import SessionLocal

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# -----------------------------------------------------------------
# キャッシュ設定
# -----------------------------------------------------------------
REDIS_KEY = "china:ch_cbs:data"
REDIS_TTL = 24 * 3600  # 24 hours

_FILE_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data", "cache", "china", "policy"
)
FILE_CACHE_PATH = os.path.join(_FILE_CACHE_DIR, "ch_cbs_cache.json")

# -----------------------------------------------------------------
# サイト定数
# -----------------------------------------------------------------
PBOC_HOST = "https://camlmac.pbc.gov.cn"
BASE_URL = f"{PBOC_HOST}/diaochatongjisi"

# 親ページ（统计数据インデックス）: 全年の货币统计概览リンクが集約されている
PARENT_INDEX_URL = f"{BASE_URL}/116219/116319/index.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://camlmac.pbc.gov.cn/",
}

# 総資産キーワード
TOTAL_ASSETS_KEYWORDS = ["总资产", "Total Assets"]


# =============================================================================
# XLS/XLSX パーサー
# =============================================================================

def _parse_date_str(date_str: Any) -> Optional[date]:
    """
    "2026.01" / "2026.1" / 2026.01 (float) などを date(2026, 1, 31) に変換
    月末日を使用（月次系列の標準）
    """
    if date_str is None:
        return None
    s = str(date_str).strip()
    # "2026.01" or "2026.1"
    m = re.match(r'^(\d{4})\.(\d{1,2})$', s)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        _, last_day = calendar.monthrange(year, month)
        return date(year, month, last_day)
    return None


def _parse_xlsx(content: bytes) -> List[Tuple[date, float]]:
    """XLSXファイルから (date, total_assets) リストを返す"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # Row 6 (1-indexed) = 日付ヘッダー行、Row 18 = 总资产
        # ただしファイルによって1-2行ずれる場合があるため動的検索
        header_row = None
        assets_row = None

        for row_idx in range(1, min(ws.max_row + 1, 30)):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a is None:
                continue
            cell_a_str = str(cell_a).strip()
            # ヘッダー行: 1列目が空で 2列目以降に "20XX.XX" パターン
            if header_row is None:
                for col_idx in range(2, min(ws.max_column + 1, 20)):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break
            # 总資産行
            if assets_row is None:
                for kw in TOTAL_ASSETS_KEYWORDS:
                    if kw in cell_a_str:
                        assets_row = row_idx
                        break

        if header_row is None or assets_row is None:
            logger.warning(f"[CBS] XLSX: header_row={header_row}, assets_row={assets_row} - parse failed")
            return []

        results = []
        for col_idx in range(2, ws.max_column + 1):
            date_val = ws.cell(row=header_row, column=col_idx).value
            asset_val = ws.cell(row=assets_row, column=col_idx).value
            d = _parse_date_str(date_val)
            if d and asset_val is not None:
                try:
                    results.append((d, float(asset_val)))
                except (ValueError, TypeError):
                    pass

        return results

    except Exception as e:
        logger.warning(f"[CBS] XLSX parse error: {e}")
        return []


def _parse_xls(content: bytes) -> List[Tuple[date, float]]:
    """XLSファイルから (date, total_assets) リストを返す"""
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)

        header_row = None
        assets_row = None

        for row_idx in range(0, min(ws.nrows, 30)):
            cell_a = ws.cell_value(row_idx, 0)
            cell_a_str = str(cell_a).strip() if cell_a else ""
            # ヘッダー行
            if header_row is None:
                for col_idx in range(1, min(ws.ncols, 20)):
                    v = ws.cell_value(row_idx, col_idx)
                    if v and re.match(r'^\d{4}\.\d{1,2}$', str(v).strip()):
                        header_row = row_idx
                        break
            # 总資産行
            if assets_row is None:
                for kw in TOTAL_ASSETS_KEYWORDS:
                    if kw in cell_a_str:
                        assets_row = row_idx
                        break

        if header_row is None or assets_row is None:
            logger.warning(f"[CBS] XLS: header_row={header_row}, assets_row={assets_row} - parse failed")
            return []

        results = []
        for col_idx in range(1, ws.ncols):
            date_val = ws.cell_value(header_row, col_idx)
            asset_val = ws.cell_value(assets_row, col_idx)
            d = _parse_date_str(date_val)
            if d and asset_val:
                try:
                    results.append((d, float(asset_val)))
                except (ValueError, TypeError):
                    pass

        return results

    except Exception as e:
        logger.warning(f"[CBS] XLS parse error: {e}")
        return []


def _download_and_parse(url: str) -> List[Tuple[date, float]]:
    """URLをダウンロードしてXLS/XLSXをパース"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        content = resp.content
        if url.lower().endswith('.xlsx'):
            return _parse_xlsx(content)
        else:
            return _parse_xls(content)
    except Exception as e:
        logger.warning(f"[CBS] Download failed {url}: {e}")
        return []


# =============================================================================
# スクレイピング: 親ページから動的に年別URLを発見
# =============================================================================

def _abs_url(href: str) -> str:
    """相対hrefを絶対URLに変換"""
    if href.startswith("http"):
        return href
    elif href.startswith("/"):
        return f"{PBOC_HOST}{href}"
    else:
        return f"{BASE_URL}/{href}"


def _fetch_soup(url: str, timeout: int = 15) -> Optional[BeautifulSoup]:
    """URLをフェッチしてBeautifulSoupを返す。失敗時はNone"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code != 200:
            return None
        resp.encoding = "utf-8"
        return BeautifulSoup(resp.text, "html.parser")
    except Exception as e:
        logger.debug(f"[CBS] Fetch failed {url}: {e}")
        return None


def _discover_year_urls(min_year: int = 2015) -> Dict[int, str]:
    """
    親インデックスページ（PARENT_INDEX_URL）をスクレイプして、
    各年の「货币统计概览」ページURLを動的に発見する。

    親ページには以下の形式でリンクが並んでいる:
      text="2026年统计数据" href="/.../{year_id}/index.html"
      text="货币统计概览"   href="/.../{year_id}/{sub_id}/index.html"
      text="2025年统计数据" href="/.../{year_id}/index.html"
      text="货币统计概览"   href="/.../{year_id}/{sub_id}/index.html"
      ...

    Returns: {year: hbtjgl_url, ...}
    """
    soup = _fetch_soup(PARENT_INDEX_URL)
    if soup is None:
        logger.warning("[CBS] Failed to fetch parent index page")
        return {}

    year_map: Dict[int, str] = {}
    current_year: Optional[int] = None

    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True)
        href = a["href"]

        # "2026年统计数据" のようなリンクで現在年を更新
        m = re.match(r'^(\d{4})年统计数据$', text)
        if m:
            current_year = int(m.group(1))
            continue

        # 直後の "货币统计概览" リンクをその年のURLとして記録
        if text == "货币统计概览" and current_year is not None:
            if current_year >= min_year and current_year not in year_map:
                year_map[current_year] = _abs_url(href)

    logger.info(f"[CBS] Discovered {len(year_map)} years from parent page: {sorted(year_map.keys())}")
    return year_map


def _find_balance_sheet_link(soup: BeautifulSoup) -> Optional[str]:
    """
    货币统计概览ページから「货币当局资产负债表」XLS/XLSXリンクを取得する。

    ページ構造:
      <tr>
        <td><div class="titp20">货币当局资产负债表<br/>Balance Sheet of Monetary Authority</div></td>
        <td><a href="...htm">htm</a></td>
        <td><a href="...xlsx">xls</a></td>  ← これを取得
        <td><a href="...pdf">pdf</a></td>
      </tr>

    注意: ネストされたテーブルがあるため、div.titp20 で正確に対象TRを特定する。
    """
    TARGET_KEYWORDS = ["货币当局资产负债表", "Balance Sheet of Monetary Authority"]

    for div in soup.find_all("div", class_="titp20"):
        if not any(kw in div.get_text() for kw in TARGET_KEYWORDS):
            continue

        tr = div.find_parent("tr")
        if tr is None:
            continue

        for a_tag in tr.find_all("a", href=True):
            href = a_tag["href"]
            if href.lower().endswith(".xlsx") or href.lower().endswith(".xls"):
                return _abs_url(href)

    return None


def _get_balance_sheet_url_for_year(hbtjgl_url: str) -> Optional[str]:
    """
    货币统计概览ページURLを受け取り、货币当局资产负债表のXLS/XLSXリンクを返す
    """
    soup = _fetch_soup(hbtjgl_url)
    if soup is None:
        return None
    return _find_balance_sheet_link(soup)


def _get_arc_url(hbtjgl_url: str) -> Optional[str]:
    """
    货币统計概覧ページから「公布日程预告（ARC）」XLSリンクを返す。

    ページ最下部の「公布日程预告 Advance Release Calendar (ARC)」行のXLSリンクを探す。
    """
    ARC_KEYWORDS = ["公布日程预告", "Advance Release Calendar", "ARC"]

    soup = _fetch_soup(hbtjgl_url)
    if soup is None:
        return None

    for div in soup.find_all("div", class_="titp20"):
        div_text = div.get_text()
        if not any(kw in div_text for kw in ARC_KEYWORDS):
            continue
        tr = div.find_parent("tr")
        if tr is None:
            continue
        for a_tag in tr.find_all("a", href=True):
            href = a_tag["href"]
            if href.lower().endswith(".xlsx") or href.lower().endswith(".xls"):
                return _abs_url(href)

    return None


def _parse_arc_next_release(arc_url: str) -> Optional[date]:
    """
    ARCファイルから「Central bank survey」の次回発表日（今日以降の最初の日付）を返す。

    ARCファイル構造:
      Row 14: "Central bank survey:"
      Row 16: "Month" + 日付列（Excelシリアル）
      Row 17: "Release Date" + 日付列（Excelシリアル）
      Row 18: "Reference Period" + 日付列（Excelシリアル）
    """
    try:
        resp = requests.get(arc_url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        content = resp.content

        # XLSまたはXLSX両対応
        if arc_url.lower().endswith(".xlsx"):
            import openpyxl
            wb_xl = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws_xl = wb_xl.active
            # openpyxlは日付セルをdatetimeに自動変換する
            rows = []
            for row in ws_xl.iter_rows(values_only=True):
                rows.append(list(row))
        else:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=content)
            ws_xls = wb_xls.sheet_by_index(0)
            rows = []
            for r in range(ws_xls.nrows):
                row = []
                for c in range(ws_xls.ncols):
                    val = ws_xls.cell_value(r, c)
                    if ws_xls.cell_type(r, c) == 3 and val > 40000:
                        val = xlrd.xldate_as_datetime(val, wb_xls.datemode).date()
                    row.append(val)
                rows.append(row)

        today = datetime.now(JST).date()

        # "Central bank survey:" 行を探す
        cbs_row_idx = None
        for i, row in enumerate(rows):
            first = str(row[0]).strip().strip("'") if row[0] else ""
            if "Central bank survey" in first:
                cbs_row_idx = i
                break

        if cbs_row_idx is None:
            return None

        # その後の "Release Date" 行を探す（最大5行以内）
        for i in range(cbs_row_idx + 1, min(cbs_row_idx + 6, len(rows))):
            first = str(rows[i][0]).strip().strip("'") if rows[i][0] else ""
            if "Release Date" in first:
                # 列1以降から日付を取得
                release_dates = []
                for cell in rows[i][1:]:
                    if isinstance(cell, date):
                        release_dates.append(cell)
                    elif isinstance(cell, float) and cell > 40000:
                        # xlrd未変換の場合のフォールバック
                        from datetime import datetime as dt
                        release_dates.append(
                            dt.fromordinal(int(cell) + 693594).date()
                        )
                # 今日以降の最初の日付
                future = [d for d in release_dates if d >= today]
                if future:
                    return min(future)
                break

    except Exception as e:
        logger.warning(f"[CBS] ARC parse error: {e}")

    return None


def get_next_release() -> Optional[date]:
    """
    最新年の货币统計概覧ページからARCを取得し、次回発表日を返す。
    """
    year_map = _discover_year_urls(min_year=datetime.now().year)
    current_year = datetime.now().year
    hbtjgl_url = year_map.get(current_year)
    if hbtjgl_url is None:
        # 前年を試す
        year_map = _discover_year_urls(min_year=current_year - 1)
        hbtjgl_url = year_map.get(current_year - 1)
    if hbtjgl_url is None:
        return None

    arc_url = _get_arc_url(hbtjgl_url)
    if arc_url is None:
        return None

    return _parse_arc_next_release(arc_url)


# =============================================================================
# DB操作
# =============================================================================

def _get_db_dates() -> set:
    """pbc_balance_sheet の既存日付セットを取得"""
    with SessionLocal() as session:
        rows = session.execute(
            text("SELECT date FROM pbc_balance_sheet")
        ).fetchall()
    return {row[0] for row in rows}


def _get_latest_db_date() -> Optional[date]:
    """pbc_balance_sheet の最終日付を取得"""
    with SessionLocal() as session:
        row = session.execute(
            text("SELECT MAX(date) FROM pbc_balance_sheet")
        ).fetchone()
    return row[0] if row and row[0] else None


def _upsert_records(records: List[Tuple[date, float]]) -> int:
    """
    pbc_balance_sheet にレコードをupsert
    Returns: 新規挿入件数
    """
    if not records:
        return 0
    inserted = 0
    with SessionLocal() as session:
        for d, val in records:
            result = session.execute(
                text("""
                    INSERT INTO pbc_balance_sheet (date, total_assets)
                    VALUES (:d, :v)
                    ON CONFLICT (date) DO UPDATE SET total_assets = EXCLUDED.total_assets
                    RETURNING (xmax = 0) AS is_insert
                """),
                {"d": d, "v": val},
            )
            row = result.fetchone()
            if row and row[0]:
                inserted += 1
        session.commit()
    return inserted


def _load_all_from_db() -> List[Dict[str, Any]]:
    """pbc_balance_sheet から全データを取得"""
    with SessionLocal() as session:
        rows = session.execute(
            text("SELECT date, total_assets FROM pbc_balance_sheet ORDER BY date ASC")
        ).fetchall()
    result = []
    for row in rows:
        result.append({
            "date": row[0].strftime("%Y-%m-%d"),
            "value": float(row[1]) if row[1] is not None else None,
        })
    return result


# =============================================================================
# 全データ取得（初回/強制更新）
# =============================================================================

def fetch_all_years() -> int:
    """
    親ページから全年のURLを動的発見してDBに保存
    Returns: 新規挿入件数
    """
    year_map = _discover_year_urls(min_year=2015)
    if not year_map:
        logger.error("[CBS] fetch_all_years: no years discovered")
        return 0

    total_inserted = 0
    for year in sorted(year_map.keys()):
        hbtjgl_url = year_map[year]
        file_url = _get_balance_sheet_url_for_year(hbtjgl_url)
        if not file_url:
            logger.warning(f"[CBS] No file URL found for year {year} ({hbtjgl_url})")
            continue

        logger.info(f"[CBS] Fetching year {year}: {file_url}")
        records = _download_and_parse(file_url)
        if records:
            n = _upsert_records(records)
            total_inserted += n
            logger.info(f"[CBS] Year {year}: {len(records)} records parsed, {n} inserted")
        else:
            logger.warning(f"[CBS] Year {year}: no records parsed from {file_url}")

    return total_inserted


def fetch_latest() -> int:
    """
    当年のデータのみ更新（月次スケジューラー用）。

    効率化方針:
    - DBの最終日付を確認し、当年のXLSから新しい月のみupsert
    - 年越し直後（1月）は前年も取得して取りこぼしを防ぐ
    - 毎回全年を再取得しない

    Returns: 新規/更新件数
    """
    current_year = datetime.now().year
    today = datetime.now(JST).date()

    # DBの最終日付を確認
    latest_db_date = _get_latest_db_date()

    # 対象年の決定: 年越し直後(1-2月)は前年も確認
    if today.month <= 2:
        target_years = {current_year - 1, current_year}
        min_year = current_year - 1
    else:
        target_years = {current_year}
        min_year = current_year

    year_map = _discover_year_urls(min_year=min_year)
    if not year_map:
        logger.error("[CBS] fetch_latest: no years discovered")
        return 0

    total_inserted = 0
    for year in sorted(year_map.keys()):
        if year not in target_years:
            continue
        hbtjgl_url = year_map[year]
        file_url = _get_balance_sheet_url_for_year(hbtjgl_url)
        if not file_url:
            logger.warning(f"[CBS] No file URL for year {year}")
            continue
        logger.info(f"[CBS] Updating year {year}: {file_url}")
        records = _download_and_parse(file_url)
        if records:
            # DBに既にある日付はupsertで上書き（値変更がある場合も対応）
            n = _upsert_records(records)
            total_inserted += n
            if n > 0:
                logger.info(f"[CBS] Year {year}: {n} records new/updated")
            else:
                logger.info(f"[CBS] Year {year}: no new records (latest DB: {latest_db_date})")
    return total_inserted


# =============================================================================
# サービスクラス
# =============================================================================

class ChCentralBankBalanceSheetService:
    """中国人民銀行 中央銀行バランスシート（総資産）サービス"""

    def __init__(self):
        self._redis = None

    def _get_redis(self):
        if self._redis is None:
            try:
                import redis
                self._redis = redis.Redis(host="localhost", port=6379, db=0, socket_timeout=2)
                self._redis.ping()
            except Exception:
                self._redis = None
        return self._redis

    def _from_redis(self) -> Optional[Dict[str, Any]]:
        r = self._get_redis()
        if not r:
            return None
        try:
            raw = r.get(REDIS_KEY)
            if raw:
                return json.loads(raw)
        except Exception:
            pass
        return None

    def _to_redis(self, payload: Dict[str, Any]) -> None:
        r = self._get_redis()
        if not r:
            return
        try:
            r.setex(REDIS_KEY, REDIS_TTL, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _from_file(self) -> Optional[Dict[str, Any]]:
        if not os.path.exists(FILE_CACHE_PATH):
            return None
        try:
            with open(FILE_CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None

    def _to_file(self, payload: Dict[str, Any]) -> None:
        os.makedirs(_FILE_CACHE_DIR, exist_ok=True)
        try:
            with open(FILE_CACHE_PATH, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"[CBS] File cache write error: {e}")

    def _build_payload(self) -> Dict[str, Any]:
        records = _load_all_from_db()
        latest = records[-1] if records else None

        # 次回発表日を取得（ARCより）
        next_release_date = get_next_release()
        next_release = None
        if next_release_date:
            next_release = {
                "date": next_release_date.isoformat(),
                "label": f"Central Bank Balance Sheet ({next_release_date.strftime('%b')})",
            }

        return {
            "data": records,
            "latest": latest,
            "metadata": {
                "unit": "億元",
                "indicator": "Total Assets of Monetary Authority",
                "source": "PBOC",
                "total_records": len(records),
                "last_fetched": datetime.now(JST).isoformat(),
            },
            "next_release": next_release,
        }

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """バランスシートデータを取得（キャッシュ優先）"""
        if not force_refresh:
            cached = self._from_redis()
            if cached:
                return cached
            cached = self._from_file()
            if cached:
                self._to_redis(cached)
                return cached

        payload = self._build_payload()
        self._to_redis(payload)
        self._to_file(payload)
        return payload

    def invalidate_cache(self) -> Dict[str, Any]:
        r = self._get_redis()
        if r:
            try:
                r.delete(REDIS_KEY)
            except Exception:
                pass
        if os.path.exists(FILE_CACHE_PATH):
            os.remove(FILE_CACHE_PATH)
        return {"success": True, "message": "CBS cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        r = self._get_redis()
        redis_exists = False
        redis_ttl = None
        if r:
            try:
                redis_exists = bool(r.exists(REDIS_KEY))
                redis_ttl = r.ttl(REDIS_KEY)
            except Exception:
                pass
        file_exists = os.path.exists(FILE_CACHE_PATH)
        file_mtime = None
        if file_exists:
            file_mtime = datetime.fromtimestamp(
                os.path.getmtime(FILE_CACHE_PATH), tz=JST
            ).isoformat()
        return {
            "redis": {"exists": redis_exists, "ttl_seconds": redis_ttl},
            "file": {"exists": file_exists, "last_modified": file_mtime},
        }


# シングルトン
ch_central_bank_balance_sheet_service = ChCentralBankBalanceSheetService()

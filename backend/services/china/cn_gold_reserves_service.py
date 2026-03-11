"""
中国人民銀行 金準備（Official Reserve Assets - Gold）サービス

データソース:
  過去データ: PBOC HTM → CSV → DB
  最新値: SAFE Excel（https://www.safe.gov.cn/en/ForexReserves/index.html）
  Row 14 = 4.黄金 (亿美元), Row 16 = 万盎司

取得方法:
  1. SAFE ForexReserves 一覧ページから当年のExcel URLを動的取得
  2. Excelダウンロード → openpyxl で Row 14(USD), Row 16(万盎司) をパース
  3. DB未登録の新月分をUPSERT

発表スケジュール: 月次（毎月7日前後、PDFからパース）
キャッシュ: Redis + ファイルキャッシュ
"""
import io
import json
import logging
import re
from datetime import date, datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
from zoneinfo import ZoneInfo
import glob as glob_module

import requests
import openpyxl

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
REDIS_KEY = "china:cn_gold_reserves:data"
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "china" / "policy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
FILE_CACHE = CACHE_DIR / "cn_gold_reserves_service_cache.json"

# PDF ディレクトリ
PDF_DIR = Path(__file__).parent.parent.parent / "data" / "pdf" / "china"

# SAFE ページURL
SAFE_INDEX_URL = "https://www.safe.gov.cn/en/ForexReserves/index.html"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
}

# 2026年の発表スケジュール（フォールバック: 官方储备资产 = 毎月7日）
FALLBACK_SCHEDULE_2026 = [
    "2026-02-07", "2026-03-07", "2026-04-07", "2026-05-07",
    "2026-06-07", "2026-07-07", "2026-08-07", "2026-09-07",
    "2026-10-07", "2026-11-07", "2026-12-07",
]


# =============================================================================
# SAFE Excel 取得・パース
# =============================================================================

def _discover_current_year_page_url() -> Optional[str]:
    """SAFE一覧ページから当年のOfficial Reserve Assetsページ URLを取得"""
    try:
        resp = requests.get(SAFE_INDEX_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        current_year = date.today().year
        pattern = re.compile(
            rf'href=["\']([^"\']*\.html)["\'][^>]*>[^<]*Official Reserve Assets[^<]*\({current_year}\)',
            re.IGNORECASE,
        )
        m = pattern.search(resp.text)
        if m:
            href = m.group(1)
            if href.startswith("/"):
                href = "https://www.safe.gov.cn" + href
            logger.info(f"[GoldReserves] Found {current_year} page: {href}")
            return href
    except Exception as e:
        logger.warning(f"[GoldReserves] Failed to scrape SAFE index: {e}")
    return None


def _discover_excel_url(page_url: str) -> Optional[str]:
    """年別ページからExcel (.xlsx) URLを取得"""
    try:
        resp = requests.get(page_url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        xlsx_links = re.findall(r'/en/file/file/\d+/[a-f0-9]+\.xlsx', resp.text)
        if xlsx_links:
            url = "https://www.safe.gov.cn" + xlsx_links[0]
            logger.info(f"[GoldReserves] Found Excel URL: {url[:80]}...")
            return url
    except Exception as e:
        logger.warning(f"[GoldReserves] Failed to get Excel URL: {e}")
    return None


def _parse_safe_excel(excel_bytes: bytes) -> List[Dict[str, Any]]:
    """SAFE Excel から金準備データを抽出

    構造:
      Row 4: ヘッダー (2026.01, 2026.02, ...) ← col B, D, F, ...
      Row 14: 4.黄金 (亿美元) ← col B=Jan USD, C=Jan SDR, D=Feb USD, ...
      Row 16: 万盎司 ← same col layout
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        logger.error(f"[GoldReserves] Excel open error: {e}")
        return []

    # Row 4 からヘッダー（月）を取得
    header_row = list(ws[4])
    col_months: Dict[int, str] = {}
    for cell in header_row:
        if cell.value is None:
            continue
        val = str(cell.value).strip()
        m = re.match(r"(\d{4})\.(\d{1,2})", val)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            if 1 <= mo <= 12:
                col_months[cell.column] = f"{y}-{mo:02d}-01"

    if not col_months:
        logger.error("[GoldReserves] No date columns found in Excel header")
        return []

    # Row 14: 黄金 (亿美元) — 各月は2列ペア (USD, SDR)
    # Row 16: 万盎司
    gold_row = list(ws[14])
    ounce_row = list(ws[16])

    records = []
    sorted_months = sorted(col_months.items(), key=lambda x: x[0])

    for header_col, date_str in sorted_months:
        # header_col は B=2, D=4, F=6, ... (1-indexed)
        # USD列 = header_col, SDR列 = header_col + 1
        usd_val = None
        ounce_val = None

        # Excel col is 1-indexed
        usd_cell = ws.cell(row=14, column=header_col)
        if usd_cell.value is not None:
            try:
                raw = str(usd_cell.value).strip().replace("\xa0", "")
                usd_val = float(raw)
            except (ValueError, TypeError):
                pass

        ounce_cell = ws.cell(row=16, column=header_col)
        if ounce_cell.value is not None:
            raw = str(ounce_cell.value).strip().replace("\xa0", "")
            m = re.search(r"([\d.]+)", raw)
            if m:
                ounce_val = float(m.group(1))

        if usd_val is not None or ounce_val is not None:
            records.append({
                "date": date_str,
                "gold_ounces_wan": ounce_val,
                "gold_value_usd_yi": usd_val,
            })

    logger.info(f"[GoldReserves] Parsed {len(records)} months from SAFE Excel")
    return records


def fetch_latest_from_safe() -> List[Dict[str, Any]]:
    """SAFEから最新データを取得"""
    page_url = _discover_current_year_page_url()
    if not page_url:
        return []

    excel_url = _discover_excel_url(page_url)
    if not excel_url:
        return []

    try:
        resp = requests.get(excel_url, headers=HEADERS, timeout=60)
        if resp.status_code != 200:
            logger.error(f"[GoldReserves] Excel download failed: HTTP {resp.status_code}")
            return []
        return _parse_safe_excel(resp.content)
    except Exception as e:
        logger.error(f"[GoldReserves] Excel download error: {e}")
        return []


def save_new_records_to_db(records: List[Dict[str, Any]]) -> int:
    """DB未登録の新しいレコードをUPSERT"""
    if not records:
        return 0

    try:
        from core.database import SessionLocal
    except ImportError:
        from backend.core.database import SessionLocal
    from sqlalchemy import text

    count = 0
    with SessionLocal() as session:
        for rec in records:
            result = session.execute(
                text("SELECT 1 FROM cn_gold_reserves WHERE date = :date"),
                {"date": rec["date"]},
            ).fetchone()

            if result is None:
                session.execute(
                    text("""
                        INSERT INTO cn_gold_reserves
                            (date, gold_ounces_wan, gold_value_usd_yi, source)
                        VALUES
                            (:date, :gold_ounces_wan, :gold_value_usd_yi, 'SAFE')
                        ON CONFLICT (date) DO UPDATE SET
                            gold_ounces_wan = EXCLUDED.gold_ounces_wan,
                            gold_value_usd_yi = EXCLUDED.gold_value_usd_yi,
                            updated_at = NOW()
                    """),
                    rec,
                )
                count += 1
        session.commit()

    return count


# =============================================================================
# DB読み込み
# =============================================================================

def _load_from_db() -> List[Dict[str, Any]]:
    """DBから全レコードを読み込み"""
    try:
        from core.database import SessionLocal
    except ImportError:
        from backend.core.database import SessionLocal
    from sqlalchemy import text

    with SessionLocal() as session:
        rows = session.execute(
            text("""
                SELECT date, gold_ounces_wan, gold_value_usd_yi
                FROM cn_gold_reserves
                ORDER BY date
            """)
        ).fetchall()

    return [
        {
            "date": row[0].strftime("%Y-%m-%d") if hasattr(row[0], "strftime") else str(row[0]),
            "gold_ounces_wan": row[1],
            "gold_value_usd_yi": row[2],
        }
        for row in rows
    ]


# =============================================================================
# PDF スケジュールパース
# =============================================================================

def _parse_schedule_from_pdf() -> List[str]:
    """PDFからリリーススケジュールを抽出（官方储备资产行）"""
    try:
        import pdfplumber
    except ImportError:
        logger.warning("[GoldReserves] pdfplumber not available, using fallback schedule")
        return FALLBACK_SCHEDULE_2026

    pattern_str = str(PDF_DIR / "中国資本フロースケジュール*.pdf")
    pdf_files = sorted(glob_module.glob(pattern_str))
    if not pdf_files:
        logger.warning("[GoldReserves] No schedule PDF found, using fallback")
        return FALLBACK_SCHEDULE_2026

    pdf_path = pdf_files[-1]
    year_match = re.search(r"(\d{4})", Path(pdf_path).name)
    year = int(year_match.group(1)) if year_match else date.today().year

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        row_text = " ".join(str(cell or "") for cell in row)
                        if "官方储备资产" not in row_text:
                            continue

                        dates = []
                        day_pattern = re.compile(r"(\d{1,2})/[一二三四五六日]")
                        for i, cell in enumerate(row):
                            if not cell:
                                continue
                            m = day_pattern.search(str(cell))
                            if m:
                                day = int(m.group(1))
                                month = i
                                if 1 <= month <= 12:
                                    dates.append(f"{year}-{month:02d}-{day:02d}")

                        if dates:
                            logger.info(f"[GoldReserves] Extracted {len(dates)} schedule dates from PDF")
                            return dates

        logger.warning("[GoldReserves] Could not extract schedule from PDF")
    except Exception as e:
        logger.warning(f"[GoldReserves] PDF parse error: {e}")

    return FALLBACK_SCHEDULE_2026


_cached_schedule: Optional[List[str]] = None


def _get_schedule() -> List[str]:
    """スケジュール取得（キャッシュ付き）"""
    global _cached_schedule
    if _cached_schedule is None:
        _cached_schedule = _parse_schedule_from_pdf()
    return _cached_schedule


def _get_next_release_date() -> Optional[Dict[str, Any]]:
    """次回発表日を返す"""
    schedule = _get_schedule()
    today_str = date.today().isoformat()

    for date_str in schedule:
        if date_str > today_str:
            return {
                "date": date_str,
                "label": "官方储备资产 (Official Reserve Assets)",
            }
    return None


# =============================================================================
# キャッシュ管理
# =============================================================================

def _load_file_cache() -> Optional[Dict[str, Any]]:
    try:
        if FILE_CACHE.exists():
            with open(FILE_CACHE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return None


def _save_file_cache(payload: Dict[str, Any]) -> None:
    try:
        with open(FILE_CACHE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _should_refresh(last_updated_str: Optional[str]) -> bool:
    """発表スケジュールベースの更新判定"""
    if not last_updated_str:
        return True

    try:
        last_updated = datetime.fromisoformat(last_updated_str)
    except (ValueError, TypeError):
        return True

    schedule = _get_schedule()
    now = datetime.now(JST)

    for date_str in schedule:
        try:
            release_dt = datetime.strptime(date_str, "%Y-%m-%d").replace(
                hour=15, minute=30, tzinfo=JST  # JTC 15:30
            )
        except ValueError:
            continue

        if now >= release_dt and last_updated < release_dt:
            logger.info(f"[GoldReserves] Refresh needed: release={date_str}, last_updated={last_updated_str}")
            return True

    return False


# =============================================================================
# メインサービスクラス
# =============================================================================

class CnGoldReservesService:
    """中国人民銀行 金準備サービス"""

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """金準備データを取得"""
        next_release = _get_next_release_date()

        # 1. Redis キャッシュ
        if not force_refresh:
            try:
                cached = redis_client.get(REDIS_KEY)
                if cached:
                    data = cached if isinstance(cached, dict) else json.loads(cached)
                    last_updated = data.get("last_updated")
                    if last_updated and not _should_refresh(last_updated):
                        data["next_release"] = next_release
                        return data
            except Exception:
                pass

        # 2. ファイルキャッシュ（スケジュール判定付き）
        if not force_refresh:
            fc = _load_file_cache()
            if fc:
                last_updated = fc.get("last_updated")
                if last_updated and not _should_refresh(last_updated):
                    # Redis に再投入して次回は高速
                    self._save_to_redis(fc)
                    fc["next_release"] = next_release
                    return fc

        # 3. DB読み込み（SAFEフェッチより先 - 高速パス）
        records = _load_from_db()
        if records:
            latest = records[-1]
            payload = {
                "data": records,
                "latest": latest,
                "metadata": {
                    "source": "PBOC / SAFE",
                    "indicator": "Official Reserve Assets - Gold",
                    "fields": {
                        "gold_ounces_wan": "万トロイオンス",
                        "gold_value_usd_yi": "億USD (2016年4月以降)",
                    },
                    "frequency": "monthly",
                },
                "next_release": next_release,
                "last_updated": datetime.now(JST).isoformat(),
            }
            self._save_to_cache(payload)

            # 4. SAFE Excelから新データ取得（バックグラウンド的 - DB/キャッシュ後）
            try:
                safe_records = fetch_latest_from_safe()
                if safe_records:
                    new_count = save_new_records_to_db(safe_records)
                    if new_count > 0:
                        logger.info(f"[GoldReserves] Added {new_count} new records from SAFE")
                        # 新データがあればキャッシュを再構築
                        records = _load_from_db()
                        if records:
                            payload["data"] = records
                            payload["latest"] = records[-1]
                            payload["last_updated"] = datetime.now(JST).isoformat()
                            self._save_to_cache(payload)
            except Exception as e:
                logger.warning(f"[GoldReserves] SAFE fetch failed: {e}")

            return payload

        # 5. ファイルキャッシュフォールバック
        fc = _load_file_cache()
        if fc:
            fc["next_release"] = next_release
            return fc

        return {
            "data": [],
            "latest": None,
            "metadata": {"source": "PBOC / SAFE", "error": "No data available"},
            "next_release": next_release,
        }

    def _save_to_redis(self, payload: Dict[str, Any]) -> None:
        try:
            redis_client.set(REDIS_KEY, payload, expire=0)
        except Exception:
            pass

    def _save_to_cache(self, payload: Dict[str, Any]) -> None:
        self._save_to_redis(payload)
        _save_file_cache(payload)

    def invalidate_cache(self) -> Dict[str, Any]:
        try:
            redis_client.delete(REDIS_KEY)
        except Exception:
            pass
        if FILE_CACHE.exists():
            FILE_CACHE.unlink()
        return {"success": True, "message": "Gold Reserves cache invalidated"}

    def get_cache_status(self) -> Dict[str, Any]:
        redis_exists = False
        try:
            redis_exists = redis_client.exists(REDIS_KEY) > 0
        except Exception:
            pass
        fc = _load_file_cache()
        return {
            "indicator": "CN Gold Reserves",
            "source": "PBOC / SAFE",
            "cache_key": REDIS_KEY,
            "redis_cached": redis_exists,
            "file_cached": FILE_CACHE.exists(),
            "file_cache_records": len(fc.get("data", [])) if fc else 0,
            "last_updated": fc.get("last_updated") if fc else None,
            "next_release": _get_next_release_date(),
        }


# シングルトン
cn_gold_reserves_service = CnGoldReservesService()

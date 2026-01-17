"""
BOE CPI Components Service
BOE MPRデータからCPI構成項目インフレ率を取得

データソース: Section 2 - Current economic conditions.xlsx from MPR Charts, Slides and Data ZIP files
- Chart sheets (Chart 2.16 to Chart 2.28)
- "Annual inflation rates for components of CPI" を検索
- サービス、コア商品、食品・非アルコール飲料のデータを抽出

発表スケジュール:
- 2月・5月・8月・11月のMPR発表月
- 発表時刻: 12:00 GMT
- ダウンロードウィンドウ: 20:00-20:10 GMT/BST

キャッシュ方式: ファイルキャッシュ + Redisキャッシュ
"""
import json
import logging
import requests
import io
import zipfile
import openpyxl
import time
from datetime import datetime
from typing import Dict, List, Optional, Any
from zoneinfo import ZoneInfo
from pathlib import Path

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")
LONDON = ZoneInfo("Europe/London")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "uk" / "economic_outlook"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "boe_cpi_components_cache.json"


class BOECPIComponentsService:
    """BOE CPI Components サービス"""

    DATA_CACHE_KEY = "uk:boe_cpi_components:data"

    # Base URL patterns for MPR data files
    MPR_URL_PATTERNS = [
        "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-charts-slides-and-data.zip",
        "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-chart-slides-and-data.zip"
    ]

    # Month mapping
    MONTH_NAMES = {
        1: "january", 2: "february", 3: "march", 4: "april",
        5: "may", 6: "june", 7: "july", 8: "august",
        9: "september", 10: "october", 11: "november", 12: "december"
    }

    # Default MPR publication months
    DEFAULT_MPR_MONTHS = [2, 5, 8, 11]

    # Excel file name patterns in ZIP (BOE changed structure in Nov 2025)
    # Old: "Section 2  - Current economic conditions.xlsx" (note: two spaces between "2" and "-")
    # New: "{Month} {Year} MPR chart data.xlsx"
    EXCEL_FILENAME_OLD = "Section 2  - Current economic conditions.xlsx"
    EXCEL_FILENAME_NEW_PATTERN = "{month} {year} MPR chart data.xlsx"

    # Chart title to look for in cell A3
    CHART_TITLE = "Annual inflation rates for components of CPI"

    # Sheet name prefixes (new format uses Chart 1.xx, old uses Chart 2.xx)
    SHEET_PREFIX_NEW = "Chart 1."
    SHEET_PREFIX_OLD = "Chart 2."

    def __init__(self):
        self.max_retries = 3
        self.retry_delay = 5

    def fetch_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """CPI構成項目データを取得"""
        # Redisキャッシュチェック
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "components": cached_data.get("components", {"data": None}),
                        "metadata": cached_data.get("metadata", {}),
                        "cached": True,
                        "source": "redis",
                    }

        # MPRからデータを取得
        data = self._fetch_cpi_components_data()

        if data:
            cache_payload = {
                "components": data.get("components", {"data": None}),
                "metadata": data.get("metadata", {}),
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "components": data.get("components", {"data": None}),
                "metadata": data.get("metadata", {}),
                "cached": False,
                "source": "boe_mpr",
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "components": file_cache.get("components", {"data": None}),
                "metadata": file_cache.get("metadata", {}),
                "cached": True,
                "source": "file (fallback)",
            }

        return self._get_fallback_data()

    def _get_latest_mpr_info(self, reference_date: Optional[datetime] = None) -> Dict[str, Any]:
        """最新と前回のMPR日を判定"""
        if reference_date is None:
            reference_date = datetime.now()

        current_year = reference_date.year
        current_month = reference_date.month
        current_day = reference_date.day

        mpr_months = self.DEFAULT_MPR_MONTHS

        # MPRは通常月の6日以降に発表
        if current_month in mpr_months and current_day < 6:
            current_month_for_search = current_month - 1
        else:
            current_month_for_search = current_month

        # 最新のMPR月を特定
        latest_mpr_month = None
        for month in reversed(mpr_months):
            if current_month_for_search >= month:
                latest_mpr_month = month
                break

        if latest_mpr_month is None:
            latest_mpr_month = mpr_months[-1]
            current_year -= 1

        # 前回のMPR月を特定
        previous_mpr_month_idx = mpr_months.index(latest_mpr_month) - 1
        if previous_mpr_month_idx < 0:
            previous_mpr_month = mpr_months[-1]
            previous_year = current_year - 1
        else:
            previous_mpr_month = mpr_months[previous_mpr_month_idx]
            previous_year = current_year

        return {
            'latest': {
                'year': current_year,
                'month': latest_mpr_month,
                'month_name': self.MONTH_NAMES[latest_mpr_month]
            },
            'previous': {
                'year': previous_year,
                'month': previous_mpr_month,
                'month_name': self.MONTH_NAMES[previous_mpr_month]
            }
        }

    def _get_previous_previous_mpr_info(self) -> Optional[Dict[str, Any]]:
        """前々回のMPR情報を取得"""
        mpr_info = self._get_latest_mpr_info()
        previous_info = mpr_info['previous']

        mpr_months = self.DEFAULT_MPR_MONTHS
        previous_mpr_month_idx = mpr_months.index(previous_info['month']) - 1

        if previous_mpr_month_idx < 0:
            prev_prev_month = mpr_months[-1]
            prev_prev_year = previous_info['year'] - 1
        else:
            prev_prev_month = mpr_months[previous_mpr_month_idx]
            prev_prev_year = previous_info['year']

        return {
            'year': prev_prev_year,
            'month': prev_prev_month,
            'month_name': self.MONTH_NAMES[prev_prev_month]
        }

    def _download_and_extract_excel(self, year: int, month: int, month_name: str) -> Optional[openpyxl.Workbook]:
        """MPR ZIPファイルをダウンロードしてExcelを抽出"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Possible Excel file names (new format first, then old)
        excel_filenames = [
            self.EXCEL_FILENAME_NEW_PATTERN.format(month=month_name.title(), year=year),
            self.EXCEL_FILENAME_OLD
        ]

        for pattern_index, url_pattern in enumerate(self.MPR_URL_PATTERNS):
            url = url_pattern.format(year=year, month=month_name)

            logger.info(f"Trying URL pattern {pattern_index + 1}/{len(self.MPR_URL_PATTERNS)}: {url}")

            for attempt in range(self.max_retries):
                try:
                    logger.info(f"Attempt {attempt + 1}/{self.max_retries}: Downloading {url}")
                    response = requests.get(url, headers=headers, timeout=60)

                    if response.status_code == 404:
                        logger.warning(f"MPR file not found (404): {url}")
                        break

                    if response.status_code == 403:
                        logger.warning(f"Access forbidden (403): {url}")
                        if attempt < self.max_retries - 1:
                            time.sleep(self.retry_delay)
                            continue
                        else:
                            break

                    if response.status_code != 200:
                        logger.warning(f"HTTP {response.status_code} on attempt {attempt + 1}")
                        if attempt < self.max_retries - 1:
                            time.sleep(self.retry_delay)
                            continue
                        else:
                            break

                    try:
                        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
                            # Try each Excel filename pattern
                            for excel_filename in excel_filenames:
                                if excel_filename in z.namelist():
                                    with z.open(excel_filename) as excel_file:
                                        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=False)
                                        logger.info(f"Successfully loaded {excel_filename} from {url}")
                                        return wb

                            # Log available files if none found
                            logger.error(f"Excel file not found in ZIP. Available files: {z.namelist()}")
                            break
                    except zipfile.BadZipFile:
                        logger.error(f"Invalid ZIP file from {url}")
                        break

                except requests.exceptions.Timeout:
                    logger.warning(f"Request timeout on attempt {attempt + 1}")
                    if attempt < self.max_retries - 1:
                        time.sleep(self.retry_delay)
                        continue
                    else:
                        break

                except Exception as e:
                    logger.error(f"Error downloading/extracting Excel from {url}: {e}")
                    if attempt < self.max_retries - 1:
                        time.sleep(self.retry_delay)
                        continue
                    else:
                        break

        logger.error("Failed to download MPR data after trying all URL patterns")
        return None

    def _find_cpi_components_sheet(self, wb: openpyxl.Workbook) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """CPI構成項目データを含むシートを検索"""
        # Try both prefixes (new format Chart 1.xx, old format Chart 2.xx)
        sheet_prefixes = [self.SHEET_PREFIX_NEW, self.SHEET_PREFIX_OLD]

        for sheet_name in wb.sheetnames:
            # Check if sheet name matches any prefix pattern
            matches_prefix = any(sheet_name.startswith(prefix) for prefix in sheet_prefixes)
            if not matches_prefix:
                continue

            try:
                ws = wb[sheet_name]
                a3_value = ws['A3'].value

                if a3_value:
                    a3_str = str(a3_value).lower()
                    if ("annual inflation" in a3_str and
                        "component" in a3_str and
                        "cpi" in a3_str):
                        logger.info(f"Found CPI components data in sheet: {sheet_name} (Title: {a3_value})")
                        return ws

            except (ValueError, IndexError):
                continue

        logger.error("Could not find sheet with CPI components inflation in cell A3")
        return None

    def _extract_data_from_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, mpr_date: str) -> Dict:
        """シートからCPI構成項目データを抽出"""
        data = {
            'date': mpr_date,
            'services': [],
            'core_goods': [],
            'food': []
        }

        header_row = None
        services_col = None
        core_goods_col = None
        food_col = None

        for row_idx in range(1, 20):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and "Date" in str(cell_val):
                header_row = row_idx
                logger.info(f"Found header row at row {row_idx}")

                for col_idx in range(1, 10):
                    col_val = ws.cell(row=header_row, column=col_idx).value
                    if col_val:
                        col_str = str(col_val).strip()
                        if "Services" == col_str:
                            services_col = col_idx
                        elif "Core goods" == col_str:
                            core_goods_col = col_idx
                        elif "Food and non-alcoholic beverages" == col_str:
                            food_col = col_idx
                break

        if not all([header_row, services_col, core_goods_col, food_col]):
            logger.error(f"Could not find all required columns. Header: {header_row}, Services: {services_col}, Core goods: {core_goods_col}, Food: {food_col}")
            return data

        data_start_row = header_row + 1
        if ws.cell(row=data_start_row, column=1).value and "average" in str(ws.cell(row=data_start_row, column=1).value).lower():
            data_start_row += 1

        if not ws.cell(row=data_start_row, column=1).value:
            data_start_row += 1

        row_idx = data_start_row
        while row_idx <= ws.max_row:
            date_cell = ws.cell(row=row_idx, column=1)
            if not date_cell.value:
                break

            if isinstance(date_cell.value, datetime):
                date_str = date_cell.value.strftime('%Y-%m')
            else:
                raw_date = str(date_cell.value).strip()
                raw_date = raw_date.replace('\n', ' ').replace('\r', '')
                raw_date = raw_date.replace('(Bank staff projections)', '').strip()

                try:
                    month_names = {
                        'january': 1, 'february': 2, 'march': 3, 'april': 4,
                        'may': 5, 'june': 6, 'july': 7, 'august': 8,
                        'september': 9, 'october': 10, 'november': 11, 'december': 12
                    }
                    parts = raw_date.split()
                    if len(parts) == 2:
                        month_name = parts[0].lower()
                        year = parts[1]
                        if month_name in month_names:
                            month_num = month_names[month_name]
                            date_str = f"{year}-{month_num:02d}"
                        else:
                            date_str = raw_date
                    else:
                        date_str = raw_date
                except Exception as e:
                    logger.warning(f"Could not parse date '{raw_date}': {e}")
                    date_str = raw_date

            services_val = ws.cell(row=row_idx, column=services_col).value
            core_goods_val = ws.cell(row=row_idx, column=core_goods_col).value
            food_val = ws.cell(row=row_idx, column=food_col).value

            if services_val is not None:
                try:
                    data['services'].append({
                        'date': date_str,
                        'value': float(services_val)
                    })
                except (ValueError, TypeError):
                    pass

            if core_goods_val is not None:
                try:
                    data['core_goods'].append({
                        'date': date_str,
                        'value': float(core_goods_val)
                    })
                except (ValueError, TypeError):
                    pass

            if food_val is not None:
                try:
                    data['food'].append({
                        'date': date_str,
                        'value': float(food_val)
                    })
                except (ValueError, TypeError):
                    pass

            row_idx += 1

        logger.info(f"Extracted {len(data['services'])} services, {len(data['core_goods'])} core goods, {len(data['food'])} food data points")
        return data

    def _fetch_cpi_components_data(self) -> Optional[Dict]:
        """最新MPRからCPI構成項目データを取得

        Note: 11月2025年MPRからCPI構成項目データが含まれなくなったため、
        最新MPRでデータが見つからない場合は前回のMPRからデータを取得する
        最新MPRのみを表示（前回比較なし）
        """
        try:
            mpr_info = self._get_latest_mpr_info()
            latest_info = mpr_info['latest']
            previous_info = mpr_info['previous']

            logger.info(f"Fetching CPI components - Latest MPR: {latest_info['month_name']} {latest_info['year']}")

            # 最新データ取得を試行
            latest_wb = self._download_and_extract_excel(
                latest_info['year'],
                latest_info['month'],
                latest_info['month_name']
            )

            latest_data = None
            actual_latest_info = latest_info
            if latest_wb:
                ws = self._find_cpi_components_sheet(latest_wb)
                if ws:
                    latest_data = self._extract_data_from_sheet(
                        ws,
                        f"{latest_info['year']}-{latest_info['month']:02d}"
                    )

            # 最新MPRでデータが見つからない場合、前回MPRをlatestとして使用
            if latest_data is None:
                logger.warning(f"CPI components data not found in {latest_info['month_name'].title()} {latest_info['year']} MPR, using previous MPR as latest")
                latest_wb = self._download_and_extract_excel(
                    previous_info['year'],
                    previous_info['month'],
                    previous_info['month_name']
                )
                actual_latest_info = previous_info
                if latest_wb:
                    ws = self._find_cpi_components_sheet(latest_wb)
                    if ws:
                        latest_data = self._extract_data_from_sheet(
                            ws,
                            f"{previous_info['year']}-{previous_info['month']:02d}"
                        )

            return {
                "components": {
                    "data": latest_data  # 最新MPRのみ表示
                },
                "metadata": {
                    "last_updated": datetime.now().isoformat(),
                    "source": "Bank of England (MPR Chart Data)",
                    "mpr_date": f"{actual_latest_info['month_name'].title()} {actual_latest_info['year']}",
                    "note": "Data source shifted due to MPR structure change" if actual_latest_info != latest_info else None
                }
            }

        except Exception as e:
            logger.error(f"Error fetching BOE CPI components: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)

            now = datetime.now(JST)

            if (now - last_updated).total_seconds() >= 7 * 24 * 3600:
                return True

            now_london = now.astimezone(LONDON)
            if now_london.month in self.DEFAULT_MPR_MONTHS:
                if now_london.hour == 20 and now_london.minute <= 10:
                    if (now - last_updated).total_seconds() >= 60:
                        return True

            return False

        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return True

    def _get_fallback_data(self) -> Dict:
        """空のフォールバックデータを返す"""
        return {
            "components": {
                "data": None
            },
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "source": "Bank of England",
                "error": "Failed to fetch BOE CPI components data"
            },
            "cached": False,
            "source": "none",
        }

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        """ファイルキャッシュを読み込む"""
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        """ファイルキャッシュを保存"""
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        """キャッシュを無効化"""
        return redis_client.delete(self.DATA_CACHE_KEY)


# シングルトンインスタンス
boe_cpi_components_service = BOECPIComponentsService()

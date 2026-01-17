"""
BOE CPI Contributions Service
BOE MPRデータからCPI寄与度データを取得

データソース: Section 2 - Current economic conditions.xlsx from MPR Charts, Slides and Data ZIP files
- "Contributions to CPI inflation" を検索
- 食品、電気・ガス、燃料、その他商品、サービス、CPI合計のデータを抽出

発表スケジュール:
- 2月・5月・8月・11月のMPR発表月
- 発表時刻: 12:00 GMT
- ダウンロードウィンドウ: 20:00-20:10 GMT/BST

キャッシュ方式: ファイルキャッシュ + Redisキャッシュ
"""
import json
import logging
import requests
import io
import zipfile
import openpyxl
import time
from datetime import datetime
from typing import Dict, List, Optional, Any
from zoneinfo import ZoneInfo
from pathlib import Path

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")
LONDON = ZoneInfo("Europe/London")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "uk" / "economic_outlook"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "boe_cpi_contributions_cache.json"


class BOECPIContributionsService:
    """BOE CPI Contributions サービス"""

    DATA_CACHE_KEY = "uk:boe_cpi_contributions:data"

    # Base URL patterns for MPR data files
    MPR_URL_PATTERNS = [
        "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-charts-slides-and-data.zip",
        "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-chart-slides-and-data.zip"
    ]

    # Month mapping
    MONTH_NAMES = {
        1: "january", 2: "february", 3: "march", 4: "april",
        5: "may", 6: "june", 7: "july", 8: "august",
        9: "september", 10: "october", 11: "november", 12: "december"
    }

    # Default MPR publication months
    DEFAULT_MPR_MONTHS = [2, 5, 8, 11]

    # Excel file name patterns in ZIP (BOE changed structure in Nov 2025)
    # Old: "Section 2  - Current economic conditions.xlsx"
    # New: "{Month} {Year} MPR chart data.xlsx"
    EXCEL_FILENAME_OLD = "Section 2  - Current economic conditions.xlsx"
    EXCEL_FILENAME_NEW_PATTERN = "{month} {year} MPR chart data.xlsx"

    # Chart title to look for in cell A3
    CHART_TITLE = "Contributions to CPI inflation"

    # Sheet name patterns
    SHEET_PREFIX_OLD = "Chart 2."
    SHEET_PREFIX_NEW = "Chart 1."

    def __init__(self):
        self.max_retries = 3
        self.retry_delay = 5

    def fetch_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """CPI寄与度データを取得"""
        # Redisキャッシュチェック
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "contributions": cached_data.get("contributions", {}),
                        "metadata": cached_data.get("metadata", {}),
                        "cached": True,
                        "source": "redis",
                    }

        # MPRからデータを取得
        data = self._fetch_cpi_contributions_data()

        if data:
            cache_payload = {
                "contributions": data.get("contributions", {}),
                "metadata": data.get("metadata", {}),
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "contributions": data.get("contributions", {}),
                "metadata": data.get("metadata", {}),
                "cached": False,
                "source": "boe_mpr",
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "contributions": file_cache.get("contributions", {}),
                "metadata": file_cache.get("metadata", {}),
                "cached": True,
                "source": "file (fallback)",
            }

        return self._get_fallback_data()

    def _get_latest_mpr_info(self, reference_date: Optional[datetime] = None) -> Dict[str, Any]:
        """最新のMPR日を判定"""
        if reference_date is None:
            reference_date = datetime.now()

        current_year = reference_date.year
        current_month = reference_date.month
        current_day = reference_date.day

        mpr_months = self.DEFAULT_MPR_MONTHS

        if current_month in mpr_months and current_day < 6:
            current_month_for_search = current_month - 1
        else:
            current_month_for_search = current_month

        latest_mpr_month = None
        for month in reversed(mpr_months):
            if current_month_for_search >= month:
                latest_mpr_month = month
                break

        if latest_mpr_month is None:
            latest_mpr_month = mpr_months[-1]
            current_year -= 1

        return {
            'latest': {
                'year': current_year,
                'month': latest_mpr_month,
                'month_name': self.MONTH_NAMES[latest_mpr_month]
            }
        }

    def _download_and_extract_excel(self, year: int, month: int, month_name: str) -> Optional[openpyxl.Workbook]:
        """MPR ZIPファイルをダウンロードしてExcelを抽出"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        # Possible Excel file names (new format first, then old)
        excel_filenames = [
            self.EXCEL_FILENAME_NEW_PATTERN.format(month=month_name.title(), year=year),
            self.EXCEL_FILENAME_OLD
        ]

        for pattern_index, url_pattern in enumerate(self.MPR_URL_PATTERNS):
            url = url_pattern.format(year=year, month=month_name)

            logger.info(f"Trying URL pattern {pattern_index + 1}/{len(self.MPR_URL_PATTERNS)}: {url}")

            for attempt in range(self.max_retries):
                try:
                    logger.info(f"Attempt {attempt + 1}/{self.max_retries}: Downloading {url}")
                    response = requests.get(url, headers=headers, timeout=60)

                    if response.status_code == 404:
                        logger.warning(f"MPR file not found (404): {url}")
                        break

                    if response.status_code == 403:
                        logger.warning(f"Access forbidden (403): {url}")
                        if attempt < self.max_retries - 1:
                            time.sleep(self.retry_delay)
                            continue
                        else:
                            break

                    if response.status_code != 200:
                        logger.warning(f"HTTP {response.status_code} on attempt {attempt + 1}")
                        if attempt < self.max_retries - 1:
                            time.sleep(self.retry_delay)
                            continue
                        else:
                            break

                    try:
                        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
                            # Try each Excel filename pattern
                            for excel_filename in excel_filenames:
                                if excel_filename in z.namelist():
                                    with z.open(excel_filename) as excel_file:
                                        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=False)
                                        logger.info(f"Successfully loaded {excel_filename} from {url}")
                                        return wb

                            # Log available files if none found
                            logger.error(f"Excel file not found in ZIP. Available files: {z.namelist()}")
                            break
                    except zipfile.BadZipFile:
                        logger.error(f"Invalid ZIP file from {url}")
                        break

                except requests.exceptions.Timeout:
                    logger.warning(f"Request timeout on attempt {attempt + 1}")
                    if attempt < self.max_retries - 1:
                        time.sleep(self.retry_delay)
                        continue
                    else:
                        break

                except Exception as e:
                    logger.error(f"Error downloading/extracting Excel from {url}: {e}")
                    if attempt < self.max_retries - 1:
                        time.sleep(self.retry_delay)
                        continue
                    else:
                        break

        logger.error("Failed to download MPR data after trying all URL patterns")
        return None

    def _find_cpi_contributions_sheet(self, workbook: openpyxl.Workbook) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """CPI寄与度データを含むシートを検索"""
        # Try both old and new sheet prefixes
        sheet_prefixes = [self.SHEET_PREFIX_NEW, self.SHEET_PREFIX_OLD]

        for sheet_name in workbook.sheetnames:
            # Check if sheet name matches any prefix
            matches_prefix = any(sheet_name.startswith(prefix) for prefix in sheet_prefixes)
            if not matches_prefix:
                continue

            try:
                ws = workbook[sheet_name]
                a3_value = ws['A3'].value

                if a3_value and self.CHART_TITLE in str(a3_value):
                    logger.info(f"Found CPI contributions sheet: {sheet_name}")
                    return ws

            except (ValueError, IndexError):
                continue

        logger.warning(f"Could not find sheet with '{self.CHART_TITLE}' in cell A3")
        return None

    def _extract_data_from_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, List]:
        """シートからCPI寄与度データを抽出"""
        data = {
            'date': [],
            'food': [],
            'electricity_gas': [],
            'fuels': [],
            'other_goods': [],
            'services': [],
            'cpi': []
        }

        # Column indices (1-based) - Updated for new BOE format (Nov 2025+)
        DATE_COL = 1
        FOOD_COL = 2
        ELEC_GAS_COL = 3
        FUELS_COL = 4
        OTHER_GOODS_COL = 5
        SERVICES_COL = 6
        CPI_COL = 8  # Changed from 13 to 8 in new format

        # Data starts at row 7
        row_num = 7

        while True:
            date_cell = ws.cell(row=row_num, column=DATE_COL)

            if date_cell.value is None:
                break

            # Parse date
            if isinstance(date_cell.value, datetime):
                date_str = date_cell.value.strftime('%Y-%m')
            else:
                raw_date = str(date_cell.value).strip()
                raw_date = raw_date.replace('\n', ' ').replace('\r', '')
                raw_date = raw_date.replace('(Bank staff projections)', '').strip()

                try:
                    month_names = {
                        'january': 1, 'february': 2, 'march': 3, 'april': 4,
                        'may': 5, 'june': 6, 'july': 7, 'august': 8,
                        'september': 9, 'october': 10, 'november': 11, 'december': 12
                    }
                    parts = raw_date.split()
                    if len(parts) == 2:
                        month_name = parts[0].lower()
                        year = parts[1]
                        if month_name in month_names:
                            month_num = month_names[month_name]
                            date_str = f"{year}-{month_num:02d}"
                        else:
                            logger.warning(f"Unknown month name in date: {raw_date}")
                            row_num += 1
                            continue
                    else:
                        logger.warning(f"Could not parse date format: {raw_date}")
                        row_num += 1
                        continue
                except Exception as e:
                    logger.warning(f"Error parsing date {raw_date}: {e}")
                    row_num += 1
                    continue

            # Extract values
            food_val = ws.cell(row=row_num, column=FOOD_COL).value
            elec_gas_val = ws.cell(row=row_num, column=ELEC_GAS_COL).value
            fuels_val = ws.cell(row=row_num, column=FUELS_COL).value
            other_goods_val = ws.cell(row=row_num, column=OTHER_GOODS_COL).value
            services_val = ws.cell(row=row_num, column=SERVICES_COL).value
            cpi_val = ws.cell(row=row_num, column=CPI_COL).value

            # Add to data lists
            data['date'].append(date_str)
            data['food'].append(float(food_val) if food_val is not None else None)
            data['electricity_gas'].append(float(elec_gas_val) if elec_gas_val is not None else None)
            data['fuels'].append(float(fuels_val) if fuels_val is not None else None)
            data['other_goods'].append(float(other_goods_val) if other_goods_val is not None else None)
            data['services'].append(float(services_val) if services_val is not None else None)
            data['cpi'].append(float(cpi_val) if cpi_val is not None else None)

            row_num += 1

        logger.info(f"Extracted {len(data['date'])} data points")
        return data

    def _fetch_cpi_contributions_data(self) -> Dict:
        """最新MPRからCPI寄与度データを取得"""
        try:
            mpr_info = self._get_latest_mpr_info()
            latest_info = mpr_info['latest']

            logger.info(f"Fetching latest MPR data for {latest_info['month_name']} {latest_info['year']}")

            wb = self._download_and_extract_excel(
                latest_info['year'],
                latest_info['month'],
                latest_info['month_name']
            )

            latest_data = None
            if wb:
                ws = self._find_cpi_contributions_sheet(wb)
                if ws:
                    latest_data = self._extract_data_from_sheet(ws)

            return {
                "contributions": {
                    "latest_data": {
                        "date": f"{latest_info['year']}{latest_info['month']:02d}",
                        "data": latest_data
                    }
                },
                "metadata": {
                    "last_updated": datetime.now().isoformat(),
                    "source": "Bank of England",
                    "mpr_date": f"{latest_info['month_name'].title()} {latest_info['year']}",
                    "latest_data_date": f"{latest_info['year']}{latest_info['month']:02d}"
                }
            }

        except Exception as e:
            logger.error(f"Error fetching CPI contributions data: {e}")
            return self._get_fallback_data()

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)

            now = datetime.now(JST)

            if (now - last_updated).total_seconds() >= 7 * 24 * 3600:
                return True

            now_london = now.astimezone(LONDON)
            if now_london.month in self.DEFAULT_MPR_MONTHS:
                if now_london.hour == 20 and now_london.minute <= 10:
                    if (now - last_updated).total_seconds() >= 60:
                        return True

            return False

        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return True

    def _get_fallback_data(self) -> Dict:
        """空のフォールバックデータを返す"""
        return {
            "contributions": {
                "latest_data": {
                    "date": "",
                    "data": {
                        "date": [],
                        "food": [],
                        "electricity_gas": [],
                        "fuels": [],
                        "other_goods": [],
                        "services": [],
                        "cpi": []
                    }
                }
            },
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "source": "Bank of England",
                "error": "Failed to fetch BOE CPI contributions data"
            },
            "cached": False,
            "source": "none",
        }

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        """ファイルキャッシュを読み込む"""
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        """ファイルキャッシュを保存"""
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        """キャッシュを無効化"""
        return redis_client.delete(self.DATA_CACHE_KEY)


# シングルトンインスタンス
boe_cpi_contributions_service = BOECPIContributionsService()

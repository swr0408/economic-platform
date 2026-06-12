"""
BOE MPR Utilities
MPRデータ取得のための共通ユーティリティ

November 2025以降、BOEはMPRファイル構造を変更:
- 旧構造: 個別のParameters for MPC...xlsxファイル + Market profiles.xlsx
- 新構造: Projections Databank - {Month} {Year} MPR.xlsx に統合

このユーティリティは両方の構造に対応
"""
import logging
import requests
import io
import zipfile
import openpyxl
import time
from datetime import datetime
from typing import Dict, Optional, Tuple, List, Any
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

LONDON = ZoneInfo("Europe/London")

# MPR publication months (静的フォールバック用)
# ※ 2026年からBOEは公表月を変更している (例: 2026年は2月の次が「4月」、5月ではない)。
#   ハードコード月だけに頼ると公表月変更で silent 404 になるため、
#   実際の公表済みレポートはサイトマップから動的に解決する (get_published_mprs)。
DEFAULT_MPR_MONTHS = [2, 5, 8, 11]

# Month name mapping
MONTH_NAMES = {
    1: "january", 2: "february", 3: "march", 4: "april",
    5: "may", 6: "june", 7: "july", 8: "august",
    9: "september", 10: "october", 11: "november", 12: "december"
}

# URL patterns for MPR data (サイトマップ解決失敗時のフォールバック)
MPR_URL_PATTERNS = [
    "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-charts-slides-and-data.zip",
    "https://www.bankofengland.co.uk/-/media/boe/files/monetary-policy-report/{year}/{month}/mpr-{month}-{year}-chart-slides-and-data.zip"
]

# MPR sitemap: 公表済み全レポートの zip URL が直接列挙されている
MPR_SITEMAP_URL = "https://www.bankofengland.co.uk/sitemap/monetary-policy-report"

_BROWSER_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# サイトマップ解決結果のメモリキャッシュ (1日)
_published_mprs_cache: Optional[Tuple[float, List[Dict[str, Any]]]] = None
_PUBLISHED_MPRS_TTL_SEC = 24 * 3600


def get_published_mprs() -> List[Dict[str, Any]]:
    """BOEサイトマップから公表済みMPRの一覧を取得 (新しい順)

    Returns:
        [{'year': 2026, 'month': 4, 'month_name': 'april', 'zip_url': '...'}, ...]
        失敗時は空リスト (呼び出し側は静的フォールバックを使うこと)
    """
    global _published_mprs_cache
    now = time.time()
    if _published_mprs_cache is not None and now - _published_mprs_cache[0] < _PUBLISHED_MPRS_TTL_SEC:
        return _published_mprs_cache[1]

    import re as _re
    month_to_num = {name: num for num, name in MONTH_NAMES.items()}
    try:
        resp = requests.get(MPR_SITEMAP_URL, headers=_BROWSER_HEADERS, timeout=30)
        resp.raise_for_status()
        # 例: .../monetary-policy-report/2026/april/mpr-april-2026-charts-slides-and-data.zip
        #     (2019-2024 は "chart-slides"、2025+ は "charts-slides" の表記ゆれあり)
        pattern = _re.compile(
            r'href="(https?://[^"]*?/monetary-policy-report/(\d{4})/([a-z]+)/'
            r'mpr-[a-z]+-\d{4}-charts?-slides-and-data\.zip)"'
        )
        results: List[Dict[str, Any]] = []
        seen = set()
        for url, year_s, month_name in pattern.findall(resp.text):
            month_num = month_to_num.get(month_name)
            if month_num is None:
                continue
            key = (int(year_s), month_num)
            if key in seen:
                continue
            seen.add(key)
            results.append({
                'year': int(year_s),
                'month': month_num,
                'month_name': month_name,
                'zip_url': url,
            })
        results.sort(key=lambda r: (r['year'], r['month']), reverse=True)
        if results:
            _published_mprs_cache = (now, results)
            logger.info(f"[BOE MPR] sitemap resolved {len(results)} published MPRs "
                        f"(latest: {results[0]['month_name']} {results[0]['year']})")
        return results
    except Exception as e:
        logger.warning(f"[BOE MPR] sitemap resolution failed: {e}")
        return []

# File patterns (new structure from Nov 2025)
PROJECTIONS_DATABANK_PATTERN = "Projections Databank - {month} {year} MPR.xlsx"
CHART_DATA_PATTERN = "{month} {year} MPR chart data.xlsx"

# Old file names (pre-Nov 2025)
OLD_FILES = {
    "cpi_projections": "Parameters for MPC CPI inflation projections from February 2004.xlsx",
    "gdp_projections": "Parameters for MPC GDP growth projections based on market interest rate expectations.xlsx",
    "unemployment_projections": "Parameters for MPC unemployment rate projections from August 2013.xlsx",
    "market_profiles": "Market profiles.xlsx",
    "section2_chart": "Section 2  - Current economic conditions.xlsx"
}


def get_mpr_info(reference_date: Optional[datetime] = None) -> Dict[str, Any]:
    """最新と前回のMPR情報を取得

    まずBOEサイトマップから実際の公表済みレポートを解決する
    (公表月のスケジュール変更に自動追従)。失敗時のみ静的月ロジックを使う。
    """
    if reference_date is None:
        reference_date = datetime.now()

    # 動的解決 (公表月変更に追従)
    published = get_published_mprs()
    if reference_date is not None and published:
        ref_key = (reference_date.year, reference_date.month)
        eligible = [p for p in published if (p['year'], p['month']) <= ref_key]
        if len(eligible) >= 1:
            latest = eligible[0]
            previous = eligible[1] if len(eligible) >= 2 else eligible[0]
            return {
                'latest': {
                    'year': latest['year'],
                    'month': latest['month'],
                    'month_name': latest['month_name'],
                },
                'previous': {
                    'year': previous['year'],
                    'month': previous['month'],
                    'month_name': previous['month_name'],
                },
            }

    current_year = reference_date.year
    current_month = reference_date.month
    current_day = reference_date.day

    # MPRは通常月の6日以降に発表
    if current_month in DEFAULT_MPR_MONTHS and current_day < 6:
        current_month_for_search = current_month - 1
    else:
        current_month_for_search = current_month

    # 最新のMPR月を特定
    latest_mpr_month = None
    for month in reversed(DEFAULT_MPR_MONTHS):
        if current_month_for_search >= month:
            latest_mpr_month = month
            break

    if latest_mpr_month is None:
        latest_mpr_month = DEFAULT_MPR_MONTHS[-1]
        current_year -= 1

    # 前回のMPR月を特定
    previous_mpr_month_idx = DEFAULT_MPR_MONTHS.index(latest_mpr_month) - 1
    if previous_mpr_month_idx < 0:
        previous_mpr_month = DEFAULT_MPR_MONTHS[-1]
        previous_year = current_year - 1
    else:
        previous_mpr_month = DEFAULT_MPR_MONTHS[previous_mpr_month_idx]
        previous_year = current_year

    return {
        'latest': {
            'year': current_year,
            'month': latest_mpr_month,
            'month_name': MONTH_NAMES[latest_mpr_month]
        },
        'previous': {
            'year': previous_year,
            'month': previous_mpr_month,
            'month_name': MONTH_NAMES[previous_mpr_month]
        }
    }


def download_mpr_zip(year: int, month_name: str, max_retries: int = 3, retry_delay: int = 5) -> Optional[bytes]:
    """MPR ZIPファイルをダウンロード

    まずサイトマップで解決した実URLを使い (ファイル名の表記ゆれに自動追従)、
    見つからなければ既知のURLパターンを順に試す。
    """
    headers = _BROWSER_HEADERS

    # サイトマップ解決URLを最優先
    candidate_urls: List[str] = []
    for p in get_published_mprs():
        if p['year'] == year and p['month_name'] == month_name.lower():
            candidate_urls.append(p['zip_url'])
            break
    candidate_urls.extend(
        url_pattern.format(year=year, month=month_name)
        for url_pattern in MPR_URL_PATTERNS
    )

    for pattern_index, url in enumerate(candidate_urls):
        logger.info(f"Trying URL {pattern_index + 1}/{len(candidate_urls)}: {url}")

        for attempt in range(max_retries):
            try:
                logger.info(f"Attempt {attempt + 1}/{max_retries}: Downloading {url}")
                response = requests.get(url, headers=headers, timeout=60)

                if response.status_code == 404:
                    logger.warning(f"MPR file not found (404): {url}")
                    break

                if response.status_code == 403:
                    logger.warning(f"Access forbidden (403): {url}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        continue
                    break

                if response.status_code != 200:
                    logger.warning(f"HTTP {response.status_code} on attempt {attempt + 1}")
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay)
                        continue
                    break

                return response.content

            except requests.exceptions.Timeout:
                logger.warning(f"Request timeout on attempt {attempt + 1}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                continue

            except Exception as e:
                logger.error(f"Error downloading from {url}: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                continue

    return None


def extract_workbook_from_zip(zip_content: bytes, filename: str) -> Optional[openpyxl.Workbook]:
    """ZIPからワークブックを抽出"""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_content)) as z:
            if filename in z.namelist():
                with z.open(filename) as excel_file:
                    wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
                    logger.info(f"Successfully loaded {filename}")
                    return wb
            else:
                logger.warning(f"{filename} not found in ZIP")
                return None
    except zipfile.BadZipFile:
        logger.error("Invalid ZIP file")
        return None
    except Exception as e:
        logger.error(f"Error extracting {filename}: {e}")
        return None


def list_zip_contents(zip_content: bytes) -> List[str]:
    """ZIP内のファイル一覧を取得"""
    try:
        with zipfile.ZipFile(io.BytesIO(zip_content)) as z:
            return z.namelist()
    except Exception as e:
        logger.error(f"Error listing ZIP contents: {e}")
        return []


def find_file_in_zip(zip_content: bytes, patterns: List[str]) -> Optional[str]:
    """ZIP内でパターンに一致するファイルを検索"""
    files = list_zip_contents(zip_content)
    for pattern in patterns:
        for f in files:
            if pattern.lower() in f.lower():
                return f
    return None


def get_projections_databank(year: int, month_name: str) -> Optional[Tuple[openpyxl.Workbook, bytes]]:
    """Projections Databankファイルを取得

    Returns:
        Tuple of (workbook, zip_content) or None
    """
    zip_content = download_mpr_zip(year, month_name)
    if zip_content is None:
        return None

    # New format: Projections Databank - {Month} {Year} MPR.xlsx
    databank_filename = PROJECTIONS_DATABANK_PATTERN.format(
        month=month_name.title(),
        year=year
    )

    wb = extract_workbook_from_zip(zip_content, databank_filename)
    if wb:
        return (wb, zip_content)

    # Try to find with flexible matching
    files = list_zip_contents(zip_content)
    for f in files:
        if 'projections databank' in f.lower() and f.endswith('.xlsx'):
            wb = extract_workbook_from_zip(zip_content, f)
            if wb:
                return (wb, zip_content)

    return None


def get_chart_data(year: int, month_name: str, zip_content: Optional[bytes] = None) -> Optional[openpyxl.Workbook]:
    """Chart Data ファイルを取得"""
    if zip_content is None:
        zip_content = download_mpr_zip(year, month_name)
        if zip_content is None:
            return None

    # New format: {Month} {Year} MPR chart data.xlsx
    chart_filename = CHART_DATA_PATTERN.format(
        month=month_name.title(),
        year=year
    )

    wb = extract_workbook_from_zip(zip_content, chart_filename)
    if wb:
        return wb

    # Try old format
    wb = extract_workbook_from_zip(zip_content, OLD_FILES["section2_chart"])
    if wb:
        return wb

    # Flexible matching
    files = list_zip_contents(zip_content)
    for f in files:
        if 'chart data' in f.lower() and f.endswith('.xlsx'):
            return extract_workbook_from_zip(zip_content, f)
        if 'current economic conditions' in f.lower() and f.endswith('.xlsx'):
            return extract_workbook_from_zip(zip_content, f)

    return None


def parse_databank_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Optional[Dict[str, Any]]:
    """Projections Databankのシートをパース

    共通フォーマット:
    - Row 5 or 12: Column headers (quarters)
    - Column 1: Date of publication
    - Data starts from row 6 or 13
    """
    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet {sheet_name} not found")
        return None

    ws = wb[sheet_name]

    # Determine header row (Bank Rate uses row 12, others use row 5)
    header_row = 12 if sheet_name == "32. Bank Rate" else 5
    data_start_row = header_row + 1

    # Get column headers (quarters)
    quarters = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            quarters.append((col, str(val).strip()))

    if not quarters:
        logger.error(f"No quarters found in {sheet_name}")
        return None

    # Find latest and previous data rows
    latest_row = None
    previous_row = None

    for row in range(ws.max_row, data_start_row - 1, -1):
        date_val = ws.cell(row=row, column=1).value
        if date_val:
            date_str = str(date_val)
            # Check for 2025-11, November 2025, etc
            if '2025-11' in date_str or 'November 2025' in date_str or '2025-11-01' in date_str:
                latest_row = row
            elif '2025-08' in date_str or 'August 2025' in date_str or '2025-08-01' in date_str:
                previous_row = row

            if latest_row and previous_row:
                break

    # If not found with specific dates, use last two rows
    if latest_row is None:
        data_rows = []
        for row in range(data_start_row, ws.max_row + 1):
            if ws.cell(row=row, column=1).value:
                data_rows.append(row)
        if len(data_rows) >= 2:
            latest_row = data_rows[-1]
            previous_row = data_rows[-2]
        elif len(data_rows) == 1:
            latest_row = data_rows[0]

    def extract_row_data(row_num: int) -> Tuple[str, List[Dict]]:
        """Extract data from a row"""
        date_val = ws.cell(row=row_num, column=1).value
        date_str = str(date_val) if date_val else ""

        data = []
        for col, quarter in quarters:
            val = ws.cell(row=row_num, column=col).value
            if val is not None:
                try:
                    data.append({
                        'date': quarter,
                        'value': float(val)
                    })
                except (ValueError, TypeError):
                    pass

        return date_str, data

    result = {
        'latest': None,
        'previous': None
    }

    if latest_row:
        date_str, data = extract_row_data(latest_row)
        result['latest'] = {
            'date': date_str,
            'data': data
        }

    if previous_row:
        date_str, data = extract_row_data(previous_row)
        result['previous'] = {
            'date': date_str,
            'data': data
        }

    return result


def parse_chart_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Optional[Dict[str, Any]]:
    """Chart Dataのシートをパース

    共通フォーマット:
    - Row 6: Column headers
    - Data starts from row 7
    - Column 1: Date
    """
    if sheet_name not in wb.sheetnames:
        logger.error(f"Sheet {sheet_name} not found")
        return None

    ws = wb[sheet_name]

    # Get column headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=6, column=col).value
        if val:
            headers.append((col, str(val).strip().replace('\n', ' ')))

    if not headers:
        logger.error(f"No headers found in {sheet_name}")
        return None

    # Extract all data
    series_data = {header: [] for _, header in headers}
    dates = []

    for row in range(7, ws.max_row + 1):
        row_has_data = False
        for col, header in headers:
            val = ws.cell(row=row, column=col).value
            if val is not None and str(val).lower() != 'n.a.':
                row_has_data = True
                if header == 'Date' or col == 1:
                    dates.append(str(val))
                else:
                    try:
                        series_data[header].append(float(val))
                    except (ValueError, TypeError):
                        series_data[header].append(None)
            else:
                if header != 'Date' and col != 1:
                    series_data[header].append(None)

        if not row_has_data:
            break

    return {
        'dates': dates,
        'series': {k: v for k, v in series_data.items() if k != 'Date'}
    }


# =============================================================================
# 共通日付パーシング・ヘッダー検出ユーティリティ
# =============================================================================

# 月名マッピング（フル名＋省略名）
MONTH_NAME_TO_NUM = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
    'jun': 6, 'jul': 7, 'aug': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}


def parse_date_to_yyyy_mm(date_val: Any) -> Optional[str]:
    """日付を YYYY-MM 形式にパース

    対応フォーマット:
    - datetime オブジェクト
    - "January 2025", "Jan 2025" (月名 年)
    - "2025-01", "2025-01-15" (ISO形式)

    Args:
        date_val: 日付値（datetime, str）

    Returns:
        YYYY-MM形式の文字列、パース失敗時はNone
    """
    if date_val is None:
        return None

    # datetime オブジェクト
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m')

    raw_date = str(date_val).strip()
    raw_date = raw_date.replace('\n', ' ').replace('\r', '')

    # "Date" ヘッダー行をスキップ
    if raw_date.lower() == 'date':
        return None

    # 特殊文字列を除去
    raw_date = raw_date.replace('(Bank staff projections)', '').strip()

    # ISO形式: 2025-01 or 2025-01-15
    import re
    iso_match = re.match(r'^(\d{4})-(\d{2})(?:-\d{2})?', raw_date)
    if iso_match:
        return f"{iso_match.group(1)}-{iso_match.group(2)}"

    # 月名 年 形式: "January 2025", "Jan 2025"
    parts = raw_date.split()
    if len(parts) >= 2:
        month_name = parts[0].lower()
        year = parts[-1]
        if month_name in MONTH_NAME_TO_NUM and year.isdigit():
            month_num = MONTH_NAME_TO_NUM[month_name]
            return f"{year}-{month_num:02d}"

    return None


def resolve_sheet_by_suffix(wb: openpyxl.Workbook, suffix: str) -> Optional[str]:
    """シート名末尾一致でシート名を解決（MPR間で番号がドリフトしても追従）

    BOE Projections Databankのシート名は番号プレフィックスが付くが、
    番号はMPR発表ごとに変動する（例: "32. Bank Rate" → "38. Bank Rate"）。
    本ヘルパーは番号を除いた本体名で照合する。

    Args:
        wb: openpyxl ワークブック
        suffix: 照合するシート名本体（大文字小文字無視）
                例: "Bank Rate", "Average weekly earnings"

    Returns:
        マッチしたシート名、なければ None
    """
    import re
    target = suffix.strip().lower()
    # 番号プレフィックス除去後の完全一致を最優先
    for sn in wb.sheetnames:
        cleaned = re.sub(r'^\s*\d+\.\s*', '', sn).strip().lower()
        if cleaned == target:
            return sn
    # フォールバック: 末尾一致
    for sn in wb.sheetnames:
        if sn.strip().lower().endswith(target):
            return sn
    return None


def parse_scenario_databank_sheet(ws, min_quarter: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """2026年4月MPR以降の転置レイアウト（シナリオ方式）シートをパースする

    レイアウト（旧来と行列が転置されている点に注意）:
      - ヘッダー行: 列1='Date'、列2以降=系列名
        例: 'February 2026 central projection', 'April 2026 Scenario A' ...
      - データ行: 列1=四半期 ('1998 Q1' ...)、列2以降=各系列の値

    系列の分類は系列名に含まれる「月 年」で行い、最も新しい日付のグループを
    最新MPR、その他から1本を前回（central/indicative projection優先）とする。
    最新グループが1本なら旧来どおり 'latest' キー、複数なら 'scenario_a' 等。
    BoEが中央見通し方式に戻った場合も自動で旧形式の出力に戻る。

    Returns:
      {
        'table_data': [{'quarter': '2026Q2', 'previous': x, 'scenario_a': y, ...}],
        'latest_forecast': 'April 2026',
        'previous_forecast': 'February 2026',
        'scenario_labels': {'scenario_a': 'April 2026 Scenario A', ...},  # 複数系列時のみ
      }
      転置レイアウトでなければ None（呼び出し側は旧パーサーへフォールバック）。
    """
    import re as _re

    # 1) ヘッダー行を特定 (列1 が 'Date')
    header_row = None
    for r in range(1, 13):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip().lower() == "date":
            header_row = r
            break
    if header_row is None:
        return None

    # 2) データ行の列1が四半期形式であることを確認 (転置レイアウトの証拠)
    first_data = ws.cell(row=header_row + 1, column=1).value
    if first_data is None or not _re.match(r"^\d{4}\s*Q\d$", str(first_data).strip()):
        return None

    # 3) 系列名ヘッダーを収集
    headers: List[Tuple[int, str]] = []
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is not None and str(v).strip():
            headers.append((col, str(v).strip()))
    if not headers:
        return None

    # 4) 系列名から (年, 月) を抽出して最新MPRグループを特定
    date_re = _re.compile(r"([A-Z][a-z]+)\s+(\d{4})")
    month_to_num = {name.title(): num for num, name in MONTH_NAMES.items()}

    def _label_date(label: str) -> Tuple[int, int]:
        m = date_re.search(label)
        if m and m.group(1) in month_to_num:
            return (int(m.group(2)), month_to_num[m.group(1)])
        return (0, 0)

    dated = [(col, label, _label_date(label)) for col, label in headers]
    latest_date = max(d for _, _, d in dated)
    latest_cols = [(col, label) for col, label, d in dated if d == latest_date]
    prev_cols = [(col, label) for col, label, d in dated if d != latest_date]

    # 前回系列: central/indicative projection を優先、無ければ先頭
    previous_col: Optional[Tuple[int, str]] = None
    if prev_cols:
        previous_col = next(
            (c for c in prev_cols if "projection" in c[1].lower()), prev_cols[0]
        )

    # 5) 最新グループのキー割り当て
    #    1本 → 'latest' (旧形式互換) / 複数 → 'scenario_a' 等
    col_keys: Dict[int, str] = {}
    scenario_labels: Dict[str, str] = {}
    if len(latest_cols) == 1:
        col_keys[latest_cols[0][0]] = "latest"
    else:
        sc_re = _re.compile(r"Scenario\s+([A-Za-z0-9]+)\s*$", _re.IGNORECASE)
        for i, (col, label) in enumerate(sorted(latest_cols, key=lambda x: x[1])):
            m = sc_re.search(label)
            key = f"scenario_{m.group(1).lower()}" if m else f"scenario_{i + 1}"
            col_keys[col] = key
            scenario_labels[key] = label
    if previous_col is not None:
        col_keys[previous_col[0]] = "previous"

    # 6) データ行を走査して table_data を構築
    if min_quarter is None:
        min_quarter = f"{datetime.now().year - 1}Q1"

    table_data: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        qv = ws.cell(row=r, column=1).value
        if qv is None:
            continue
        q = str(qv).strip().replace(" ", "")  # '2026 Q2' -> '2026Q2'
        if not _re.match(r"^\d{4}Q\d$", q) or q < min_quarter:
            continue
        row: Dict[str, Any] = {"quarter": q}
        has_value = False
        for col, key in col_keys.items():
            v = ws.cell(row=r, column=col).value
            try:
                row[key] = float(v) if v is not None else None
            except (ValueError, TypeError):
                row[key] = None
            if row[key] is not None:
                has_value = True
        if has_value:
            table_data.append(row)

    if not table_data:
        return None

    latest_label = f"{MONTH_NAMES[latest_date[1]].title()} {latest_date[0]}" if latest_date != (0, 0) else ""
    prev_date = _label_date(previous_col[1]) if previous_col else (0, 0)
    prev_label = f"{MONTH_NAMES[prev_date[1]].title()} {prev_date[0]}" if prev_date != (0, 0) else ""

    result: Dict[str, Any] = {
        "table_data": table_data,
        "latest_forecast": latest_label,
        "previous_forecast": prev_label,
    }
    if scenario_labels:
        result["scenario_labels"] = scenario_labels
    return result


def resolve_sheet(wb: openpyxl.Workbook, name: str) -> Optional[str]:
    """シート名を解決（完全一致 → 番号プレフィックス無視のサフィックス一致）

    BOEはMPRごとにシート番号を振り直す（例: "1. CPI inflation" →
    2026年4月版では "2. CPI inflation"）。ハードコード名が外れても本体名で追従する。
    シート自体が廃止された場合（2026年4月版で distribution / Bank Rate /
    UK import prices 等が消えた）は None を返すので、呼び出し側は
    旧キャッシュ温存などのフォールバックを行うこと。
    """
    if name in wb.sheetnames:
        return name
    import re as _re
    suffix = _re.sub(r'^\s*\d+\.\s*', '', name).strip()
    return resolve_sheet_by_suffix(wb, suffix)


def detect_multi_row_header(ws) -> bool:
    """ワークシートがマルチ行ヘッダー構造かどうかを検出

    Nov 2025+ MPR構造:
    - Row 5: グループヘッダー
    - Row 6: シリーズ名（Col 1 = "Date"）
    - Row 7+: データ

    Args:
        ws: openpyxl ワークシート

    Returns:
        マルチ行ヘッダー構造の場合True
    """
    row6_col1 = ws.cell(row=6, column=1).value
    return bool(row6_col1 and str(row6_col1).strip().lower() == 'date')


def find_chart_sheet_by_keywords(
    wb: openpyxl.Workbook,
    primary_keywords: List[str],
    fallback_keywords: Optional[List[str]] = None,
    exclude_keywords: Optional[List[str]] = None,
    prefer_multi_column: bool = True
) -> Optional[Any]:
    """キーワードでチャートシートを検索

    A3セルのチャートタイトルに基づいてシートを検索。
    複数候補がある場合は列数が多いシートを優先。

    Args:
        wb: ワークブック
        primary_keywords: 優先キーワードリスト
        fallback_keywords: フォールバックキーワードリスト
        exclude_keywords: 除外キーワードリスト
        prefer_multi_column: True の場合、列数が多いシートを優先

    Returns:
        見つかったワークシート、なければNone
    """
    candidates = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Chart"):
            continue

        try:
            ws = wb[sheet_name]
            a3_value = ws['A3'].value

            if not a3_value:
                continue

            a3_lower = str(a3_value).lower()

            # 除外キーワードチェック
            if exclude_keywords and any(kw in a3_lower for kw in exclude_keywords):
                continue

            # 列数をカウント（row 5とrow 6の両方をチェック）
            col_count_row5 = sum(1 for col in range(2, 20) if ws.cell(row=5, column=col).value is not None)
            col_count_row6 = sum(1 for col in range(2, 20) if ws.cell(row=6, column=col).value is not None)
            col_count = max(col_count_row5, col_count_row6)

            # キーワードマッチング
            is_primary = any(kw in a3_lower for kw in primary_keywords)
            is_fallback = fallback_keywords and any(kw in a3_lower for kw in fallback_keywords)

            if is_primary or is_fallback:
                candidates.append({
                    'sheet': ws,
                    'name': sheet_name,
                    'title': a3_value,
                    'col_count': col_count,
                    'is_primary': is_primary,
                })

        except Exception as e:
            logger.warning(f"Error checking sheet {sheet_name}: {e}")

    if not candidates:
        return None

    # ソート: primary優先、列数優先
    def sort_key(c):
        score = 0
        if c['is_primary']:
            score += 1000
        if prefer_multi_column:
            score += c['col_count'] * 10
        return -score

    candidates.sort(key=sort_key)
    best = candidates[0]
    logger.info(f"Selected sheet: {best['name']} with {best['col_count']} columns, title: {best['title']}")
    return best['sheet']

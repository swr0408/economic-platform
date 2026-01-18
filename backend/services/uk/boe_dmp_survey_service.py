"""
BOE DMP Survey Service
BOE Decision Maker Panel (意思決定者パネル) サーベイデータを取得

指標:
- CPI Expectations: CPIインフレ期待（1年先・3年先）
- Price Growth: 価格成長（実績・期待）
- Wage Growth: 賃金成長（実績・期待）
- Employment Growth: 雇用成長（実績・期待）

データソース:
- BOE DMP月次データExcelファイル
- URL: https://www.bankofengland.co.uk/-/media/boe/files/decision-maker-panel-survey/{year}/monthly-dmp-data-{month}-{year}.xlsx

発表スケジュール:
- 毎月1日〜15日頃
- ダウンロードウィンドウ: 20:00-20:10 ロンドン時間

キャッシュ方式: ファイルキャッシュ + Redisキャッシュ
"""
import io
import json
import logging
import requests
import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")
LONDON = ZoneInfo("Europe/London")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "uk" / "dmp"
CACHE_DIR.mkdir(parents=True, exist_ok=True)


class BOEDMPSurveyService:
    """BOE DMP Survey サービス"""

    DATA_CACHE_KEY = "uk:boe_dmp_survey:data"
    CACHE_FILE_PATH = CACHE_DIR / "boe_dmp_survey_cache.json"

    # Base URL pattern for DMP data files
    DMP_URL_PATTERN = "https://www.bankofengland.co.uk/-/media/boe/files/decision-maker-panel-survey/{year}/monthly-dmp-data-{month}-{year}.xlsx"

    # Month mapping (number to name)
    MONTH_NAMES = {
        1: "january", 2: "february", 3: "march", 4: "april",
        5: "may", 6: "june", 7: "july", 8: "august",
        9: "september", 10: "october", 11: "november", 12: "december"
    }

    # Month name to number mapping (for parsing)
    MONTH_NAME_TO_NUM = {
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4,
        'may': 5, 'jun': 6, 'jul': 7, 'aug': 8,
        'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
        'january': 1, 'february': 2, 'march': 3, 'april': 4,
        'june': 6, 'july': 7, 'august': 8,
        'september': 9, 'october': 10, 'november': 11, 'december': 12
    }

    def fetch_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """データを取得（キャッシュ優先）"""
        next_release = self._get_next_release()

        # Redisキャッシュチェック
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "survey_data": cached_data.get("survey_data", {}),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": next_release,
                        "cached": True,
                        "source": "redis",
                    }

        # データソースから取得
        data = self._fetch_data_from_source()

        if data and data.get("survey_data"):
            cache_payload = {
                "survey_data": data.get("survey_data", {}),
                "metadata": data.get("metadata", {}),
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "survey_data": data.get("survey_data", {}),
                "metadata": data.get("metadata", {}),
                "next_release": next_release,
                "cached": False,
                "source": "boe_dmp",
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "survey_data": file_cache.get("survey_data", {}),
                "metadata": file_cache.get("metadata", {}),
                "next_release": next_release,
                "cached": True,
                "source": "file (fallback)",
            }

        return self._get_fallback_data()

    def _fetch_data_from_source(self) -> Optional[Dict[str, Any]]:
        """BOEからDMPサーベイデータを取得"""
        now = datetime.now()

        # Try current month - 1, then current month - 2, then current month - 3
        for months_back in range(1, 4):
            try:
                prev_month_date = now - relativedelta(months=months_back)
                year = prev_month_date.year
                month = prev_month_date.month
                month_name = self.MONTH_NAMES[month]

                url = self.DMP_URL_PATTERN.format(year=year, month=month_name)

                logger.info(f"Trying to download DMP data for {month_name} {year}")

                # Download Excel file
                excel_content = self._download_excel(url)

                if excel_content:
                    # Extract data
                    data = self._extract_all_dmp_data(excel_content)

                    return {
                        "survey_data": data,
                        "metadata": {
                            "last_updated": datetime.now().isoformat(),
                            "source": "Bank of England Decision Maker Panel Survey",
                            "url": url,
                            "data_month": f"{year}-{month:02d}"
                        }
                    }

            except Exception as e:
                logger.warning(f"Failed to fetch DMP data for {month_name} {year}: {e}")
                if months_back == 3:
                    logger.error("All attempts to fetch DMP survey data failed")

        return None

    def _download_excel(self, url: str) -> Optional[bytes]:
        """ExcelファイルをダウンロードSPACER"""
        try:
            logger.info(f"Downloading DMP data from {url}")
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }

            response = requests.get(url, headers=headers, timeout=60)
            response.raise_for_status()

            logger.info(f"Successfully downloaded DMP data ({len(response.content)} bytes)")
            return response.content

        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to download from {url}: {e}")
            return None

    def _extract_all_dmp_data(self, excel_content: bytes) -> Dict[str, Any]:
        """ExcelファイルからすべてのDMPデータを抽出"""
        wb = openpyxl.load_workbook(io.BytesIO(excel_content))

        # Extract CPI expectations
        # Row 5+, Col A (date), Col E (1-year ahead 3mo avg), Col G (3-year ahead 3mo avg)
        cpi_data = self._extract_sheet_data(
            wb, 'CPI expectations', 5, 1, 5, 7,
            'one_year_ahead', 'three_year_ahead'
        )

        # Extract Price growth
        # Row 5+, Col A (date), Col C (realised 3mo avg), Col N (expected 3mo avg)
        price_data = self._extract_sheet_data(
            wb, 'Price growth', 5, 1, 3, 14,
            'realised_3mo_avg', 'expected_3mo_avg'
        )

        # Extract Wage growth
        # Row 3+, Col A (date), Col C (realised 3mo avg), Col F (expected 3mo avg)
        wage_data = self._extract_sheet_data(
            wb, 'Wage growth', 3, 1, 3, 6,
            'realised_3mo_avg', 'expected_3mo_avg'
        )

        # Extract Employment growth
        # Row 5+, Col A (date), Col C (realised 3mo avg), Col N (expected 3mo avg)
        employment_data = self._extract_sheet_data(
            wb, 'Employment growth', 5, 1, 3, 14,
            'realised_3mo_avg', 'expected_3mo_avg'
        )

        return {
            'cpi_expectations': cpi_data,
            'price_growth': price_data,
            'wage_growth': wage_data,
            'employment_growth': employment_data
        }

    def _extract_sheet_data(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        data_start_row: int,
        date_col: int,
        col1: int,
        col2: int,
        key1: str,
        key2: str
    ) -> Dict[str, List]:
        """シートからデータを抽出"""
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in Excel file")
            return {'date': [], key1: [], key2: []}

        ws = wb[sheet_name]

        data = {
            'date': [],
            key1: [],
            key2: []
        }

        row_num = data_start_row
        consecutive_empty_rows = 0
        max_consecutive_empty = 3

        while consecutive_empty_rows < max_consecutive_empty:
            date_cell = ws.cell(row=row_num, column=date_col)

            if date_cell.value is None:
                consecutive_empty_rows += 1
                row_num += 1
                continue

            consecutive_empty_rows = 0

            # Parse date
            date_str = self._parse_date(date_cell.value)
            if date_str:
                # Get data values
                val1_cell = ws.cell(row=row_num, column=col1)
                val2_cell = ws.cell(row=row_num, column=col2)

                data['date'].append(date_str)

                # Handle None or invalid values
                try:
                    data[key1].append(float(val1_cell.value) if val1_cell.value is not None else None)
                except (ValueError, TypeError):
                    data[key1].append(None)

                try:
                    data[key2].append(float(val2_cell.value) if val2_cell.value is not None else None)
                except (ValueError, TypeError):
                    data[key2].append(None)

            row_num += 1

        logger.info(f"Extracted {len(data['date'])} data points from {sheet_name} sheet")
        return data

    def _parse_date(self, date_val: Any) -> Optional[str]:
        """日付を YYYY-MM 形式にパース

        対応フォーマット:
        - datetime オブジェクト
        - "May-22", "Jun-22" (Mon-YY形式)
        - "May 2022", "June 2022" (Month Year形式)
        """
        if date_val is None:
            return None

        # datetime オブジェクト
        if isinstance(date_val, datetime):
            return date_val.strftime('%Y-%m')

        raw_date = str(date_val).strip()

        # "Mon-YY" 形式 (例: May-22, Jun-22)
        if '-' in raw_date and len(raw_date) <= 7:
            parts = raw_date.split('-')
            if len(parts) == 2:
                month_name = parts[0].lower()
                year_short = parts[1]
                if month_name in self.MONTH_NAME_TO_NUM and year_short.isdigit():
                    month_num = self.MONTH_NAME_TO_NUM[month_name]
                    # Handle 2-digit year
                    year = int(year_short)
                    if year < 100:
                        year = 2000 + year if year < 50 else 1900 + year
                    return f"{year}-{month_num:02d}"

        # "Month Year" 形式 (例: May 2022)
        parts = raw_date.split()
        if len(parts) >= 2:
            month_name = parts[0].lower()
            year = parts[-1]
            if month_name in self.MONTH_NAME_TO_NUM and year.isdigit():
                month_num = self.MONTH_NAME_TO_NUM[month_name]
                return f"{year}-{month_num:02d}"

        return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定

        更新条件:
        - 前回更新から7日以上経過
        - 毎月1日〜15日の20:00-20:10（ロンドン時間）かつ60秒以上経過
        """
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)

            now = datetime.now(JST)

            # 7日以上経過
            if (now - last_updated).total_seconds() >= 7 * 24 * 3600:
                return True

            # DMP発表タイミング（ロンドン時間20:00-20:10、毎月1日〜15日）
            now_london = now.astimezone(LONDON)
            if 1 <= now_london.day <= 15:
                if now_london.hour == 20 and now_london.minute <= 10:
                    if (now - last_updated).total_seconds() >= 60:
                        return True

            return False

        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return True

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """次回発表日を取得

        FMPカレンダーから取得を試み、取得できない場合は推定日を返す。
        DMPサーベイは毎月発表されるが、FMPカレンダーに登録されていない場合がある。
        """
        try:
            from services.uk.fmp_next_release_utils import get_next_release_by_pattern

            # FMPカレンダーから取得を試みる
            fmp_result = get_next_release_by_pattern("BoE Decision Maker Panel", country="GB")
            if fmp_result:
                return fmp_result

            # FMPに登録されていない場合は推定日を計算
            return self._estimate_next_release()

        except Exception as e:
            logger.error(f"Error getting next release date: {e}")
            return self._estimate_next_release()

    def _estimate_next_release(self) -> Optional[Dict[str, Any]]:
        """次回発表日を推定

        DMPサーベイは毎月発表される。
        正確な日付は公開されていないため、翌月の15日を推定日とする。
        """
        try:
            now = datetime.now(LONDON)

            # 翌月の15日を次回発表日と推定
            if now.day >= 15:
                # 今月15日以降なら、翌月15日
                next_month = now + relativedelta(months=1)
            else:
                # 今月15日以前なら、今月15日
                next_month = now

            next_release_date = next_month.replace(day=15, hour=20, minute=0, second=0, microsecond=0)

            # JST時刻を計算
            next_release_jst = next_release_date.astimezone(JST)

            return {
                "date": next_release_date.strftime("%Y-%m-%d"),
                "time_jst": next_release_jst.strftime("%H:%M"),
                "estimated": True,  # 推定日であることを示す
            }

        except Exception as e:
            logger.error(f"Error estimating next release date: {e}")
            return None

    def _get_fallback_data(self) -> Dict[str, Any]:
        """空のフォールバックデータを返す"""
        empty_cpi = {'date': [], 'one_year_ahead': [], 'three_year_ahead': []}
        empty_growth = {'date': [], 'realised_3mo_avg': [], 'expected_3mo_avg': []}

        return {
            "survey_data": {
                'cpi_expectations': empty_cpi,
                'price_growth': empty_growth.copy(),
                'wage_growth': empty_growth.copy(),
                'employment_growth': empty_growth.copy()
            },
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "source": "Bank of England Decision Maker Panel Survey",
                "error": "Failed to fetch DMP survey data"
            },
            "next_release": self._get_next_release(),
            "cached": False,
            "source": "none",
        }

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        """ファイルキャッシュを読み込む"""
        try:
            if not self.CACHE_FILE_PATH.exists():
                return None
            with open(self.CACHE_FILE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        """ファイルキャッシュを保存"""
        try:
            self.CACHE_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(self.CACHE_FILE_PATH, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        """キャッシュを無効化"""
        return redis_client.delete(self.DATA_CACHE_KEY)


# シングルトンインスタンス
boe_dmp_survey_service = BOEDMPSurveyService()

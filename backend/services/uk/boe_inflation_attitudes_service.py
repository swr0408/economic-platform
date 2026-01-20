"""
BOE Inflation Attitudes Survey（インフレ期待調査）サービス
Bank of EnglandのExcelファイルからインフレ期待調査データを取得

指標:
- Q. 2a: Expected price changes over next 12 months (Median) - 今後12ヶ月
- Q. 2b: Expected price changes 12 months after that (Median) - その後12ヶ月
- Q. 2c: Expected price changes in 5 years (Median) - 5年先

データソース:
- BOE Inflation Attitudes Survey Excel (Long-run file)
- https://www.bankofengland.co.uk/-/media/boe/files/inflation-attitudes-survey/long-run.xlsx

発表スケジュール:
- 四半期毎（3月、6月、9月、12月）
- 7日〜17日頃
- 20:00-20:10 ロンドン時間

キャッシュ方式: 90日経過で更新チェック
"""
import json
import logging
import requests
import openpyxl
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path
from io import BytesIO

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")
LONDON = ZoneInfo("Europe/London")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "uk" / "inflation"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "boe_inflation_attitudes_cache.json"


class BOEInflationAttitudesService:
    """BOE Inflation Attitudes Survey サービス"""

    DATA_CACHE_KEY = "uk:boe_inflation_attitudes:data"

    # BOE Excel file URL
    SURVEY_URL = "https://www.bankofengland.co.uk/-/media/boe/files/inflation-attitudes-survey/long-run.xlsx"

    def __init__(self):
        pass

    def get_inflation_attitudes_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """インフレ期待調査データを取得"""
        # Redisキャッシュチェック
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "series": cached_data.get("series", {}),
                        "latest": cached_data.get("latest"),
                        "next_release": self._calculate_next_release(),
                        "cached": True,
                        "source": "redis",
                        "last_updated": last_updated_str,
                    }

        # BOEからデータを取得
        api_result = self._fetch_from_boe()

        if api_result:
            # 最新値を決定
            latest = self._get_latest_values(api_result.get("series", {}))

            cache_payload = {
                "series": api_result.get("series", {}),
                "latest": latest,
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "series": api_result.get("series", {}),
                "latest": latest,
                "next_release": self._calculate_next_release(),
                "cached": False,
                "source": "boe_excel",
                "last_updated": datetime.now(JST).isoformat(),
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "series": file_cache.get("series", {}),
                "latest": file_cache.get("latest"),
                "next_release": self._calculate_next_release(),
                "cached": True,
                "source": "file (fallback)",
                "last_updated": file_cache.get("last_updated"),
            }

        return {
            "series": {},
            "latest": None,
            "next_release": self._calculate_next_release(),
            "cached": False,
            "source": "none",
            "last_updated": None,
            "error": "No data available",
        }

    def _fetch_from_boe(self) -> Optional[Dict[str, Any]]:
        """BOEからExcelデータを取得してパース"""
        try:
            logger.info(f"[BOE Inflation Attitudes] Fetching from {self.SURVEY_URL}")
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }

            response = requests.get(self.SURVEY_URL, headers=headers, timeout=120)
            response.raise_for_status()

            logger.info(f"[BOE Inflation Attitudes] Downloaded {len(response.content)} bytes")

            # Excelをパース
            return self._parse_excel(response.content)

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error fetching from BOE: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _parse_excel(self, excel_data: bytes) -> Optional[Dict[str, Any]]:
        """Excelファイルをパースしてデータを抽出"""
        try:
            wb = openpyxl.load_workbook(BytesIO(excel_data), data_only=True)

            # シートを探す
            sheet_name = None
            for name in ['Data', 'Sheet1', wb.sheetnames[0]]:
                if name in wb.sheetnames:
                    sheet_name = name
                    break

            if sheet_name is None:
                logger.error("[BOE Inflation Attitudes] Could not find data sheet")
                return None

            ws = wb[sheet_name]
            logger.info(f"[BOE Inflation Attitudes] Processing sheet: {sheet_name}")

            # 各質問のデータを抽出
            q2a_text = "Q. 2a How much would you expect prices in the"
            q2b_text = "Q. 2b And how about the 12 months after that?"
            q2c_text = "Q. 2c And how about the longer term"

            q2a_row = self._find_median_row(ws, q2a_text)
            q2b_row = self._find_median_row(ws, q2b_text)
            q2c_row = self._find_median_row(ws, q2c_text)

            if not q2a_row and not q2b_row and not q2c_row:
                logger.error("[BOE Inflation Attitudes] Could not find any questions")
                return None

            result = {"series": {}}

            # Q2a: 今後12ヶ月
            if q2a_row:
                q2a_data = self._extract_time_series(ws, q2a_row, header_row=4)
                if q2a_data:
                    result["series"]["next_12_months"] = {
                        "data": q2a_data,
                        "metadata": {"name": "今後12ヶ月", "name_en": "Next 12 months"}
                    }
                    logger.info(f"[BOE Inflation Attitudes] Extracted {len(q2a_data)} Q2a data points")

            # Q2b: その後12ヶ月
            if q2b_row:
                q2b_data = self._extract_time_series(ws, q2b_row, header_row=4)
                if q2b_data:
                    result["series"]["following_12_months"] = {
                        "data": q2b_data,
                        "metadata": {"name": "その後12ヶ月", "name_en": "Following 12 months"}
                    }
                    logger.info(f"[BOE Inflation Attitudes] Extracted {len(q2b_data)} Q2b data points")

            # Q2c: 5年先
            if q2c_row:
                q2c_data = self._extract_time_series(ws, q2c_row, header_row=4)
                if q2c_data:
                    result["series"]["five_years"] = {
                        "data": q2c_data,
                        "metadata": {"name": "5年先", "name_en": "5 years ahead"}
                    }
                    logger.info(f"[BOE Inflation Attitudes] Extracted {len(q2c_data)} Q2c data points")

            return result

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error parsing Excel: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _find_median_row(self, ws, question_text: str) -> Optional[int]:
        """質問テキストを探してMedian行を返す"""
        try:
            # 質問テキストを検索
            question_row = None
            for row in range(1, min(ws.max_row + 1, 100)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and question_text in str(cell_value):
                    question_row = row
                    logger.info(f"[BOE Inflation Attitudes] Found question at row {row}")
                    break

            if question_row is None:
                return None

            # Median行を検索
            for row in range(question_row, min(question_row + 20, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "Median" in str(cell_value):
                    logger.info(f"[BOE Inflation Attitudes] Found Median at row {row}")
                    return row

            return None

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error finding median row: {e}")
            return None

    def _extract_time_series(self, ws, row_num: int, header_row: int = 4) -> List[Dict[str, Any]]:
        """指定行から時系列データを抽出"""
        data = []

        try:
            col = 2  # B列から開始
            max_cols = min(ws.max_column, 1000)
            consecutive_empty = 0

            while col <= max_cols:
                date_cell = ws.cell(row=header_row, column=col).value
                value_cell = ws.cell(row=row_num, column=col).value

                # 連続空白でループ終了
                if date_cell is None:
                    consecutive_empty += 1
                    if consecutive_empty >= 3:
                        break
                    col += 1
                    continue
                else:
                    consecutive_empty = 0

                # 日付をパース
                date_str = None
                if isinstance(date_cell, datetime):
                    date_str = date_cell.strftime("%Y-%m-%d")
                elif isinstance(date_cell, str):
                    if "*" in date_cell or len(date_cell) < 4:
                        col += 1
                        continue
                    try:
                        parsed_date = datetime.strptime(date_cell[:10], "%Y-%m-%d")
                        date_str = parsed_date.strftime("%Y-%m-%d")
                    except:
                        try:
                            if "/" in date_cell:
                                parts = date_cell.split("/")
                                if len(parts) == 3:
                                    parsed_date = datetime(int(parts[2]), int(parts[0]), int(parts[1]))
                                    date_str = parsed_date.strftime("%Y-%m-%d")
                        except:
                            col += 1
                            continue

                # 値をパース
                value = None
                if value_cell is not None:
                    try:
                        value = float(value_cell)
                    except (ValueError, TypeError):
                        pass

                if date_str and value is not None:
                    data.append({
                        "date": date_str,
                        "value": round(value, 2)
                    })

                col += 1

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error extracting time series: {e}")

        return data

    def _get_latest_values(self, series: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """各系列の最新値を取得"""
        try:
            latest = {}
            latest_date = None

            for key, series_data in series.items():
                if series_data and series_data.get("data"):
                    data = series_data["data"]
                    if data:
                        last_point = data[-1]
                        latest[key] = last_point.get("value")
                        if latest_date is None or last_point.get("date", "") > latest_date:
                            latest_date = last_point.get("date")

            if latest_date:
                latest["date"] = latest_date

            return latest if latest else None

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error getting latest values: {e}")
            return None

    def _calculate_next_release(self) -> Optional[Dict[str, Any]]:
        """次回発表日を計算（四半期: 3月、6月、9月、12月の7日〜17日頃）"""
        try:
            now = datetime.now(LONDON)
            year = now.year
            month = now.month

            # 次の発表月を決定
            release_months = [3, 6, 9, 12]
            next_month = None
            next_year = year

            for m in release_months:
                if m > month or (m == month and now.day < 17):
                    next_month = m
                    break

            if next_month is None:
                next_month = 3
                next_year = year + 1

            # 発表日は中旬（12日頃）と仮定
            release_date = datetime(next_year, next_month, 12, 20, 0, tzinfo=LONDON)
            release_jst = release_date.astimezone(JST)

            return {
                "date": release_date.strftime("%Y-%m-%d"),
                "datetime_utc": release_date.astimezone(ZoneInfo("UTC")).isoformat(),
                "datetime_jst": release_jst.isoformat(),
                "time_jst": release_jst.strftime("%H:%M"),
                "label": f"BOE Inflation Attitudes Survey ({release_date.strftime('%b %Y')})",
            }

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error calculating next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（90日経過で更新）"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)

            now = datetime.now(JST)
            days_since_update = (now - last_updated).days

            # 90日以上経過していれば更新
            if days_since_update >= 90:
                return True

            # 発表期間中（3月、6月、9月、12月の7日〜17日）は毎日チェック
            london_now = now.astimezone(LONDON)
            if london_now.month in [3, 6, 9, 12] and 7 <= london_now.day <= 17:
                # 1日以上経過していれば更新
                if days_since_update >= 1:
                    return True

            return False

        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Error in should_refresh: {e}")
            return True

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        """ファイルキャッシュを読み込む"""
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        """ファイルキャッシュを保存"""
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"[BOE Inflation Attitudes] Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        """キャッシュを無効化"""
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        """キャッシュ状態を取得"""
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None

        series_counts = {}
        if cached_data and cached_data.get("series"):
            for key, series in cached_data["series"].items():
                if series and series.get("data"):
                    series_counts[key] = len(series["data"])

        return {
            "indicator": "BOE Inflation Attitudes Survey",
            "source": "BOE Excel",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "series_counts": series_counts,
            "latest": cached_data.get("latest") if cached_data else None,
            "next_release": self._calculate_next_release(),
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
boe_inflation_attitudes_service = BOEInflationAttitudesService()

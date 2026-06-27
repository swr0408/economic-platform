"""
Australian Bureau of Statistics (ABS) 四半期CPI サービス

データソース:
- ABS SDMX API
  - CPI dataflow (FREQ=Q): All groups CPI (10001) — Index Numbers + QoQ (1949-Q3～)
  - CPI_M dataflow (FREQ=M): 四半期月(3,6,9,12月)の SA YoY / Trimmed Mean YoY (2018-09～)
  - CPI dataflow (FREQ=M, SA): SA YoY(999901) / Trimmed Mean YoY(999902) / Weighted Median YoY(999903) (2025-04～)
    ※ CPI_Mより先に四半期CPI発表で更新されるため、最新データの補完に使用
- RBA G1 Historical Data (Excel)
  - Trimmed Mean QoQ / Weighted Median QoQ (1982-Q2～)
  - URL: https://www.rba.gov.au/statistics/tables/xls/g01hist.xlsx

7系列:
- All groups CPI (原系列, 10001) — QoQ + YoY(Index計算)
- All groups CPI SA (999901) — YoY
- Trimmed Mean (999905/999902) — YoY
- Weighted Median (999903) — YoY
- Trimmed Mean — QoQ (RBA G1)
- Weighted Median — QoQ (RBA G1)

発表スケジュール: 四半期（1,4,7,10月下旬）
"""
import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests
from openpyxl import load_workbook

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "inflation"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "abs_quarterly_cpi_cache.json"

# ABS API設定
ABS_API_BASE = "https://data.api.abs.gov.au/rest/data"

# 1. CPI Quarterly: All groups CPI (10001), Index + QoQ
CPI_Q_KEY = "1+2.10001.10.50.Q"
CPI_Q_URL = f"{ABS_API_BASE}/ABS,CPI,/{CPI_Q_KEY}?dimensionAtObservation=AllDimensions"

# 2. CPI_M: SA YoY + Trimmed Mean YoY (四半期月のみ使用, 2018-09～)
CPI_M_YOY_KEY = "3.999901+999905..50.M"
CPI_M_YOY_URL = f"{ABS_API_BASE}/ABS,CPI_M,/{CPI_M_YOY_KEY}?dimensionAtObservation=AllDimensions"

# 3. CPI Monthly SA: SA YoY(999901) + Trimmed Mean(999902) + Weighted Median(999903) (2025-04～)
# CPI_Mより先に更新されるため、最新四半期のデータ補完に使用
CPI_SA_YOY_KEY = "3.999901+999902+999903.20.50.M"
CPI_SA_YOY_URL = f"{ABS_API_BASE}/ABS,CPI,/{CPI_SA_YOY_KEY}?dimensionAtObservation=AllDimensions"

# 四半期月のマッピング
QUARTER_MONTH_MAP = {
    "Q1": "03",
    "Q2": "06",
    "Q3": "09",
    "Q4": "12",
}
MONTH_TO_QUARTER = {
    "03": "Q1",
    "06": "Q2",
    "09": "Q3",
    "12": "Q4",
}

# RBA G1 Historical Data (Trimmed Mean QoQ, Weighted Median QoQ)
RBA_G1_URL = "https://www.rba.gov.au/statistics/tables/xls/g01hist.xlsx"

# RBA G1 長期履歴シードCSV (1982-06～)
# RBAサイトはAkamaiで当サーバIPを403ブロックするため、長期履歴は静的CSVでシード。
# 直近四半期は RBA G1 ライブ取得(成功時) と ABS API(Phase3 WM YoY) が上書き補完。
CSV_DIR = Path(__file__).parent.parent.parent / "data" / "csv_import"
RBA_G1_SEED_CSV = CSV_DIR / "au_cpi_underlying_rba_g1.csv"

# FMPイベントパターン
FMP_EVENT_PATTERNS = ["CPI", "Consumer Price Index"]


class AbsQuarterlyCpiService:
    """ABS Quarterly CPI Service"""

    DATA_CACHE_KEY = "australia:abs_quarterly_cpi:data"

    def __init__(self):
        pass

    def _fetch_sdmx(self, url: str) -> Optional[Dict[str, Any]]:
        """ABS SDMX APIからデータを取得"""
        try:
            logger.info(f"Fetching ABS data from {url}")
            response = requests.get(
                url,
                headers={
                    "Accept": "application/vnd.sdmx.data+json",
                    "User-Agent": "Mozilla/5.0 (economic-platform)",
                },
                timeout=30,
            )

            if response.status_code != 200:
                logger.error(f"ABS API returned HTTP {response.status_code} for {url}")
                return None

            return response.json()

        except Exception as e:
            logger.error(f"Error fetching ABS data: {e}")
            return None

    def _parse_sdmx_observations(
        self, raw_data: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """SDMX JSONレスポンスをパースして観測値リストを返す"""
        result = []
        try:
            datasets = raw_data.get("data", {}).get("dataSets", [])
            if not datasets:
                return []

            observations = datasets[0].get("observations", {})
            if not observations:
                return []

            structures = raw_data.get("data", {}).get("structures", [])
            if not structures:
                return []

            dimensions = structures[0].get("dimensions", {}).get("observation", [])

            dim_lookups = {}
            for dim in dimensions:
                dim_id = dim.get("id", "")
                values = dim.get("values", [])
                dim_lookups[dim_id] = {
                    str(i): v.get("id", "") for i, v in enumerate(values)
                }

            dim_ids = [d.get("id", "") for d in dimensions]

            for obs_key, obs_value in observations.items():
                if obs_value is None or obs_value[0] is None:
                    continue

                value = obs_value[0]
                indices = obs_key.split(":")

                codes = {}
                for i, idx in enumerate(indices):
                    if i < len(dim_ids):
                        dim_id = dim_ids[i]
                        lookup = dim_lookups.get(dim_id, {})
                        codes[dim_id] = lookup.get(idx, idx)

                result.append(
                    {
                        "measure": codes.get("MEASURE", ""),
                        "index": codes.get("INDEX", ""),
                        "time_period": codes.get("TIME_PERIOD", ""),
                        "value": value,
                    }
                )

        except Exception as e:
            logger.error(f"Error parsing SDMX data: {e}")

        return result

    def _load_seed_csv(self) -> Dict[str, Dict[str, float]]:
        """RBA G1 長期履歴シードCSVを読み込む

        RBAサイトがIPブロック(403)のため、RBA G1の長期TM/WM(QoQ+YoY, 1982-06～)を
        静的CSVから供給する。_fetch_rba_g1_qoq と同じ形式で返す。

        Returns:
            { "2025-12-01": {"trimmed_mean_qoq":.., "weighted_median_qoq":..,
                             "trimmed_mean_yoy":.., "weighted_median_yoy":..}, ... }
        """
        result: Dict[str, Dict[str, float]] = {}
        try:
            if not RBA_G1_SEED_CSV.exists():
                logger.warning(f"RBA G1 seed CSV not found: {RBA_G1_SEED_CSV}")
                return result

            import csv
            with open(RBA_G1_SEED_CSV, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    date_str = (row.get("date") or "").strip()
                    if not date_str:
                        continue
                    point: Dict[str, float] = {}
                    for key, decimals in [
                        ("trimmed_mean_qoq", 2),
                        ("weighted_median_qoq", 2),
                        ("trimmed_mean_yoy", 1),
                        ("weighted_median_yoy", 1),
                    ]:
                        raw_val = (row.get(key) or "").strip()
                        if raw_val == "":
                            continue
                        try:
                            point[key] = round(float(raw_val), decimals)
                        except (ValueError, TypeError):
                            pass
                    if point:
                        result[date_str] = point

            logger.info(f"RBA G1 seed CSV: loaded {len(result)} quarterly data points")
        except Exception as e:
            logger.error(f"Error loading RBA G1 seed CSV: {e}")

        return result

    def _fetch_rba_g1_qoq(self) -> Dict[str, Dict[str, float]]:
        """RBA G1 Excelから TM/WM の QoQ + YoY を取得

        Column mapping (0-indexed):
          J(9)=WM YoY, K(10)=TM YoY, T(19)=WM QoQ, U(20)=TM QoQ

        Returns:
            { "2025-12-01": {
                "trimmed_mean_qoq": 0.5, "weighted_median_qoq": 0.4,
                "trimmed_mean_yoy": 3.3, "weighted_median_yoy": 3.6
              }, ... }
        """
        result: Dict[str, Dict[str, float]] = {}
        try:
            logger.info(f"Fetching RBA G1 Excel from {RBA_G1_URL}")
            resp = requests.get(
                RBA_G1_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=30,
            )
            if resp.status_code != 200:
                logger.error(f"RBA G1 returned HTTP {resp.status_code}")
                return result

            wb = load_workbook(BytesIO(resp.content), data_only=True, read_only=True)
            ws = wb["Data"]

            # Row 12 onwards: data rows
            # Column A (0) = date (end-of-quarter datetime)
            for row in ws.iter_rows(min_row=12, values_only=True):
                date_val = row[0]
                if date_val is None:
                    continue

                # Convert end-of-quarter date to first-of-month format
                if isinstance(date_val, datetime):
                    dt = date_val
                else:
                    try:
                        dt = datetime.strptime(str(date_val), "%Y-%m-%d %H:%M:%S")
                    except (ValueError, TypeError):
                        continue

                # End-of-quarter (e.g. 2025-12-31) → first-of-quarter-month (2025-12-01)
                date_str = f"{dt.year}-{dt.month:02d}-01"

                wm_yoy = row[9] if len(row) > 9 else None
                tm_yoy = row[10] if len(row) > 10 else None
                wm_qoq = row[19] if len(row) > 19 else None
                tm_qoq = row[20] if len(row) > 20 else None

                point: Dict[str, float] = {}
                for raw_val, key, decimals in [
                    (tm_qoq, "trimmed_mean_qoq", 2),
                    (wm_qoq, "weighted_median_qoq", 2),
                    (tm_yoy, "trimmed_mean_yoy", 1),
                    (wm_yoy, "weighted_median_yoy", 1),
                ]:
                    if raw_val is not None:
                        try:
                            point[key] = round(float(raw_val), decimals)
                        except (ValueError, TypeError):
                            pass
                if point:
                    result[date_str] = point

            wb.close()
            logger.info(f"RBA G1: parsed {len(result)} quarterly data points")

        except Exception as e:
            logger.error(f"Error fetching RBA G1 data: {e}")

        return result

    def _quarter_to_date(self, quarter_str: str) -> str:
        """'2025-Q4' -> '2025-12-01'"""
        parts = quarter_str.split("-")
        if len(parts) == 2:
            year = parts[0]
            q = parts[1]
            month = QUARTER_MONTH_MAP.get(q, "12")
            return f"{year}-{month}-01"
        return quarter_str

    def _month_to_quarter_date(self, month_str: str) -> Optional[str]:
        """'2025-12' -> '2025-12-01' (四半期月のみ、それ以外はNone)"""
        parts = month_str.split("-")
        if len(parts) == 2:
            month = parts[1]
            if month in MONTH_TO_QUARTER:
                return f"{month_str}-01"
        return None

    def _merge_data(
        self,
        cpi_q_obs: List[Dict[str, Any]],
        cpi_m_yoy_obs: List[Dict[str, Any]],
        cpi_sa_yoy_obs: List[Dict[str, Any]],
        rba_g1_qoq: Optional[Dict[str, Dict[str, float]]] = None,
    ) -> List[Dict[str, Any]]:
        """四半期CPI + 月次CPIの四半期月データ + RBA G1 QoQデータをマージ"""
        parsed_points: Dict[str, Dict[str, Any]] = {}

        def ensure_point(date_str: str) -> Dict[str, Any]:
            if date_str not in parsed_points:
                parsed_points[date_str] = {
                    "date": date_str,
                    "cpi_qoq": None,
                    "cpi_yoy": None,
                    "cpi_sa_yoy": None,
                    "trimmed_mean_yoy": None,
                    "weighted_median_yoy": None,
                    "trimmed_mean_qoq": None,
                    "weighted_median_qoq": None,
                }
            return parsed_points[date_str]

        # Phase 1: Quarterly CPI (10001) — Index Numbers + QoQ
        index_map: Dict[str, float] = {}
        for obs in cpi_q_obs:
            tp = obs["time_period"]
            if not tp or obs["index"] != "10001":
                continue
            date_str = self._quarter_to_date(tp)

            if obs["measure"] == "1":
                # Index Numbers → YoY計算用に保存
                index_map[tp] = obs["value"]
            elif obs["measure"] == "2":
                # QoQ
                point = ensure_point(date_str)
                point["cpi_qoq"] = round(obs["value"], 2)

        # Index Numbers からYoY計算（前年同期比）
        for tp, curr_val in index_map.items():
            parts = tp.split("-")
            if len(parts) == 2:
                prev_year = str(int(parts[0]) - 1)
                prev_tp = f"{prev_year}-{parts[1]}"
                prev_val = index_map.get(prev_tp)
                if prev_val and prev_val > 0:
                    yoy = ((curr_val / prev_val) - 1) * 100
                    date_str = self._quarter_to_date(tp)
                    point = ensure_point(date_str)
                    point["cpi_yoy"] = round(yoy, 1)

        # Phase 2: Monthly CPI_M (SA YoY + Trimmed Mean YoY) → 四半期月のみ
        # ベースデータ (2018-09～)
        for obs in cpi_m_yoy_obs:
            if obs["measure"] != "3":
                continue
            tp = obs["time_period"]
            if not tp:
                continue
            date_str = self._month_to_quarter_date(tp)
            if date_str is None:
                continue
            point = ensure_point(date_str)
            val = round(obs["value"], 1)

            if obs["index"] == "999901":
                point["cpi_sa_yoy"] = val
            elif obs["index"] == "999905":
                point["trimmed_mean_yoy"] = val

        # Phase 3: CPI dataflow SA YoY (999901/999902/999903) → 四半期月のみ
        # CPI_Mより先に更新されるため、Noneの場合のみ上書き（補完）
        for obs in cpi_sa_yoy_obs:
            if obs["measure"] != "3":
                continue
            tp = obs["time_period"]
            if not tp:
                continue
            date_str = self._month_to_quarter_date(tp)
            if date_str is None:
                continue
            point = ensure_point(date_str)
            val = round(obs["value"], 1)

            if obs["index"] == "999901" and point["cpi_sa_yoy"] is None:
                point["cpi_sa_yoy"] = val
            elif obs["index"] == "999902" and point["trimmed_mean_yoy"] is None:
                # CPI dataflow uses 999902 for Trimmed Mean (vs CPI_M's 999905)
                point["trimmed_mean_yoy"] = val
            elif obs["index"] == "999903" and point["weighted_median_yoy"] is None:
                point["weighted_median_yoy"] = val

        # Phase 4: RBA G1 Excel (TM/WM QoQ + YoY backfill)
        if rba_g1_qoq:
            for date_str, g1_data in rba_g1_qoq.items():
                point = ensure_point(date_str)
                # QoQ: always use RBA G1 (sole source)
                if "trimmed_mean_qoq" in g1_data:
                    point["trimmed_mean_qoq"] = g1_data["trimmed_mean_qoq"]
                if "weighted_median_qoq" in g1_data:
                    point["weighted_median_qoq"] = g1_data["weighted_median_qoq"]
                # YoY: backfill where ABS API didn't provide data
                if "trimmed_mean_yoy" in g1_data and point["trimmed_mean_yoy"] is None:
                    point["trimmed_mean_yoy"] = g1_data["trimmed_mean_yoy"]
                if "weighted_median_yoy" in g1_data and point["weighted_median_yoy"] is None:
                    point["weighted_median_yoy"] = g1_data["weighted_median_yoy"]

        result = sorted(parsed_points.values(), key=lambda x: x["date"])
        logger.info(f"Merged {len(result)} quarterly CPI data points")
        return result

    def get_quarterly_cpi_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """四半期CPIデータを取得（キャッシュ付き）"""
        # Redisキャッシュ
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        # ABS APIからデータ取得（3段階フェッチ + RBA G1 Excel）
        cpi_q_obs = []
        cpi_m_yoy_obs = []
        cpi_sa_yoy_obs = []
        rba_g1_qoq: Dict[str, Dict[str, float]] = {}

        # 1. CPI Quarterly (10001, Index + QoQ)
        raw_q = self._fetch_sdmx(CPI_Q_URL)
        if raw_q:
            cpi_q_obs = self._parse_sdmx_observations(raw_q)
            logger.info(f"CPI Q: {len(cpi_q_obs)} observations")

        # 2. CPI_M (SA YoY + Trimmed Mean YoY) — ベースデータ (2018-09～)
        raw_m_yoy = self._fetch_sdmx(CPI_M_YOY_URL)
        if raw_m_yoy:
            cpi_m_yoy_obs = self._parse_sdmx_observations(raw_m_yoy)
            logger.info(f"CPI_M YoY: {len(cpi_m_yoy_obs)} observations")

        # 3. CPI SA YoY (999901/999902/999903) — 最新データ補完 (2025-04～)
        raw_sa_yoy = self._fetch_sdmx(CPI_SA_YOY_URL)
        if raw_sa_yoy:
            cpi_sa_yoy_obs = self._parse_sdmx_observations(raw_sa_yoy)
            logger.info(f"CPI SA YoY: {len(cpi_sa_yoy_obs)} observations")

        # 4. RBA G1 (Trimmed Mean QoQ/YoY / Weighted Median QoQ/YoY)
        #    長期履歴は静的シードCSV(1982-06～)をベースに、ライブExcel取得が
        #    成功すれば直近分を上書き(RBAは現状Akamaiで403のためシードが主経路)。
        rba_g1_qoq = self._load_seed_csv()
        live_g1 = self._fetch_rba_g1_qoq()
        if live_g1:
            for date_str, vals in live_g1.items():
                rba_g1_qoq.setdefault(date_str, {}).update(vals)
            logger.info(f"RBA G1 live overlay: {len(live_g1)} points merged onto seed")

        if cpi_q_obs or cpi_m_yoy_obs or cpi_sa_yoy_obs or rba_g1_qoq:
            data_points = self._merge_data(cpi_q_obs, cpi_m_yoy_obs, cpi_sa_yoy_obs, rba_g1_qoq)

            if data_points:
                latest = data_points[-1] if data_points else None
                next_release = self._get_next_release()

                result = {
                    "data": data_points,
                    "latest": latest,
                    "metadata": {
                        "source": "Australian Bureau of Statistics",
                        "indicator": "Consumer Price Index (Quarterly)",
                        "frequency": "quarterly",
                        "unit": "%",
                    },
                    "next_release": next_release,
                }

                cache_payload = {
                    **result,
                    "last_updated": datetime.now(JST).isoformat(),
                }
                redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                self._save_file_cache(cache_payload)

                return {
                    **result,
                    "cached": False,
                    "source": "abs_api",
                }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [],
            "latest": None,
            "metadata": {
                "source": "Australian Bureau of Statistics",
                "error": "No data available",
            },
            "next_release": None,
            "cached": False,
            "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        """FMPから次回発表日を取得"""
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern

            return get_next_release_by_pattern(FMP_EVENT_PATTERNS[0], "AU")
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュ更新判定"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)
            now = datetime.now(JST)
            if (now - last_updated).total_seconds() >= 7 * 24 * 3600:
                return True
            return False
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return True

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU Quarterly CPI",
            "source": "ABS",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
abs_quarterly_cpi_service = AbsQuarterlyCpiService()

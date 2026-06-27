"""
オーストラリア 住宅ローン延滞率サービス

データソース:
- APRA Quarterly ADI Property Exposures Statistics
- Tab 1b (All ADIs, Residential Property)
- https://www.apra.gov.au/quarterly-authorised-deposit-taking-institution-statistics

取得系列（行はラベルで動的検出。APRAは版により行を前後挿入するため固定番号は不可）:
- "Total credit oustanding" 行: Total credit outstanding ($M)
- "Loans 30-89 days past due" 行: 30-89日延滞 ($M)
- "Non-performing loans" 行: 不良債権合計 ($M) (= 90+ days past due or impaired)
→ 計算: past_due_30_89_pct, non_performing_pct, total_arrears_pct

発表スケジュール: 四半期（2/5/8/11月頃）
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Callable
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "policy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_housing_loan_arrears_cache.json"

# APRA Property Exposures Excel URL
# Note: ファイル名・パスは四半期ごとに変わる（/sites/default/files/ → /system/files/ 等）。
# 最新版は APRA の統計ページをスクレイプして動的に解決し、失敗時のみ下記フォールバックを使う。
APRA_STATS_PAGE_URL = "https://www.apra.gov.au/quarterly-authorised-deposit-taking-institution-statistics"
APRA_BASE = "https://www.apra.gov.au"
APRA_PROPERTY_EXPOSURES_URLS = [
    "https://www.apra.gov.au/system/files/2026-05/Quarterly authorised deposit-taking institution property exposures statistics December 2025.xlsx",
    "https://www.apra.gov.au/sites/default/files/2025-09/Quarterly authorised deposit-taking institution property exposures statistics June 2025.xlsx",
]

# 対象シート
SHEET_NAME = "Tab 1b"

# 行ラベル検出パターン（固定行番号はAPRAの行挿入で容易にズレるため使わない）
def _norm_label(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""

def _is_total_credit(lbl: str) -> bool:
    # "Total credit oustanding"（APRA原文の綴り誤りを含む）。"Total credit limits" は除外。
    return "total credit" in lbl and ("outstanding" in lbl or "oustanding" in lbl)

def _is_past_due_30_89(lbl: str) -> bool:
    return "30-89" in lbl and "past due" in lbl

def _is_non_performing(lbl: str) -> bool:
    # 合計行のみ（"Non-performing loans - selected characteristics" や内訳 "Term loans" は除外）
    return lbl == "non-performing loans"


class AuHousingLoanArrearsService:
    """オーストラリア住宅ローン延滞率サービス"""

    DATA_CACHE_KEY = "australia:au_housing_loan_arrears:data"

    def __init__(self):
        pass

    def _discover_urls(self) -> List[str]:
        """APRA統計ページから最新のProperty Exposures Excel URLを発見

        ファイル名・パスは四半期ごとに変わるため、ページ内のxlsxリンクを走査して
        "property exposures" を含み "historical" を含まないものを返す。
        """
        try:
            response = requests.get(
                APRA_STATS_PAGE_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=60,
            )
            if response.status_code != 200:
                logger.warning(f"APRA stats page HTTP {response.status_code}")
                return []
            links = re.findall(r'href=["\']([^"\']+\.xlsx)["\']', response.text, re.I)
            found: List[str] = []
            for href in links:
                low = href.lower()
                if "property" not in low or "exposure" not in low:
                    continue
                if "historical" in low:
                    continue
                # 相対パスを絶対URL化（%20等のエンコードはそのままrequestsが扱える）
                url = href if href.startswith("http") else APRA_BASE + href
                if url not in found:
                    found.append(url)
            if found:
                logger.info(f"Discovered {len(found)} APRA property exposures URL(s): {found[0]}")
            return found
        except Exception as e:
            logger.warning(f"Failed to discover APRA URLs: {e}")
            return []

    def _fetch_excel(self) -> Optional[bytes]:
        """APRA Property Exposures Excelをダウンロード（発見URL優先・フォールバック付き）"""
        urls = self._discover_urls() + [
            u for u in APRA_PROPERTY_EXPOSURES_URLS
        ]
        for url in urls:
            try:
                logger.info(f"Downloading APRA Property Exposures from {url}")
                response = requests.get(
                    url,
                    headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                    timeout=120,
                )
                if response.status_code == 200:
                    logger.info(f"Downloaded {len(response.content)} bytes")
                    return response.content
                else:
                    logger.warning(f"HTTP {response.status_code} for {url}")
            except Exception as e:
                logger.warning(f"Error downloading from {url}: {e}")
                continue
        logger.error("All APRA download URLs failed")
        return None

    def _parse_excel(self, excel_bytes: bytes) -> List[Dict[str, Any]]:
        """Tab 1bから延滞データを抽出"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

            if SHEET_NAME not in wb.sheetnames:
                logger.error(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
                return []

            ws = wb[SHEET_NAME]

            # Row 4 に日付がある（Col 3以降）
            dates = []
            date_cols = []
            for col in range(3, ws.max_column + 1):
                date_val = ws.cell(row=4, column=col).value
                if date_val is not None:
                    if isinstance(date_val, datetime):
                        dates.append(date_val.strftime("%Y-%m-%d"))
                        date_cols.append(col)
                    elif isinstance(date_val, str):
                        try:
                            dt = datetime.strptime(date_val.strip(), "%d/%m/%Y")
                            dates.append(dt.strftime("%Y-%m-%d"))
                            date_cols.append(col)
                        except ValueError:
                            pass

            if not dates:
                logger.error("No dates found in Row 4")
                return []

            logger.info(f"Found {len(dates)} quarters: {dates[0]} to {dates[-1]}")

            # ラベルで対象行を動的検出（col1のラベルを走査）
            def find_row(pred: Callable[[str], bool]) -> Optional[int]:
                for r in range(1, ws.max_row + 1):
                    if pred(_norm_label(ws.cell(row=r, column=1).value)):
                        return r
                return None

            row_total = find_row(_is_total_credit)
            row_past_due = find_row(_is_past_due_30_89)
            row_npl = find_row(_is_non_performing)

            if not (row_total and row_past_due and row_npl):
                logger.error(
                    f"Required rows not found by label: "
                    f"total_credit={row_total}, past_due={row_past_due}, npl={row_npl}"
                )
                return []
            logger.info(
                f"Detected rows: total_credit={row_total}, "
                f"past_due_30_89={row_past_due}, non_performing={row_npl}"
            )

            # 各列（四半期）のデータを読み取り
            data_points = []
            for i, (date_str, col) in enumerate(zip(dates, date_cols)):
                total_credit = ws.cell(row=row_total, column=col).value
                past_due_30_89 = ws.cell(row=row_past_due, column=col).value
                non_performing = ws.cell(row=row_npl, column=col).value

                # 数値変換（APRAの "*"（秘匿）等の非数値はNone扱い）
                def _num(v: Any) -> Optional[float]:
                    if isinstance(v, (int, float)):
                        return float(v)
                    return None

                total_credit_f = _num(total_credit)
                if not total_credit_f:  # None または 0
                    continue
                past_due_30_89_f = _num(past_due_30_89) or 0.0
                non_performing_f = _num(non_performing) or 0.0

                # パーセンテージ計算
                past_due_30_89_pct = round((past_due_30_89_f / total_credit_f) * 100, 3)
                non_performing_pct = round((non_performing_f / total_credit_f) * 100, 3)
                total_arrears_pct = round(((past_due_30_89_f + non_performing_f) / total_credit_f) * 100, 3)

                data_points.append({
                    "date": date_str,
                    "past_due_30_89": past_due_30_89_pct,
                    "non_performing": non_performing_pct,
                    "total_arrears": total_arrears_pct,
                })

            logger.info(f"Built {len(data_points)} arrears data points")
            if data_points:
                latest = data_points[-1]
                logger.info(
                    f"Latest ({latest['date']}): "
                    f"30-89d={latest['past_due_30_89']}%, "
                    f"NPL={latest['non_performing']}%, "
                    f"Total={latest['total_arrears']}%"
                )

            return data_points

        except Exception as e:
            logger.error(f"Error parsing APRA Excel: {e}")
            import traceback
            traceback.print_exc()
            return []

    def get_au_housing_loan_arrears_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """住宅ローン延滞率データを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        excel_bytes = self._fetch_excel()
        if excel_bytes:
            data_points = self._parse_excel(excel_bytes)
            if data_points:
                latest = data_points[-1]
                next_release = self._get_next_release(latest.get("date"))
                result = {
                    "data": data_points,
                    "latest": latest,
                    "metadata": {
                        "source": "Australian Prudential Regulation Authority",
                        "indicator": "Housing Loan Arrears (ADI Property Exposures)",
                        "frequency": "quarterly",
                        "unit": "%",
                    },
                    "next_release": next_release,
                }
                cache_payload = {**result, "last_updated": datetime.now(JST).isoformat()}
                redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                self._save_file_cache(cache_payload)
                return {**result, "cached": False, "source": "apra_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "APRA", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self, latest_date_str: Optional[str] = None) -> Optional[Dict[str, str]]:
        """次回発表日を推定（APRA固有 - FMPイベントなし）

        APRAは四半期データ（3/6/9/12月末）を約2.5ヶ月後に公表。
        直近データの期末から次の四半期末+2.5ヶ月を推定する。
        公表パターン: 3月末データ→6月頃、6月末→9月頃、9月末→12月頃、12月末→3月頃
        """
        try:
            if not latest_date_str:
                return None

            latest_date = datetime.strptime(latest_date_str, "%Y-%m-%d")

            # 次の四半期末を算出
            quarter_ends = [(3, 31), (6, 30), (9, 30), (12, 31)]
            next_quarter_end = None
            for m, d in quarter_ends:
                qe = datetime(latest_date.year, m, d)
                if qe > latest_date:
                    next_quarter_end = qe
                    break
            if next_quarter_end is None:
                next_quarter_end = datetime(latest_date.year + 1, 3, 31)

            # 発表はデータ期末の約2.5ヶ月後
            # 例: 12月末データ → 3月中旬発表
            release_month = next_quarter_end.month + 2
            release_year = next_quarter_end.year
            if release_month > 12:
                release_month -= 12
                release_year += 1

            # 概算日（月の中旬）
            estimated_date = f"{release_year:04d}-{release_month:02d}-15"

            # 次の四半期ラベル
            quarter_labels = {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}
            q_label = quarter_labels.get(next_quarter_end.month, "")

            return {
                "date": estimated_date,
                "label": f"APRA Property Exposures ({q_label} {next_quarter_end.year})",
            }
        except Exception as e:
            logger.warning(f"Failed to estimate next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)
            now = datetime.now(JST)
            # 7日以上経過していたら更新
            return (now - last_updated).days >= 7
        except Exception:
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU Housing Loan Arrears",
            "source": "APRA (ADI Property Exposures)",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
au_housing_loan_arrears_service = AuHousingLoanArrearsService()

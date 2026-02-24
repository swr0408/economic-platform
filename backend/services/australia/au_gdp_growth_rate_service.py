"""
オーストラリア GDP成長率（GDP Growth Rate）サービス

データソース:
- ABS 5206.0 Table 1 (Key Aggregates, Chain volume measures)
- Excel URL: latest-release 経由で取得
- Series ID: A2304370T (GDP, Chain volume measures, Seasonally Adjusted, % Change from Previous Period)
- QoQ%は直接取得、YoY%はレベル値から算出

発表スケジュール: 四半期（National Accounts）
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "economy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_gdp_growth_rate_cache.json"

# ABS Excel URL (固定: latest-release は常に最新版を指す)
EXCEL_URL = "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/latest-release/5206001_key_aggregates.xlsx"

# 対象Series ID
# A2304370T = GDP, Chain volume measures, Seasonally Adjusted, % Change from Previous Period (QoQ)
QOQ_SERIES_ID = "A2304370T"
# A2304402X = GDP, Chain volume measures, Seasonally Adjusted (レベル値、YoY算出用)
LEVEL_SERIES_ID = "A2304402X"


class AuGdpGrowthRateService:
    """オーストラリアGDP成長率サービス"""

    DATA_CACHE_KEY = "australia:au_gdp_growth_rate:data"
    FMP_EVENT_PATTERN = "GDP Growth Rate QoQ"
    FMP_COUNTRY = "AU"

    def __init__(self):
        pass

    def _fetch_excel(self) -> Optional[bytes]:
        """ABS ExcelファイルをダウンロードしてBytesを返す"""
        try:
            logger.info(f"Downloading ABS 5206.0 Key Aggregates Excel from {EXCEL_URL}")
            response = requests.get(
                EXCEL_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=60,
            )
            if response.status_code != 200:
                logger.error(f"ABS Excel download returned HTTP {response.status_code}")
                return None
            return response.content
        except Exception as e:
            logger.error(f"Error downloading ABS Excel: {e}")
            return None

    def _parse_excel(self, excel_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
        """Excelファイルから QoQ% と レベル値 を抽出"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            ws = wb["Data1"]

            # 対象列を特定（Series ID行 = Row 10）
            qoq_col = None
            level_col = None
            for col in range(1, ws.max_column + 1):
                series_id = ws.cell(row=10, column=col).value
                if series_id == QOQ_SERIES_ID:
                    qoq_col = col
                if series_id == LEVEL_SERIES_ID:
                    level_col = col
                if qoq_col and level_col:
                    break

            if qoq_col is None:
                logger.error(f"Could not find QoQ Series ID {QOQ_SERIES_ID}")
                return {"qoq": [], "level": []}

            # データ行は Row 11 から
            qoq_observations = []
            level_observations = []
            for row in range(11, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    continue
                if not isinstance(date_val, datetime):
                    continue

                year = date_val.year
                month = date_val.month
                quarter = (month - 1) // 3 + 1
                time_period = f"{year}-Q{quarter}"

                # QoQ%
                qoq_val = ws.cell(row=row, column=qoq_col).value
                if qoq_val is not None:
                    qoq_observations.append({
                        "time_period": time_period,
                        "value": float(qoq_val),
                    })

                # レベル値（YoY算出用）
                if level_col:
                    level_val = ws.cell(row=row, column=level_col).value
                    if level_val is not None:
                        level_observations.append({
                            "time_period": time_period,
                            "value": float(level_val),
                        })

            logger.info(f"Parsed {len(qoq_observations)} QoQ observations, {len(level_observations)} Level observations")
            return {"qoq": qoq_observations, "level": level_observations}

        except Exception as e:
            logger.error(f"Error parsing ABS Excel: {e}")
            return {"qoq": [], "level": []}

    def _build_data_points(self, parsed: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """QoQ% と レベル値 から時系列データを構築"""
        qoq_list = parsed["qoq"]
        level_list = parsed["level"]

        # QoQをdictにマッピング
        qoq_map: Dict[str, float] = {obs["time_period"]: obs["value"] for obs in qoq_list}

        # レベル値をdictにマッピング（YoY算出用）
        level_map: Dict[str, float] = {obs["time_period"]: obs["value"] for obs in level_list}

        # 全ての時期を統合
        all_periods = sorted(set(list(qoq_map.keys()) + list(level_map.keys())))

        data_points = []
        for tp in all_periods:
            qoq = qoq_map.get(tp)
            yoy = None

            # YoY%: レベル値から前年同期比を算出
            current_level = level_map.get(tp)
            if current_level is not None:
                # 前年同期のperiodを算出
                year = int(tp.split("-Q")[0])
                quarter = int(tp.split("-Q")[1])
                prev_year_tp = f"{year - 1}-Q{quarter}"
                prev_level = level_map.get(prev_year_tp)
                if prev_level and prev_level != 0:
                    yoy = round((current_level - prev_level) / prev_level * 100, 2)

            # 日付を YYYY-MM-01 形式に変換
            year = int(tp.split("-Q")[0])
            quarter = int(tp.split("-Q")[1])
            month = (quarter - 1) * 3 + 1
            date_str = f"{year}-{month:02d}-01"

            data_points.append({
                "date": date_str,
                "qoq": round(qoq, 2) if qoq is not None else None,
                "yoy": yoy,
            })

        logger.info(f"Built {len(data_points)} GDP Growth Rate data points")
        return data_points

    def get_au_gdp_growth_rate_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """GDP成長率データを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        excel_bytes = self._fetch_excel()
        if excel_bytes:
            parsed = self._parse_excel(excel_bytes)
            if parsed["qoq"]:
                data_points = self._build_data_points(parsed)
                if data_points:
                    latest = data_points[-1]
                    next_release = self._get_next_release()
                    result = {
                        "data": data_points,
                        "latest": latest,
                        "metadata": {
                            "source": "Australian Bureau of Statistics",
                            "indicator": "GDP Growth Rate (Chain Volume Measures, Seasonally Adjusted)",
                            "frequency": "quarterly",
                            "unit": "%",
                        },
                        "next_release": next_release,
                    }
                    cache_payload = {**result, "last_updated": datetime.now(JST).isoformat()}
                    redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                    self._save_file_cache(cache_payload)
                    return {**result, "cached": False, "source": "abs_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "Australian Bureau of Statistics", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern
            result = get_next_release_by_pattern(self.FMP_EVENT_PATTERN, country=self.FMP_COUNTRY)
            if result:
                label = result.get("label", "")
                q_match = re.search(r"\(Q[1-4]\)", label)
                quarter = q_match.group(0) if q_match else ""
                result["label"] = f"GDP成長率 {quarter}".strip()
            return result
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（FMP発表日ベース）"""
        try:
            from services.australia.fmp_next_release_utils import should_refresh_by_pattern
            return should_refresh_by_pattern(
                self.FMP_EVENT_PATTERN,
                last_updated_str,
                country=self.FMP_COUNTRY,
            )
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU GDP Growth Rate", "source": "ABS (5206.0 Key Aggregates)",
            "cache_key": self.DATA_CACHE_KEY, "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


au_gdp_growth_rate_service = AuGdpGrowthRateService()

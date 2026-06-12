"""
RBA SMP (Statement on Monetary Policy) 経済予測サービス

データソース:
- RBA Forecast Archive Excel
- URL: https://www.rba.gov.au/statistics/xls/smp-forecast-archive.xlsx

7指標:
- CR:    政策金利 (Cash Rate, Level %)
- RGDP:  GDP (Year-ended growth %)
- RHFCE: 世帯消費 (Household Consumption, YoY %)
- Emp:   雇用 (Employment, YoY %)
- UR:    失業率 (Unemployment Rate, Level %)
- CPI:   消費者物価指数 (CPI, YoY %)
- TMI:   CPI トリム平均 (Trimmed Mean Inflation, YoY %)

Excel構造:
- Row 1: タイトル
- Row 4: 出版日ヘッダー (Col B以降: "Nov-2018", "Feb-2019", ... "Feb-2026")
- Row 6+: Col A = 予測対象期 ("Dec-2018", "Jun-2019", ...), Col B+ = 予測値
- 疎行列: 最新列は下方に値、過去は上方に値

発表スケジュール: 2月・5月・8月・11月
"""
import io
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import openpyxl

from core.redis_client import redis_client
from services.australia._rba_fetch import fetch_rba_xlsx

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")
AEST = ZoneInfo("Australia/Sydney")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "policy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "rba_smp_forecast_cache.json"

# RBA SMP Forecast Excel URL
SMP_FORECAST_URL = "https://www.rba.gov.au/statistics/xls/smp-forecast-archive.xlsx"

# シートマッピング
SHEET_CONFIGS = {
    "cash_rate": {
        "sheet_name": "CR",
        "name_jp": "政策金利",
        "name_en": "Cash Rate",
        "unit": "%",
    },
    "gdp": {
        "sheet_name": "RGDP",
        "name_jp": "GDP",
        "name_en": "Gross Domestic Product",
        "unit": "% yoy",
    },
    "household_consumption": {
        "sheet_name": "RHFCE",
        "name_jp": "世帯消費",
        "name_en": "Household Consumption",
        "unit": "% yoy",
    },
    "employment": {
        "sheet_name": "Emp",
        "name_jp": "雇用",
        "name_en": "Employment",
        "unit": "% yoy",
    },
    "unemployment_rate": {
        "sheet_name": "UR",
        "name_jp": "失業率",
        "name_en": "Unemployment Rate",
        "unit": "%",
    },
    "cpi": {
        "sheet_name": "CPI",
        "name_jp": "CPI",
        "name_en": "Consumer Price Index",
        "unit": "% yoy",
    },
    "trimmed_mean": {
        "sheet_name": "TMI",
        "name_jp": "CPI(トリム平均)",
        "name_en": "Trimmed Mean Inflation",
        "unit": "% yoy",
    },
}

# SMP発表月
SMP_MONTHS = [2, 5, 8, 11]


class RbaSmpForecastService:
    """RBA SMP 経済予測サービス"""

    DATA_CACHE_KEY = "australia:rba_smp_forecast:data"

    def __init__(self):
        pass

    def _download_excel(self) -> Optional[bytes]:
        """RBA SMP forecast Excel をダウンロード

        注意: ブラウザ UA 偽装は RBA WAF に 403 で弾かれる（2026-06確認）。
        共通ヘルパー fetch_rba_xlsx を使用すること。
        """
        logger.info(f"Downloading RBA SMP forecast from {SMP_FORECAST_URL}")
        return fetch_rba_xlsx(SMP_FORECAST_URL, timeout=60)

    def _parse_sheet(
        self, wb: openpyxl.Workbook, sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """
        1シートをパースして latest / previous の予測データを抽出

        Returns:
            {
                "latest": [{"date": "Jun-2026", "value": 2.1}, ...],
                "previous": [{"date": "Jun-2026", "value": 1.9}, ...],
                "latest_publication": "Feb-2026",
                "previous_publication": "Nov-2025"
            }
        """
        try:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found")
                return None

            ws = wb[sheet_name]

            # Row 4: 出版日ヘッダーを読み取り（Col B以降）
            publication_dates = []
            for col in range(2, ws.max_column + 1):
                val = ws.cell(row=4, column=col).value
                if val is not None:
                    publication_dates.append((col, str(val).strip()))

            if len(publication_dates) < 2:
                logger.warning(
                    f"Sheet '{sheet_name}': Not enough publication dates ({len(publication_dates)})"
                )
                return None

            # 最終2列: latest (最右端) と previous (その1つ左)
            latest_col, latest_pub_date = publication_dates[-1]
            previous_col, previous_pub_date = publication_dates[-2]

            logger.info(
                f"Sheet '{sheet_name}': latest={latest_pub_date} (col {latest_col}), "
                f"previous={previous_pub_date} (col {previous_col})"
            )

            # Row 6以降: 予測対象期 (Col A) と値を抽出
            latest_data = []
            previous_data = []

            for row in range(6, ws.max_row + 1):
                horizon_date = ws.cell(row=row, column=1).value
                if horizon_date is None:
                    continue

                horizon_str = str(horizon_date).strip()

                # latest列の値
                latest_val = ws.cell(row=row, column=latest_col).value
                if latest_val is not None:
                    try:
                        latest_data.append({
                            "date": horizon_str,
                            "value": round(float(latest_val), 2),
                        })
                    except (ValueError, TypeError):
                        pass

                # previous列の値
                previous_val = ws.cell(row=row, column=previous_col).value
                if previous_val is not None:
                    try:
                        previous_data.append({
                            "date": horizon_str,
                            "value": round(float(previous_val), 2),
                        })
                    except (ValueError, TypeError):
                        pass

            logger.info(
                f"Sheet '{sheet_name}': latest={len(latest_data)} points, "
                f"previous={len(previous_data)} points"
            )

            return {
                "latest": latest_data,
                "previous": previous_data,
                "latest_publication": latest_pub_date,
                "previous_publication": previous_pub_date,
            }

        except Exception as e:
            logger.error(f"Error parsing sheet '{sheet_name}': {e}")
            return None

    def _fetch_all_forecasts(self) -> Optional[Dict[str, Any]]:
        """全シートからデータを取得"""
        excel_data = self._download_excel()
        if not excel_data:
            return None

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(excel_data), data_only=True, read_only=True
            )
        except Exception as e:
            logger.error(f"Error loading Excel workbook: {e}")
            return None

        indicators = {}
        latest_pub = None
        previous_pub = None

        for indicator_key, config in SHEET_CONFIGS.items():
            sheet_name = config["sheet_name"]
            parsed = self._parse_sheet(wb, sheet_name)

            if parsed:
                indicators[indicator_key] = {
                    "latest": parsed["latest"],
                    "previous": parsed["previous"],
                }
                # メタデータ取得（最初のシートから）
                if latest_pub is None:
                    latest_pub = parsed["latest_publication"]
                    previous_pub = parsed["previous_publication"]
            else:
                indicators[indicator_key] = {
                    "latest": [],
                    "previous": [],
                }

        try:
            wb.close()
        except Exception:
            pass

        return {
            "indicators": indicators,
            "metadata": {
                "latest_publication": latest_pub or "",
                "previous_publication": previous_pub or "",
                "source": "Reserve Bank of Australia",
                "last_updated": datetime.now(JST).isoformat(),
            },
        }

    def get_smp_forecast_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """SMP予測データを取得（キャッシュ付き）"""
        # Redisキャッシュ
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "indicators": cached_data.get("indicators", {}),
                        "metadata": cached_data.get("metadata", {}),
                        "cached": True,
                        "source": "redis",
                    }

        # データ取得
        data = self._fetch_all_forecasts()

        if data and any(
            len(v.get("latest", [])) > 0
            for v in data.get("indicators", {}).values()
        ):
            cache_payload = {
                "indicators": data["indicators"],
                "metadata": data["metadata"],
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "indicators": data["indicators"],
                "metadata": data["metadata"],
                "cached": False,
                "source": "rba",
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "indicators": file_cache.get("indicators", {}),
                "metadata": file_cache.get("metadata", {}),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "indicators": {},
            "metadata": {
                "latest_publication": "",
                "previous_publication": "",
                "source": "Reserve Bank of Australia",
                "last_updated": datetime.now(JST).isoformat(),
                "error": "No data available",
            },
            "cached": False,
            "source": "none",
        }

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュ更新判定: 7日TTL + SMP発表月チェック"""
        try:
            last_updated = datetime.fromisoformat(last_updated_str)
            if last_updated.tzinfo is None:
                last_updated = last_updated.replace(tzinfo=JST)

            now = datetime.now(JST)

            # 7日以上経過していれば更新
            if (now - last_updated).total_seconds() >= 7 * 24 * 3600:
                return True

            # SMP発表月（2,5,8,11月）は1日以上経過で更新チェック
            if now.month in SMP_MONTHS:
                if (now - last_updated).total_seconds() >= 24 * 3600:
                    return True

            return False

        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return True

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        """ファイルキャッシュ読み込み"""
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        """ファイルキャッシュ保存"""
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"Cache saved to {DATA_CACHE_FILE}")
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        """キャッシュ無効化"""
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        """キャッシュ状態取得"""
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None

        return {
            "indicator": "RBA SMP Forecast",
            "source": "RBA",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "metadata": cached_data.get("metadata") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
rba_smp_forecast_service = RbaSmpForecastService()

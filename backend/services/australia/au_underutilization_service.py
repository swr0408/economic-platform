"""
オーストラリア アンダー・ユーティライゼーション（不足雇用含む）サービス

データソース:
- ABS 6202.0 Labour Force, Australia, Table 23
- Excel URL: latest-release 経由で取得
- Series ID: A85255726K (Underutilisation rate, Persons, Seasonally Adjusted)
- Series ID: A85255725J (Underemployment rate, Persons, Seasonally Adjusted)
- Series ID: A84423050A (Unemployment rate, Persons, Seasonally Adjusted)
- 3系列を1つのExcelファイルから取得（Data2, Data3, Data4シート）

発表スケジュール: 毎月（Labour Force Survey, 通常第3木曜日）
"""
import io
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "employment"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_underutilization_cache.json"

# ABS Excel URL (固定: latest-release は常に最新版を指す)
EXCEL_URL = "https://www.abs.gov.au/statistics/labour/employment-and-unemployment/labour-force-australia/latest-release/6202023.xlsx"

# 対象Series ID（すべて Seasonally Adjusted, Persons）
SERIES_CONFIG = {
    "underutilisation": {
        "series_id": "A85255726K",
        "sheet": "Data4",
        "label": "Underutilisation rate",
    },
    "underemployment": {
        "series_id": "A85255725J",
        "sheet": "Data2",
        "label": "Underemployment rate",
    },
    "unemployment": {
        "series_id": "A84423050A",
        "sheet": "Data3",
        "label": "Unemployment rate",
    },
}

# FMPイベントパターン（失業率と同じ発表日）
FMP_EVENT_PATTERN = "Unemployment Rate"
FMP_COUNTRY = "AU"


class AuUnderutilizationService:
    """オーストラリア アンダー・ユーティライゼーション サービス"""

    DATA_CACHE_KEY = "australia:au_underutilization:data"

    def __init__(self):
        pass

    def _fetch_excel(self) -> Optional[bytes]:
        """ABS ExcelファイルをダウンロードしてBytesを返す"""
        try:
            logger.info(f"Downloading ABS 6202.0 Table 23 Excel from {EXCEL_URL}")
            response = requests.get(
                EXCEL_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=90,
            )
            if response.status_code != 200:
                logger.error(f"ABS Excel download returned HTTP {response.status_code}")
                return None
            return response.content
        except Exception as e:
            logger.error(f"Error downloading ABS Excel: {e}")
            return None

    def _parse_series_from_sheet(self, wb, sheet_name: str, series_id: str) -> List[Dict[str, Any]]:
        """指定シートからSeries IDに対応するデータを抽出"""
        try:
            ws = wb[sheet_name]

            # 対象列を特定（Series ID行 = Row 10）
            target_col = None
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=10, column=col).value
                if cell_val == series_id:
                    target_col = col
                    break

            if target_col is None:
                logger.warning(f"Could not find Series ID {series_id} in sheet {sheet_name}")
                return []

            # データ行は Row 11 から
            observations = []
            for row in range(11, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    continue
                if not isinstance(date_val, datetime):
                    continue

                value = ws.cell(row=row, column=target_col).value
                if value is not None:
                    date_str = f"{date_val.year}-{date_val.month:02d}-01"
                    observations.append({
                        "date": date_str,
                        "value": round(float(value), 2),
                    })

            return observations

        except Exception as e:
            logger.error(f"Error parsing sheet {sheet_name} for series {series_id}: {e}")
            return []

    def _parse_excel(self, excel_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
        """Excelファイルから3系列を抽出"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

            result = {}
            for key, config in SERIES_CONFIG.items():
                observations = self._parse_series_from_sheet(
                    wb, config["sheet"], config["series_id"]
                )
                result[key] = observations
                logger.info(f"Parsed {len(observations)} {config['label']} observations from {config['sheet']}")

            return result

        except Exception as e:
            logger.error(f"Error parsing ABS Excel: {e}")
            return {"underutilisation": [], "underemployment": [], "unemployment": []}

    def _build_data_points(self, parsed: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """3系列のデータを日付でマージ"""
        underutil_map = {obs["date"]: obs["value"] for obs in parsed.get("underutilisation", [])}
        underempl_map = {obs["date"]: obs["value"] for obs in parsed.get("underemployment", [])}
        unempl_map = {obs["date"]: obs["value"] for obs in parsed.get("unemployment", [])}

        # 全日付を統合
        all_dates = sorted(set(
            list(underutil_map.keys()) +
            list(underempl_map.keys()) +
            list(unempl_map.keys())
        ))

        data_points = []
        for date_str in all_dates:
            underutil = underutil_map.get(date_str)
            underempl = underempl_map.get(date_str)
            unempl = unempl_map.get(date_str)

            # メイン系列（underutilisation）がある場合のみ追加
            if underutil is not None:
                data_points.append({
                    "date": date_str,
                    "underutilisation": underutil,
                    "underemployment": underempl,
                    "unemployment": unempl,
                })

        logger.info(f"Built {len(data_points)} underutilization data points")
        return data_points

    def get_au_underutilization_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """アンダー・ユーティライゼーションデータを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        excel_bytes = self._fetch_excel()
        if excel_bytes:
            parsed = self._parse_excel(excel_bytes)
            if parsed.get("underutilisation"):
                data_points = self._build_data_points(parsed)
                if data_points:
                    latest = data_points[-1]
                    next_release = self._get_next_release()
                    result = {
                        "data": data_points,
                        "latest": latest,
                        "metadata": {
                            "source": "Australian Bureau of Statistics",
                            "indicator": "Underutilisation Rate (Labour Force, Seasonally Adjusted)",
                            "frequency": "monthly",
                            "unit": "%",
                        },
                        "next_release": next_release,
                    }
                    cache_payload = {**result, "last_updated": datetime.now(JST).isoformat()}
                    redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                    self._save_file_cache(cache_payload)
                    return {**result, "cached": False, "source": "abs_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "Australian Bureau of Statistics", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern
            result = get_next_release_by_pattern(FMP_EVENT_PATTERN, country=FMP_COUNTRY)
            return result
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（FMP発表日ベース）"""
        try:
            from services.australia.fmp_next_release_utils import should_refresh_by_pattern
            return should_refresh_by_pattern(
                FMP_EVENT_PATTERN,
                last_updated_str,
                country=FMP_COUNTRY,
            )
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU Underutilization Rate",
            "source": "ABS (6202.0 Labour Force, Table 23)",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
au_underutilization_service = AuUnderutilizationService()

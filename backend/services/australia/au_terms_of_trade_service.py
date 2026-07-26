"""
オーストラリア 交易条件（Terms of Trade）サービス

データソース:
- ABS 5206.0 Table 1 (Key Aggregates, Seasonally Adjusted)
- Excel URL: latest-release 経由で取得
- Series ID: A2304200A (Terms of trade: Index)
- Series ID: A2304400V (Terms of trade: Index - Percentage changes, QoQ%)
- YoY%はIndex値から算出

発表スケジュール: 四半期（National Accounts, GDP発表と同時）
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "economy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_terms_of_trade_cache.json"

# ABS Excel URL (固定: latest-release は常に最新版を指す)
EXCEL_URL = "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/latest-release/5206001_key_aggregates.xlsx"

# 対象Series ID
# A2304200A = Terms of trade: Index, Seasonally Adjusted
INDEX_SERIES_ID = "A2304200A"
# A2304400V = Terms of trade: Index - Percentage changes (QoQ%), Seasonally Adjusted
QOQ_SERIES_ID = "A2304400V"


class AuTermsOfTradeService:
    """オーストラリア交易条件サービス"""

    DATA_CACHE_KEY = "australia:au_terms_of_trade:data"
    FMP_EVENT_PATTERN = "GDP Growth Rate QoQ"
    FMP_COUNTRY = "AU"

    def __init__(self):
        pass

    def _fetch_excel(self) -> Optional[bytes]:
        """ABS ExcelファイルをダウンロードしてBytesを返す"""
        try:
            logger.info(f"Downloading ABS 5206.0 Key Aggregates Excel from {EXCEL_URL}")
            response = requests.get(
                EXCEL_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=60,
            )
            if response.status_code != 200:
                logger.error(f"ABS Excel download returned HTTP {response.status_code}")
                return None
            return response.content
        except Exception as e:
            logger.error(f"Error downloading ABS Excel: {e}")
            return None

    def _parse_excel(self, excel_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
        """Excelファイルから Index と QoQ% を抽出"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            ws = wb["Data1"]

            # 対象列を特定（Series ID行 = Row 10）
            index_col = None
            qoq_col = None
            for col in range(1, ws.max_column + 1):
                series_id = ws.cell(row=10, column=col).value
                if series_id == INDEX_SERIES_ID:
                    index_col = col
                if series_id == QOQ_SERIES_ID:
                    qoq_col = col
                if index_col and qoq_col:
                    break

            if index_col is None:
                logger.error(f"Could not find Index Series ID {INDEX_SERIES_ID}")
                return {"index": [], "qoq": []}

            # データ行は Row 11 から
            index_observations = []
            qoq_observations = []
            for row in range(11, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None:
                    continue
                if not isinstance(date_val, datetime):
                    continue

                year = date_val.year
                month = date_val.month
                quarter = (month - 1) // 3 + 1
                time_period = f"{year}-Q{quarter}"

                # Index値
                index_val = ws.cell(row=row, column=index_col).value
                if index_val is not None:
                    index_observations.append({
                        "time_period": time_period,
                        "value": float(index_val),
                    })

                # QoQ%
                if qoq_col:
                    qoq_val = ws.cell(row=row, column=qoq_col).value
                    if qoq_val is not None:
                        qoq_observations.append({
                            "time_period": time_period,
                            "value": float(qoq_val),
                        })

            logger.info(f"Parsed {len(index_observations)} Index observations, {len(qoq_observations)} QoQ observations")
            return {"index": index_observations, "qoq": qoq_observations}

        except Exception as e:
            logger.error(f"Error parsing ABS Excel: {e}")
            return {"index": [], "qoq": []}

    def _build_data_points(self, parsed: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """Index と QoQ% から時系列データを構築（YoY%も算出）"""
        index_list = parsed["index"]
        qoq_list = parsed["qoq"]

        # マッピング
        index_map: Dict[str, float] = {obs["time_period"]: obs["value"] for obs in index_list}
        qoq_map: Dict[str, float] = {obs["time_period"]: obs["value"] for obs in qoq_list}

        # 全ての時期を統合
        all_periods = sorted(set(list(index_map.keys()) + list(qoq_map.keys())))

        data_points = []
        for tp in all_periods:
            index_val = index_map.get(tp)
            qoq = qoq_map.get(tp)
            yoy = None

            # YoY%: Index値から前年同期比を算出
            if index_val is not None:
                year = int(tp.split("-Q")[0])
                quarter = int(tp.split("-Q")[1])
                prev_year_tp = f"{year - 1}-Q{quarter}"
                prev_index = index_map.get(prev_year_tp)
                if prev_index and prev_index != 0:
                    yoy = round((index_val - prev_index) / prev_index * 100, 2)

            # 日付を YYYY-MM-01 形式に変換
            year = int(tp.split("-Q")[0])
            quarter = int(tp.split("-Q")[1])
            month = (quarter - 1) * 3 + 1
            date_str = f"{year}-{month:02d}-01"

            data_points.append({
                "date": date_str,
                "value": round(index_val, 2) if index_val is not None else None,
                "qoq": round(qoq, 2) if qoq is not None else None,
                "yoy": yoy,
            })

        logger.info(f"Built {len(data_points)} Terms of Trade data points")
        return data_points

    def get_au_terms_of_trade_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """交易条件データを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        excel_bytes = self._fetch_excel()
        if excel_bytes:
            parsed = self._parse_excel(excel_bytes)
            if parsed["index"]:
                data_points = self._build_data_points(parsed)
                if data_points:
                    latest = data_points[-1]
                    next_release = self._get_next_release()
                    result = {
                        "data": data_points,
                        "latest": latest,
                        "metadata": {
                            "source": "Australian Bureau of Statistics",
                            "indicator": "Terms of Trade (Seasonally Adjusted)",
                            "frequency": "quarterly",
                            "unit_index": "Index",
                            "unit_change": "%",
                        },
                        "next_release": next_release,
                    }
                    from services.usa.fmp_next_release_utils import guarded_last_updated
                    cache_payload = {**result, "last_updated": guarded_last_updated(self.DATA_CACHE_KEY, latest.get("date") if latest else None, datetime.now(JST).isoformat())}
                    redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                    self._save_file_cache(cache_payload)
                    return {**result, "cached": False, "source": "abs_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "Australian Bureau of Statistics", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern
            result = get_next_release_by_pattern(self.FMP_EVENT_PATTERN, country=self.FMP_COUNTRY)
            if result:
                label = result.get("label", "")
                q_match = re.search(r"\(Q[1-4]\)", label)
                quarter = q_match.group(0) if q_match else ""
                result["label"] = f"交易条件 {quarter}".strip()
            return result
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（FMP発表日ベース）"""
        try:
            from services.australia.fmp_next_release_utils import should_refresh_by_pattern
            return should_refresh_by_pattern(
                self.FMP_EVENT_PATTERN,
                last_updated_str,
                country=self.FMP_COUNTRY,
            )
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU Terms of Trade",
            "source": "ABS (5206.0 Key Aggregates)",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
au_terms_of_trade_service = AuTermsOfTradeService()

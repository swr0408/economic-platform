"""
オーストラリア 可処分所得（Gross Disposable Income）サービス

データソース:
- ABS 5206.0 Table 20 (Household Income Account, Current prices)
- Excel URL: Time Series Directory 経由で取得
- Column: GROSS DISPOSABLE INCOME, Seasonally Adjusted (Series ID: A2302939L)
- 生値（AUD百万）からQoQ%とYoY%を算出

発表スケジュール: 四半期（National Accounts と同日）
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "consumer"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_disposable_personal_income_cache.json"

# ABS Excel URL (固定: latest-release は常に最新版を指す)
EXCEL_URL = "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/latest-release/5206020_household_income.xlsx"

# 対象Series ID (Seasonally Adjusted, GROSS DISPOSABLE INCOME)
TARGET_SERIES_ID = "A2302939L"


class AuDisposablePersonalIncomeService:
    """オーストラリア可処分所得サービス"""

    DATA_CACHE_KEY = "australia:au_disposable_personal_income:data"
    FMP_EVENT_PATTERN = "GDP Growth Rate QoQ"
    FMP_COUNTRY = "AU"

    def __init__(self):
        pass

    def _fetch_excel(self) -> Optional[bytes]:
        """ABS ExcelファイルをダウンロードしてBytesを返す"""
        try:
            logger.info(f"Downloading ABS Table 20 Excel from {EXCEL_URL}")
            response = requests.get(
                EXCEL_URL,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=60,
            )
            if response.status_code != 200:
                logger.error(f"ABS Excel download returned HTTP {response.status_code}")
                return None
            return response.content
        except Exception as e:
            logger.error(f"Error downloading ABS Excel: {e}")
            return None

    def _parse_excel(self, excel_bytes: bytes) -> List[Dict[str, Any]]:
        """Excelファイルから対象系列のデータを抽出"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            ws = wb["Data1"]

            # 対象列を特定（Series ID行 = Row 10）
            target_col = None
            for col in range(1, ws.max_column + 1):
                series_id = ws.cell(row=10, column=col).value
                if series_id == TARGET_SERIES_ID:
                    target_col = col
                    break

            if target_col is None:
                # フォールバック: ヘッダーから Seasonally Adjusted の GROSS DISPOSABLE INCOME を探す
                for col in range(1, ws.max_column + 1):
                    header = str(ws.cell(row=1, column=col).value or "").upper()
                    series_type = str(ws.cell(row=3, column=col).value or "").lower()
                    if "GROSS DISPOSABLE INCOME" in header and "seasonally" in series_type:
                        target_col = col
                        break

            if target_col is None:
                logger.error("Could not find GROSS DISPOSABLE INCOME (Seasonally Adjusted) column")
                return []

            # データ行は Row 11 から
            observations = []
            for row in range(11, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                value = ws.cell(row=row, column=target_col).value
                if date_val is None or value is None:
                    continue
                # 日付をYYYY-QN形式に変換
                if isinstance(date_val, datetime):
                    year = date_val.year
                    month = date_val.month
                    quarter = (month - 1) // 3 + 1
                    time_period = f"{year}-Q{quarter}"
                else:
                    continue
                observations.append({
                    "time_period": time_period,
                    "value": float(value),
                })

            logger.info(f"Parsed {len(observations)} observations from Excel (col {target_col})")
            return observations

        except Exception as e:
            logger.error(f"Error parsing ABS Excel: {e}")
            return []

    def _build_data_points(self, observations: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """生値（AUD百万）からQoQ%とYoY%を算出"""
        observations.sort(key=lambda x: x["time_period"])

        data_points = []
        for i, obs in enumerate(observations):
            tp = obs["time_period"]
            val = obs["value"]
            qoq = None
            yoy = None

            # QoQ%: 前四半期比
            if i >= 1:
                prev = observations[i - 1]["value"]
                if prev and prev != 0:
                    qoq = round((val - prev) / prev * 100, 2)

            # YoY%: 前年同期比（4四半期前）
            if i >= 4:
                prev4 = observations[i - 4]["value"]
                if prev4 and prev4 != 0:
                    yoy = round((val - prev4) / prev4 * 100, 2)

            data_points.append({
                "date": tp,
                "qoq": qoq,
                "yoy": yoy,
            })

        logger.info(f"Built {len(data_points)} Disposable Personal Income data points")
        return data_points

    def get_au_disposable_personal_income_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """可処分所得データを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        excel_bytes = self._fetch_excel()
        if excel_bytes:
            observations = self._parse_excel(excel_bytes)
            if observations:
                data_points = self._build_data_points(observations)
                if data_points:
                    latest = data_points[-1]
                    next_release = self._get_next_release()
                    result = {
                        "data": data_points,
                        "latest": latest,
                        "metadata": {
                            "source": "Australian Bureau of Statistics",
                            "indicator": "Gross Disposable Income (Household, Seasonally Adjusted)",
                            "frequency": "quarterly",
                            "unit": "%",
                        },
                        "next_release": next_release,
                    }
                    cache_payload = {**result, "last_updated": datetime.now(JST).isoformat()}
                    redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                    self._save_file_cache(cache_payload)
                    return {**result, "cached": False, "source": "abs_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "Australian Bureau of Statistics", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern
            result = get_next_release_by_pattern(self.FMP_EVENT_PATTERN, country=self.FMP_COUNTRY)
            if result:
                label = result.get("label", "")
                q_match = re.search(r"\(Q[1-4]\)", label)
                quarter = q_match.group(0) if q_match else ""
                result["label"] = f"可処分所得 {quarter}".strip()
            return result
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（FMP発表日ベース）"""
        try:
            from services.australia.fmp_next_release_utils import should_refresh_by_pattern
            return should_refresh_by_pattern(
                self.FMP_EVENT_PATTERN,
                last_updated_str,
                country=self.FMP_COUNTRY,
            )
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU Disposable Personal Income", "source": "ABS (5206.0 Table 20)",
            "cache_key": self.DATA_CACHE_KEY, "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


au_disposable_personal_income_service = AuDisposablePersonalIncomeService()

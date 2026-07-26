"""
オーストラリア GDP詳細分析サービス

GDPデフレーター＋GDP支出項目（純輸出寄与・設備投資・家計消費）を統合

データソース:
- ABS 5206.0 Table 5 (Expenditure, Implicit Price Deflators)
  - A2303730T: GDP Deflator 指数
  - A2303803V: GDP Deflator 前期比%
- ABS 5206.0 Table 2 (Expenditure, Chain Volume Measures)
  - A2304024A: Exports Contribution to GDP growth (ppt)
  - A2304026F: Imports Contribution to GDP growth (ppt)
  - A2304110W: GFCF Level ($m, SA)
  - A2304181F: GFCF QoQ (%)
  - A2304081W: Household Consumption Level ($m, SA)
  - A2304127T: Household Consumption QoQ (%)

発表スケジュール: 四半期（National Accounts と同日）
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

# キャッシュ設定
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "australia" / "economy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "au_gdp_price_related_cache.json"

# ABS Excel URLs
DEFLATOR_EXCEL_URL = "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/latest-release/5206005_expenditure_implicit_price_deflators.xlsx"
EXPENDITURE_EXCEL_URL = "https://www.abs.gov.au/statistics/economy/national-accounts/australian-national-accounts-national-income-expenditure-and-product/latest-release/5206002_Expenditure_Volume_Measures.xlsx"

# Series IDs
# GDP Deflator (Table 5)
DEFLATOR_INDEX_SID = "A2303730T"     # GDP IPD 指数
DEFLATOR_QOQ_SID = "A2303803V"       # GDP IPD 前期比%
# Net Exports Contribution (Table 2)
EXPORTS_CONTRIBUTION_SID = "A2304024A"   # Exports contribution (ppt)
IMPORTS_CONTRIBUTION_SID = "A2304026F"   # Imports contribution (ppt)
# Capital Expenditure - GFCF (Table 2)
GFCF_LEVEL_SID = "A2304110W"            # GFCF ($m, SA)
GFCF_QOQ_SID = "A2304181F"              # GFCF QoQ (%)
# Household Consumption (Table 2)
CONSUMPTION_LEVEL_SID = "A2304081W"      # Household consumption ($m, SA)
CONSUMPTION_QOQ_SID = "A2304127T"        # Household consumption QoQ (%)


class AuGdpPriceRelatedService:
    """オーストラリアGDP詳細分析サービス"""

    DATA_CACHE_KEY = "australia:au_gdp_price_related:data"
    FMP_EVENT_PATTERN = "GDP Growth Rate QoQ"
    FMP_COUNTRY = "AU"

    def __init__(self):
        pass

    def _fetch_excel(self, url: str, label: str) -> Optional[bytes]:
        """ABS ExcelファイルをダウンロードしてBytesを返す"""
        try:
            logger.info(f"Downloading ABS {label} Excel from {url}")
            response = requests.get(
                url,
                headers={"User-Agent": "Mozilla/5.0 (economic-platform)"},
                timeout=60,
            )
            if response.status_code != 200:
                logger.error(f"ABS {label} Excel download returned HTTP {response.status_code}")
                return None
            return response.content
        except Exception as e:
            logger.error(f"Error downloading ABS {label} Excel: {e}")
            return None

    def _extract_series_from_excel(
        self, excel_bytes: bytes, series_ids: Dict[str, str]
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Excelから複数のSeries IDのデータを抽出"""
        result: Dict[str, List[Dict[str, Any]]] = {label: [] for label in series_ids}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            ws = wb["Data1"]

            # Series ID → 列番号 のマッピング（Row 10）
            col_map: Dict[str, int] = {}
            for col in range(1, ws.max_column + 1):
                sid = ws.cell(row=10, column=col).value
                for label, target_sid in series_ids.items():
                    if sid == target_sid:
                        col_map[label] = col

            if not col_map:
                logger.warning("No matching Series IDs found in Excel")
                return result

            logger.info(f"Found {len(col_map)}/{len(series_ids)} series: {list(col_map.keys())}")

            # データ行は Row 11 から
            for row in range(11, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                if date_val is None or not isinstance(date_val, datetime):
                    continue

                year = date_val.year
                month = date_val.month
                quarter = (month - 1) // 3 + 1
                time_period = f"{year}-Q{quarter}"

                for label, col_idx in col_map.items():
                    val = ws.cell(row=row, column=col_idx).value
                    if val is not None:
                        result[label].append({
                            "time_period": time_period,
                            "value": float(val),
                        })

            for label, obs in result.items():
                logger.info(f"Parsed {len(obs)} observations for {label}")

        except Exception as e:
            logger.error(f"Error parsing ABS Excel: {e}")

        return result

    def _calc_yoy_from_index(self, index_data: List[Dict[str, Any]]) -> Dict[str, float]:
        """指数/レベルデータから前年比%を算出"""
        index_map = {obs["time_period"]: obs["value"] for obs in index_data}
        yoy_map: Dict[str, float] = {}
        for tp, current in index_map.items():
            year = int(tp.split("-Q")[0])
            quarter = int(tp.split("-Q")[1])
            prev_tp = f"{year - 1}-Q{quarter}"
            prev = index_map.get(prev_tp)
            if prev and prev != 0:
                yoy_map[tp] = round((current - prev) / prev * 100, 2)
        return yoy_map

    def _build_data_points(
        self,
        deflator_data: Dict[str, List[Dict[str, Any]]],
        expenditure_data: Dict[str, List[Dict[str, Any]]],
    ) -> List[Dict[str, Any]]:
        """全系列を統合してデータポイントを構築"""
        # デフレーター
        deflator_qoq_map = {obs["time_period"]: obs["value"] for obs in deflator_data.get("deflator_qoq", [])}
        deflator_yoy_map = self._calc_yoy_from_index(deflator_data.get("deflator_index", []))

        # 支出項目
        exports_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("exports_contribution", [])}
        imports_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("imports_contribution", [])}
        gfcf_level_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("gfcf_level", [])}
        gfcf_qoq_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("gfcf_qoq", [])}
        consumption_level_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("consumption_level", [])}
        consumption_qoq_map = {obs["time_period"]: obs["value"] for obs in expenditure_data.get("consumption_qoq", [])}

        # レベルデータからYoYを算出
        gfcf_yoy_map = self._calc_yoy_from_index(expenditure_data.get("gfcf_level", []))
        consumption_yoy_map = self._calc_yoy_from_index(expenditure_data.get("consumption_level", []))

        # 全期間を統合
        all_periods = sorted(set(
            list(deflator_qoq_map.keys()) +
            list(deflator_yoy_map.keys()) +
            list(exports_map.keys()) +
            list(imports_map.keys()) +
            list(gfcf_qoq_map.keys()) +
            list(consumption_qoq_map.keys())
        ))

        data_points = []
        for tp in all_periods:
            year = int(tp.split("-Q")[0])
            quarter = int(tp.split("-Q")[1])
            month = (quarter - 1) * 3 + 1
            date_str = f"{year}-{month:02d}-01"

            # 純輸出寄与 = 輸出寄与 + 輸入寄与
            exp_val = exports_map.get(tp)
            imp_val = imports_map.get(tp)
            net_exports = None
            if exp_val is not None and imp_val is not None:
                net_exports = round(exp_val + imp_val, 2)

            dp = {
                "date": date_str,
                "deflator_qoq": round(deflator_qoq_map[tp], 2) if tp in deflator_qoq_map else None,
                "deflator_yoy": deflator_yoy_map.get(tp),
                "net_exports_contribution": net_exports,
                "exports_contribution": round(exp_val, 2) if exp_val is not None else None,
                "imports_contribution": round(imp_val, 2) if imp_val is not None else None,
                "gfcf_qoq": round(gfcf_qoq_map[tp], 2) if tp in gfcf_qoq_map else None,
                "gfcf_yoy": gfcf_yoy_map.get(tp),
                "gfcf_level": round(gfcf_level_map[tp], 0) if tp in gfcf_level_map else None,
                "consumption_qoq": round(consumption_qoq_map[tp], 2) if tp in consumption_qoq_map else None,
                "consumption_yoy": consumption_yoy_map.get(tp),
                "consumption_level": round(consumption_level_map[tp], 0) if tp in consumption_level_map else None,
            }
            data_points.append(dp)

        logger.info(f"Built {len(data_points)} GDP detail data points")
        return data_points

    def get_au_gdp_price_related_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """GDP詳細分析データを取得（キャッシュ付き）"""
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "metadata": cached_data.get("metadata", {}),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                    }

        # 2つのExcelファイルをダウンロード
        deflator_bytes = self._fetch_excel(DEFLATOR_EXCEL_URL, "Deflator")
        expenditure_bytes = self._fetch_excel(EXPENDITURE_EXCEL_URL, "Expenditure Volume Measures")

        deflator_data: Dict[str, List[Dict[str, Any]]] = {"deflator_index": [], "deflator_qoq": []}
        expenditure_data: Dict[str, List[Dict[str, Any]]] = {
            "exports_contribution": [], "imports_contribution": [],
            "gfcf_level": [], "gfcf_qoq": [],
            "consumption_level": [], "consumption_qoq": [],
        }

        if deflator_bytes:
            deflator_data = self._extract_series_from_excel(deflator_bytes, {
                "deflator_index": DEFLATOR_INDEX_SID,
                "deflator_qoq": DEFLATOR_QOQ_SID,
            })

        if expenditure_bytes:
            expenditure_data = self._extract_series_from_excel(expenditure_bytes, {
                "exports_contribution": EXPORTS_CONTRIBUTION_SID,
                "imports_contribution": IMPORTS_CONTRIBUTION_SID,
                "gfcf_level": GFCF_LEVEL_SID,
                "gfcf_qoq": GFCF_QOQ_SID,
                "consumption_level": CONSUMPTION_LEVEL_SID,
                "consumption_qoq": CONSUMPTION_QOQ_SID,
            })

        has_data = bool(
            deflator_data.get("deflator_qoq") or
            expenditure_data.get("exports_contribution") or
            expenditure_data.get("gfcf_qoq") or
            expenditure_data.get("consumption_qoq")
        )
        if has_data:
            data_points = self._build_data_points(deflator_data, expenditure_data)
            if data_points:
                latest = data_points[-1]
                next_release = self._get_next_release()
                result = {
                    "data": data_points,
                    "latest": latest,
                    "metadata": {
                        "source": "Australian Bureau of Statistics",
                        "indicator": "GDP Detail (Deflator, Net Exports, GFCF, Consumption)",
                        "frequency": "quarterly",
                        "unit": "% / ppt / $m",
                    },
                    "next_release": next_release,
                }
                from services.usa.fmp_next_release_utils import guarded_last_updated
                cache_payload = {**result, "last_updated": guarded_last_updated(self.DATA_CACHE_KEY, latest.get("date") if latest else None, datetime.now(JST).isoformat())}
                redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
                self._save_file_cache(cache_payload)
                return {**result, "cached": False, "source": "abs_excel"}

        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "metadata": file_cache.get("metadata", {}),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
            }

        return {
            "data": [], "latest": None,
            "metadata": {"source": "Australian Bureau of Statistics", "error": "No data available"},
            "next_release": None, "cached": False, "source": "none",
        }

    def _get_next_release(self) -> Optional[Dict[str, str]]:
        try:
            from services.australia.fmp_next_release_utils import get_next_release_by_pattern
            result = get_next_release_by_pattern(self.FMP_EVENT_PATTERN, country=self.FMP_COUNTRY)
            if result:
                label = result.get("label", "")
                q_match = re.search(r"\(Q[1-4]\)", label)
                quarter = q_match.group(0) if q_match else ""
                result["label"] = f"GDP詳細 {quarter}".strip()
            return result
        except Exception as e:
            logger.warning(f"Failed to get next release: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        try:
            from services.australia.fmp_next_release_utils import should_refresh_by_pattern
            return should_refresh_by_pattern(
                self.FMP_EVENT_PATTERN,
                last_updated_str,
                country=self.FMP_COUNTRY,
            )
        except Exception as e:
            logger.error(f"Error in should_refresh: {e}")
            return False

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None
        return {
            "indicator": "AU GDP Detail",
            "source": "ABS (5206.0 Table 2 & 5)",
            "cache_key": self.DATA_CACHE_KEY, "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


au_gdp_price_related_service = AuGdpPriceRelatedService()

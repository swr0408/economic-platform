"""
短期エネルギー見通し（天然ガス） (Short-Term Energy Outlook - Natural Gas) サービス

データソース:
  - EIA API v2 (STEO): 月次データ
  - エンドポイント: /v2/steo/data/
  - ファセットキー: seriesId

系列:
  - NGPRPUS: U.S. Natural Gas Total Dry Production (bcf/d)
  - NGTCPUS: U.S. Natural Gas Consumption (bcf/d)
  - NGWGPUS: Natural Gas Working Inventory U.S. Total (bcf)
  - NGEXPUS_LNG: Natural Gas LNG Gross Exports (bcf/d)
  - ZWHDPUS: Heating Degree Days U.S. Average (degree days)
  - ZWCDPUS: Cooling Degree Days U.S. Average (degree days)
  - NGHHUUS: Natural Gas Henry Hub Spot Price ($/MMBtu)

前回見通し比較:
  - アーカイブExcel: https://www.eia.gov/outlooks/steo/archives/{mon}{yy}_base.xlsx
  - 2世代保持: 最新(API) + 前回(Redis)

次回発表日: https://www.eia.gov/outlooks/steo/ からスクレイピング

更新: 6時間TTL (Redis + ファイル)
"""
import io
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

import requests

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "market"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "steo_natural_gas_cache.json"

REDIS_KEY = "market:steo_natural_gas:data"
REDIS_PREV_KEY = "market:steo_ng:previous_forecast"
REDIS_NEXT_RELEASE_KEY = "market:steo:next_release"  # Shared with crude oil STEO

EIA_API_URL = "https://api.eia.gov/v2/steo/data/"
STEO_PAGE_URL = "https://www.eia.gov/outlooks/steo/"
STEO_ARCHIVE_URL = "https://www.eia.gov/outlooks/steo/archives/{mon}{yy}_base.xlsx"

# Month abbreviations for archive URL
MONTH_ABBR = {
    1: "jan", 2: "feb", 3: "mar", 4: "apr",
    5: "may", 6: "jun", 7: "jul", 8: "aug",
    9: "sep", 10: "oct", 11: "nov", 12: "dec",
}

# Excel series ID → field name mapping (case-insensitive)
EXCEL_SERIES_MAP = {
    "ngprpus": "us_production",
    "ngtcpus": "us_consumption",
    "ngwgpus": "us_storage",
    "ngexpus_lng": "lng_exports",
    "zwhdpus": "hdd",
    "zwcdpus": "cdd",
    "nghhuus": "henry_hub_price",
}

PREV_CACHE_FILE = CACHE_DIR / "steo_ng_previous_forecast_cache.json"

# Series IDs → output field names and labels
SERIES_MAP = {
    "NGPRPUS": {"field": "us_production", "label": "米国天然ガス生産量", "unit": "bcf/d"},
    "NGTCPUS": {"field": "us_consumption", "label": "米国天然ガス消費量", "unit": "bcf/d"},
    "NGWGPUS": {"field": "us_storage", "label": "米国天然ガス貯蔵量", "unit": "bcf"},
    "NGEXPUS_LNG": {"field": "lng_exports", "label": "米国LNG輸出", "unit": "bcf/d"},
    "ZWHDPUS": {"field": "hdd", "label": "暖房稼働日 (HDD)", "unit": "degree days"},
    "ZWCDPUS": {"field": "cdd", "label": "冷房稼働日 (CDD)", "unit": "degree days"},
    "NGHHUUS": {"field": "henry_hub_price", "label": "Henry Hub天然ガス価格", "unit": "$/MMBtu"},
}

# Month name → number mapping
MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


class SteoNaturalGasService:
    """短期エネルギー見通し（天然ガス）サービス"""

    def _should_refresh(self) -> bool:
        try:
            cached = redis_client.get(REDIS_KEY)
            if not cached:
                return True
            last_updated_str = cached.get("last_updated")
            if not last_updated_str:
                return True
            last_updated = datetime.fromisoformat(last_updated_str)
            now = datetime.now(JST)
            return (now - last_updated).total_seconds() >= 6 * 3600
        except Exception:
            return True

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        if not force_refresh and not self._should_refresh():
            cached = self._load_from_redis()
            if cached and cached.get("data"):
                cached["next_release"] = self._get_next_release()
                cached["previous_forecast"] = self._load_previous_forecast()
                return cached

        try:
            data = self._build_data()
            if data and data.get("data"):
                self._save_to_cache(data)
                data["next_release"] = self._get_next_release()
                data["previous_forecast"] = self._load_previous_forecast()
                return data
        except Exception as e:
            logger.error(f"[STEO-NG] Build error: {e}")
            import traceback
            traceback.print_exc()

        cached = self._load_from_redis()
        if cached and cached.get("data"):
            cached["next_release"] = self._get_next_release()
            cached["previous_forecast"] = self._load_previous_forecast()
            return cached

        cached = self._load_from_file()
        if cached and cached.get("data"):
            cached["next_release"] = self._get_next_release()
            cached["previous_forecast"] = self._load_previous_forecast()
            return cached

        return {
            "data": [],
            "latest": None,
            "next_release": None,
            "previous_forecast": None,
            "metadata": {"source": "EIA STEO"},
            "cached": False,
            "source": "error",
            "last_updated": None,
        }

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """STEOページから次回発表日を取得（キャッシュ付き・原油STEOと共有）"""
        try:
            cached = redis_client.get(REDIS_NEXT_RELEASE_KEY)
            if cached:
                return cached
        except Exception:
            pass

        try:
            resp = requests.get(
                STEO_PAGE_URL,
                timeout=15,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            if resp.status_code != 200:
                logger.warning(f"[STEO-NG] Page fetch error: HTTP {resp.status_code}")
                return None

            text = resp.text
            match = re.search(
                r"Next\s+Release\s+Date[:\s<>/strong]+\s*(\w+)\s+(\d{1,2}),?\s+(\d{4})",
                text,
                re.IGNORECASE,
            )
            if not match:
                logger.warning("[STEO-NG] Could not find Next Release Date on page")
                return None

            month_name = match.group(1).lower()
            day = int(match.group(2))
            year = int(match.group(3))

            month_num = MONTH_MAP.get(month_name)
            if not month_num:
                logger.warning(f"[STEO-NG] Unknown month: {month_name}")
                return None

            date_str = f"{year:04d}-{month_num:02d}-{day:02d}"
            result = {
                "date": date_str,
                "label": f"STEO ({match.group(1)} {year})",
            }

            try:
                redis_client.set(REDIS_NEXT_RELEASE_KEY, result, expire=86400)
            except Exception:
                pass

            return result

        except Exception as e:
            logger.error(f"[STEO-NG] Next release scrape error: {e}")
            return None

    def _build_data(self) -> Optional[Dict[str, Any]]:
        """EIA API v2 (STEO) からデータを取得"""
        logger.info("[STEO-NG] Building data from EIA STEO API v2...")

        api_key = os.environ.get("EIA_API_KEY", "")
        if not api_key:
            logger.warning("[STEO-NG] EIA_API_KEY not set")
            return None

        # Before fetching new data, rotate current → previous
        self._rotate_previous_forecast()

        all_rows = self._fetch_from_eia(api_key)
        if not all_rows:
            return None

        # Group by date, merge series
        date_map: Dict[str, Dict[str, Any]] = {}
        for row in all_rows:
            period = row.get("period", "")
            series_id = row.get("seriesId", "")
            value = row.get("value")

            if not period or series_id not in SERIES_MAP or value is None:
                continue

            try:
                val = round(float(value), 3)
            except (ValueError, TypeError):
                continue

            date_str = f"{period}-01" if len(period) == 7 else period
            field = SERIES_MAP[series_id]["field"]

            if date_str not in date_map:
                date_map[date_str] = {"date": date_str}
            date_map[date_str][field] = val

        result_data = sorted(date_map.values(), key=lambda x: x["date"])

        # Fill missing fields with None
        all_fields = [info["field"] for info in SERIES_MAP.values()]
        for item in result_data:
            for field in all_fields:
                if field not in item:
                    item[field] = None

        # Filter out rows with no data at all
        result_data = [
            d for d in result_data
            if any(d.get(f) is not None for f in all_fields)
        ]

        if not result_data:
            logger.error("[STEO-NG] No data from EIA API")
            return None

        latest = result_data[-1].copy()
        now_str = datetime.now(JST).isoformat()

        logger.info(
            f"[STEO-NG] {len(result_data)} data points "
            f"({result_data[0]['date']} ~ {result_data[-1]['date']})"
        )

        return {
            "data": result_data,
            "latest": latest,
            "metadata": {
                "source": "EIA STEO",
                "frequency": "monthly",
                "series": {
                    info["field"]: {
                        "label": info["label"],
                        "unit": info["unit"],
                        "series_id": sid,
                    }
                    for sid, info in SERIES_MAP.items()
                },
                "data_count": len(result_data),
                "start_date": result_data[0]["date"],
                "end_date": result_data[-1]["date"],
            },
            "cached": False,
            "source": "model",
            "last_updated": now_str,
        }

    def _fetch_from_eia(self, api_key: str) -> Optional[List[Dict[str, Any]]]:
        """EIA API v2 (STEO) から月次データを取得"""
        all_rows: List[Dict[str, Any]] = []
        series_ids = list(SERIES_MAP.keys())
        offset = 0
        length = 5000

        while True:
            try:
                resp = requests.get(EIA_API_URL, params={
                    "api_key": api_key,
                    "frequency": "monthly",
                    "data[0]": "value",
                    "facets[seriesId][]": series_ids,
                    "sort[0][column]": "period",
                    "sort[0][direction]": "asc",
                    "offset": offset,
                    "length": length,
                }, timeout=60, headers={"User-Agent": "Mozilla/5.0"})

                if resp.status_code != 200:
                    logger.error(f"[STEO-NG] EIA API error: HTTP {resp.status_code}")
                    return all_rows if all_rows else None

                data = resp.json()
                rows = data.get("response", {}).get("data", [])
                total = int(data.get("response", {}).get("total", 0))
                all_rows.extend(rows)

                logger.info(
                    f"[STEO-NG] Fetched {len(rows)} rows "
                    f"(offset={offset}, total={total})"
                )

                if len(all_rows) >= total or len(rows) == 0:
                    break
                offset += length

            except Exception as e:
                logger.error(f"[STEO-NG] EIA API error: {e}")
                return all_rows if all_rows else None

        logger.info(f"[STEO-NG] Total rows fetched: {len(all_rows)}")
        return all_rows

    # ── Previous Forecast (前回見通し比較) ──────────────────────

    def _rotate_previous_forecast(self) -> None:
        """現在のSTEOデータを「前回見通し」としてRedisに保存（世代交代）"""
        try:
            current = redis_client.get(REDIS_KEY)
            if not current or not current.get("data"):
                return

            prev_data = {
                "data": current["data"],
                "forecast_month": current.get("metadata", {}).get("end_date", ""),
                "saved_at": datetime.now(JST).isoformat(),
            }

            redis_client.set(REDIS_PREV_KEY, prev_data, expire=30 * 86400)

            try:
                with open(PREV_CACHE_FILE, "w", encoding="utf-8") as f:
                    json.dump(prev_data, f, ensure_ascii=False, indent=2, default=str)
            except Exception as e:
                logger.error(f"[STEO-NG] Previous forecast file save error: {e}")

            logger.info(
                f"[STEO-NG] Rotated current forecast to previous "
                f"(forecast_month={prev_data['forecast_month']})"
            )
        except Exception as e:
            logger.error(f"[STEO-NG] Rotate previous forecast error: {e}")

    def _load_previous_forecast(self) -> Optional[Dict[str, Any]]:
        """前回見通しを読み込む（Redis → ファイル → アーカイブExcel）"""
        # 1. Redis
        try:
            prev = redis_client.get(REDIS_PREV_KEY)
            if prev and prev.get("data"):
                return prev
        except Exception:
            pass

        # 2. ファイル
        try:
            if PREV_CACHE_FILE.exists():
                with open(PREV_CACHE_FILE, "r", encoding="utf-8") as f:
                    prev = json.load(f)
                if prev and prev.get("data"):
                    try:
                        redis_client.set(REDIS_PREV_KEY, prev, expire=30 * 86400)
                    except Exception:
                        pass
                    return prev
        except Exception:
            pass

        # 3. アーカイブExcelからフェッチ（初回のみ）
        try:
            prev = self._fetch_previous_from_archive()
            if prev and prev.get("data"):
                try:
                    redis_client.set(REDIS_PREV_KEY, prev, expire=30 * 86400)
                except Exception:
                    pass
                try:
                    with open(PREV_CACHE_FILE, "w", encoding="utf-8") as f:
                        json.dump(prev, f, ensure_ascii=False, indent=2, default=str)
                except Exception:
                    pass
                return prev
        except Exception as e:
            logger.error(f"[STEO-NG] Archive fetch error: {e}")

        return None

    def _fetch_previous_from_archive(self) -> Optional[Dict[str, Any]]:
        """前月のアーカイブExcelをダウンロードしてパース"""
        try:
            import openpyxl
        except ImportError:
            logger.warning("[STEO-NG] openpyxl not installed, cannot parse archive Excel")
            return None

        now = datetime.now(JST)
        prev_month = now.month - 1
        prev_year = now.year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1

        mon_abbr = MONTH_ABBR[prev_month]
        yy = str(prev_year)[-2:]
        url = STEO_ARCHIVE_URL.format(mon=mon_abbr, yy=yy)

        logger.info(f"[STEO-NG] Fetching archive Excel: {url}")
        resp = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code != 200:
            logger.warning(
                f"[STEO-NG] Archive Excel not found: HTTP {resp.status_code} for {url}"
            )
            prev_month2 = prev_month - 1
            prev_year2 = prev_year
            if prev_month2 < 1:
                prev_month2 = 12
                prev_year2 -= 1
            mon_abbr2 = MONTH_ABBR[prev_month2]
            yy2 = str(prev_year2)[-2:]
            url2 = STEO_ARCHIVE_URL.format(mon=mon_abbr2, yy=yy2)
            logger.info(f"[STEO-NG] Trying 2 months back: {url2}")
            resp = requests.get(url2, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200:
                logger.error(f"[STEO-NG] Archive Excel not found: HTTP {resp.status_code}")
                return None
            mon_abbr = mon_abbr2
            yy = yy2

        return self._parse_archive_excel(resp.content, f"{mon_abbr}{yy}")

    def _parse_archive_excel(
        self, content: bytes, label: str
    ) -> Optional[Dict[str, Any]]:
        """アーカイブExcelをパースしてSTEOデータを抽出"""
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        # Datesシートから日付マッピングを取得
        dates_sheet = wb["Dates"]
        date_row = None
        for row in dates_sheet.iter_rows(min_row=1, max_row=15, values_only=True):
            row_list = list(row)
            if len(row_list) > 2 and row_list[1] and "DATEX" in str(row_list[1]):
                date_row = row_list
                break

        if not date_row:
            logger.error("[STEO-NG] Could not find date row in Dates sheet")
            wb.close()
            return None

        col_dates: Dict[int, str] = {}
        for i in range(2, len(date_row)):
            val = date_row[i]
            if val is not None:
                s = str(int(val)) if isinstance(val, (int, float)) else str(val)
                if len(s) == 6 and s.isdigit():
                    col_dates[i] = f"{s[:4]}-{s[4:]}-01"

        if not col_dates:
            logger.error("[STEO-NG] No date columns found in Dates sheet")
            wb.close()
            return None

        # Parse data from key sheets
        target_sheets = ["1tab", "2tab", "3atab", "3ctab", "3dtab", "3etab", "4tab", "5atab", "5btab", "6tab"]
        series_data: Dict[str, Dict[str, float]] = {}

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, values_only=True):
                row_list = list(row)
                if not row_list or not row_list[0]:
                    continue
                series_id = str(row_list[0]).strip().lower()
                if series_id not in EXCEL_SERIES_MAP:
                    continue
                field = EXCEL_SERIES_MAP[series_id]
                if field in series_data:
                    continue

                field_data: Dict[str, float] = {}
                for col_idx, date_str in col_dates.items():
                    if col_idx < len(row_list):
                        val = row_list[col_idx]
                        if val is not None:
                            try:
                                v = float(val)
                                field_data[date_str] = round(v, 3)
                            except (ValueError, TypeError):
                                pass
                if field_data:
                    series_data[field] = field_data

        wb.close()

        if not series_data:
            logger.error("[STEO-NG] No series data found in archive Excel")
            return None

        # Merge into date-indexed records
        all_dates = sorted(
            set(d for field_data in series_data.values() for d in field_data.keys())
        )
        all_fields = list(EXCEL_SERIES_MAP.values())

        result_data: List[Dict[str, Any]] = []
        for date_str in all_dates:
            item: Dict[str, Any] = {"date": date_str}
            for field in all_fields:
                item[field] = series_data.get(field, {}).get(date_str)
            if any(item.get(f) is not None for f in all_fields):
                result_data.append(item)

        if not result_data:
            return None

        logger.info(
            f"[STEO-NG] Parsed archive {label}: {len(result_data)} data points "
            f"({result_data[0]['date']} ~ {result_data[-1]['date']}), "
            f"series found: {list(series_data.keys())}"
        )

        return {
            "data": result_data,
            "forecast_month": label,
            "saved_at": datetime.now(JST).isoformat(),
        }

    def _save_to_cache(self, data: Dict[str, Any]) -> None:
        try:
            redis_client.set(REDIS_KEY, data, expire=86400)
        except Exception as e:
            logger.error(f"[STEO-NG] Redis save error: {e}")
        try:
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            logger.error(f"[STEO-NG] File save error: {e}")

    def _load_from_redis(self) -> Optional[Dict[str, Any]]:
        try:
            data = redis_client.get(REDIS_KEY)
            if data:
                data["cached"] = True
                data["source"] = "redis"
                return data
        except Exception as e:
            logger.error(f"[STEO-NG] Redis load error: {e}")
        return None

    def _load_from_file(self) -> Optional[Dict[str, Any]]:
        try:
            if DATA_CACHE_FILE.exists():
                with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                data["cached"] = True
                data["source"] = "file"
                return data
        except Exception as e:
            logger.error(f"[STEO-NG] File load error: {e}")
        return None


steo_natural_gas_service = SteoNaturalGasService()

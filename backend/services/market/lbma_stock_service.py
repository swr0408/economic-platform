"""
LBMA在庫 (London Vault Holdings) サービス

データソース:
  - LBMA CDN: 月次Excelファイル
  - URL: https://cdn.lbma.org.uk/downloads/LBMA-London-Vault-Holdings-Data-{Month}-{yyyy}.xlsx
  - 毎月5営業日目に更新 (3:00 AM JST)

更新: 6時間TTL (Redis + ファイル)
"""
import json
import logging
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "market"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "lbma_stock_cache.json"

REDIS_KEY = "market:lbma_stock:data"

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

BASE_URL = "https://cdn.lbma.org.uk/downloads/LBMA-London-Vault-Holdings-Data-{month}-{year}.xlsx"


class LbmaStockService:
    """LBMA London Vault Holdings 月次サービス"""

    def _should_refresh(self) -> bool:
        try:
            cached = redis_client.get(REDIS_KEY)
            if not cached:
                return True
            last_updated_str = cached.get("last_updated")
            if not last_updated_str:
                return True
            last_updated = datetime.fromisoformat(last_updated_str)
            now = datetime.now(JST)
            return (now - last_updated).total_seconds() >= 6 * 3600
        except Exception:
            return True

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        if not force_refresh and not self._should_refresh():
            cached = self._load_from_redis()
            if cached and cached.get("data"):
                return cached

        try:
            data = self._build_data()
            if data and data.get("data"):
                self._save_to_cache(data)
                return data
        except Exception as e:
            logger.error(f"[LbmaStock] Build error: {e}")
            import traceback
            traceback.print_exc()

        cached = self._load_from_redis()
        if cached and cached.get("data"):
            return cached

        cached = self._load_from_file()
        if cached and cached.get("data"):
            return cached

        return {
            "data": [],
            "latest": None,
            "metadata": {"source": "LBMA"},
            "cached": False,
            "source": "error",
            "last_updated": None,
        }

    def _build_data(self) -> Optional[Dict[str, Any]]:
        """LBMAからExcelをダウンロードしてパース"""
        logger.info("[LbmaStock] Building data from LBMA Excel...")

        excel_bytes = self._download_latest_excel()
        if not excel_bytes:
            logger.error("[LbmaStock] Failed to download Excel")
            return None

        return self._parse_excel(excel_bytes)

    def _download_latest_excel(self) -> Optional[bytes]:
        """最新のExcelをダウンロード。当月→前月→前々月の順で試行"""
        now = datetime.now(JST)

        # Try current month, then previous 2 months
        for offset in range(3):
            year = now.year
            month = now.month - offset
            if month <= 0:
                month += 12
                year -= 1

            month_name = MONTH_NAMES[month - 1]
            url = BASE_URL.format(month=month_name, year=year)

            try:
                logger.info(f"[LbmaStock] Trying {url}")
                resp = requests.get(url, timeout=30, headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                })
                if resp.status_code == 200 and len(resp.content) > 1000:
                    logger.info(
                        f"[LbmaStock] Downloaded {month_name} {year} "
                        f"({len(resp.content)} bytes)"
                    )
                    return resp.content
                else:
                    logger.info(
                        f"[LbmaStock] {month_name} {year}: "
                        f"status={resp.status_code}, size={len(resp.content)} (skipped)"
                    )
            except Exception as e:
                logger.warning(f"[LbmaStock] Download error for {month_name} {year}: {e}")

        return None

    def _parse_excel(self, excel_bytes: bytes) -> Optional[Dict[str, Any]]:
        """Excelファイルをパース

        構造:
          - Sheet: "London Vault Holdings Data"
          - Row 1: タイトル
          - Row 2: ヘッダー ("Month End", "Gold", "Silver")
          - Row 3: 単位行 ("Troy Ounces ('000s)")
          - Row 4+: データ（降順、最新が先頭）
          - Column A: 日付 (YYYY-MM文字列 or datetimeオブジェクト)
          - Column B: Gold Troy Ounces ('000s)
          - Column C: Silver Troy Ounces ('000s)
        """
        try:
            wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                logger.error("[LbmaStock] No active sheet")
                return None
        except Exception as e:
            logger.error(f"[LbmaStock] Excel parse error: {e}")
            return None

        result_data: List[Dict[str, Any]] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            if not row or row[0] is None:
                continue

            # Parse date
            raw_date = row[0]
            date_str = self._parse_date(raw_date)
            if not date_str:
                continue

            # Parse values
            gold_val = self._parse_number(row[1]) if len(row) > 1 else None
            silver_val = self._parse_number(row[2]) if len(row) > 2 else None

            if gold_val is None and silver_val is None:
                continue

            result_data.append({
                "date": date_str,
                "gold_thousands_oz": gold_val,
                "silver_thousands_oz": silver_val,
            })

        wb.close()

        if not result_data:
            logger.error("[LbmaStock] No data parsed from Excel")
            return None

        # Sort ascending by date
        result_data.sort(key=lambda x: x["date"])

        latest = result_data[-1].copy()
        now_str = datetime.now(JST).isoformat()

        logger.info(
            f"[LbmaStock] Parsed {len(result_data)} data points "
            f"({result_data[0]['date']} ~ {result_data[-1]['date']}), "
            f"latest gold={latest.get('gold_thousands_oz')}, "
            f"silver={latest.get('silver_thousands_oz')}"
        )

        return {
            "data": result_data,
            "latest": latest,
            "metadata": {
                "source": "LBMA",
                "frequency": "monthly",
                "unit": "Troy Ounces ('000s)",
                "data_count": len(result_data),
                "start_date": result_data[0]["date"],
                "end_date": result_data[-1]["date"],
            },
            "cached": False,
            "source": "model",
            "last_updated": now_str,
        }

    @staticmethod
    def _parse_date(raw) -> Optional[str]:
        """日付セルをYYYY-MM形式に変換"""
        if isinstance(raw, datetime):
            return raw.strftime("%Y-%m")
        if isinstance(raw, date):
            return raw.strftime("%Y-%m")
        if isinstance(raw, str):
            raw = raw.strip()
            # "YYYY-MM" format
            if len(raw) == 7 and raw[4] == "-":
                return raw
            # "YYYY-MM-DD" format
            if len(raw) >= 10 and raw[4] == "-" and raw[7] == "-":
                return raw[:7]
        return None

    @staticmethod
    def _parse_number(raw) -> Optional[float]:
        """数値セルを解析"""
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return round(float(raw), 2)
        if isinstance(raw, str):
            try:
                return round(float(raw.replace(",", "")), 2)
            except ValueError:
                return None
        return None

    def _save_to_cache(self, data: Dict[str, Any]) -> None:
        try:
            redis_client.set(REDIS_KEY, data, expire=86400)
        except Exception as e:
            logger.error(f"[LbmaStock] Redis save error: {e}")
        try:
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            logger.error(f"[LbmaStock] File save error: {e}")

    def _load_from_redis(self) -> Optional[Dict[str, Any]]:
        try:
            data = redis_client.get(REDIS_KEY)
            if data:
                data["cached"] = True
                data["source"] = "redis"
                return data
        except Exception as e:
            logger.error(f"[LbmaStock] Redis load error: {e}")
        return None

    def _load_from_file(self) -> Optional[Dict[str, Any]]:
        try:
            if DATA_CACHE_FILE.exists():
                with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                data["cached"] = True
                data["source"] = "file"
                return data
        except Exception as e:
            logger.error(f"[LbmaStock] File load error: {e}")
        return None


lbma_stock_service = LbmaStockService()

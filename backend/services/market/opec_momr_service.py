"""
OPEC MOMR (Monthly Oil Market Report) サービス

データソース:
  - 手動配置Excel: backend/data/manual_update/monthly/opec/momr-appendix-{month}-{year}.xlsx
  - Table 11-1: World oil demand and production balance (mb/d)
  - Table 11-2: Changes from last month's table (mb/d)
  - Table 11-3: OECD oil stocks and oil on water (mb)

抽出系列 (Table 11-1):
  - (a) Total world demand (mb/d)
  - Total Non-DoC liquids production and DoC NGLs (mb/d)
  - OPEC crude oil production (mb/d)
  - Total liquids production (mb/d)
  - Balance (stock change and miscellaneous) (mb/d)
  - Y-o-y change (demand / supply)

抽出系列 (Table 11-3):
  - OECD onland commercial (mb)
  - OECD SPR (mb)
  - OECD total (mb)
  - Oil-on-water (mb)
  - Days of forward consumption (commercial, SPR, total)

前月比較:
  - 最新ファイルと1つ前のファイルを自動検出して比較
  - Table 11-2 の change データも提供

更新: ファイル検出 + 6時間TTL (Redis + ファイル)
"""
import glob
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from core.redis_client import redis_client

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

EXCEL_DIR = Path(__file__).parent.parent.parent / "data" / "manual_update" / "monthly" / "opec"
CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "market"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "opec_momr_cache.json"

REDIS_KEY = "market:opec_momr:data"

# Month name → number
MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# Quarter → date mapping for converting "1Q26" → "2026-Q1"
QUARTER_MAP = {"1Q": "Q1", "2Q": "Q2", "3Q": "Q3", "4Q": "Q4"}


def _parse_period(val: Any) -> Optional[str]:
    """Excel列ヘッダーを期間文字列に変換
    Examples: 2023 → "2023", "1Q26" → "2026-Q1", "2026" → "2026"
    """
    if val is None:
        return None
    s = str(val).strip()

    # Quarterly: "1Q26", "2Q26" etc.
    m = re.match(r"^([1-4])Q(\d{2})$", s)
    if m:
        q = int(m.group(1))
        yr = 2000 + int(m.group(2))
        return f"{yr}-Q{q}"

    # Annual: "2023", "2024" etc.
    if re.match(r"^\d{4}(\.0)?$", s):
        return str(int(float(s)))

    return None


class OpecMomrService:
    """OPEC MOMR サービス"""

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """FMPから次回OPEC Monthly Report発表日を取得
        週1回（土曜 7:00 JST）にFMPを確認し、結果をRedisにキャッシュ。
        取得できない場合は空欄（None）を返す。"""
        redis_key = "market:opec_momr:next_release"
        try:
            # Redisキャッシュを確認（7日間有効）
            cached = redis_client.get(redis_key)
            if cached is not None:
                # cached == {} は「取得できなかった」を意味
                return cached if cached else None

            # キャッシュがない場合、初回は取得を試みる
            return self._fetch_and_cache_next_release(redis_key)
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Next release error: {e}")
            return None

    def _fetch_and_cache_next_release(self, redis_key: str) -> Optional[Dict[str, Any]]:
        """FMPから次回発表日を取得してRedisにキャッシュ（TTL=7日）"""
        try:
            try:
                from backend.services.usa.fmp_next_release_utils import get_next_release_from_fmp
            except ImportError:
                from services.usa.fmp_next_release_utils import get_next_release_from_fmp

            result = get_next_release_from_fmp(
                "opec_monthly_report",
                patterns=["%OPEC Monthly Report%"],
                country="US",
            )

            # 結果をキャッシュ（取得できなくても空dictをキャッシュして再問い合わせを防止）
            # TTL = 7日（604800秒）— 次の土曜チェックで更新される
            redis_client.set(redis_key, result if result else {}, expire=604800)
            return result
        except Exception as e:
            logger.error(f"[OPEC-MOMR] FMP fetch error: {e}")
            return None

    def _should_refresh(self) -> bool:
        try:
            cached = redis_client.get(REDIS_KEY)
            if not cached:
                return True
            last_updated_str = cached.get("last_updated")
            if not last_updated_str:
                return True
            last_updated = datetime.fromisoformat(last_updated_str)
            now = datetime.now(JST)
            # 6 hour TTL
            if (now - last_updated).total_seconds() >= 6 * 3600:
                return True
            # Also check if a newer file exists
            cached_file = cached.get("metadata", {}).get("latest_file", "")
            latest = self._find_latest_files()
            if latest and latest[0][1] != cached_file:
                return True
            return False
        except Exception:
            return True

    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        next_release = self._get_next_release()

        if not force_refresh and not self._should_refresh():
            cached = self._load_from_redis()
            if cached and cached.get("current"):
                cached["next_release"] = next_release
                return cached

        try:
            data = self._build_data()
            if data and data.get("current"):
                self._save_to_cache(data)
                data["next_release"] = next_release
                return data
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Build error: {e}")
            import traceback
            traceback.print_exc()

        cached = self._load_from_redis()
        if cached and cached.get("current"):
            cached["next_release"] = next_release
            return cached

        cached = self._load_from_file()
        if cached and cached.get("current"):
            cached["next_release"] = next_release
            return cached

        return {
            "current": None,
            "previous": None,
            "changes": None,
            "next_release": next_release,
            "metadata": {"source": "OPEC MOMR"},
            "cached": False,
            "source": "error",
            "last_updated": None,
        }

    def _find_latest_files(self) -> List[Tuple[datetime, str]]:
        """momr-appendix-{month}-{year}.xlsx ファイルを検索し、日付順にソート"""
        pattern = str(EXCEL_DIR / "momr-appendix-*.xlsx")
        files = glob.glob(pattern)

        result: List[Tuple[datetime, str]] = []
        for f in files:
            basename = os.path.basename(f).lower()
            # momr-appendix-march-2026.xlsx
            m = re.match(r"momr-appendix-(\w+)-(\d{4})\.xlsx$", basename)
            if not m:
                continue
            month_name = m.group(1)
            year = int(m.group(2))
            month_num = MONTH_MAP.get(month_name)
            if not month_num:
                continue
            dt = datetime(year, month_num, 1)
            result.append((dt, f))

        result.sort(key=lambda x: x[0], reverse=True)
        return result

    def _build_data(self) -> Optional[Dict[str, Any]]:
        """最新と前月のExcelファイルをパースしてデータを構築"""
        logger.info("[OPEC-MOMR] Building data from Excel files...")

        files = self._find_latest_files()
        if not files:
            logger.error("[OPEC-MOMR] No MOMR Excel files found")
            return None

        latest_dt, latest_file = files[0]
        logger.info(f"[OPEC-MOMR] Latest file: {os.path.basename(latest_file)} ({latest_dt.strftime('%Y-%m')})")

        # Parse latest
        current = self._parse_excel(latest_file)
        if not current:
            logger.error("[OPEC-MOMR] Failed to parse latest Excel")
            return None

        # Parse changes (Table 11-2) from latest file
        changes = self._parse_changes(latest_file)

        # Parse previous (if available)
        previous = None
        prev_file_name = None
        if len(files) >= 2:
            prev_dt, prev_file = files[1]
            prev_file_name = os.path.basename(prev_file)
            logger.info(f"[OPEC-MOMR] Previous file: {prev_file_name} ({prev_dt.strftime('%Y-%m')})")
            previous = self._parse_excel(prev_file)

        latest_basename = os.path.basename(latest_file)
        report_month = latest_dt.strftime("%B %Y")
        now_str = datetime.now(JST).isoformat()

        return {
            "current": current,
            "previous": previous,
            "changes": changes,
            "report_month": report_month,
            "metadata": {
                "source": "OPEC MOMR",
                "latest_file": latest_file,
                "latest_file_name": latest_basename,
                "previous_file_name": prev_file_name,
                "report_month": report_month,
            },
            "cached": False,
            "source": "model",
            "last_updated": now_str,
        }

    def _parse_excel(self, filepath: str) -> Optional[Dict[str, Any]]:
        """Table 11-1 と Table 11-3 からデータを抽出"""
        try:
            import openpyxl
        except ImportError:
            logger.error("[OPEC-MOMR] openpyxl not installed")
            return None

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Failed to open {filepath}: {e}")
            return None

        result: Dict[str, Any] = {}

        # ── Table 11-1: World oil demand and production balance ──
        try:
            ws1 = wb["Table 11 - 1"]
            t1_data = self._parse_table_11_1(ws1)
            if t1_data:
                result["table_11_1"] = t1_data
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Table 11-1 parse error: {e}")

        # ── Table 11-3: OECD oil stocks and oil on water ──
        try:
            ws3 = wb["Table 11 - 3"]
            t3_data = self._parse_table_11_3(ws3)
            if t3_data:
                result["table_11_3"] = t3_data
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Table 11-3 parse error: {e}")

        wb.close()

        if not result:
            return None

        return result

    def _get_label(self, row: list) -> Optional[str]:
        """Get label from column B (index 1), falling back to column A"""
        if len(row) > 1 and row[1] is not None:
            return str(row[1]).strip()
        if row[0] is not None:
            return str(row[0]).strip()
        return None

    def _extract_values(self, row: list, periods: Dict[int, str]) -> Dict[str, float]:
        """Extract numeric values from row for each period column"""
        result: Dict[str, float] = {}
        for col_idx, period in periods.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = round(float(row[col_idx]), 3)
                    result[period] = val
                except (ValueError, TypeError):
                    pass
        return result

    def _parse_table_11_1(self, ws) -> Optional[Dict[str, Any]]:
        """Table 11-1 パース: 需給バランス
        Labels are in column B (index 1), periods start at column C (index 2)"""
        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=65, values_only=True):
            rows_data.append(list(row))

        if len(rows_data) < 50:
            return None

        # Get period headers from row 5 (index 4) — periods start at col 2
        header_row = rows_data[4]
        periods: Dict[int, str] = {}
        for col_idx in range(2, len(header_row)):
            p = _parse_period(header_row[col_idx])
            if p:
                periods[col_idx] = p

        if not periods:
            logger.error("[OPEC-MOMR] No periods found in Table 11-1 header")
            return None

        # Extract key series by row label matching (labels in col B)
        SERIES_MAP = {
            "(a) total world demand": "world_demand",
            "(b) total non-doc liquids production and doc ngls": "non_doc_production",
            "total non-doc liquids production and doc ngls": "non_doc_production",
            "opec crude oil production": "opec_production",
            "non-opec doc crude production": "non_opec_doc_production",
            "doc crude oil production": "doc_production",
            "total liquids production": "total_production",
            "balance (stock change and miscellaneous)": "balance",
            "(a) - (b)": "call_on_opec",
        }

        # OECD stocks at bottom of Table 11-1 (labels have leading spaces)
        STOCK_LABELS = {
            "commercial": "oecd_commercial",
            "spr": "oecd_spr",
            "oil-on-water": "oil_on_water",
        }

        series_data: Dict[str, Dict[str, float]] = {}
        in_stock_section = False  # Track when we're in the OECD stock section

        yoy_demand_found = False
        yoy_supply_found = False

        for row in rows_data:
            label = self._get_label(row)
            if not label:
                continue

            label_lower = label.lower()

            # Detect OECD stock section
            if "oecd closing stock" in label_lower:
                in_stock_section = True
                continue

            if "days of forward consumption" in label_lower:
                in_stock_section = False
                continue

            # Y-o-y change rows
            if "y-o-y change" in label_lower:
                if not yoy_demand_found:
                    yoy_demand_found = True
                    field = "demand_yoy_change"
                elif not yoy_supply_found:
                    yoy_supply_found = True
                    field = "supply_yoy_change"
                else:
                    continue
                series_data[field] = self._extract_values(row, periods)
                continue

            # Main series
            matched = False
            for pattern, field in SERIES_MAP.items():
                if pattern in label_lower:
                    if field not in series_data:
                        series_data[field] = self._extract_values(row, periods)
                    matched = True
                    break

            # Stock section labels (e.g. "  Commercial", "  SPR", "Oil-on-water")
            if not matched and in_stock_section:
                clean = label.strip().lower()
                for pattern, field in STOCK_LABELS.items():
                    if clean == pattern:
                        if field not in series_data:
                            series_data[field] = self._extract_values(row, periods)
                        break

        if not series_data:
            return None

        # Build time-series array
        all_periods = sorted(set(p for d in series_data.values() for p in d.keys()))
        all_fields = list(series_data.keys())

        data_array: List[Dict[str, Any]] = []
        for period in all_periods:
            item: Dict[str, Any] = {"period": period}
            for field in all_fields:
                item[field] = series_data.get(field, {}).get(period)
            data_array.append(item)

        logger.info(
            f"[OPEC-MOMR] Table 11-1: {len(data_array)} periods, "
            f"fields: {all_fields}"
        )

        return {
            "data": data_array,
            "fields": all_fields,
            "periods": all_periods,
        }

    def _parse_table_11_3(self, ws) -> Optional[Dict[str, Any]]:
        """Table 11-3 パース: OECD在庫
        Labels in column B (index 1), periods start at col 3 (annual) and col 7 (quarterly)"""
        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
            rows_data.append(list(row))

        if len(rows_data) < 20:
            return None

        # Get period headers from row 5 (index 4) — scan all columns
        header_row = rows_data[4]
        periods: Dict[int, str] = {}
        for col_idx in range(2, len(header_row)):
            p = _parse_period(header_row[col_idx])
            if p:
                periods[col_idx] = p

        if not periods:
            return None

        # Series to extract (labels in col B)
        SERIES_MAP = {
            "oecd onland commercial": "oecd_commercial",
            "oecd spr": "oecd_spr",
            "oecd total": "oecd_total",
            "oil-on-water": "oil_on_water",
        }

        DAYS_MAP = {
            "oecd onland commercial": "days_commercial",
            "oecd spr": "days_spr",
            "oecd total": "days_total",
        }

        series_data: Dict[str, Dict[str, float]] = {}
        in_days_section = False

        for row in rows_data:
            label = self._get_label(row)
            if not label:
                continue

            label_lower = label.lower()

            if "days of forward consumption" in label_lower:
                in_days_section = True
                continue

            if not in_days_section:
                for pattern, field in SERIES_MAP.items():
                    if pattern == label_lower:
                        if field not in series_data:
                            series_data[field] = self._extract_values(row, periods)
                        break
            else:
                for pattern, field in DAYS_MAP.items():
                    if pattern == label_lower:
                        if field not in series_data:
                            series_data[field] = self._extract_values(row, periods)
                        break

        if not series_data:
            return None

        # Build time-series
        all_periods = sorted(set(p for d in series_data.values() for p in d.keys()))
        all_fields = list(series_data.keys())

        data_array: List[Dict[str, Any]] = []
        for period in all_periods:
            item: Dict[str, Any] = {"period": period}
            for field in all_fields:
                item[field] = series_data.get(field, {}).get(period)
            data_array.append(item)

        logger.info(
            f"[OPEC-MOMR] Table 11-3: {len(data_array)} periods, "
            f"fields: {all_fields}"
        )

        return {
            "data": data_array,
            "fields": all_fields,
            "periods": all_periods,
        }

    def _parse_changes(self, filepath: str) -> Optional[Dict[str, Any]]:
        """Table 11-2 パース: 前月からの変更
        Same column layout as Table 11-1 (labels in col B, periods from col 2)"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb["Table 11 - 2"]
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Table 11-2 open error: {e}")
            return None

        rows_data = []
        for row in ws.iter_rows(min_row=1, max_row=65, values_only=True):
            rows_data.append(list(row))

        wb.close()

        if len(rows_data) < 50:
            return None

        # Get period headers from row 5 (index 4) — periods start at col 2
        header_row = rows_data[4]
        periods: Dict[int, str] = {}
        for col_idx in range(2, len(header_row)):
            p = _parse_period(header_row[col_idx])
            if p:
                periods[col_idx] = p

        if not periods:
            return None

        SERIES_MAP = {
            "(a) total world demand": "world_demand_chg",
            "(b) total non-doc liquids production and doc ngls": "non_doc_production_chg",
            "total non-doc liquids production and doc ngls": "non_doc_production_chg",
            "balance (stock change and miscellaneous)": "balance_chg",
            "(a) - (b)": "call_on_opec_chg",
        }

        series_data: Dict[str, Dict[str, float]] = {}

        for row in rows_data:
            label = self._get_label(row)
            if not label:
                continue

            label_lower = label.lower()
            for pattern, field in SERIES_MAP.items():
                if pattern in label_lower:
                    if field not in series_data:
                        series_data[field] = self._extract_values(row, periods)
                    break

        if not series_data:
            return None

        all_periods = sorted(set(p for d in series_data.values() for p in d.keys()))
        all_fields = list(series_data.keys())

        data_array: List[Dict[str, Any]] = []
        for period in all_periods:
            item: Dict[str, Any] = {"period": period}
            for field in all_fields:
                item[field] = series_data.get(field, {}).get(period)
            data_array.append(item)

        return {
            "data": data_array,
            "fields": all_fields,
        }

    def _save_to_cache(self, data: Dict[str, Any]) -> None:
        try:
            redis_client.set(REDIS_KEY, data, expire=86400)
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Redis save error: {e}")
        try:
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
        except Exception as e:
            logger.error(f"[OPEC-MOMR] File save error: {e}")

    def _load_from_redis(self) -> Optional[Dict[str, Any]]:
        try:
            data = redis_client.get(REDIS_KEY)
            if data:
                data["cached"] = True
                data["source"] = "redis"
                return data
        except Exception as e:
            logger.error(f"[OPEC-MOMR] Redis load error: {e}")
        return None

    def _load_from_file(self) -> Optional[Dict[str, Any]]:
        try:
            if DATA_CACHE_FILE.exists():
                with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                data["cached"] = True
                data["source"] = "file"
                return data
        except Exception as e:
            logger.error(f"[OPEC-MOMR] File load error: {e}")
        return None


opec_momr_service = OpecMomrService()

"""
日本 企業の物価見通し（日銀短観 表7）サービス

全規模合計・全産業・1年後の「販売価格の見通し」と「物価全般の見通し」を
四半期時系列で提供する。

データソース:
- 日銀短観 概要（gaiyo）ZIP に同梱される Excel(GA_J1.xlsx) の
  「表7 企業の物価見通し」シート
- ZIP: https://www.boj.or.jp/statistics/tk/gaiyo/{group}/tka{YY}{QQ}.zip
  （年グループ別ディレクトリ。各 ZIP は当該調査の「今回」値を含む）

注:
- 標準の単独PDF(tkc{YY}{QQ}.pdf)は 2019Q4 で公開停止し、以降は概要に統合。
- 「企業の物価見通し」調査自体は 2014Q1(3月調査)開始。それ以前のExcelには
  表7が無いため自動スキップする。
- 各ZIPは「今回」(=当該調査月)と「前回」を持つ。本サービスは各ZIPの「今回」を
  採用し、ZIPファイル名の調査月で日付を決める。

発表スケジュール: 四半期（3/6/9/12月調査、翌月初〜中旬公表）
キャッシュ方式: FMP発表日時ベース（INDICATOR_ID=boj_tankan）。初回のみ全ZIPから
履歴を構築し、以降は最新ZIPのみ取得して差分マージする。
"""
import io
import json
import re
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import requests

from core.redis_client import redis_client
from zoneinfo import ZoneInfo

JST = ZoneInfo("Asia/Tokyo")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "japan" / "price"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "japan_inflation_outlook_cache.json"

GAIYO_INDEX_URL = "https://www.boj.or.jp/statistics/tk/gaiyo/index.htm"
BOJ_ORIGIN = "https://www.boj.or.jp"
HTTP_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; econ-platform/1.0)"}

# 調査開始（企業の物価見通し）。これより前の概要Excelには表7が無い。
SURVEY_START = "2014-03-01"


class BOJInflationOutlookService:
    """日銀短観 企業の物価見通し（全規模合計・全産業・1年後）サービス"""

    DATA_CACHE_KEY = "japan:boj_inflation_outlook:data"
    CACHE_TTL = 60 * 60 * 24 * 7  # 7日
    INDICATOR_ID = "boj_tankan"  # FMP発表スケジュール共用

    def __init__(self):
        self._fmp_utils = None

    # ------------------------------------------------------------------ #
    # FMP next release
    # ------------------------------------------------------------------ #
    @property
    def fmp_utils(self):
        if self._fmp_utils is None:
            from services.japan.fmp_next_release_utils import (
                get_next_release_from_fmp,
                should_refresh_by_fmp_schedule,
            )
            self._fmp_utils = {
                "get_next_release": get_next_release_from_fmp,
                "should_refresh": should_refresh_by_fmp_schedule,
            }
        return self._fmp_utils

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        try:
            return self.fmp_utils["get_next_release"](self.INDICATOR_ID)
        except Exception as e:
            print(f"[inflation_outlook] next_release error: {e}")
            return None

    def _should_refresh(self, last_updated_str: str) -> bool:
        try:
            return self.fmp_utils["should_refresh"](self.INDICATOR_ID, last_updated_str)
        except Exception as e:
            print(f"[inflation_outlook] _should_refresh error: {e}")
            try:
                last_updated = datetime.fromisoformat(last_updated_str)
                if last_updated.tzinfo is None:
                    last_updated = last_updated.replace(tzinfo=JST)
                return (datetime.now(JST) - last_updated).days >= 7
            except Exception:
                return True

    # ------------------------------------------------------------------ #
    # Public API
    # ------------------------------------------------------------------ #
    def get_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """物価見通しデータ（四半期時系列）を取得"""
        if not force_refresh:
            cached = redis_client.get(self.DATA_CACHE_KEY)
            if cached and cached.get("last_updated") and not self._should_refresh(cached["last_updated"]):
                cached["cached"] = True
                cached["source"] = "redis"
                cached["next_release"] = self._get_next_release()
                return cached

            file_cache = self._load_file_cache()
            if file_cache and file_cache.get("last_updated") and not self._should_refresh(file_cache["last_updated"]):
                redis_client.set(self.DATA_CACHE_KEY, file_cache, expire=0)
                file_cache["cached"] = True
                file_cache["source"] = "file"
                file_cache["next_release"] = self._get_next_release()
                return file_cache

        # 既存履歴をベースに最新分だけ差分取得（初回は全ZIPバックフィル）
        existing = self._existing_history()
        data = self._refresh_history(existing)

        if data:
            latest = data[-1]
            next_release = self._get_next_release()
            payload = {
                "data": data,
                "latest": latest,
                "metadata": {
                    "source": "Bank of Japan (BOJ) Tankan Survey - 企業の物価見通し（表7）",
                    "description": "全規模合計・全産業・1年後 / 今回調査値",
                    "category": "全規模合計・全産業",
                    "horizon": "1年後",
                    "unit": "%",
                    "frequency": "Quarterly",
                },
                "next_release": next_release,
                "last_updated": datetime.now(JST).isoformat(),
            }
            redis_client.set(self.DATA_CACHE_KEY, payload, expire=0)
            self._save_file_cache(payload)
            result = dict(payload)
            result.update({"cached": False, "source": "boj"})
            return result

        # 取得失敗時はファイルキャッシュにフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            file_cache["cached"] = True
            file_cache["source"] = "file (fallback)"
            file_cache["next_release"] = self._get_next_release()
            return file_cache

        return {
            "data": [],
            "latest": None,
            "metadata": {},
            "next_release": self._get_next_release(),
            "cached": False,
            "source": "none",
            "last_updated": None,
            "error": "No data available",
        }

    # ------------------------------------------------------------------ #
    # History build / refresh
    # ------------------------------------------------------------------ #
    def _existing_history(self) -> List[Dict[str, Any]]:
        cached = redis_client.get(self.DATA_CACHE_KEY)
        if cached and cached.get("data"):
            return list(cached["data"])
        file_cache = self._load_file_cache()
        if file_cache and file_cache.get("data"):
            return list(file_cache["data"])
        return []

    def _refresh_history(self, existing: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """既存履歴に差分マージ。履歴が無ければ全ZIPから構築。

        取りこぼし対策: キャッシュに無い全四半期(=欠損ギャップ)を埋めつつ、
        最新2四半期は毎回取得して公表直後の改定も反映する。通常は1〜2本のみの
        取得で済むため高速、長期間未更新でもギャップを自動補完する。
        """
        zip_urls = self._discover_zip_urls()
        if not zip_urls:
            return existing

        if not existing:
            return self._build_full_history(zip_urls)

        by_date = {d["date"]: d for d in existing}
        available = {d for d in zip_urls if d >= SURVEY_START}
        latest = set(sorted(available, reverse=True)[:2])
        missing = available - set(by_date)
        to_fetch = sorted(missing | latest)

        if not to_fetch:
            self._seed_manual(by_date)
            return sorted(by_date.values(), key=lambda x: x["date"])

        def work(date_str: str) -> Optional[Dict[str, Any]]:
            return self._fetch_point(date_str, zip_urls[date_str])

        with ThreadPoolExecutor(max_workers=4) as ex:
            for point in ex.map(work, to_fetch):
                if point:
                    by_date[point["date"]] = point
        self._seed_manual(by_date)
        return sorted(by_date.values(), key=lambda x: x["date"])

    def _seed_manual(self, by_date: Dict[str, Dict[str, Any]]) -> None:
        """手動CSV種（BOJ時系列検索のネイティブCSV）で2021Q1より前の期間を補完。
        gaiyo ZIP は2021Q1以降のみのため、それ以前(調査開始2014Q1〜)を埋める。"""
        try:
            from services.japan.boj_tankan_manual_seed import seed_into
            seed_into("inflation_outlook", by_date)
        except Exception as e:
            print(f"[inflation_outlook] manual seed failed: {e}")

    def _build_full_history(self, zip_urls: Dict[str, str]) -> List[Dict[str, Any]]:
        targets = {d: u for d, u in zip_urls.items() if d >= SURVEY_START}
        print(f"[inflation_outlook] full backfill: {len(targets)} surveys")

        def work(item: Tuple[str, str]) -> Optional[Dict[str, Any]]:
            date_str, url = item
            return self._fetch_point(date_str, url)

        results: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=6) as ex:
            for point in ex.map(work, sorted(targets.items())):
                if point:
                    results.append(point)
        by_date = {p["date"]: p for p in results}
        self._seed_manual(by_date)
        merged = sorted(by_date.values(), key=lambda x: x["date"])
        print(f"[inflation_outlook] built {len(merged)} points")
        return merged

    def _fetch_point(self, date_str: str, url: str) -> Optional[Dict[str, Any]]:
        try:
            resp = requests.get(url, headers=HTTP_HEADERS, timeout=40)
            if resp.status_code != 200 or len(resp.content) < 5000:
                return None
            parsed = self._parse_outlook_zip(resp.content)
            if not parsed:
                return None
            selling, general = parsed
            if selling is None and general is None:
                return None
            return {
                "date": date_str,
                "selling_price_outlook": selling,
                "general_price_outlook": general,
            }
        except Exception as e:
            print(f"[inflation_outlook] fetch {url} failed: {e}")
            return None

    # ------------------------------------------------------------------ #
    # Discovery & parsing
    # ------------------------------------------------------------------ #
    def _discover_zip_urls(self) -> Dict[str, str]:
        """概要indexから全 tka{YY}{QQ}.zip を {date_str: url} で収集。"""
        urls: Dict[str, str] = {}
        try:
            idx = requests.get(GAIYO_INDEX_URL, headers=HTTP_HEADERS, timeout=30)
            dir_paths = sorted(set(re.findall(r'href="(/statistics/tk/gaiyo/\d{4}/index\.htm)"', idx.text)))
            for dp in dir_paths:
                try:
                    sub = requests.get(BOJ_ORIGIN + dp, headers=HTTP_HEADERS, timeout=30)
                    for m in re.finditer(r'href="(/statistics/tk/gaiyo/\d{4}/tka(\d{2})(\d{2})\.zip)"', sub.text):
                        path, yy, qq = m.group(1), m.group(2), m.group(3)
                        date_str = f"20{yy}-{qq}-01"
                        urls[date_str] = BOJ_ORIGIN + path
                except Exception as e:
                    print(f"[inflation_outlook] dir {dp} failed: {e}")
        except Exception as e:
            print(f"[inflation_outlook] discover failed: {e}")
        return urls

    @staticmethod
    def _norm(value: Any) -> Optional[str]:
        if isinstance(value, str):
            return re.sub(r"\s+", "", value)
        return None

    def _parse_outlook_zip(self, content: bytes) -> Optional[Tuple[Optional[float], Optional[float]]]:
        """概要ZIP内Excelの「表7 企業の物価見通し」から
        全規模合計・全産業・1年後・今回の (販売価格, 物価全般) を返す。"""
        try:
            z = zipfile.ZipFile(io.BytesIO(content))
            xlsx_names = [n for n in z.namelist() if n.lower().endswith(".xlsx")]
            if not xlsx_names:
                return None
            wb = openpyxl.load_workbook(io.BytesIO(z.read(xlsx_names[0])), data_only=True, read_only=True)
            try:
                ws = None
                for sn in wb.sheetnames:
                    title = wb[sn].cell(1, 1).value or ""
                    if "物価見通し" in str(title):
                        ws = wb[sn]
                        break
                if ws is None:
                    return None

                rows = list(ws.iter_rows(min_row=1, max_row=80, max_col=10, values_only=True))

                # ヘッダから値列を検出（販売価格の見通し / 物価全般の見通し）
                sell_col = gen_col = None
                for r in rows[:10]:
                    for j, c in enumerate(r):
                        if isinstance(c, str):
                            if sell_col is None and "販売価格" in c:
                                sell_col = j
                            if gen_col is None and "物価全般" in c:
                                gen_col = j
                if sell_col is None or gen_col is None:
                    sell_col, gen_col = 4, 6  # フォールバック（既知レイアウト E/G 列）

                cat = ind = hor = None
                selling = general = None
                for r in rows:
                    a = self._norm(r[0]) if len(r) > 0 else None
                    b = self._norm(r[1]) if len(r) > 1 else None
                    cc = self._norm(r[2]) if len(r) > 2 else None
                    d = self._norm(r[3]) if len(r) > 3 else None
                    if a:
                        cat = a
                    if b:
                        ind = b
                    if cc:
                        hor = cc
                    if cat == "全規模合計" and ind == "全産業" and hor and ("1年後" in hor or "１年後" in hor):
                        if d == "今回":
                            sv = r[sell_col] if len(r) > sell_col else None
                            gv = r[gen_col] if len(r) > gen_col else None
                            selling = round(float(sv), 2) if isinstance(sv, (int, float)) else None
                            general = round(float(gv), 2) if isinstance(gv, (int, float)) else None
                return (selling, general)
            finally:
                wb.close()
        except Exception as e:
            print(f"[inflation_outlook] parse failed: {e}")
            return None

    # ------------------------------------------------------------------ #
    # Cache helpers
    # ------------------------------------------------------------------ #
    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[inflation_outlook] load file cache failed: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[inflation_outlook] save file cache failed: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached = redis_client.get(self.DATA_CACHE_KEY) if exists else None
        return {
            "indicator": "Japan BOJ Tankan Inflation Outlook (企業の物価見通し)",
            "source": "Bank of Japan (BOJ) Tankan Survey - 表7",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": exists,
            "last_updated": cached.get("last_updated") if cached else None,
            "data_count": len(cached.get("data", [])) if cached else 0,
            "latest": cached.get("latest") if cached else None,
            "next_release": self._get_next_release(),
            "file_cache_exists": DATA_CACHE_FILE.exists(),
        }


# シングルトンインスタンス
boj_inflation_outlook_service = BOJInflationOutlookService()

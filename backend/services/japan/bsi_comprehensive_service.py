"""
法人企業景気予測調査 包括サービス (Business Survey Index - BSI Comprehensive)
財務省・e-StatからBSI全シートデータを取得

データソース: e-Stat (https://www.e-stat.go.jp/)
シート:
- 参考-1: 貴社の景況 - Business conditions (current judgment)
- 参考-2: 国内の景況 - Domestic business conditions
- 参考-3: 設備投資 - Capital investment
- 参考-4: 雇用 - Employment

各シートには以下が含まれる:
- 実績 (actual)
- 翌期見通し (forecast1)
- 翌々期見通し (forecast2)
"""

import json
import requests
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from pathlib import Path
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

from io import BytesIO

try:
    from backend.core.redis_client import redis_client
except ImportError:
    from core.redis_client import redis_client


class BSIComprehensiveService:
    """法人企業景気予測調査 包括サービス"""

    CACHE_KEY_PREFIX = "japan:bsi_comprehensive"
    CACHE_TTL = 60 * 60 * 24 * 7  # 7 days

    # Data source URL
    DATA_URL = "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040283549&fileKind=0"

    # Sheet information
    SHEET_INFO = {
        0: {"name": "ref1", "title": "貴社の景況", "description": "Business Conditions Index"},
        1: {"name": "ref2", "title": "国内の景況", "description": "Domestic Business Conditions"},
        2: {"name": "ref3", "title": "設備投資", "description": "Capital Investment"},
        3: {"name": "ref4", "title": "雇用", "description": "Employment"}
    }

    # Column mapping for all sheets
    COLUMN_MAP = {
        "large_all_industries": {"actual": 2, "forecast1": 3, "forecast2": 4},
        "large_manufacturing": {"actual": 5, "forecast1": 6, "forecast2": 7},
        "large_non_manufacturing": {"actual": 8, "forecast1": 9, "forecast2": 10},
        "medium_all_industries": {"actual": 11, "forecast1": 12, "forecast2": 13},
        "small_all_industries": {"actual": 20, "forecast1": 21, "forecast2": 22},
    }

    def __init__(self):
        self.cache_dir = Path(__file__).parent.parent.parent / "data" / "cache" / "bsi"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _get_cache_key(self, sheet_name: str, period_type: str) -> str:
        """Get cache key for specific sheet and period type"""
        return f"{self.CACHE_KEY_PREFIX}:{sheet_name}:{period_type}"

    def _get_from_cache(self, sheet_name: str, period_type: str) -> Optional[Dict[str, Any]]:
        """Get data from Redis cache"""
        try:
            cached = redis_client.get(self._get_cache_key(sheet_name, period_type))
            if cached:
                return cached
        except Exception as e:
            print(f"Redis get error: {e}")
        return None

    def _set_to_cache(self, sheet_name: str, period_type: str, data: Dict[str, Any]) -> bool:
        """Set data to Redis cache"""
        try:
            redis_client.set(self._get_cache_key(sheet_name, period_type), data, expire=self.CACHE_TTL)
            return True
        except Exception as e:
            print(f"Redis set error: {e}")
        return False

    def _get_file_cache(self, sheet_name: str, period_type: str) -> Optional[Dict[str, Any]]:
        """Get data from file cache"""
        cache_file = self.cache_dir / f"bsi_{sheet_name}_{period_type}.json"
        if cache_file.exists():
            try:
                mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
                if datetime.now() - mtime < timedelta(days=7):
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
            except Exception as e:
                print(f"File cache read error: {e}")
        return None

    def _set_file_cache(self, sheet_name: str, period_type: str, data: Dict[str, Any]) -> bool:
        """Set data to file cache"""
        cache_file = self.cache_dir / f"bsi_{sheet_name}_{period_type}.json"
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            return True
        except Exception as e:
            print(f"File cache write error: {e}")
        return False

    def _parse_quarter(self, quarter_str: str) -> str:
        """Parse quarter string to Q1-Q4 format"""
        quarter_map = {
            "1~3月": "Q1", "4~6月": "Q2", "7~9月": "Q3", "10~12月": "Q4",
            "1～3月": "Q1", "4～6月": "Q2", "7～9月": "Q3", "10～12月": "Q4",
            "3月末": "Q1", "6月末": "Q2", "9月末": "Q3", "12月末": "Q4"
        }
        return quarter_map.get(quarter_str, quarter_str)

    def _quarter_to_date(self, year: int, quarter: str) -> str:
        """Convert year and quarter to date string"""
        if not quarter.startswith("Q"):
            return f"{year}-03-01"

        quarter_num = quarter[1]
        month_map = {"1": "03", "2": "06", "3": "09", "4": "12"}
        month = month_map.get(quarter_num, "03")
        return f"{year}-{month}-01"

    def _get_sheet_index(self, sheet_name: str) -> Optional[int]:
        """Get sheet index from sheet name"""
        for idx, info in self.SHEET_INFO.items():
            if info["name"] == sheet_name:
                return idx
        return None

    def _fetch_from_estat(self, sheet_index: int, period_type: str) -> List[Dict[str, Any]]:
        """Fetch data from e-Stat for specific sheet and period type"""
        if not openpyxl:
            raise ImportError("openpyxl is required for parsing Excel files")

        print(f"Fetching BSI data from {self.DATA_URL} (sheet={sheet_index}, period={period_type})")

        response = requests.get(self.DATA_URL, timeout=30)
        response.raise_for_status()

        wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)

        if sheet_index >= len(wb.worksheets):
            raise ValueError(f"Invalid sheet index: {sheet_index}")

        sheet = wb.worksheets[sheet_index]

        data = []
        current_year = None

        def safe_float(val):
            if val is None:
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        # Data starts from row 7
        for row_idx in range(7, sheet.max_row):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Skip footer or empty rows
            if row[1] is None or "※" in str(row[1]):
                break

            # Extract year if present
            if row[0] is not None:
                year_str = str(row[0])
                year_match = re.search(r'(\d{4})', year_str)
                if year_match:
                    current_year = int(year_match.group(1))

            # Extract quarter
            quarter_str = str(row[1])
            quarter = self._parse_quarter(quarter_str)

            if current_year is None:
                continue

            # Create date string
            date_str = self._quarter_to_date(current_year, quarter)

            try:
                data_point = {
                    "date": date_str,
                    "year": current_year,
                    "quarter": quarter,
                }

                # Add all industry/size combinations for the specified period
                for category, col_info in self.COLUMN_MAP.items():
                    col_idx = col_info[period_type]
                    value = row[col_idx] if col_idx < len(row) else None
                    data_point[category] = safe_float(value)

                data.append(data_point)

            except (ValueError, TypeError, IndexError) as e:
                print(f"Error parsing row {row_idx}: {e}")
                continue

        wb.close()
        return data

    def get_comprehensive_data(
        self,
        sheet_name: str,
        period_type: str,
        force_refresh: bool = False
    ) -> Dict[str, Any]:
        """Get BSI data for specific sheet and period type"""
        # Validate inputs
        sheet_index = self._get_sheet_index(sheet_name)
        if sheet_index is None:
            return {
                "data": [],
                "error": f"Invalid sheet_name: {sheet_name}. Must be one of: ref1, ref2, ref3, ref4"
            }

        if period_type not in ["actual", "forecast1", "forecast2"]:
            return {
                "data": [],
                "error": f"Invalid period_type: {period_type}. Must be: actual, forecast1, or forecast2"
            }

        sheet_info = self.SHEET_INFO[sheet_index]

        # Check cache
        if not force_refresh:
            cached_data = self._get_from_cache(sheet_name, period_type)
            if cached_data:
                cached_data["cached"] = True
                cached_data["source"] = "redis"
                return cached_data

            file_cached = self._get_file_cache(sheet_name, period_type)
            if file_cached:
                self._set_to_cache(sheet_name, period_type, file_cached)
                file_cached["cached"] = True
                file_cached["source"] = "file"
                return file_cached

        # Fetch from e-Stat
        try:
            data_points = self._fetch_from_estat(sheet_index, period_type)

            # Sort by date
            data_points.sort(key=lambda x: x["date"])

            result = {
                "data": data_points,
                "sheet_name": sheet_info["name"],
                "sheet_title": sheet_info["title"],
                "sheet_description": sheet_info["description"],
                "period_type": period_type,
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "e-stat"
            }

            # Cache the result
            self._set_to_cache(sheet_name, period_type, result)
            self._set_file_cache(sheet_name, period_type, result)

            return result

        except Exception as e:
            print(f"Error fetching BSI comprehensive data: {e}")
            import traceback
            traceback.print_exc()

            # Try file cache as fallback
            file_cached = self._get_file_cache(sheet_name, period_type)
            if file_cached:
                file_cached["cached"] = True
                file_cached["source"] = "file_fallback"
                file_cached["error"] = str(e)
                return file_cached

            return {
                "data": [],
                "sheet_name": sheet_info["name"],
                "sheet_title": sheet_info["title"],
                "period_type": period_type,
                "error": str(e),
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "error"
            }

    def get_chart_data(
        self,
        sheet_name: str,
        period_type: str,
        force_refresh: bool = False
    ) -> Dict[str, Any]:
        """Get BSI data formatted for chart display"""
        raw_data = self.get_comprehensive_data(sheet_name, period_type, force_refresh)

        if "error" in raw_data and not raw_data.get("data"):
            return raw_data

        data_points = raw_data.get("data", [])

        # Deduplicate by year and quarter
        seen_quarters = set()
        unique_data = []
        for point in data_points:
            key = (point.get("year"), point.get("quarter"))
            if key not in seen_quarters:
                seen_quarters.add(key)
                unique_data.append(point)

        # Format for chart
        chart_data = {
            "dates": [],
            "large_all_industries": [],
            "large_manufacturing": [],
            "large_non_manufacturing": [],
            "medium_all_industries": [],
            "small_all_industries": [],
        }

        for point in unique_data:
            chart_data["dates"].append(point["date"])
            chart_data["large_all_industries"].append(point.get("large_all_industries"))
            chart_data["large_manufacturing"].append(point.get("large_manufacturing"))
            chart_data["large_non_manufacturing"].append(point.get("large_non_manufacturing"))
            chart_data["medium_all_industries"].append(point.get("medium_all_industries"))
            chart_data["small_all_industries"].append(point.get("small_all_industries"))

        period_labels = {
            "actual": "実績（当期）",
            "forecast1": "見通し（翌期）",
            "forecast2": "見通し（翌々期）"
        }

        return {
            "chart_data": chart_data,
            "sheet_name": raw_data.get("sheet_name"),
            "sheet_title": raw_data.get("sheet_title"),
            "sheet_description": raw_data.get("sheet_description"),
            "period_type": period_type,
            "period_label": period_labels.get(period_type, period_type),
            "metadata": {
                "title": f"法人企業景気予測調査 {raw_data.get('sheet_title', '')} - {period_labels.get(period_type, period_type)}",
                "source": "e-Stat（財務省）",
                "series": [
                    {"key": "large_all_industries", "name": "大企業 全産業", "color": "#1890ff"},
                    {"key": "large_manufacturing", "name": "大企業 製造業", "color": "#52c41a"},
                    {"key": "large_non_manufacturing", "name": "大企業 非製造業", "color": "#fa8c16"},
                    {"key": "medium_all_industries", "name": "中堅企業 全産業", "color": "#722ed1"},
                    {"key": "small_all_industries", "name": "中小企業 全産業", "color": "#eb2f96"},
                ]
            },
            "last_updated": raw_data.get("last_updated"),
            "cached": raw_data.get("cached", False),
            "source": raw_data.get("source", "unknown")
        }

    def get_sheet_info(self) -> Dict[str, Any]:
        """Get available sheet information"""
        return {
            "sheets": [
                {
                    "name": info["name"],
                    "title": info["title"],
                    "description": info["description"]
                }
                for info in self.SHEET_INFO.values()
            ],
            "period_types": [
                {"name": "actual", "label": "実績（当期）"},
                {"name": "forecast1", "label": "見通し（翌期）"},
                {"name": "forecast2", "label": "見通し（翌々期）"}
            ]
        }

    def invalidate_cache(self, sheet_name: Optional[str] = None, period_type: Optional[str] = None) -> bool:
        """Invalidate cache"""
        success = True

        if sheet_name and period_type:
            # Invalidate specific cache
            try:
                redis_client.delete(self._get_cache_key(sheet_name, period_type))
            except Exception as e:
                print(f"Redis cache invalidation error: {e}")
                success = False

            cache_file = self.cache_dir / f"bsi_{sheet_name}_{period_type}.json"
            if cache_file.exists():
                try:
                    cache_file.unlink()
                except Exception as e:
                    print(f"File cache invalidation error: {e}")
                    success = False
        else:
            # Invalidate all caches
            for info in self.SHEET_INFO.values():
                for pt in ["actual", "forecast1", "forecast2"]:
                    try:
                        redis_client.delete(self._get_cache_key(info["name"], pt))
                    except Exception as e:
                        print(f"Redis cache invalidation error: {e}")
                        success = False

                    cache_file = self.cache_dir / f"bsi_{info['name']}_{pt}.json"
                    if cache_file.exists():
                        try:
                            cache_file.unlink()
                        except Exception as e:
                            print(f"File cache invalidation error: {e}")
                            success = False

        return success


# Singleton instance
bsi_comprehensive_service = BSIComprehensiveService()

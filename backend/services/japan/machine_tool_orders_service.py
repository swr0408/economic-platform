"""
工作機械受注サービス
日本工作機械工業会 (JMTBA) / 日工販からExcel/PDFを取得

指標:
- 工作機械受注 (Machine Tool Orders)
- 前年比 (YoY)

データソース:
- DB: 過去データ（2009年〜、PDFアーカイブからインポート）
- Excel: 日工販（確報データ、2021年〜）
- PDF: 日本工作機械工業会（速報データ）

発表スケジュール:
- 月次 (毎月10日頃)

キャッシュ方式: FMP発表日時ベース判定方式 + 直接取得
"""
import io
import json
import logging
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from zoneinfo import ZoneInfo
from pathlib import Path

import requests
from openpyxl import load_workbook
import PyPDF2

try:
    from backend.core.redis_client import redis_client
    from backend.core.database import SessionLocal
    from backend.services.japan.fmp_next_release_utils import (
        get_next_release_from_fmp,
        should_refresh_by_fmp_schedule,
    )
except ImportError:
    from core.redis_client import redis_client
    from core.database import SessionLocal
    from services.japan.fmp_next_release_utils import (
        get_next_release_from_fmp,
        should_refresh_by_fmp_schedule,
    )

from sqlalchemy import text

logger = logging.getLogger(__name__)

JST = ZoneInfo("Asia/Tokyo")

CACHE_DIR = Path(__file__).parent.parent.parent / "data" / "cache" / "japan" / "economy"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
DATA_CACHE_FILE = CACHE_DIR / "machine_tool_orders_cache.json"


class MachineToolOrdersService:
    """工作機械受注サービス - 直接取得方式"""

    DATA_CACHE_KEY = "japan:machine_tool_orders:data"
    ECONALPHA_ID = "japan_machine_tool_orders"

    # Excel URL: 日工販（確報データ）
    # パターン: /YYYY/MM/確報1工作機械受注統計YYYY.M.xlsx
    # ディレクトリ: current_month - 1, ファイル名: current_month - 2
    BASE_EXCEL_URL = "https://www.nikkohan.or.jp/wp-content/uploads/{dir_year}/{dir_month:02d}/%E7%A2%BA%E5%A0%B11%E5%B7%A5%E4%BD%9C%E6%A9%9F%E6%A2%B0%E5%8F%97%E6%B3%A8%E7%B5%B1%E8%A8%88{data_year}.{data_month}.xlsx"

    # PDF URL: 日本工作機械工業会（速報データ）
    # パターン: /YYYY/MM/sokuhouYYMM.pdf
    BASE_PDF_URL = "https://www.jmtba.or.jp/wjmtbap/wp-content/uploads/{year}/{month:02d}/sokuhou{year_short}{month:02d}.pdf"

    def __init__(self):
        pass

    def _get_current_urls(self) -> Dict[str, str]:
        """
        現在日時に基づいてExcelとPDFのURLを生成

        Returns:
            Dictionary with excel_url, pdf_url, excel_data_month, pdf_data_month
        """
        now = datetime.now(JST)

        # Excel URL: directory = current_month - 1, filename data = current_month - 2
        excel_dir_year = now.year
        excel_dir_month = now.month - 1
        excel_data_year = now.year
        excel_data_month = now.month - 2

        # 年跨ぎ調整（ディレクトリ）
        if excel_dir_month <= 0:
            excel_dir_month += 12
            excel_dir_year -= 1

        # 年跨ぎ調整（データ）
        if excel_data_month <= 0:
            excel_data_month += 12
            excel_data_year -= 1

        excel_url = self.BASE_EXCEL_URL.format(
            dir_year=excel_dir_year,
            dir_month=excel_dir_month,
            data_year=excel_data_year,
            data_month=excel_data_month
        )

        # PDF URL: current_year/current_month
        pdf_year = now.year
        pdf_month = now.month
        pdf_year_short = pdf_year % 100

        pdf_url = self.BASE_PDF_URL.format(
            year=pdf_year,
            month=pdf_month,
            year_short=pdf_year_short
        )

        return {
            'excel_url': excel_url,
            'pdf_url': pdf_url,
            'excel_data_month': f"{excel_data_year}-{excel_data_month:02d}",
            'pdf_data_month': f"{pdf_year}-{pdf_month:02d}"
        }

    def _fetch_excel_data(self, url: str) -> List[Dict[str, Any]]:
        """
        Excelファイルを取得してパース

        Args:
            url: ExcelファイルのURL

        Returns:
            List of data points
        """
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()

            excel_file = io.BytesIO(response.content)
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active

            data = []
            current_year = None

            # Row 5からデータ開始
            # B=Year(21年), C=Month(2021/1 or just "2"), D=Total Orders, E=YoY Index
            for row in sheet.iter_rows(min_row=5):
                try:
                    year_value = row[1].value  # Column B: 年表示
                    date_value = row[2].value  # Column C: 日付/月
                    total_value = row[3].value  # Column D: 受注総額（百万円）
                    yoy_value = row[4].value  # Column E: 前年比指数

                    if not date_value or total_value is None:
                        continue

                    # 年が変わった場合、Column Bから年を更新
                    if year_value:
                        year_str = str(year_value).strip()
                        year_match = re.search(r'(\d+)年', year_str)
                        if year_match:
                            year_short = int(year_match.group(1))
                            current_year = 2000 + year_short

                    # 日付パース
                    date_str = str(date_value).strip()

                    # フルフォーマット (e.g., "2021/1")
                    match = re.search(r'(\d{4})[/\-](\d{1,2})', date_str)
                    if match:
                        year = int(match.group(1))
                        month = int(match.group(2))
                        current_year = year
                        formatted_date = f"{year}-{month:02d}-01"
                    else:
                        # 月のみフォーマット (e.g., "2", "3", "10")
                        month_match = re.match(r'^(\d{1,2})$', date_str)
                        if month_match and current_year:
                            month = int(month_match.group(1))
                            formatted_date = f"{current_year}-{month:02d}-01"
                        else:
                            continue

                    # 値のパース
                    total = float(total_value) if total_value is not None else None

                    # YoYは指数形式（例: 109.7）なので成長率に変換
                    yoy = None
                    if yoy_value is not None:
                        try:
                            yoy_raw = float(yoy_value)
                            yoy = yoy_raw - 100  # 指数から成長率に変換
                        except (ValueError, TypeError):
                            pass

                    data.append({
                        'date': formatted_date,
                        'total_orders': total,
                        'value': yoy  # フロントエンド用にvalueとして設定
                    })

                except (IndexError, ValueError, TypeError):
                    continue

            logger.info(f"Fetched {len(data)} records from Excel")
            return data

        except requests.exceptions.RequestException as e:
            logger.error(f"Failed to fetch Excel file: {e}")
            return []
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            return []

    def _fetch_pdf_data(self, url: str, target_month: str) -> Optional[Dict[str, Any]]:
        """
        PDFファイルを取得して最新月のデータを抽出

        Args:
            url: PDFファイルのURL
            target_month: ターゲット月 (format: "YYYY-MM")

        Returns:
            Single data point or None
        """
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()

            pdf_file = io.BytesIO(response.content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            full_text = ""
            for page in pdf_reader.pages:
                full_text += page.extract_text()

            # YoY指数を探す (80-150の範囲の数値)
            all_numbers = re.findall(r'(\d+\.\d+|\d+,\d+)', full_text)

            yoy_growth = None
            indices_found = []

            for num_str in all_numbers:
                try:
                    num_str_clean = num_str.replace(',', '')
                    value = float(num_str_clean)

                    if 80 <= value <= 150:
                        indices_found.append(value)
                        # 2番目の指数がYoY
                        if len(indices_found) == 2:
                            yoy_growth = value - 100
                            break
                except (ValueError, IndexError):
                    continue

            if yoy_growth is None:
                logger.warning(f"Could not extract YoY growth from PDF")
                return None

            date = f"{target_month}-01"

            logger.info(f"Fetched PDF data for {target_month}: YoY={yoy_growth:.1f}%")
            return {
                'date': date,
                'total_orders': None,
                'value': yoy_growth
            }

        except requests.exceptions.RequestException as e:
            logger.warning(f"Failed to fetch PDF file: {e}")
            return None
        except Exception as e:
            logger.warning(f"Failed to parse PDF file: {e}")
            return None

    def _fetch_from_db(self) -> List[Dict[str, Any]]:
        """
        DBから過去データを取得（2009年〜2020年、PDFアーカイブからインポート済み）

        Returns:
            List of data points from DB
        """
        try:
            with SessionLocal() as session:
                query = text("""
                    SELECT date, total_orders, total_yoy as value
                    FROM japan_machine_tool_orders_history
                    ORDER BY date ASC
                """)
                result = session.execute(query).fetchall()

                data = []
                for row in result:
                    data.append({
                        'date': row[0].strftime('%Y-%m-%d') if row[0] else None,
                        'total_orders': float(row[1]) if row[1] else None,
                        'value': float(row[2]) if row[2] else None
                    })

                logger.info(f"Fetched {len(data)} records from DB")
                return data

        except Exception as e:
            logger.warning(f"Failed to fetch data from DB: {e}")
            return []

    def _fetch_from_source(self) -> List[Dict[str, Any]]:
        """
        日工販/JMTBAからデータを直接取得 + DBから過去データを結合

        Returns:
            Combined list of data points
        """
        urls = self._get_current_urls()

        # DBから過去データ取得（2009年〜）
        db_data = self._fetch_from_db()

        # Excelデータ取得（2021年〜）
        excel_data = self._fetch_excel_data(urls['excel_url'])

        # PDFデータ取得（最新月の速報）
        pdf_data = self._fetch_pdf_data(urls['pdf_url'], urls['pdf_data_month'])

        # データ結合（DB -> Excel -> PDF）
        # dateをキーにした辞書で重複を排除（後のデータが優先）
        data_dict = {}

        # DBデータを追加
        for item in db_data:
            if item['date']:
                data_dict[item['date']] = item

        # Excelデータを追加（DBデータを上書き）
        for item in excel_data:
            if item['date']:
                data_dict[item['date']] = item

        # PDFデータを追加（最新月のみ）
        if pdf_data and pdf_data.get('date'):
            existing = data_dict.get(pdf_data['date'])
            if existing:
                # 既存エントリを更新（valueのみ更新）
                if pdf_data['value'] is not None:
                    existing['value'] = pdf_data['value']
            else:
                # 新規エントリを追加
                data_dict[pdf_data['date']] = pdf_data

        # 日付でソート
        combined_data = sorted(data_dict.values(), key=lambda x: x['date'])

        return combined_data

    def get_machine_tool_orders_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """
        工作機械受注データを取得

        Returns:
            {
                "data": [...],
                "latest": {...},
                "next_release": str,
                "cached": bool,
                "source": str,
                "last_updated": str
            }
        """
        # Redisキャッシュチェック
        if not force_refresh:
            cached_data = redis_client.get(self.DATA_CACHE_KEY)
            if cached_data:
                last_updated_str = cached_data.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    logger.info("Returning cached machine tool orders data from Redis")
                    return {
                        "data": cached_data.get("data", []),
                        "latest": cached_data.get("latest"),
                        "next_release": cached_data.get("next_release"),
                        "cached": True,
                        "source": "redis",
                        "last_updated": last_updated_str
                    }

            # ファイルキャッシュチェック
            file_cache = self._load_file_cache()
            if file_cache:
                last_updated_str = file_cache.get("last_updated")
                if last_updated_str and not self._should_refresh(last_updated_str):
                    logger.info("Returning cached machine tool orders data from file")
                    redis_client.set(self.DATA_CACHE_KEY, file_cache, expire=0)
                    return {
                        "data": file_cache.get("data", []),
                        "latest": file_cache.get("latest"),
                        "next_release": file_cache.get("next_release"),
                        "cached": True,
                        "source": "file",
                        "last_updated": last_updated_str
                    }

        # 直接取得
        source_data = self._fetch_from_source()
        if source_data:
            next_release = get_next_release_from_fmp(self.ECONALPHA_ID)
            latest = source_data[-1] if source_data else None

            cache_payload = {
                "data": source_data,
                "latest": latest,
                "next_release": next_release,
                "last_updated": datetime.now(JST).isoformat()
            }
            redis_client.set(self.DATA_CACHE_KEY, cache_payload, expire=0)
            self._save_file_cache(cache_payload)

            return {
                "data": source_data,
                "latest": latest,
                "next_release": next_release,
                "cached": False,
                "source": "direct (Excel/PDF)",
                "last_updated": datetime.now(JST).isoformat()
            }

        # ファイルキャッシュフォールバック
        file_cache = self._load_file_cache()
        if file_cache:
            return {
                "data": file_cache.get("data", []),
                "latest": file_cache.get("latest"),
                "next_release": file_cache.get("next_release"),
                "cached": True,
                "source": "file (fallback)",
                "last_updated": file_cache.get("last_updated")
            }

        return {
            "data": [],
            "latest": None,
            "next_release": None,
            "cached": False,
            "source": "none",
            "last_updated": None,
            "error": "No data available"
        }

    def _should_refresh(self, last_updated_str: str) -> bool:
        """キャッシュを更新すべきかどうかを判定（FMP発表日時ベース）"""
        return should_refresh_by_fmp_schedule(self.ECONALPHA_ID, last_updated_str, max_age_hours=72)

    def _load_file_cache(self) -> Optional[Dict[str, Any]]:
        try:
            if not DATA_CACHE_FILE.exists():
                return None
            with open(DATA_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Failed to load file cache: {e}")
            return None

    def _save_file_cache(self, data: Dict[str, Any]) -> None:
        try:
            CACHE_DIR.mkdir(parents=True, exist_ok=True)
            with open(DATA_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Failed to save file cache: {e}")

    def invalidate_cache(self) -> bool:
        return redis_client.delete(self.DATA_CACHE_KEY)

    def get_cache_status(self) -> Dict[str, Any]:
        data_exists = redis_client.exists(self.DATA_CACHE_KEY)
        cached_data = redis_client.get(self.DATA_CACHE_KEY) if data_exists else None

        urls = self._get_current_urls()

        return {
            "indicator": "Machine Tool Orders (Japan)",
            "source": "Direct (Excel/PDF from JMTBA/日工販)",
            "cache_key": self.DATA_CACHE_KEY,
            "exists": data_exists,
            "last_updated": cached_data.get("last_updated") if cached_data else None,
            "data_count": len(cached_data.get("data", [])) if cached_data else 0,
            "latest": cached_data.get("latest") if cached_data else None,
            "next_release": get_next_release_from_fmp(self.ECONALPHA_ID),
            "file_cache_exists": DATA_CACHE_FILE.exists(),
            "excel_url": urls['excel_url'],
            "pdf_url": urls['pdf_url']
        }


# シングルトンインスタンス
machine_tool_orders_service = MachineToolOrdersService()

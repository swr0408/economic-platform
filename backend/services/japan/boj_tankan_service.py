"""
BOJ Tankan Service
日銀短観 大企業業況判断指数 (DI)
Data source: Bank of Japan
"""
import json
import requests
import io
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from backend.core.redis_client import redis_client
    from backend.services.japan.fmp_next_release_utils import get_next_release_from_fmp
except ImportError:
    from core.redis_client import redis_client
    from services.japan.fmp_next_release_utils import get_next_release_from_fmp


class BOJTankanService:
    """Service for fetching BOJ Tankan DI data"""

    CACHE_KEY = "japan:boj_tankan_di"
    CACHE_TTL = 60 * 60 * 24 * 7  # 7 days (quarterly data)

    # FMP event pattern for next release (matches indicator_event_mapping.econalpha_id)
    INDICATOR_ID = "jp_tankan"

    # BOJ Excel URL pattern - year-based path (e.g., /2026/data/ for 2026)
    BASE_URL_TEMPLATE = "https://www.boj.or.jp/statistics/tk/zenyo/{year}/data/"

    # Quarter to month mapping
    QUARTER_MONTHS = {1: "03", 2: "06", 3: "09", 4: "12"}

    # Sheet and row configuration for DI data
    SHEET_NAME = "A1(p2,3)"

    # Row numbers for data extraction
    # 日銀短観Excelの構造:
    # 「表」行: 前回調査の「先行き」予測値（参考値）
    # 「裏」行: 今回調査の「最近」実績値
    #
    # Row 18: 大企業製造業「表」（前回先行き）
    # Row 19: 大企業製造業「裏」（今回最近）← Current
    # Row 62: 大企業非製造業「表」（前回先行き）
    # Row 63: 大企業非製造業「裏」（今回最近）← Current
    #
    # 「先行き」は翌四半期列の「表」行から取得
    MANUFACTURING_CURRENT_ROW = 19      # 「裏」行 = 最近の実績
    MANUFACTURING_OUTLOOK_ROW = 18      # 「表」行 = 先行き（翌四半期列で取得）
    NON_MANUFACTURING_CURRENT_ROW = 63  # 「裏」行 = 最近の実績
    NON_MANUFACTURING_OUTLOOK_ROW = 62  # 「表」行 = 先行き（翌四半期列で取得）

    # Date header rows
    YEAR_ROW = 12
    MONTH_ROW = 15

    # Data columns (L=12 to S=19)
    DATA_START_COL = 12
    DATA_END_COL = 19

    def __init__(self):
        self.cache_dir = Path(__file__).parent.parent.parent / "data" / "cache" / "boj_tankan"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """Get next release date from FMP"""
        try:
            return get_next_release_from_fmp(self.INDICATOR_ID)
        except Exception as e:
            print(f"Error getting next release for {self.INDICATOR_ID}: {e}")
            return None

    def _get_from_cache(self) -> Optional[Dict[str, Any]]:
        """Get data from Redis cache"""
        try:
            cached = redis_client.get(self.CACHE_KEY)
            if cached:
                return cached
        except Exception as e:
            print(f"Redis get error: {e}")
        return None

    def _set_to_cache(self, data: Dict[str, Any]) -> bool:
        """Set data to Redis cache"""
        try:
            redis_client.set(self.CACHE_KEY, data, expire=self.CACHE_TTL)
            return True
        except Exception as e:
            print(f"Redis set error: {e}")
        return False

    def _get_file_cache(self) -> Optional[Dict[str, Any]]:
        """Get data from file cache"""
        cache_file = self.cache_dir / "boj_tankan_di.json"
        if cache_file.exists():
            try:
                mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
                if datetime.now() - mtime < timedelta(days=7):
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
            except Exception as e:
                print(f"File cache read error: {e}")
        return None

    def _set_file_cache(self, data: Dict[str, Any]) -> bool:
        """Set data to file cache"""
        cache_file = self.cache_dir / "boj_tankan_di.json"
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            return True
        except Exception as e:
            print(f"File cache write error: {e}")
        return False

    def _needs_release_refetch(self, cached: Dict[str, Any]) -> bool:
        """調査全容(zenyo)公表スケジュール基準で、保有データが最新四半期に未達なら
        再取得を促す（レート制限付き）。判定失敗時は False（従来通りキャッシュ返却）。"""
        try:
            from services.japan.boj_tankan_schedule import needs_release_refetch
            return needs_release_refetch(cached)
        except Exception:
            return False

    def _get_latest_excel_url(self) -> str:
        """Get URL for the latest Tankan Excel file

        BOJ URL pattern: /statistics/tk/zenyo/{full_year}/data/all{YY}{QQ}a.xlsx
        e.g. /statistics/tk/zenyo/2026/data/all2603a.xlsx for Q1 2026 (March)
        """
        now = datetime.now()
        full_year = now.year
        year = full_year % 100  # Get last 2 digits

        # Determine the most recent quarter based on release schedule
        # Release months: 4, 7, 10, 12 (for Q1, Q2, Q3, Q4 respectively)
        current_month = now.month

        if current_month >= 12:
            quarter = 12
        elif current_month >= 10:
            quarter = 9
        elif current_month >= 7:
            quarter = 6
        elif current_month >= 4:
            quarter = 3
        else:
            # If before April, use previous year's Q4
            full_year -= 1
            year = full_year % 100
            quarter = 12

        # Try current quarter first, then previous quarters
        for q_offset in range(8):
            try_full_year = full_year
            try_year = year
            try_quarter = quarter

            # Adjust quarter with offset
            for _ in range(q_offset):
                try_quarter -= 3
                if try_quarter <= 0:
                    try_quarter += 12
                    try_full_year -= 1
                    try_year = try_full_year % 100

            base_url = self.BASE_URL_TEMPLATE.format(year=try_full_year)
            url = f"{base_url}all{try_year:02d}{try_quarter:02d}a.xlsx"

            try:
                response = requests.head(url, timeout=5)
                if response.status_code == 200:
                    return url
            except:
                continue

        # Fallback
        return self.BASE_URL_TEMPLATE.format(year=2024) + "all2412a.xlsx"

    def _parse_date_from_header(self, ws, col: int) -> Optional[str]:
        """
        Parse date from header rows
        Row 12: Year (2024, 2025, etc.)
        Row 15: Month in English (Mar., Jun., Sept., Dec.*)
        """
        try:
            year_val = ws.cell(row=self.YEAR_ROW, column=col).value
            month_text = ws.cell(row=self.MONTH_ROW, column=col).value

            if not month_text:
                return None

            # Parse month (remove asterisk if present for forecast data)
            month_text = str(month_text).strip().replace('*', '')
            month_map = {
                'Mar.': 3,
                'Jun.': 6,
                'Sept.': 9,
                'Dec.': 12
            }
            month = month_map.get(month_text)
            if not month:
                return None

            # Parse year - Excel structure has year at Jun column of each 4-quarter group
            # Pattern: Mar(col), Jun(col+1, has year), Sept(col+2), Dec(col+3)
            # For Mar: look at Jun column (+1)
            # For Sept/Dec: look backward at Jun column
            if not year_val or not isinstance(year_val, (int, float)) or year_val < 2000:
                # Determine position within the quarter group based on month
                # Mar=3, Jun=6, Sept=9, Dec=12
                if month == 3:
                    # Mar: year is at next column (Jun)
                    candidate = ws.cell(row=self.YEAR_ROW, column=col + 1).value
                    if candidate and isinstance(candidate, (int, float)) and candidate > 2000:
                        year_val = candidate
                elif month in [9, 12]:
                    # Sept/Dec: year is at Jun column (search backward)
                    for search_col in range(col - 1, max(0, col - 4), -1):
                        candidate = ws.cell(row=self.YEAR_ROW, column=search_col).value
                        if candidate and isinstance(candidate, (int, float)) and candidate > 2000:
                            year_val = candidate
                            break

            if not year_val or not isinstance(year_val, (int, float)):
                return None

            year = int(year_val)

            # Use end of quarter for date
            day_map = {3: 31, 6: 30, 9: 30, 12: 31}
            day = day_map.get(month, 1)

            return f"{year}-{month:02d}-{day:02d}"
        except Exception as e:
            print(f"Error parsing date from header at column {col}: {e}")
            return None

    def _parse_excel_data(self, content: bytes) -> List[Dict[str, Any]]:
        """Parse Excel file and extract DI data"""
        if not openpyxl:
            raise ImportError("openpyxl is required for parsing Excel files")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Find the sheet
        ws = None
        if self.SHEET_NAME in wb.sheetnames:
            ws = wb[self.SHEET_NAME]
        else:
            for sheet_name in wb.sheetnames:
                if "A1" in sheet_name:
                    ws = wb[sheet_name]
                    break

        if ws is None:
            ws = wb.active

        data_points = []

        def to_float(val):
            if val is None or val == '-':
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        # Extract data from columns L (12) to S (19)
        # Skip forecast columns (marked with *) - they don't have current values
        for col in range(self.DATA_START_COL, self.DATA_END_COL + 1):
            month_text = ws.cell(row=self.MONTH_ROW, column=col).value
            if month_text and '*' in str(month_text):
                # This is a forecast column, skip it
                continue

            date_str = self._parse_date_from_header(ws, col)

            if date_str is None:
                continue

            # Extract "最近" (Current) values from 「裏」rows
            mfg_current_val = ws.cell(row=self.MANUFACTURING_CURRENT_ROW, column=col).value
            non_mfg_current_val = ws.cell(row=self.NON_MANUFACTURING_CURRENT_ROW, column=col).value

            # Extract "先行き" (Outlook) values
            # 日銀短観の構造:
            # - 「裏」行: 今回調査の「最近」実績値
            # - 「表」行: 前回調査時点での予測値（参考）
            #
            # 各調査の「先行き」は次の四半期を予測した値なので、
            # 常に次の列の「表」行から取得する
            # 例: Sept.調査の先行き = Dec.列の表行（Sept.調査で発表された12月予測）
            next_col = col + 1
            if next_col <= self.DATA_END_COL:
                mfg_outlook_val = ws.cell(row=self.MANUFACTURING_OUTLOOK_ROW, column=next_col).value
                non_mfg_outlook_val = ws.cell(row=self.NON_MANUFACTURING_OUTLOOK_ROW, column=next_col).value
            else:
                mfg_outlook_val = None
                non_mfg_outlook_val = None

            mfg_current = to_float(mfg_current_val)
            mfg_outlook = to_float(mfg_outlook_val)
            non_mfg_current = to_float(non_mfg_current_val)
            non_mfg_outlook = to_float(non_mfg_outlook_val)

            # Skip if all values are None
            if all(v is None for v in [mfg_current, mfg_outlook, non_mfg_current, non_mfg_outlook]):
                continue

            data_point = {
                "date": date_str,
                "large_manufacturing_current": mfg_current,
                "large_manufacturing_outlook": mfg_outlook,
                "large_non_manufacturing_current": non_mfg_current,
                "large_non_manufacturing_outlook": non_mfg_outlook
            }

            data_points.append(data_point)

        wb.close()
        return data_points

    ZENYO_BACKFILL_START_YEAR = 2019  # zenyo xlsx 遡及範囲。以前は手動CSV種で補完

    def _zenyo_url_for_quarter(self, full_year: int, quarter: int) -> str:
        """特定四半期の業況判断(A1含む 'a' ファイル) URL。過去は5年グループdir配下。"""
        group_start = 2001 + 5 * ((full_year - 2001) // 5)
        return (f"https://www.boj.or.jp/statistics/tk/zenyo/{group_start}/data/"
                f"all{full_year % 100:02d}{quarter:02d}a.xlsx")

    def _build_long_history(self) -> List[Dict[str, Any]]:
        """zenyo 過去ファイルを年次蓄積して業況判断DIの長期履歴を構築。"""
        now = datetime.now()
        cur_q = 12 if now.month >= 12 else 9 if now.month >= 10 else 6 if now.month >= 7 else 3 if now.month >= 4 else 12
        cur_year = now.year if now.month >= 4 else now.year - 1
        merged: Dict[str, Dict[str, Any]] = {}
        for year in range(self.ZENYO_BACKFILL_START_YEAR, cur_year + 1):
            quarters = [cur_q] if year == cur_year else [3]
            if year == cur_year and cur_q != 3:
                quarters.append(3)
            for q in quarters:
                url = self._zenyo_url_for_quarter(year, q)
                try:
                    resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
                    if resp.status_code != 200 or len(resp.content) < 10000:
                        continue
                    for p in self._parse_excel_data(resp.content):
                        if p.get("date"):
                            merged[p["date"]] = p  # 昇順処理で新ファイル後勝ち
                except Exception as e:
                    print(f"[tankan-di] long-history {year}Q{q} failed: {e}")
        self._seed_from_manual_csv(merged)
        return sorted(merged.values(), key=lambda x: x.get("date", ""))

    def _seed_from_manual_csv(self, merged: Dict[str, Dict[str, Any]]) -> None:
        """手動CSV種（BOJ時系列検索のネイティブCSV）で古い期間を補完する。
        `data/manual_update/japan/tankan/*.csv` を解析（zenyo優先・未充足のみ）。"""
        try:
            from services.japan.boj_tankan_manual_seed import seed_into
            seed_into("business_conditions", merged)
        except Exception as e:
            print(f"[tankan-di] manual CSV seed failed: {e}")

    def get_tankan_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BOJ Tankan DI data"""
        # Check cache if not forcing refresh
        # 調査全容(zenyo)が概要の翌日公表のため、保有データが公表済みであるべき
        # 四半期に未達ならキャッシュを返さず再取得（レート制限付き）。
        if not force_refresh:
            cached_data = self._get_from_cache()
            if cached_data and not self._needs_release_refetch(cached_data):
                cached_data["cached"] = True
                cached_data["source"] = "redis"
                cached_data["next_release"] = self._get_next_release()
                return cached_data

            file_cached = self._get_file_cache()
            if file_cached and not self._needs_release_refetch(file_cached):
                self._set_to_cache(file_cached)
                file_cached["cached"] = True
                file_cached["source"] = "file"
                file_cached["next_release"] = self._get_next_release()
                return file_cached

        # Fetch from BOJ（過去ファイルを年次蓄積して長期履歴を構築）
        try:
            print("Building long-term BOJ Tankan DI (zenyo backfill)")
            data_points = self._build_long_history()

            result = {
                "data": data_points,
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "boj",
                "next_release": self._get_next_release()
            }

            # Cache the result
            self._set_to_cache(result)
            self._set_file_cache(result)

            return result

        except Exception as e:
            print(f"Error fetching BOJ Tankan data: {e}")
            import traceback
            traceback.print_exc()

            # Try file cache as fallback
            file_cached = self._get_file_cache()
            if file_cached:
                file_cached["cached"] = True
                file_cached["source"] = "file_fallback"
                file_cached["error"] = str(e)
                file_cached["next_release"] = self._get_next_release()
                return file_cached

            return {
                "data": [],
                "error": str(e),
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "error",
                "next_release": None
            }

    def get_tankan_chart_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BOJ Tankan data formatted for chart display"""
        raw_data = self.get_tankan_data(force_refresh)

        if "error" in raw_data and not raw_data.get("data"):
            return raw_data

        data_points = raw_data.get("data", [])

        # Sort by date ascending for chart
        sorted_data = sorted(data_points, key=lambda x: x["date"])

        # Format for chart
        chart_data = []
        for point in sorted_data:
            chart_data.append({
                "date": point["date"],
                "large_manufacturing_current": point.get("large_manufacturing_current"),
                "large_manufacturing_outlook": point.get("large_manufacturing_outlook"),
                "large_non_manufacturing_current": point.get("large_non_manufacturing_current"),
                "large_non_manufacturing_outlook": point.get("large_non_manufacturing_outlook")
            })

        return {
            "data": chart_data,
            "metadata": {
                "title": "日銀短観 大企業業況判断DI",
                "unit": "%ポイント",
                "series": [
                    {"key": "large_manufacturing_current", "name": "大企業製造業（業況判断）", "color": "#1890ff"},
                    {"key": "large_manufacturing_outlook", "name": "大企業製造業（先行き）", "color": "#40a9ff"},
                    {"key": "large_non_manufacturing_current", "name": "大企業非製造業（業況判断）", "color": "#fa8c16"},
                    {"key": "large_non_manufacturing_outlook", "name": "大企業非製造業（先行き）", "color": "#ffc069"}
                ]
            },
            "last_updated": raw_data.get("last_updated"),
            "cached": raw_data.get("cached", False),
            "source": raw_data.get("source", "unknown"),
            "next_release": raw_data.get("next_release")
        }

    def invalidate_cache(self) -> bool:
        """Invalidate all caches"""
        success = True

        try:
            redis_client.delete(self.CACHE_KEY)
        except Exception as e:
            print(f"Redis cache invalidation error: {e}")
            success = False

        cache_file = self.cache_dir / "boj_tankan_di.json"
        if cache_file.exists():
            try:
                cache_file.unlink()
            except Exception as e:
                print(f"File cache invalidation error: {e}")
                success = False

        return success

    def get_cache_status(self) -> Dict[str, Any]:
        """Get cache status"""
        redis_cached = False
        file_cached = False

        try:
            redis_cached = redis_client.exists(self.CACHE_KEY)
        except:
            pass

        cache_file = self.cache_dir / "boj_tankan_di.json"
        file_cached = cache_file.exists()

        return {
            "redis_cached": redis_cached,
            "file_cached": file_cached,
            "cache_key": self.CACHE_KEY,
            "cache_ttl_days": self.CACHE_TTL // (60 * 60 * 24)
        }


# Singleton instance
boj_tankan_service = BOJTankanService()

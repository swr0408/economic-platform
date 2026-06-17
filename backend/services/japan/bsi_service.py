"""
法人企業景気予測調査サービス (Business Survey Index - BSI)
財務省・e-StatからBSIデータを取得

データソース: e-Stat (https://www.e-stat.go.jp/)
参考-１シート: 大企業・中堅企業・中小企業の景気判断BSI
"""

import json
import requests
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from pathlib import Path
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

from io import BytesIO

try:
    from backend.core.redis_client import redis_client
    from backend.services.japan.fmp_next_release_utils import get_next_release_from_fmp
    from backend.services.japan.bsi_estat_source import download_bsi_excel
except ImportError:
    from core.redis_client import redis_client
    from services.japan.fmp_next_release_utils import get_next_release_from_fmp
    from services.japan.bsi_estat_source import download_bsi_excel


class BSIService:
    """法人企業景気予測調査サービス"""

    CACHE_KEY = "japan:bsi"
    CACHE_TTL = 60 * 60 * 24 * 7  # 7 days (quarterly data)

    # FMP event pattern for next release
    INDICATOR_ID = "jp_bsi"

    # 読み取り対象シート（参考-１ 時系列表：景気判断 BSI 企業規模別）。
    # 現行ファイルは多シート構成で worksheets[0] は「目次」のため、シート名で選ぶ。
    TARGET_SHEET_NAMES = ("参考-１", "参考‐１", "参考-1")

    def __init__(self):
        self.cache_dir = Path(__file__).parent.parent.parent / "data" / "cache" / "bsi"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _get_next_release(self) -> Optional[Dict[str, Any]]:
        """Get next release date from FMP"""
        try:
            return get_next_release_from_fmp(self.INDICATOR_ID)
        except Exception as e:
            print(f"Error getting next release for {self.INDICATOR_ID}: {e}")
            return None

    def _get_from_cache(self) -> Optional[Dict[str, Any]]:
        """Get data from Redis cache"""
        try:
            cached = redis_client.get(self.CACHE_KEY)
            if cached:
                return cached
        except Exception as e:
            print(f"Redis get error: {e}")
        return None

    def _set_to_cache(self, data: Dict[str, Any]) -> bool:
        """Set data to Redis cache"""
        try:
            redis_client.set(self.CACHE_KEY, data, expire=self.CACHE_TTL)
            return True
        except Exception as e:
            print(f"Redis set error: {e}")
        return False

    def _get_file_cache(self) -> Optional[Dict[str, Any]]:
        """Get data from file cache"""
        cache_file = self.cache_dir / "bsi_data.json"
        if cache_file.exists():
            try:
                mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
                if datetime.now() - mtime < timedelta(days=7):
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
            except Exception as e:
                print(f"File cache read error: {e}")
        return None

    def _set_file_cache(self, data: Dict[str, Any]) -> bool:
        """Set data to file cache"""
        cache_file = self.cache_dir / "bsi_data.json"
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            return True
        except Exception as e:
            print(f"File cache write error: {e}")
        return False

    def _parse_quarter(self, quarter_str: str) -> str:
        """Parse quarter string to Q1-Q4 format"""
        quarter_map = {
            "1~3月": "Q1", "4~6月": "Q2", "7~9月": "Q3", "10~12月": "Q4",
            "1～3月": "Q1", "4～6月": "Q2", "7～9月": "Q3", "10～12月": "Q4",
        }
        return quarter_map.get(quarter_str, quarter_str)

    def _quarter_to_date(self, year: int, quarter: str) -> str:
        """Convert year and quarter to date string"""
        if not quarter.startswith("Q"):
            return f"{year}-03-01"

        quarter_num = quarter[1]
        month_map = {"1": "03", "2": "06", "3": "09", "4": "12"}
        month = month_map.get(quarter_num, "03")
        return f"{year}-{month}-01"

    def _fetch_from_estat(self) -> List[Dict[str, Any]]:
        """Fetch data from e-Stat"""
        if not openpyxl:
            raise ImportError("openpyxl is required for parsing Excel files")

        # statInfId を動的解決して Excel を取得（旧固定 URL は 404 になるため）
        content = download_bsi_excel()

        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

        # 現行ファイルは多シート構成。景気判断 BSI 時系列は「参考-１」シート。
        # worksheets[0]（目次）ではデータが取れないため、シート名で選択する。
        # 旧来の単一シートファイルにも対応できるよう worksheets[0] にフォールバック。
        sheet = None
        for name in self.TARGET_SHEET_NAMES:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.worksheets[0]
            print(
                f"BSI: target sheet {self.TARGET_SHEET_NAMES} not found; "
                f"falling back to first sheet '{sheet.title}' "
                f"(available: {wb.sheetnames})"
            )

        data = []
        current_year = None

        # Data starts from row 7
        for row_idx in range(7, sheet.max_row):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Skip footer or empty rows
            if row[1] is None or "※" in str(row[1]):
                break

            # Extract year if present
            if row[0] is not None:
                year_str = str(row[0])
                year_match = re.search(r'(\d{4})', year_str)
                if year_match:
                    current_year = int(year_match.group(1))

            # Extract quarter
            quarter_str = str(row[1])
            quarter = self._parse_quarter(quarter_str)

            if current_year is None:
                continue

            # Create date string
            date_str = self._quarter_to_date(current_year, quarter)

            # Column indices:
            # 大企業全産業-実績: 2, 判断: 3, 見通し: 4
            # 大企業製造業-実績: 5, 判断: 6, 見通し: 7
            # 大企業非製造業-実績: 8, 判断: 9, 見通し: 10
            # 中堅企業全産業-実績: 11, 判断: 12, 見通し: 13
            # 中小企業全産業-実績: 20, 判断: 21, 見通し: 22

            def safe_float(val):
                if val is None:
                    return None
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return None

            try:
                data_point = {
                    "date": date_str,
                    "year": current_year,
                    "quarter": quarter,
                    # 大企業 (Large enterprises)
                    "large_all_industries": safe_float(row[2]),
                    "large_manufacturing": safe_float(row[5]),
                    "large_non_manufacturing": safe_float(row[8]),
                    # 中堅企業 (Medium enterprises)
                    "medium_all_industries": safe_float(row[11]),
                    # 中小企業 (Small enterprises)
                    "small_all_industries": safe_float(row[20]),
                }
                data.append(data_point)
            except (ValueError, TypeError, IndexError) as e:
                print(f"Error parsing row {row_idx}: {e}")
                continue

        wb.close()
        return data

    def get_bsi_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BSI data"""
        # Check cache if not forcing refresh
        if not force_refresh:
            cached_data = self._get_from_cache()
            if cached_data:
                cached_data["cached"] = True
                cached_data["source"] = "redis"
                cached_data["next_release"] = self._get_next_release()
                return cached_data

            file_cached = self._get_file_cache()
            if file_cached:
                self._set_to_cache(file_cached)
                file_cached["cached"] = True
                file_cached["source"] = "file"
                file_cached["next_release"] = self._get_next_release()
                return file_cached

        # Fetch from e-Stat
        try:
            data_points = self._fetch_from_estat()

            # Sort by date
            data_points.sort(key=lambda x: x["date"])

            result = {
                "data": data_points,
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "e-stat",
                "description": "法人企業景気予測調査 BSI（企業規模別）",
                "next_release": self._get_next_release()
            }

            # Cache the result
            self._set_to_cache(result)
            self._set_file_cache(result)

            return result

        except Exception as e:
            print(f"Error fetching BSI data: {e}")
            import traceback
            traceback.print_exc()

            # Try file cache as fallback
            file_cached = self._get_file_cache()
            if file_cached:
                file_cached["cached"] = True
                file_cached["source"] = "file_fallback"
                file_cached["error"] = str(e)
                file_cached["next_release"] = self._get_next_release()
                return file_cached

            return {
                "data": [],
                "error": str(e),
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "error",
                "next_release": None
            }

    def get_chart_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BSI data formatted for chart display"""
        raw_data = self.get_bsi_data(force_refresh)

        if "error" in raw_data and not raw_data.get("data"):
            return raw_data

        data_points = raw_data.get("data", [])

        # Deduplicate by year and quarter
        seen_quarters = set()
        unique_data = []
        for point in data_points:
            key = (point.get("year"), point.get("quarter"))
            if key not in seen_quarters:
                seen_quarters.add(key)
                unique_data.append(point)

        # Format for chart
        chart_data = {
            "dates": [],
            "large_all_industries": [],
            "large_manufacturing": [],
            "large_non_manufacturing": [],
            "medium_all_industries": [],
            "small_all_industries": [],
        }

        for point in unique_data:
            chart_data["dates"].append(point["date"])
            chart_data["large_all_industries"].append(point.get("large_all_industries"))
            chart_data["large_manufacturing"].append(point.get("large_manufacturing"))
            chart_data["large_non_manufacturing"].append(point.get("large_non_manufacturing"))
            chart_data["medium_all_industries"].append(point.get("medium_all_industries"))
            chart_data["small_all_industries"].append(point.get("small_all_industries"))

        return {
            "chart_data": chart_data,
            "metadata": {
                "title": "法人企業景気予測調査 BSI（企業規模別）",
                "source": "e-Stat（財務省）",
                "series": [
                    {"key": "large_all_industries", "name": "大企業 全産業", "color": "#1890ff"},
                    {"key": "large_manufacturing", "name": "大企業 製造業", "color": "#52c41a"},
                    {"key": "large_non_manufacturing", "name": "大企業 非製造業", "color": "#fa8c16"},
                    {"key": "medium_all_industries", "name": "中堅企業 全産業", "color": "#722ed1"},
                    {"key": "small_all_industries", "name": "中小企業 全産業", "color": "#eb2f96"},
                ]
            },
            "last_updated": raw_data.get("last_updated"),
            "cached": raw_data.get("cached", False),
            "source": raw_data.get("source", "unknown"),
            "next_release": raw_data.get("next_release")
        }

    def get_table_data(self, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BSI data formatted for table display"""
        raw_data = self.get_bsi_data(force_refresh)

        if "error" in raw_data and not raw_data.get("data"):
            return raw_data

        data_points = raw_data.get("data", [])

        # Sort by date descending for table
        sorted_data = sorted(data_points, key=lambda x: x.get("date", ""), reverse=True)

        table_data = []
        for point in sorted_data:
            table_data.append({
                "date": point["date"],
                "quarter": f"{point.get('year', '')} {point.get('quarter', '')}",
                "large_all": point.get("large_all_industries"),
                "large_mfg": point.get("large_manufacturing"),
                "large_non_mfg": point.get("large_non_manufacturing"),
                "medium_all": point.get("medium_all_industries"),
                "small_all": point.get("small_all_industries"),
            })

        return {
            "table_data": table_data,
            "metadata": {
                "title": "法人企業景気予測調査 BSI",
                "source": "e-Stat（財務省）"
            },
            "last_updated": raw_data.get("last_updated"),
            "cached": raw_data.get("cached", False),
            "source": raw_data.get("source", "unknown"),
            "next_release": raw_data.get("next_release")
        }

    def invalidate_cache(self) -> bool:
        """Invalidate all caches"""
        success = True

        try:
            redis_client.delete(self.CACHE_KEY)
        except Exception as e:
            print(f"Redis cache invalidation error: {e}")
            success = False

        cache_file = self.cache_dir / "bsi_data.json"
        if cache_file.exists():
            try:
                cache_file.unlink()
            except Exception as e:
                print(f"File cache invalidation error: {e}")
                success = False

        return success


# Singleton instance
bsi_service = BSIService()

"""
BOJ Tankan Comprehensive Service (日銀短観 - 包括的データ)
Fetches comprehensive Tankan survey data from Bank of Japan:
1. Capital Investment (設備投資額) - from all{YY}{QQ}c.xlsx
2. Production Facilities (生産・営業用設備) - from all{YY}{QQ}a.xlsx
3. Employment (雇用人員) - from all{YY}{QQ}a.xlsx
4. Purchase Price (仕入価格) - from all{YY}{QQ}a.xlsx
5. Selling Price (販売価格) - from all{YY}{QQ}a.xlsx

Source: https://www.boj.or.jp/statistics/tk/zenyo/{year}/data/
Updated: 4th, 7th, 10th, 12th month of each quarter
"""

import json
import requests
import io
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from backend.core.redis_client import redis_client
except ImportError:
    from core.redis_client import redis_client


class BOJTankanComprehensiveService:
    """Service for fetching comprehensive BOJ Tankan survey data"""

    CACHE_KEY_PREFIX = "japan:boj_tankan_comprehensive"
    CACHE_TTL = 60 * 60 * 24 * 7  # 7 days (quarterly data)

    # BOJ Excel URL patterns - year-based path (e.g., /2026/data/ for 2026)
    BASE_URL_TEMPLATE = "https://www.boj.or.jp/statistics/tk/zenyo/{year}/data/"

    # Data configurations for each data type
    # Based on old environment configuration
    DATA_CONFIGS = {
        "capital_investment": {
            "name": "設備投資額",
            "url_type": "c",
            "sheet_name": "A31(p80,83)",
            "data_series": {
                "large_manufacturing": {"row": 19, "cols": "S,V,Y"},
                "large_manufacturing_revision": {"row": 19, "cols": "AE"},
                "large_non_manufacturing": {"row": 63, "cols": "S,V,Y"},
                "large_non_manufacturing_revision": {"row": 63, "cols": "AE"}
            },
            "unit": "前年度比%"
        },
        "production_facilities": {
            "name": "生産・営業用設備判断指数（DI）",
            "url_type": "a",
            "sheet_name": "A6(p12,13)",
            "data_series": {
                "all_industries_current": {"row": 17, "cols": "L:S"},
                "all_industries_forecast": {"row": 16, "cols": "L:S"},
                "large_manufacturing_current": {"row": 19, "cols": "L:S"},
                "large_manufacturing_forecast": {"row": 18, "cols": "L:S"}
            },
            "unit": "%ポイント"
        },
        "employment": {
            "name": "雇用人員判断指数（DI）",
            "url_type": "a",
            "sheet_name": "A7(p14,15)",
            "data_series": {
                "all_industries_current": {"row": 17, "cols": "L:S"},
                "all_industries_forecast": {"row": 16, "cols": "L:S"},
                "large_manufacturing_current": {"row": 19, "cols": "L:S"},
                "large_manufacturing_forecast": {"row": 18, "cols": "L:S"}
            },
            "unit": "%ポイント"
        },
        "purchase_price": {
            "name": "仕入価格判断指数（DI）",
            "url_type": "a",
            "sheet_name": "A13(p26,27)",
            "data_series": {
                "all_industries_current": {"row": 17, "cols": "L:S"},
                "all_industries_forecast": {"row": 16, "cols": "L:S"},
                "large_manufacturing_current": {"row": 19, "cols": "L:S"},
                "large_manufacturing_forecast": {"row": 18, "cols": "L:S"}
            },
            "unit": "%ポイント"
        },
        "selling_price": {
            "name": "販売価格判断指数（DI）",
            "url_type": "a",
            "sheet_name": "A12(p24,25)",
            "data_series": {
                "all_industries_current": {"row": 17, "cols": "L:S"},
                "all_industries_forecast": {"row": 16, "cols": "L:S"},
                "large_manufacturing_current": {"row": 19, "cols": "L:S"},
                "large_manufacturing_forecast": {"row": 18, "cols": "L:S"}
            },
            "unit": "%ポイント"
        }
    }

    # Date header rows (same as boj_tankan_service.py)
    YEAR_ROW = 12
    MONTH_ROW = 15

    # Data columns (L=12 to S=19)
    DATA_START_COL = 12
    DATA_END_COL = 19

    def __init__(self):
        self.cache_dir = Path(__file__).parent.parent.parent / "data" / "cache" / "boj_tankan"
        self.cache_dir.mkdir(parents=True, exist_ok=True)

    def _get_cache_key(self, data_type: str) -> str:
        """Get cache key for specific data type"""
        return f"{self.CACHE_KEY_PREFIX}:{data_type}"

    def _get_from_cache(self, data_type: str) -> Optional[Dict[str, Any]]:
        """Get data from Redis cache"""
        try:
            cached = redis_client.get(self._get_cache_key(data_type))
            if cached:
                return cached
        except Exception as e:
            print(f"Redis get error: {e}")
        return None

    def _set_to_cache(self, data_type: str, data: Dict[str, Any]) -> bool:
        """Set data to Redis cache"""
        try:
            redis_client.set(self._get_cache_key(data_type), data, expire=self.CACHE_TTL)
            return True
        except Exception as e:
            print(f"Redis set error: {e}")
        return False

    def _get_file_cache(self, data_type: str) -> Optional[Dict[str, Any]]:
        """Get data from file cache"""
        cache_file = self.cache_dir / f"boj_tankan_{data_type}.json"
        if cache_file.exists():
            try:
                mtime = datetime.fromtimestamp(cache_file.stat().st_mtime)
                if datetime.now() - mtime < timedelta(days=7):
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
            except Exception as e:
                print(f"File cache read error: {e}")
        return None

    def _set_file_cache(self, data_type: str, data: Dict[str, Any]) -> bool:
        """Set data to file cache"""
        cache_file = self.cache_dir / f"boj_tankan_{data_type}.json"
        try:
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            return True
        except Exception as e:
            print(f"File cache write error: {e}")
        return False

    def _get_excel_url(self, url_type: str) -> str:
        """Get URL for the latest Tankan Excel file

        BOJ URL pattern: /statistics/tk/zenyo/{full_year}/data/all{YY}{QQ}{type}.xlsx
        e.g. /statistics/tk/zenyo/2026/data/all2603c.xlsx for Q1 2026 capital_investment
        """
        now = datetime.now()
        full_year = now.year
        year = full_year % 100  # Get last 2 digits

        # Determine the most recent quarter based on release schedule
        current_month = now.month

        if current_month >= 12:
            quarter = 12
        elif current_month >= 10:
            quarter = 9
        elif current_month >= 7:
            quarter = 6
        elif current_month >= 4:
            quarter = 3
        else:
            # If before April, use previous year's Q4
            full_year -= 1
            year = full_year % 100
            quarter = 12

        # Try current quarter first, then previous quarters
        for q_offset in range(8):
            try_full_year = full_year
            try_year = year
            try_quarter = quarter

            # Adjust quarter with offset
            for _ in range(q_offset):
                try_quarter -= 3
                if try_quarter <= 0:
                    try_quarter += 12
                    try_full_year -= 1
                    try_year = try_full_year % 100

            base_url = self.BASE_URL_TEMPLATE.format(year=try_full_year)
            url = f"{base_url}all{try_year:02d}{try_quarter:02d}{url_type}.xlsx"

            try:
                response = requests.head(url, timeout=5)
                if response.status_code == 200:
                    return url
            except:
                continue

        # Fallback
        return self.BASE_URL_TEMPLATE.format(year=2024) + f"all2412{url_type}.xlsx"

    def _zenyo_url_for_quarter(self, full_year: int, quarter: int, url_type: str) -> str:
        """特定の調査年・四半期の zenyo Excel URL を構築する。

        過去ファイルは「5年グループ」ディレクトリ配下に置かれる
        (例: 2021-2025 → /zenyo/2021/、2026- → /zenyo/2026/)。
        """
        group_start = 2001 + 5 * ((full_year - 2001) // 5)
        yy = full_year % 100
        return (
            f"https://www.boj.or.jp/statistics/tk/zenyo/{group_start}/data/"
            f"all{yy:02d}{quarter:02d}{url_type}.xlsx"
        )

    def fetch_di_quarter(self, data_type: str, full_year: int, quarter: int) -> List[Dict[str, Any]]:
        """指定四半期の zenyo ファイルを取得し DI 系列を解析して返す（履歴バックフィル用）。

        取得不能・該当無しは空リストを返す。1ファイルに直近数四半期分が含まれるため、
        年次ストライドで呼び出して結果を日付マージすれば長期履歴を構築できる。
        """
        config = self.DATA_CONFIGS.get(data_type)
        if not config:
            return []
        url = self._zenyo_url_for_quarter(full_year, quarter, config["url_type"])
        try:
            resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            if resp.status_code != 200 or len(resp.content) < 10000:
                return []
            return self._parse_di_data(resp.content, config)
        except Exception as e:
            print(f"[tankan] fetch_di_quarter {data_type} {full_year}Q{quarter} failed: {e}")
            return []

    ZENYO_BACKFILL_START_YEAR = 2019  # zenyo xlsx が遡れる範囲。以前は手動CSV種で補完

    def _build_long_history(self, data_type: str, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """zenyo 過去ファイルを年次ストライドで取得し長期履歴を構築する。

        1ファイルに直近数四半期が含まれるため、年1本＋当年最新で全期間を被覆できる。
        """
        now = datetime.now()
        cur_q = 12 if now.month >= 12 else 9 if now.month >= 10 else 6 if now.month >= 7 else 3 if now.month >= 4 else 12
        cur_year = now.year if now.month >= 4 else now.year - 1
        is_capital = (data_type == "capital_investment")
        merged: Dict[str, Dict[str, Any]] = {}

        def parse(content: bytes) -> List[Dict[str, Any]]:
            return (self._parse_capital_investment_data(content, config) if is_capital
                    else self._parse_di_data(content, config))

        for year in range(self.ZENYO_BACKFILL_START_YEAR, cur_year + 1):
            quarters = [cur_q] if year == cur_year else [3]
            if year == cur_year and cur_q != 3:
                quarters.append(3)  # 当年Q1も取り当年内の被覆を確保
            for q in quarters:
                url = self._zenyo_url_for_quarter(year, q, config["url_type"])
                try:
                    resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
                    if resp.status_code != 200 or len(resp.content) < 10000:
                        continue
                    for p in parse(resp.content):
                        if p.get("date"):
                            # 昇順処理のため新しいファイルが後勝ち＝最新改定値が反映される
                            merged[p["date"]] = p
                except Exception as e:
                    print(f"[tankan] long-history {data_type} {year}Q{q} failed: {e}")
        self._seed_from_manual_csv(data_type, merged)
        return sorted(merged.values(), key=lambda x: x.get("date", ""))

    def _seed_from_manual_csv(self, data_type: str, merged: Dict[str, Dict[str, Any]]) -> None:
        """手動CSV種（BOJ時系列検索のネイティブCSV）で古い期間を補完する。
        `data/manual_update/japan/tankan/*.csv` を解析（zenyo優先・未充足のみ）。"""
        try:
            from services.japan.boj_tankan_manual_seed import seed_into
            seed_into(data_type, merged)
        except Exception as e:
            print(f"[tankan] manual CSV seed failed for {data_type}: {e}")

    def _col_to_index(self, col_letter: str) -> int:
        """Convert column letter(s) to index (A=1, B=2, ..., AA=27, etc.)"""
        result = 0
        for char in col_letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result

    def _parse_column_range(self, col_range: str) -> List[int]:
        """Parse column range string to list of column indices"""
        if ',' in col_range:
            # Comma-separated list (e.g., "S,V,Y")
            cols = col_range.split(',')
            return [self._col_to_index(col.strip()) for col in cols]
        elif ':' in col_range:
            # Range (e.g., "L:S")
            start_col, end_col = col_range.split(':')
            start_idx = self._col_to_index(start_col)
            end_idx = self._col_to_index(end_col)
            return list(range(start_idx, end_idx + 1))
        else:
            # Single column (e.g., "AE")
            return [self._col_to_index(col_range)]

    def _parse_date_from_header(self, ws, col: int, is_capital_investment: bool = False) -> Optional[str]:
        """Parse date from header rows"""
        try:
            if is_capital_investment:
                # Row 13: Year info (2023, 2024, 2025) for fiscal year
                year_val = ws.cell(row=13, column=col).value

                # Search backwards for year if merged cell
                if not year_val or not isinstance(year_val, (int, float)) or year_val < 2000:
                    for search_col in range(col, max(0, col - 10), -1):
                        candidate = ws.cell(row=13, column=search_col).value
                        if candidate and isinstance(candidate, (int, float)) and candidate >= 2000:
                            year_val = candidate
                            break

                if not year_val or not isinstance(year_val, (int, float)):
                    return None

                year = int(year_val)

                # Fiscal year ends in March of the following year
                actual_year = year + 1
                month = 3
                day = 31

                return f"{actual_year}-{month:02d}-{day:02d}"
            else:
                # For DI type data (same as boj_tankan_service.py)
                year_val = ws.cell(row=self.YEAR_ROW, column=col).value
                month_text = ws.cell(row=self.MONTH_ROW, column=col).value

                if not month_text:
                    return None

                # Parse month (remove asterisk if present for forecast data)
                month_text = str(month_text).strip().replace('*', '')
                month_map = {
                    'Mar.': 3,
                    'Jun.': 6,
                    'Sept.': 9,
                    'Dec.': 12
                }
                month = month_map.get(month_text)
                if not month:
                    return None

                # Parse year - Excel structure has year at Jun column of each 4-quarter group
                # Pattern: Mar(col), Jun(col+1, has year), Sept(col+2), Dec(col+3)
                if not year_val or not isinstance(year_val, (int, float)) or year_val < 2000:
                    if month == 3:
                        # Mar: year is at next column (Jun)
                        candidate = ws.cell(row=self.YEAR_ROW, column=col + 1).value
                        if candidate and isinstance(candidate, (int, float)) and candidate > 2000:
                            year_val = candidate
                    elif month in [9, 12]:
                        # Sept/Dec: year is at Jun column (search backward)
                        for search_col in range(col - 1, max(0, col - 4), -1):
                            candidate = ws.cell(row=self.YEAR_ROW, column=search_col).value
                            if candidate and isinstance(candidate, (int, float)) and candidate > 2000:
                                year_val = candidate
                                break

                if not year_val or not isinstance(year_val, (int, float)):
                    return None

                year = int(year_val)

                # Use end of quarter for date
                day_map = {3: 31, 6: 30, 9: 30, 12: 31}
                day = day_map.get(month, 1)

                return f"{year}-{month:02d}-{day:02d}"
        except Exception as e:
            print(f"Error parsing date from header at column {col}: {e}")
            return None

    def _parse_di_data(self, content: bytes, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Parse DI type data (production_facilities, employment, purchase_price, selling_price)

        日銀短観Excelの構造:
        - 「裏」行 (Row 17/19): 今回調査の「最近」実績値 (_current)
        - 「表」行 (Row 16/18): 前回調査の「先行き」予測値 (_forecast)

        先行き（forecast）の取得方法:
        - 次の列が予測列（*マーク）の場合: その列の「表」行から取得（最新調査の先行き）
        - それ以外: 同列の「表」行から取得（前回調査の先行き予測）
        """
        if not openpyxl:
            raise ImportError("openpyxl is required for parsing Excel files")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Find the sheet
        sheet_name = config["sheet_name"]
        ws = None
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Try to find a sheet containing part of the name
            for sn in wb.sheetnames:
                if sheet_name.split('(')[0] in sn:
                    ws = wb[sn]
                    break

        if ws is None:
            ws = wb.active

        data_series = config["data_series"]
        data_points = []

        def to_float(val):
            if val is None or val == '-':
                return None
            try:
                return float(val)
            except (ValueError, TypeError):
                return None

        # Extract data from columns L (12) to S (19)
        # Skip forecast columns (marked with *) - they don't have current values
        for col in range(self.DATA_START_COL, self.DATA_END_COL + 1):
            month_text = ws.cell(row=self.MONTH_ROW, column=col).value
            if month_text and '*' in str(month_text):
                # This is a forecast column, skip it
                continue

            date_str = self._parse_date_from_header(ws, col, is_capital_investment=False)

            if date_str is None:
                continue

            data_point = {"date": date_str}

            # 日銀短観の構造:
            # - 「裏」行: 今回調査の「最近」実績値
            # - 「表」行: 前回調査時点での予測値（参考）
            #
            # 各調査の「先行き」は次の四半期を予測した値なので、
            # 常に次の列の「表」行から取得する
            next_col = col + 1

            for series_name, series_config in data_series.items():
                row = series_config["row"]

                if "_forecast" in series_name and next_col <= self.DATA_END_COL:
                    # 先行きデータ → 次の列の「表」行から取得
                    val = ws.cell(row=row, column=next_col).value
                else:
                    # 実績データ → 同じ列から取得
                    val = ws.cell(row=row, column=col).value

                data_point[series_name] = to_float(val)

            # Only add if we have at least some data
            if any(v is not None for k, v in data_point.items() if k != "date"):
                data_points.append(data_point)

        wb.close()
        return data_points

    def _parse_capital_investment_data(self, content: bytes, config: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Parse capital investment data with fiscal year structure"""
        if not openpyxl:
            raise ImportError("openpyxl is required for parsing Excel files")

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Find the sheet
        sheet_name = config["sheet_name"]
        ws = None
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            for sn in wb.sheetnames:
                if sheet_name.split('(')[0] in sn:
                    ws = wb[sn]
                    break

        if ws is None:
            ws = wb.active

        data_series = config["data_series"]
        data_points = []

        # For capital investment, first get revision values
        revision_values = {}
        for series_name, series_config in data_series.items():
            if "_revision" in series_name:
                row = series_config["row"]
                col_idx = self._parse_column_range(series_config["cols"])[0]
                val = ws.cell(row=row, column=col_idx).value
                if val is not None and val != '-':
                    try:
                        revision_values[series_name] = float(val)
                    except (ValueError, TypeError):
                        revision_values[series_name] = None

        # Get main data columns (S,V,Y = columns 19,22,25)
        first_series = None
        for series_name, series_config in data_series.items():
            if "_revision" not in series_name:
                first_series = series_config
                break

        if first_series:
            col_indices = self._parse_column_range(first_series["cols"])

            for col in col_indices:
                date_str = self._parse_date_from_header(ws, col, is_capital_investment=True)

                if date_str is None:
                    continue

                data_point = {"date": date_str}

                def to_float(val):
                    if val is None or val == '-':
                        return None
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return None

                for series_name, series_config in data_series.items():
                    if "_revision" in series_name:
                        continue

                    row = series_config["row"]
                    val = ws.cell(row=row, column=col).value
                    data_point[series_name] = to_float(val)

                # Add revision values to all data points
                for rev_name, rev_value in revision_values.items():
                    data_point[rev_name] = rev_value

                if any(v is not None for k, v in data_point.items() if k != "date" and "_revision" not in k):
                    data_points.append(data_point)

        wb.close()
        return data_points

    def _needs_release_refetch(self, data_type: str, cached: Dict[str, Any]) -> bool:
        """調査全容(zenyo)公表スケジュール基準で、保有データが最新四半期に未達なら
        再取得を促す（レート制限付き）。capital_investment(年度)は対象外。
        判定失敗時は False（従来通りキャッシュ返却）。"""
        if data_type == "capital_investment":
            return False
        try:
            from services.japan.boj_tankan_schedule import needs_release_refetch
            return needs_release_refetch(cached)
        except Exception:
            return False

    def get_comprehensive_data(self, data_type: str, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BOJ Tankan comprehensive data for specific type"""
        if data_type not in self.DATA_CONFIGS:
            return {
                "data": [],
                "error": f"Unknown data type: {data_type}",
                "available_types": list(self.DATA_CONFIGS.keys())
            }

        config = self.DATA_CONFIGS[data_type]

        # Check cache
        # 調査全容(zenyo)が概要の翌日公表のため、保有データが公表済みであるべき
        # 四半期に未達ならキャッシュを返さず再取得（レート制限付き）。
        # capital_investment は年度(FY)ベースで四半期スケジュールに乗らないため対象外。
        if not force_refresh:
            cached_data = self._get_from_cache(data_type)
            if cached_data and not self._needs_release_refetch(data_type, cached_data):
                cached_data["cached"] = True
                cached_data["source"] = "redis"
                return cached_data

            file_cached = self._get_file_cache(data_type)
            if file_cached and not self._needs_release_refetch(data_type, file_cached):
                self._set_to_cache(data_type, file_cached)
                file_cached["cached"] = True
                file_cached["source"] = "file"
                return file_cached

        # Fetch from BOJ（過去ファイルを年次蓄積して長期履歴を構築）
        try:
            print(f"Building long-term BOJ Tankan {data_type} (zenyo backfill)")
            data_points = self._build_long_history(data_type, config)

            result = {
                "data": data_points,
                "data_type": data_type,
                "name": config["name"],
                "unit": config["unit"],
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "boj"
            }

            self._set_to_cache(data_type, result)
            self._set_file_cache(data_type, result)

            return result

        except Exception as e:
            print(f"Error fetching BOJ Tankan {data_type} data: {e}")
            import traceback
            traceback.print_exc()

            file_cached = self._get_file_cache(data_type)
            if file_cached:
                file_cached["cached"] = True
                file_cached["source"] = "file_fallback"
                file_cached["error"] = str(e)
                return file_cached

            return {
                "data": [],
                "data_type": data_type,
                "name": config["name"],
                "error": str(e),
                "last_updated": datetime.now().isoformat(),
                "cached": False,
                "source": "error"
            }

    def get_table_data(self, data_type: str, force_refresh: bool = False) -> Dict[str, Any]:
        """Get BOJ Tankan data formatted for table display"""
        raw_data = self.get_comprehensive_data(data_type, force_refresh)

        if "error" in raw_data and not raw_data.get("data"):
            return raw_data

        data_points = raw_data.get("data", [])

        # Sort by date descending for table
        sorted_data = sorted(data_points, key=lambda x: x.get("date", ""), reverse=True)

        # Build series metadata based on data type
        config = self.DATA_CONFIGS.get(data_type, {})
        series = []
        if data_type == "capital_investment":
            series = [
                {"key": "large_manufacturing", "name": "大企業製造業", "color": "#1890ff"},
                {"key": "large_non_manufacturing", "name": "大企業非製造業", "color": "#fa8c16"}
            ]
        else:
            series = [
                {"key": "all_industries_current", "name": "大企業全産業（実績）", "color": "#1890ff"},
                {"key": "all_industries_forecast", "name": "大企業全産業（予測）", "color": "#40a9ff"},
                {"key": "large_manufacturing_current", "name": "大企業製造業（実績）", "color": "#fa8c16"},
                {"key": "large_manufacturing_forecast", "name": "大企業製造業（予測）", "color": "#ffc069"}
            ]

        return {
            "table_data": sorted_data,
            "metadata": {
                "title": config.get("name", data_type),
                "source": "Bank of Japan",
                "source_url": raw_data.get("excel_url", ""),
                "last_updated": raw_data.get("last_updated"),
                "data_type": data_type,
                "series": series
            },
            "cached": raw_data.get("cached", False),
            "source": raw_data.get("source", "unknown")
        }

    def get_all_data_types(self) -> Dict[str, Any]:
        """Get list of available data types"""
        return {
            "data_types": [
                {
                    "key": key,
                    "name": config["name"],
                    "unit": config["unit"]
                }
                for key, config in self.DATA_CONFIGS.items()
            ]
        }

    def invalidate_cache(self, data_type: Optional[str] = None) -> bool:
        """Invalidate cache for specific type or all types"""
        success = True
        types_to_clear = [data_type] if data_type else list(self.DATA_CONFIGS.keys())

        for dt in types_to_clear:
            try:
                redis_client.delete(self._get_cache_key(dt))
            except Exception as e:
                print(f"Redis cache invalidation error for {dt}: {e}")
                success = False

            cache_file = self.cache_dir / f"boj_tankan_{dt}.json"
            if cache_file.exists():
                try:
                    cache_file.unlink()
                except Exception as e:
                    print(f"File cache invalidation error for {dt}: {e}")
                    success = False

        return success


# Singleton instance
boj_tankan_comprehensive_service = BOJTankanComprehensiveService()

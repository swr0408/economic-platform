"""
e-Stat 統計表ファイル(Excel)取得 汎用ヘルパー

METI 等が e-Stat 上で公開している統計表ファイルは、リリース毎に statInfId が
ローテートし、旧 ID はやがて 404 になる。また直接の公式サイト(例 www.meti.go.jp)が
サーバ IP をブロックして取得不能になる場合がある。

本ヘルパーは e-Stat Data Catalog API で現行 statInfId を動的解決し、Excel を
ダウンロードする。複数の指標サービス(IIP / 稼働率 / 予測 等)で共通利用する。
"""
from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import List, Optional, Sequence

import requests

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    from backend.services.estat_catalog_service import get_estat_catalog_service
except ImportError:
    from services.estat_catalog_service import get_estat_catalog_service


ESTAT_FILE_DOWNLOAD_URL = "https://www.e-stat.go.jp/stat-search/file-download"


def resolve_stat_inf_ids(
    stats_code: str,
    table_name_filter: str,
    fallback_stat_inf_ids: Sequence[str] = (),
    updated_days: int = 460,
) -> List[str]:
    """e-Stat Data Catalog API で現行 statInfId を新しい順に解決する。

    Args:
        stats_code: 政府統計コード(例 "00550300" 鉱工業生産・出荷・在庫指数)
        table_name_filter: 対象テーブル名に含まれる文字列(例 "品目別／月次／季節調整済指数")
        fallback_stat_inf_ids: API 失敗時に使う既知 ID(新しい順)
        updated_days: 公開日の絞り込み窓(日)。古い順返却で最新が埋もれるのを防ぐ
    """
    ids: List[str] = []
    try:
        catalog = get_estat_catalog_service()
        today = datetime.now()
        start = (today - timedelta(days=updated_days)).strftime("%Y%m%d")
        end = today.strftime("%Y%m%d")
        urls = catalog.get_excel_download_urls(
            survey_code=stats_code,
            limit=80,
            table_name_filter=table_name_filter,
            updated_date=f"{start}-{end}",
        )
        for url_info in urls:  # 公開日 降順
            sid = catalog.extract_stat_inf_id_from_url(url_info.get("url", ""))
            if sid and sid not in ids:
                ids.append(sid)
        if ids:
            print(f"e-Stat[{stats_code}/{table_name_filter}]: discovered {len(ids)} statInfId(s): {ids[:3]}")
        else:
            print(f"e-Stat[{stats_code}/{table_name_filter}]: none discovered, using fallback")
    except Exception as e:  # noqa: BLE001
        print(f"e-Stat[{stats_code}/{table_name_filter}]: discovery failed ({e}), using fallback")

    for fid in fallback_stat_inf_ids:
        if fid not in ids:
            ids.append(fid)
    return ids


def download_estat_excel(
    stats_code: str,
    table_name_filter: str,
    fallback_stat_inf_ids: Sequence[str] = (),
    updated_days: int = 460,
    timeout: int = 30,
) -> Optional[bytes]:
    """現行 statInfId を順に試し、最初に取得できた Excel の内容(bytes)を返す。

    取得不能なら None(呼び出し側はキャッシュフォールバックする)。
    """
    last_error: Optional[str] = None
    ids = resolve_stat_inf_ids(stats_code, table_name_filter, fallback_stat_inf_ids, updated_days)
    for sid in ids:
        try:
            resp = requests.get(
                ESTAT_FILE_DOWNLOAD_URL,
                params={"statInfId": sid, "fileKind": "0"},
                timeout=timeout,
            )
            if resp.status_code == 200 and len(resp.content) > 1000:
                print(f"e-Stat[{stats_code}]: downloaded via statInfId={sid} ({len(resp.content)} bytes)")
                return resp.content
            last_error = f"statInfId={sid} HTTP {resp.status_code}"
            print(f"e-Stat[{stats_code}]: {last_error}, trying next...")
        except Exception as e:  # noqa: BLE001
            last_error = f"statInfId={sid}: {e}"
            print(f"e-Stat[{stats_code}]: {last_error}, trying next...")
    print(f"e-Stat[{stats_code}]: all statInfIds failed. Last error: {last_error}")
    return None


def download_estat_excel_matching_title(
    stats_code: str,
    table_name_filter: str,
    title_substrings: Sequence[str],
    fallback_stat_inf_ids: Sequence[str] = (),
    updated_days: int = 460,
    timeout: int = 30,
) -> Optional[bytes]:
    """複数の statInfId が同一 table_name を共有する場合に、ファイル内タイトル
    (データシート A1)で目的のファイルを選別してダウンロードする。

    例: 第3次産業活動指数の「時系列データ」は 原指数/季調済/月次/四半期 が
    すべて同じ table_name で、区別はファイル先頭の項目名タイトルにしかない。

    Args:
        title_substrings: データシート A1 の項目名にすべて含まれるべき文字列
            (例 ["月次", "季調済"] / ["月次", "原指数"])
    """
    if openpyxl is None:
        print("e-Stat: openpyxl 未導入のためタイトル選別不可、先頭候補を返す")
        return download_estat_excel(stats_code, table_name_filter, fallback_stat_inf_ids, updated_days, timeout)

    ids = resolve_stat_inf_ids(stats_code, table_name_filter, fallback_stat_inf_ids, updated_days)
    last_error: Optional[str] = None
    for sid in ids:
        try:
            resp = requests.get(
                ESTAT_FILE_DOWNLOAD_URL,
                params={"statInfId": sid, "fileKind": "0"},
                timeout=timeout,
            )
            if resp.status_code != 200 or len(resp.content) <= 1000:
                last_error = f"statInfId={sid} HTTP {resp.status_code}"
                continue
            # データシート(「利用上の注意」以外)の A1 タイトルを確認
            wb = openpyxl.load_workbook(BytesIO(resp.content), data_only=True, read_only=True)
            try:
                title = ""
                for name in wb.sheetnames:
                    if name != "利用上の注意":
                        title = str(wb[name].cell(1, 1).value or "")
                        break
            finally:
                wb.close()
            if all(s in title for s in title_substrings):
                print(f"e-Stat[{stats_code}]: matched statInfId={sid} title='{title[:45]}'")
                return resp.content
            print(f"e-Stat[{stats_code}]: statInfId={sid} title mismatch ('{title[:30]}'), trying next...")
        except Exception as e:  # noqa: BLE001
            last_error = f"statInfId={sid}: {e}"
            print(f"e-Stat[{stats_code}]: {last_error}, trying next...")
    print(f"e-Stat[{stats_code}]: no file matched {list(title_substrings)}. Last error: {last_error}")
    return None
